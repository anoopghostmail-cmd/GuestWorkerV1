from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, Response, Request, Cookie, Body
from fastapi.responses import StreamingResponse, HTMLResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from datetime import datetime, timezone, timedelta
from jose import JWTError, jwt
from passlib.context import CryptContext
from pydantic import BaseModel, Field, ConfigDict, EmailStr, field_validator
from typing import List, Optional, Dict, Any
import os
import logging
import uuid
import io
import asyncio
import csv
import bleach
import re
from pathlib import Path
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from attendance_analytics import (
    analyze_day_of_week_trends,
    predict_worker_absences,
    detect_absence_patterns,
    calculate_worker_leaderboard,
    get_overall_attendance_stats
)

# Date utility functions for standardizing date formats
# NOTE: Date formats in this application:
# - Commission dates: DD-MM-YYYY (legacy format, kept for backward compatibility)
# - Attendance dates: YYYY-MM-DD (ISO format, recommended)
# - Datetime fields: ISO format with timezone
# For new features, use ISO format (YYYY-MM-DD) for consistency
try:
    from date_utils import parse_date_string, to_iso_date, date_to_comparable, is_date_in_range
except ImportError:
    # Fallback if date_utils not available
    def date_to_comparable(date_str):
        try:
            day, month, year = date_str.split('-')
            return int(year) * 10000 + int(month) * 100 + int(day)
        except:
            return 0

ROOT_DIR = Path(__file__).parent

# Load .env file only if it exists (for local development)
# On production platforms like Render, environment variables are set via dashboard
env_file = ROOT_DIR / '.env'
if env_file.exists():
    load_dotenv(env_file)
    print("Loaded environment variables from .env file")
else:
    print("No .env file found - using system environment variables")

# MongoDB connection - Production ready
# Supports both MONGODB_URI (standard) and MONGO_URL (legacy)
mongo_url = os.environ.get('MONGODB_URI') or os.environ.get('MONGO_URL') or 'mongodb://localhost:27017/guestworker'
db_name = os.environ.get('DB_NAME', 'guestworker')

print(f"Environment check - MONGODB_URI exists: {bool(os.environ.get('MONGODB_URI'))}")
print(f"Connecting to MongoDB: {mongo_url[:50]}...")

# Detect if using MongoDB Atlas (requires SSL) or local MongoDB (no SSL)
is_atlas = 'mongodb+srv://' in mongo_url or 'mongodb.net' in mongo_url
is_local = 'localhost' in mongo_url or '127.0.0.1' in mongo_url

if is_atlas:
    # MongoDB Atlas SSL/TLS configuration for production
    # Fix for SSL handshake errors on some hosting platforms
    import ssl
    import certifi
    
    print("Detected MongoDB Atlas - using SSL/TLS configuration")
    
    # Create SSL context that bypasses certificate verification
    ssl_context = ssl.create_default_context(cafile=certifi.where())
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE
    
    client = AsyncIOMotorClient(
        mongo_url,
        tls=True,
        tlsAllowInvalidCertificates=True,
        tlsAllowInvalidHostnames=True,
        serverSelectionTimeoutMS=10000,
        connectTimeoutMS=10000,
        socketTimeoutMS=10000
    )
else:
    # Local MongoDB - no SSL required
    print("Detected local MongoDB - no SSL configuration needed")
    client = AsyncIOMotorClient(
        mongo_url,
        serverSelectionTimeoutMS=5000
    )

db = client[db_name]
print("MongoDB client initialized successfully")

# Security - JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'guestworker-secret-key-change-in-production')
JWT_SECRET_KEY = SECRET_KEY  # Alias for consistency
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 30  # 30 days for monthly subscription

# Security Configuration for Cookies
# Production: COOKIE_SECURE=true, COOKIE_SAMESITE=none (for cross-domain)
# Development: COOKIE_SECURE=false, COOKIE_SAMESITE=lax
COOKIE_SECURE = os.environ.get('COOKIE_SECURE', 'false').lower() == 'true'
COOKIE_SAMESITE = os.environ.get('COOKIE_SAMESITE', 'lax')

print(f"Cookie settings - Secure: {COOKIE_SECURE}, SameSite: {COOKIE_SAMESITE}")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

MONTHLY_PLAN_PRICE = 799  # Rupees

app = FastAPI()

# ============ RATE LIMITING & BRUTE FORCE PROTECTION ============

# Initialize rate limiter
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Brute force protection configuration
MAX_LOGIN_ATTEMPTS = 5  # Max failed attempts before lockout
LOCKOUT_DURATION_MINUTES = 15  # Lockout duration
FAILED_ATTEMPT_WINDOW_MINUTES = 30  # Time window for counting failed attempts

# CORS Configuration - allow ANY origin with credentials (proxy may rewrite headers, but
# allow_origin_regex makes FastAPI reflect the requesting Origin back, which is compatible
# with allow_credentials=True for browsers when same/diff-origin).
ALLOWED_ORIGINS_STR = os.environ.get(
    'ALLOWED_ORIGINS',
    'http://localhost:3000,http://127.0.0.1:3000,http://localhost:3001'
)
ALLOWED_ORIGINS = [origin.strip() for origin in ALLOWED_ORIGINS_STR.split(',')]
print(f"CORS allowed origins (legacy list, ignored when regex active): {ALLOWED_ORIGINS}")

# Add production domains if not already in the list
PRODUCTION_ORIGINS = [
    "https://guestworker.app",
    "https://app.guestworker.app",
    "https://guestworker.in",
]

for origin in PRODUCTION_ORIGINS:
    if origin not in ALLOWED_ORIGINS:
        ALLOWED_ORIGINS.append(origin)

app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=".*",  # Allow any origin (reflected in response, compatible with credentials)
    allow_credentials=True,   # ✅ Important for cookies
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],     # Allow frontend to read all response headers
)

# ========================= Attendance Alerts Scheduler =========================
async def send_attendance_alerts_once():
    try:
        today_str = datetime.now(timezone.utc).astimezone().date().isoformat()
        users = await db.users.find({}, {"_id": 0, "id": 1}).to_list(1000)
        if not users:
            return
        user_ids = [u["id"] for u in users]
        for uid in user_ids:
            # Consider any attendance where status is not 'Absent'
            has_att = await db.attendance.count_documents({
                "contractor_id": uid,
                "date": today_str,
                "status": {"$ne": "Absent"}
            })
            if has_att == 0:
                now = datetime.now(timezone.utc)
                # Check last alert time on user doc to prevent re-send after deletion
                user_doc = await db.users.find_one({"id": uid}, {"_id": 0, "last_attendance_alert_at": 1})
                last_sent = user_doc.get("last_attendance_alert_at") if user_doc else None
                if isinstance(last_sent, str):
                    try:
                        last_sent = datetime.fromisoformat(last_sent)
                    except Exception:
                        last_sent = None
                should_send = not last_sent or (now - last_sent) >= timedelta(hours=12)
                if should_send:
                    await db.notifications.insert_one({
                        "id": str(uuid.uuid4()),
                        "user_id": uid,
                        "title": "Reminder: Mark Today's Attendance",
                        "message": "You have not marked worker attendance for today. Please update attendance to keep records accurate.",
                        "type": "alert",
                        "read": False,
                        "created_at": now
                    })
                    await db.users.update_one({"id": uid}, {"$set": {"last_attendance_alert_at": now}})
    except Exception as e:
        logging.error(f"Attendance alert job error: {e}")

async def attendance_alerts_scheduler():
    """Run attendance alerts every 12 hours with exception handling and graceful shutdown."""
    while True:
        try:
            await send_attendance_alerts_once()
        except asyncio.CancelledError:
            # Graceful shutdown - task was cancelled
            logging.info("Attendance alerts scheduler shutting down gracefully")
            break
        except Exception as e:
            # Log error but continue running
            logging.error(f"Error in attendance alerts scheduler: {e}", exc_info=True)
        
        try:
            await asyncio.sleep(12 * 60 * 60)  # 12 hours
        except asyncio.CancelledError:
            logging.info("Attendance alerts scheduler cancelled during sleep")
            break

@app.on_event("startup")
async def _start_schedulers():
    try:
        asyncio.create_task(attendance_alerts_scheduler())
    except Exception as e:
        logging.error(f"Failed to start scheduler: {e}")

api_router = APIRouter(prefix="/api")

# ============ SECURITY HELPERS ============

def sanitize_input(text: str) -> str:
    """
    Sanitize user input to prevent XSS, script injection, and other attacks.
    Removes ALL HTML tags, JavaScript, and potentially malicious content.
    """
    if not text:
        return text
    
    # Remove all HTML tags and strip whitespace
    # bleach.clean with tags=[] removes all HTML
    # strip=True removes the tags completely rather than escaping them
    sanitized = bleach.clean(
        text, 
        tags=[],  # No tags allowed
        attributes={},  # No attributes allowed
        strip=True,  # Strip tags instead of escaping
        strip_comments=True  # Remove HTML comments
    )
    
    # Additional protection: remove any null bytes
    sanitized = sanitized.replace('\x00', '')
    
    # Normalize whitespace but preserve newlines
    sanitized = sanitized.strip()
    
    return sanitized

def validate_uuid(uuid_string: str) -> bool:
    """Validate UUID format to prevent injection attacks"""
    import re
    uuid_pattern = re.compile(
        r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
        re.IGNORECASE
    )
    return bool(uuid_pattern.match(uuid_string))

# ============ EMAIL & PASSWORD VALIDATION ============

# List of common disposable email domains (extend as needed)
DISPOSABLE_EMAIL_DOMAINS = {
    'tempmail.com', 'throwaway.email', 'guerrillamail.com', 'mailinator.com',
    '10minutemail.com', 'temp-mail.org', 'fakeinbox.com', 'trashmail.com',
    'yopmail.com', 'getnada.com', 'maildrop.cc', 'spamgourmet.com',
    'dispostable.com', 'mintemail.com', 'mytemp.email', 'sharklasers.com',
    'guerrillamail.info', 'guerrillamail.biz', 'guerrillamail.de', 'grr.la',
    'guerrillamailblock.com', 'pokemail.net', 'spam4.me', 'mailnesia.com',
    '33mail.com', 'getairmail.com', 'emailondeck.com', 'tempinbox.com'
}

# Common weak passwords (add more as needed)
COMMON_WEAK_PASSWORDS = {
    'password', 'password123', '12345678', '123456789', 'qwerty', 'abc123',
    'password1', 'admin', 'letmein', 'welcome', 'monkey', '1234567890',
    'password!', 'Password1', 'guest', 'user', 'test', 'demo', '00000000'
}

def validate_email_format(email: str):
    """
    Validate email format and check for disposable email providers.
    Returns (is_valid, error_message)
    """
    if not email or not isinstance(email, str):
        return False, "Email is required"
    
    # Basic format validation
    email = email.lower().strip()
    email_regex = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
    
    if not email_regex.match(email):
        return False, "Invalid email format"
    
    # Check for consecutive dots
    if '..' in email:
        return False, "Invalid email format"
    
    # Extract domain
    try:
        domain = email.split('@')[1]
    except IndexError:
        return False, "Invalid email format"
    
    # Check for disposable email
    if domain in DISPOSABLE_EMAIL_DOMAINS:
        return False, "Temporary/disposable email addresses are not allowed"
    
    # Check for common typos in popular domains
    common_domains = {
        'gmail.com': ['gmial.com', 'gmai.com', 'gmal.com'],
        'yahoo.com': ['yaho.com', 'yahooo.com'],
        'outlook.com': ['outlok.com', 'outloo.com'],
        'hotmail.com': ['hotmial.com', 'hotmal.com']
    }
    
    for correct_domain, typos in common_domains.items():
        if domain in typos:
            return False, f"Did you mean @{correct_domain}?"
    
    return True, ""

def validate_password_strength(password: str, email: str = ""):
    """
    Validate password strength with comprehensive checks.
    Returns (is_valid, error_message)
    
    Requirements:
    - Minimum 6 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one number
    """
    if not password:
        return False, "Password is required"
    
    # Length check - Minimum 6 characters (matches frontend)
    if len(password) < 6:
        return False, "Password must be at least 6 characters long"
    
    if len(password) > 128:
        return False, "Password is too long (max 128 characters)"
    
    # Check for common weak passwords
    if password.lower() in COMMON_WEAK_PASSWORDS:
        return False, "This password is too common. Please choose a stronger password"
    
    # Check if password contains part of email
    if email and len(email) > 3:
        email_parts = email.lower().split('@')[0].split('.')
        for part in email_parts:
            if len(part) >= 4 and part in password.lower():
                return False, "Password should not contain parts of your email address"
    
    # Complexity checks - All are required (matches frontend)
    has_upper = any(c.isupper() for c in password)
    has_lower = any(c.islower() for c in password)
    has_digit = any(c.isdigit() for c in password)
    
    # Build specific error messages for missing requirements
    missing_requirements = []
    if not has_upper:
        missing_requirements.append("uppercase letter")
    if not has_lower:
        missing_requirements.append("lowercase letter")
    if not has_digit:
        missing_requirements.append("number")
    
    if missing_requirements:
        if len(missing_requirements) == 1:
            return False, f"Password must include at least one {missing_requirements[0]}"
        elif len(missing_requirements) == 2:
            return False, f"Password must include {missing_requirements[0]} and {missing_requirements[1]}"
        else:
            return False, "Password must include uppercase letter, lowercase letter, and number"
    
    return True, ""

# ============ BRUTE FORCE PROTECTION ============

async def check_login_attempts(email: str, ip_address: str):
    """
    Check if the account or IP is currently locked due to too many failed attempts.
    Returns (is_allowed, error_message)
    """
    now = datetime.now(timezone.utc)
    
    # Check failed attempts collection (create if doesn't exist)
    failed_attempts_collection = db.failed_login_attempts
    
    # Check account-level lockout
    account_attempts = await failed_attempts_collection.find_one({
        "email": email,
        "lockout_until": {"$gt": now}
    })
    
    if account_attempts:
        lockout_until = account_attempts['lockout_until']
        minutes_remaining = int((lockout_until - now).total_seconds() / 60)
        return False, f"Account temporarily locked. Try again in {minutes_remaining} minutes"
    
    # Check IP-level lockout
    ip_attempts = await failed_attempts_collection.find_one({
        "ip_address": ip_address,
        "lockout_until": {"$gt": now}
    })
    
    if ip_attempts:
        lockout_until = ip_attempts['lockout_until']
        minutes_remaining = int((lockout_until - now).total_seconds() / 60)
        return False, f"Too many failed attempts from this IP. Try again in {minutes_remaining} minutes"
    
    return True, ""

async def record_failed_login(email: str, ip_address: str):
    """Record a failed login attempt and apply lockout if necessary"""
    now = datetime.now(timezone.utc)
    window_start = now - timedelta(minutes=FAILED_ATTEMPT_WINDOW_MINUTES)
    
    failed_attempts_collection = db.failed_login_attempts
    
    # Count recent failed attempts for this email
    email_attempts = await failed_attempts_collection.count_documents({
        "email": email,
        "timestamp": {"$gte": window_start},
        "lockout_until": {"$exists": False}
    })
    
    # Count recent failed attempts from this IP
    ip_attempts = await failed_attempts_collection.count_documents({
        "ip_address": ip_address,
        "timestamp": {"$gte": window_start},
        "lockout_until": {"$exists": False}
    })
    
    # Record this failed attempt
    attempt_record = {
        "id": str(uuid.uuid4()),
        "email": email,
        "ip_address": ip_address,
        "timestamp": now
    }
    
    # Apply lockout if max attempts exceeded
    if email_attempts + 1 >= MAX_LOGIN_ATTEMPTS:
        attempt_record["lockout_until"] = now + timedelta(minutes=LOCKOUT_DURATION_MINUTES)
        attempt_record["lockout_reason"] = "max_account_attempts"
    
    if ip_attempts + 1 >= MAX_LOGIN_ATTEMPTS * 2:  # More lenient for IP
        attempt_record["lockout_until"] = now + timedelta(minutes=LOCKOUT_DURATION_MINUTES)
        attempt_record["lockout_reason"] = "max_ip_attempts"
    
    await failed_attempts_collection.insert_one(attempt_record)

async def clear_failed_login_attempts(email: str, ip_address: str):
    """Clear failed login attempts after successful login"""
    failed_attempts_collection = db.failed_login_attempts
    
    # Remove all failed attempts for this email and IP
    await failed_attempts_collection.delete_many({
        "$or": [
            {"email": email},
            {"ip_address": ip_address}
        ]
    })

# ============ SECURITY LOGGING & ANOMALY DETECTION ============

class SecurityEventType:
    """
    Security event types for logging and monitoring
    
    Note: Only logs suspicious activities and failures to minimize database storage.
    Successful user logins are NOT logged. Admin logins are always logged for audit trail.
    """
    # Authentication Events (Failures & Suspicious)
    FAILED_LOGIN = "failed_login"
    ACCOUNT_LOCKED = "account_locked"
    LOCKOUT_EXPIRED = "lockout_expired"
    
    # Rate Limiting Events
    RATE_LIMIT_EXCEEDED = "rate_limit_exceeded"
    
    # Registration Events (Suspicious)
    INVALID_EMAIL = "invalid_email"
    DISPOSABLE_EMAIL = "disposable_email"
    WEAK_PASSWORD = "weak_password"
    REGISTRATION_BLOCKED = "registration_blocked"
    
    # Suspicious Activity
    TIMING_ATTACK_ATTEMPT = "timing_attack_attempt"
    MULTIPLE_ACCOUNTS_SAME_IP = "multiple_accounts_same_ip"
    RAPID_REQUESTS = "rapid_requests"
    INVALID_TOKEN = "invalid_token"
    TOKEN_TAMPERING = "token_tampering"
    
    # Admin Events
    ADMIN_LOGIN_FAILED = "admin_login_failed"
    ADMIN_LOGIN_SUCCESS = "admin_login_success"
    UNAUTHORIZED_ADMIN_ACCESS = "unauthorized_admin_access"
    
    # Subscription & Limit Violations
    PLAN_LIMIT_VIOLATION = "plan_limit_violation"
    SUBSCRIPTION_TAMPER_ATTEMPT = "subscription_tamper_attempt"
    PLAN_CHANGE = "plan_change"

class SecuritySeverity:
    """Severity levels for security events"""
    LOW = "low"
    MEDIUM = "medium"
    HIGH = "high"
    CRITICAL = "critical"

async def log_security_event(
    event_type: str,
    severity: str,
    ip_address: str,
    user_agent: str = None,
    email: str = None,
    user_id: str = None,
    details: Dict[str, Any] = None,
    endpoint: str = None
):
    """
    Log security events and anomalies to MongoDB for admin monitoring.
    
    Args:
        event_type: Type of security event (from SecurityEventType)
        severity: Severity level (from SecuritySeverity)
        ip_address: Client IP address
        user_agent: Client User-Agent header
        email: Associated email (if applicable)
        user_id: Associated user ID (if applicable)
        details: Additional event details
        endpoint: API endpoint that triggered the event
    """
    security_logs_collection = db.security_logs
    
    event = {
        "id": str(uuid.uuid4()),
        "event_type": event_type,
        "severity": severity,
        "ip_address": ip_address,
        "user_agent": user_agent,
        "email": email,
        "user_id": user_id,
        "details": details or {},
        "endpoint": endpoint,
        "timestamp": datetime.now(timezone.utc),
        "resolved": False,  # Admin can mark as resolved
        "notes": ""  # Admin can add notes
    }
    
    await security_logs_collection.insert_one(event)
    
    # Log critical events to console for immediate attention
    if severity == SecuritySeverity.CRITICAL:
        logging.warning(f"🚨 CRITICAL SECURITY EVENT: {event_type} from {ip_address} ({email})")

async def detect_anomalies(ip_address: str, email: str = None) -> Dict[str, Any]:
    """
    Detect suspicious patterns and anomalies.
    Returns dict with anomaly flags and details.
    """
    anomalies = {
        "is_suspicious": False,
        "reasons": [],
        "threat_score": 0  # 0-100 scale
    }
    
    now = datetime.now(timezone.utc)
    last_hour = now - timedelta(hours=1)
    last_24h = now - timedelta(hours=24)
    
    security_logs = db.security_logs
    
    # Check 1: Multiple failed logins from same IP in last hour
    failed_logins_ip = await security_logs.count_documents({
        "ip_address": ip_address,
        "event_type": {"$in": [SecurityEventType.FAILED_LOGIN, SecurityEventType.ADMIN_LOGIN_FAILED]},
        "timestamp": {"$gte": last_hour}
    })
    
    if failed_logins_ip >= 10:
        anomalies["is_suspicious"] = True
        anomalies["reasons"].append(f"Multiple failed logins ({failed_logins_ip}) from this IP in last hour")
        anomalies["threat_score"] += 30
    
    # Check 2: Rate limit violations in last 24 hours
    rate_limit_violations = await security_logs.count_documents({
        "ip_address": ip_address,
        "event_type": SecurityEventType.RATE_LIMIT_EXCEEDED,
        "timestamp": {"$gte": last_24h}
    })
    
    if rate_limit_violations >= 5:
        anomalies["is_suspicious"] = True
        anomalies["reasons"].append(f"Multiple rate limit violations ({rate_limit_violations}) in last 24 hours")
        anomalies["threat_score"] += 20
    
    # Check 3: Multiple accounts from same IP (if email provided)
    if email:
        accounts_from_ip = await security_logs.count_documents({
            "ip_address": ip_address,
            "event_type": SecurityEventType.SUCCESSFUL_LOGIN,
            "email": {"$ne": email},
            "timestamp": {"$gte": last_24h}
        })
        
        if accounts_from_ip >= 5:
            anomalies["is_suspicious"] = True
            anomalies["reasons"].append(f"Multiple different accounts ({accounts_from_ip}) logged in from this IP")
            anomalies["threat_score"] += 25
    
    # Check 4: Disposable email attempts
    disposable_attempts = await security_logs.count_documents({
        "ip_address": ip_address,
        "event_type": SecurityEventType.DISPOSABLE_EMAIL,
        "timestamp": {"$gte": last_24h}
    })
    
    if disposable_attempts >= 3:
        anomalies["is_suspicious"] = True
        anomalies["reasons"].append(f"Multiple disposable email attempts ({disposable_attempts})")
        anomalies["threat_score"] += 15
    
    # Check 5: Token tampering attempts
    token_tampering = await security_logs.count_documents({
        "ip_address": ip_address,
        "event_type": {"$in": [SecurityEventType.TOKEN_TAMPERING, SecurityEventType.INVALID_TOKEN]},
        "timestamp": {"$gte": last_24h}
    })
    
    if token_tampering >= 3:
        anomalies["is_suspicious"] = True
        anomalies["reasons"].append(f"Token tampering attempts detected ({token_tampering})")
        anomalies["threat_score"] += 40
    
    # Cap threat score at 100
    anomalies["threat_score"] = min(anomalies["threat_score"], 100)
    
    return anomalies

# ==============================================
# Helper function to find already assigned workers
# ==============================================
async def get_assigned_worker_ids(contractor_id: str, date: str) -> list[str]:
    """
    Returns only workers who are assigned to an employer (not blanks)
    and are not marked absent.
    """
    assigned = await db.attendance.distinct("worker_id", {
        "contractor_id": contractor_id,
        "date": date,
        "mode": "worker",
        "employer_id": {"$ne": ""},  # ✅ only those linked to employer
        "status": {"$ne": "Absent"}  # ✅ exclude absentees
    })
    return [wid for wid in assigned if wid]


# ============ MODELS ============


class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    phone: Optional[str] = None
    role: str = "contractor"
    is_active: bool = True
    subscription_plan: str = "none"
    subscription_status: str = "inactive"
    plan_start_date: Optional[datetime] = None
    plan_end_date: Optional[datetime] = None
    trial_activated_at: Optional[str] = None  # ISO datetime string when trial was activated (never removed)
    payment_method: Optional[str] = None  # 'trial', 'razorpay', 'razorpay_subscription', 'activation_key', etc.
    razorpay_subscription_id: Optional[str] = None  # Razorpay subscription ID for auto-renewal
    auto_renew: bool = False  # Whether subscription auto-renews
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str = Field(..., min_length=6, max_length=128)
    name: str = Field(..., max_length=200)
    phone: Optional[str] = Field(None, max_length=20)

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class ActivatePlan(BaseModel):
    activation_key: str

class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    title: str
    message: str
    type: str = "info"  # info | alert | promo
    link: Optional[str] = None  # URL to redirect when notification is clicked
    metadata: Optional[dict] = None  # Additional data (e.g., message_id for opening specific conversation)
    read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotificationCreate(BaseModel):
    title: str = Field(..., max_length=200)
    message: str = Field(..., max_length=2000)
    type: str = "info"  # info | alert | promo

class AdminBroadcastRequest(BaseModel):
    title: str = Field(..., max_length=200)
    message: str = Field(..., max_length=2000)
    type: str = "promo"
    filter: str = "all"  # all | free | trial | paid
    preset: Optional[str] = Field(None, max_length=100)  # optional preset identifier
    cta_type: Optional[str] = None  # pricing | subscription | help | attendance | custom
    cta_label: Optional[str] = Field(None, max_length=50)  # Custom button text
    cta_url: Optional[str] = Field(None, max_length=200)  # Custom URL for custom CTA type

# ========================= Promotions/Offers =========================
class PromotionCreate(BaseModel):
    name: str = Field(..., max_length=200)
    description: str = Field(..., max_length=500)
    discount_type: str  # percentage | fixed_amount | custom_price
    discount_value: float  # percentage (e.g., 20 for 20%), fixed amount, or custom price
    target_category: str  # all | free_users | trial_users | paid_users | cancelled_users | expired_trial | low_activity | custom_users
    custom_user_ids: Optional[List[str]] = []  # For custom_users category
    plan_targets: Optional[List[str]] = []  # Which plans this applies to (empty = all plans)
    valid_from: datetime
    valid_until: datetime
    max_uses: Optional[int] = None  # Maximum number of users who can use this
    send_notification: bool = True
    notification_title: Optional[str] = None
    notification_message: Optional[str] = None
    active: bool = True

class PromotionUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    valid_until: Optional[datetime] = None
    max_uses: Optional[int] = None
    active: Optional[bool] = None

# ========================= Site-Wide Offers (Festival/Seasonal) =========================
class SiteWideOfferCreate(BaseModel):
    """Site-wide offers visible to ALL visitors (registered and non-registered)"""
    name: str = Field(..., max_length=200)  # e.g., "Black Friday Sale", "Diwali Offer"
    description: str = Field(..., max_length=500)  # Offer description
    offer_reason: str = Field(..., max_length=100)  # e.g., "Black Friday", "Festival Season"
    discount_type: str  # percentage | fixed_amount
    discount_value: float  # percentage (e.g., 30 for 30%) or fixed amount
    plan_targets: List[str] = []  # Which plans this applies to (empty = all plans)
    valid_from: datetime
    valid_until: datetime
    badge_text: Optional[str] = Field(None, max_length=50)  # e.g., "30% OFF", "Limited Time"
    badge_color: Optional[str] = Field("red", max_length=20)  # red, orange, green, blue, purple
    show_countdown: bool = False  # Show countdown timer
    active: bool = True

class SiteWideOfferUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    valid_until: Optional[datetime] = None
    active: Optional[bool] = None
    badge_text: Optional[str] = None
    show_countdown: Optional[bool] = None

# ========================= Trial Settings & Activation =========================
class TrialSettings(BaseModel):
    duration_days: int = 14

class TrialActivateRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    plan_name: Optional[str] = None  # "Contractor Plus" or "Contractor Pro"

async def _get_current_user_dep(auth_token: Optional[str] = Cookie(None)):
    # Thin wrapper so this module can reference the auth dependency
    # before the full `get_current_user` implementation is defined below.
    return await get_current_user(auth_token)

@api_router.post("/subscription/activate-trial")
async def activate_trial(
    payload: TrialActivateRequest = Body(default_factory=TrialActivateRequest),
    current_user: User = Depends(_get_current_user_dep),
):
    # Get full user record from DB to check trial history
    user_dict = await db.users.find_one({"id": current_user.id})
    if not user_dict:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Reject if user has already activated a trial (check if trial_activated_at exists)
    if user_dict.get("trial_activated_at"):
        raise HTTPException(
            status_code=400, 
            detail="You have already used your free trial. Please choose a subscription plan to continue."
        )
    
    # Reject if already active subscription or already on trial
    if current_user.subscription_status in ["active", "trial"]:
        raise HTTPException(status_code=400, detail="Trial not available for current plan")
    
    # Resolve trial plan: honour the plan the user actually picked, fall back to Plus.
    # Enterprise plan does not offer a free trial.
    TRIAL_ELIGIBLE_PLANS = {"Contractor Plus", "Contractor Pro"}
    requested_plan = (payload.plan_name or "").strip()
    if requested_plan and requested_plan not in TRIAL_ELIGIBLE_PLANS:
        raise HTTPException(
            status_code=400,
            detail=f"Free trial is not available for the {requested_plan} plan."
        )
    trial_plan = requested_plan if requested_plan in TRIAL_ELIGIBLE_PLANS else "Contractor Plus"

    # Get trial settings (default to 7 days)
    settings = await db.trial_settings.find_one({"is_active": True})
    if not settings:
        duration_days = 7
    else:
        duration_days = int(settings.get("duration_days", 7))
    
    now = datetime.now(timezone.utc)
    end = now + timedelta(days=duration_days)
    await db.users.update_one({"id": current_user.id}, {"$set": {
        "subscription_plan": trial_plan,
        "subscription_status": "active",
        "plan_start_date": now.isoformat(),
        "plan_end_date": end.isoformat(),
        "subscription_start_date": now.isoformat(),
        "subscription_end_date": end.isoformat(),
        "trial_activated_at": now.isoformat(),
        "trial_duration_days": duration_days,
        "payment_method": "trial"
    }})
    
    # Create notification to prompt user to add payment method
    notification = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.id,
        "type": "trial_payment_setup",
        "title": "Add Payment Method for Seamless Renewal",
        "message": f"Your {duration_days}-day trial is active! Add your payment method now to ensure uninterrupted service when your trial ends. No charge until trial completes.",
        "action_url": "/pricing?trial_conversion=true",
        "action_label": "Add Payment Method",
        "created_at": now.isoformat(),
        "read": False,
        "priority": "high"
    }
    await db.notifications.insert_one(notification)
    
    return {"message": "Trial activated", "plan": trial_plan, "status": "active", "end": end.isoformat()}

class Worker(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    name: str
    phone_number: str
    address: Optional[str] = None
    wage_per_day: float = 450.0
    wage_from_employer: float = 500.0  # New field
    date_of_joining: datetime
    status: str = "Active"
    notes: Optional[str] = None
    advance_paid: float = 0.0  # Renamed from advance_balance
    pending_settlement: float = 0.0
    charges: List[Dict[str, Any]] = []  # Worker charges (electricity, etc)
    wage_history: List[Dict[str, Any]] = []
    room_id: Optional[str] = None  # ✅ Optional room assignment
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class WorkerCreate(BaseModel):
    name: str = Field(..., max_length=200)
    phone_number: str = Field(..., max_length=20)
    address: Optional[str] = Field(None, max_length=500)
    wage_per_day: float = 450.0
    wage_from_employer: float = 500.0
    date_of_joining: str
    notes: Optional[str] = Field(None, max_length=2000)
    initial_pending_settlement: Optional[float] = 0.0  # Initial pending wage when migrating to platform
    initial_advance_paid: Optional[float] = 0.0  # Initial advance paid when migrating to platform
    room_id: Optional[str] = None  # ✅ Optional room assignment at creation time

class WorkerUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    phone_number: Optional[str] = Field(None, max_length=20)
    address: Optional[str] = Field(None, max_length=500)
    wage_per_day: Optional[float] = None
    wage_from_employer: Optional[float] = None
    status: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=2000)
    room_id: Optional[str] = None

class WorkerCharge(BaseModel):
    worker_id: str
    amount: float
    reason: str = Field(..., max_length=500)
    date: Optional[str] = None

class ExtraCharge(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    contractor_id: str
    worker_id: str
    amount: float
    reason: str
    date: str
    is_settled: Optional[bool] = False  # Track if charge has been deducted in a settlement
    settled_at: Optional[str] = None  # Settlement ID where this charge was used

class Employer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    name: str
    address: Optional[str] = None
    phone_number: str
    work_location: Optional[str] = None
    status: str = "Active"
    notes: Optional[str] = None
    advance_paid: float = 0.0
    advance_received: float = 0.0  # Advance received from employer (to deduct from payments)
    pending_payment: float = 0.0
    payment_history: List[Dict[str, Any]] = []
    work_history: List[Dict[str, Any]] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployerCreate(BaseModel):
    name: str = Field(..., max_length=200)
    address: Optional[str] = Field(None, max_length=500)
    phone_number: str = Field(..., max_length=20)
    work_location: Optional[str] = Field(None, max_length=500)
    notes: Optional[str] = Field(None, max_length=2000)
    initial_pending_payment: Optional[float] = 0.0  # Initial pending payment when migrating to platform

class EmployerUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=200)
    address: Optional[str] = Field(None, max_length=500)
    phone_number: Optional[str] = Field(None, max_length=20)
    work_location: Optional[str] = Field(None, max_length=500)
    status: Optional[str] = None
    notes: Optional[str] = Field(None, max_length=2000)

class EmployerAttendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    employer_id: str
    date: str
    workers_count: int
    selected_workers: List[str] = []  # Worker IDs
    payment_per_worker: float = 500.0
    additional_charges: float = 0.0
    charge_description: Optional[str] = None
    total_amount: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployerAttendanceCreate(BaseModel):
    employer_id: str
    date: str
    workers_count: int
    selected_workers: List[str] = []
    payment_per_worker: float = 500.0
    additional_charges: float = 0.0
    charge_description: Optional[str] = Field(None, max_length=500)

class WorkerAttendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    worker_id: str
    employer_id: str
    date: str
    status: str
    wage_earned: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class WorkerAttendanceCreate(BaseModel):
    worker_id: str
    employer_id: str
    date: str
    status: str

class PaymentCollection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    employer_id: str
    amount: float
    advance_deducted: float = 0.0
    payment_mode: str
    remarks: Optional[str] = None
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentCollectionCreate(BaseModel):
    employer_id: str
    amount: float
    advance_deducted: float = 0.0
    payment_mode: str = Field(..., max_length=50)
    remarks: Optional[str] = Field(None, max_length=1000)
    payment_date: Optional[str] = None  # ISO format date string

class WageSettlement(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    worker_id: str
    start_date: str
    end_date: str
    days_worked: int
    total_wage: float
    advance_deducted: float
    charges_deducted: float
    amount_paid: float  # Actual amount given
    settlement_type: str  # "full" or "partial"
    settlement_date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class WageSettlementCreate(BaseModel):
    worker_id: str
    start_date: str
    end_date: str
    advance_deducted: float = 0.0
    charges_deducted: float = 0.0
    amount_paid: float
    settlement_type: str = "full"
    extra_charge_amount: Optional[float] = 0.0  # Extra charge to add during settlement
    extra_charge_reason: Optional[str] = None  # Reason for extra charge

class Advance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    worker_id: str
    amount: float
    purpose: Optional[str] = None
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AdvanceCreate(BaseModel):
    worker_id: str
    amount: float
    purpose: Optional[str] = Field(None, max_length=500)
    date: Optional[str] = None  # YYYY-MM-DD format, defaults to today

class Activity(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    type: str
    description: str
    date: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Commission(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    date: str  # DD-MM-YYYY format (legacy - kept for backward compatibility)
    employer_id: str
    worker_id: str
    payment_from_employer: float  # Amount employer pays per worker
    wage_to_worker: float  # Amount worker receives
    commission_amount: float  # Difference (payment - wage)
    attendance_id: str  # Link to employer attendance record
    workers_count: Optional[int] = None  # Number of workers (for SUMMARY records when workers not selected)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ============ AUTH HELPERS ============

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    """
    Create JWT access token with expiration and issued-at timestamp.
    Tokens include: sub (user_id), role, exp (expiration), iat (issued at)
    """
    to_encode = data.copy()
    now = datetime.now(timezone.utc)
    
    if expires_delta:
        expire = now + expires_delta
    else:
        expire = now + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    
    # Add security claims
    to_encode.update({
        "exp": expire,  # Expiration time
        "iat": now,     # Issued at time
        "nbf": now      # Not before time (token not valid before this)
    })
    
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(auth_token: Optional[str] = Cookie(None)):
    """
    Validate JWT token and return current user.
    
    Security checks:
    1. Token exists
    2. Token signature valid
    3. Token not expired
    4. User exists in database
    5. User account is active
    """
    if not auth_token:
        raise HTTPException(
            status_code=401, 
            detail="Authentication required",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    try:
        # Decode and validate JWT
        # This automatically checks signature, expiration (exp), and not-before (nbf)
        payload = jwt.decode(
            auth_token, 
            SECRET_KEY, 
            algorithms=[ALGORITHM],
            options={
                "verify_signature": True,
                "verify_exp": True,
                "verify_nbf": True,
                "require": ["sub", "exp", "iat"]  # Required claims
            }
        )
        
        user_id: str = payload.get("sub")
        role: str = payload.get("role")
        
        if not user_id:
            raise HTTPException(
                status_code=401, 
                detail="Invalid token format",
                headers={"WWW-Authenticate": "Bearer"}
            )
            
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=401, 
            detail="Token has expired. Please login again.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    except JWTError:
        # Catches all JWT errors including invalid tokens, signature errors, etc.
        raise HTTPException(
            status_code=401, 
            detail="Invalid token. Please login again.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    # Fetch user from database
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(
            status_code=401, 
            detail="User not found. Token invalid.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    # Check if account is active
    if not user.get('is_active', True):
        raise HTTPException(
            status_code=403, 
            detail="Account is deactivated. Contact support."
        )
    
    # Validate role hasn't changed (security check)
    if role and user.get('role') != role:
        raise HTTPException(
            status_code=401, 
            detail="Token invalid. Please login again.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    return User(**user)

async def get_active_subscription_user(current_user: User = Depends(get_current_user)):
    """Check if user has active subscription (includes cancelled subscriptions that are still within billing period)"""
    if current_user.role == "admin":
        return current_user
    
    # Allow access if subscription is "active" or "cancelled" (cancelled subscriptions still have access until end date)
    allowed_statuses = ["active", "cancelled"]
    if current_user.subscription_status not in allowed_statuses:
        raise HTTPException(
            status_code=403, 
            detail="Active subscription required. Please activate your plan."
        )
    
    # Check if plan expired (except for admin key users)
    if current_user.plan_end_date:
        if isinstance(current_user.plan_end_date, str):
            end_date = datetime.fromisoformat(current_user.plan_end_date)
        else:
            end_date = current_user.plan_end_date
        # Normalize to timezone-aware UTC to avoid naive vs aware comparison errors
        if getattr(end_date, 'tzinfo', None) is None:
            end_date = end_date.replace(tzinfo=timezone.utc)
        
        now = datetime.now(timezone.utc)
        if end_date < now:
            # Mark subscription as expired
            await db.users.update_one(
                {"id": current_user.id},
                {"$set": {"subscription_status": "expired"}}
            )
            raise HTTPException(
                status_code=403,
                detail="Your subscription has expired. Please renew."
            )
    
    return current_user

# ============ PLAN LIMITS HELPER FUNCTIONS ============

async def get_plan_limits(plan_name: str) -> dict:
    """
    Get plan limits (max_workers, max_employers) from subscription_plans collection.
    Returns default limits if plan not found or limits not set.
    
    Returns:
        dict: {"max_workers": int or None, "max_employers": int or None}
        None means unlimited
    """
    if not plan_name:
        return {"max_workers": None, "max_employers": None}
    
    # Try to find plan in subscription_plans collection
    plan = await db.subscription_plans.find_one({"name": plan_name}, {"_id": 0})
    
    if plan:
        return {
            "max_workers": plan.get("max_workers"),
            "max_employers": plan.get("max_employers")
        }
    
    # Fallback: Hardcoded limits based on plan name (for backward compatibility)
    plan_name_lower = plan_name.lower()
    
    if "contractor plus" in plan_name_lower or "plus" in plan_name_lower:
        return {"max_workers": 50, "max_employers": 25}
    elif "contractor pro" in plan_name_lower or ("pro" in plan_name_lower and "plus" not in plan_name_lower):
        return {"max_workers": 250, "max_employers": 100}
    elif "enterprise" in plan_name_lower or "estate" in plan_name_lower:
        return {"max_workers": None, "max_employers": None}  # Unlimited
    else:
        # Default/unknown plan - assume unlimited for backward compatibility
        return {"max_workers": None, "max_employers": None}

async def check_worker_limit(current_user: User, allow_exactly_at_limit: bool = True) -> None:
    """
    SECURITY: Check if user can add more workers based on their plan limit.
    Fetches fresh user data from database to prevent token tampering.
    Raises HTTPException if limit reached.
    
    Args:
        current_user: User object (may be from token, will be refreshed)
        allow_exactly_at_limit: If True, allows creation when at limit (for new creations)
                               If False, requires being below limit (for reactivations)
    """
    # SECURITY: Always fetch fresh user data from database to prevent token/role tampering
    fresh_user = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not fresh_user:
        raise HTTPException(status_code=401, detail="User not found")
    
    # SECURITY: Verify admin role from database, not token
    if fresh_user.get("role") == "admin":
        return  # Admins have unlimited access
    
    # SECURITY: Use subscription_plan from database, not from potentially tampered token
    subscription_plan = fresh_user.get("subscription_plan", "none")
    limits = await get_plan_limits(subscription_plan)
    max_workers = limits.get("max_workers")
    
    # None means unlimited
    if max_workers is None:
        return
    
    # SECURITY: Count ALL workers (including inactive) to prevent bypass via status manipulation
    # Users must delete workers to free up slots, not just mark them inactive
    worker_count = await db.workers.count_documents({"contractor_id": current_user.id})
    
    # Check limit based on operation type
    if allow_exactly_at_limit:
        if worker_count >= max_workers:
            # SECURITY: Log limit violation attempt
            await log_security_event(
                event_type=SecurityEventType.PLAN_LIMIT_VIOLATION,
                severity=SecuritySeverity.MEDIUM,
                ip_address="",  # Will be set by caller if available
                user_id=current_user.id,
                email=fresh_user.get("email"),
                details={
                    "resource_type": "worker",
                    "current_count": worker_count,
                    "limit": max_workers,
                    "plan": subscription_plan,
                    "operation": "create"
                },
                endpoint="/api/workers"
            )
            raise HTTPException(
                status_code=403,
                detail=f"Worker limit reached. Your plan ({subscription_plan}) allows maximum {max_workers} workers. Please upgrade to add more."
            )
    else:
        # For reactivations, must be strictly below limit
        if worker_count >= max_workers:
            # SECURITY: Log limit violation attempt
            await log_security_event(
                event_type=SecurityEventType.PLAN_LIMIT_VIOLATION,
                severity=SecuritySeverity.MEDIUM,
                ip_address="",
                user_id=current_user.id,
                email=fresh_user.get("email"),
                details={
                    "resource_type": "worker",
                    "current_count": worker_count,
                    "limit": max_workers,
                    "plan": subscription_plan,
                    "operation": "reactivate"
                },
                endpoint="/api/workers"
            )
            raise HTTPException(
                status_code=403,
                detail=f"Cannot reactivate worker. Your plan ({subscription_plan}) limit of {max_workers} workers is already reached. Please delete inactive workers or upgrade."
            )

async def check_employer_limit(current_user: User, allow_exactly_at_limit: bool = True) -> None:
    """
    SECURITY: Check if user can add more employers based on their plan limit.
    Fetches fresh user data from database to prevent token tampering.
    Raises HTTPException if limit reached.
    
    Args:
        current_user: User object (may be from token, will be refreshed)
        allow_exactly_at_limit: If True, allows creation when at limit (for new creations)
                               If False, requires being below limit (for reactivations)
    """
    # SECURITY: Always fetch fresh user data from database to prevent token/role tampering
    fresh_user = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not fresh_user:
        raise HTTPException(status_code=401, detail="User not found")
    
    # SECURITY: Verify admin role from database, not token
    if fresh_user.get("role") == "admin":
        return  # Admins have unlimited access
    
    # SECURITY: Use subscription_plan from database, not from potentially tampered token
    subscription_plan = fresh_user.get("subscription_plan", "none")
    limits = await get_plan_limits(subscription_plan)
    max_employers = limits.get("max_employers")
    
    # None means unlimited
    if max_employers is None:
        return
    
    # SECURITY: Count ALL employers (including inactive) to prevent bypass via status manipulation
    # Users must delete employers to free up slots, not just mark them inactive
    employer_count = await db.employers.count_documents({"contractor_id": current_user.id})
    
    # Check limit based on operation type
    if allow_exactly_at_limit:
        if employer_count >= max_employers:
            # SECURITY: Log limit violation attempt
            await log_security_event(
                event_type=SecurityEventType.PLAN_LIMIT_VIOLATION,
                severity=SecuritySeverity.MEDIUM,
                ip_address="",
                user_id=current_user.id,
                email=fresh_user.get("email"),
                details={
                    "resource_type": "employer",
                    "current_count": employer_count,
                    "limit": max_employers,
                    "plan": subscription_plan,
                    "operation": "create"
                },
                endpoint="/api/employers"
            )
            raise HTTPException(
                status_code=403,
                detail=f"Employer limit reached. Your plan ({subscription_plan}) allows maximum {max_employers} employers. Please upgrade to add more."
            )
    else:
        # For reactivations, must be strictly below limit
        if employer_count >= max_employers:
            # SECURITY: Log limit violation attempt
            await log_security_event(
                event_type=SecurityEventType.PLAN_LIMIT_VIOLATION,
                severity=SecuritySeverity.MEDIUM,
                ip_address="",
                user_id=current_user.id,
                email=fresh_user.get("email"),
                details={
                    "resource_type": "employer",
                    "current_count": employer_count,
                    "limit": max_employers,
                    "plan": subscription_plan,
                    "operation": "reactivate"
                },
                endpoint="/api/employers"
            )
            raise HTTPException(
                status_code=403,
                detail=f"Cannot reactivate employer. Your plan ({subscription_plan}) limit of {max_employers} employers is already reached. Please delete inactive employers or upgrade."
            )

async def get_admin_user(current_user: User = Depends(get_current_user)):
    """
    SECURITY: Verify admin role from database, not from token.
    Prevents role tampering via JWT token manipulation.
    """
    # SECURITY: Always verify role from database, not from token
    fresh_user = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not fresh_user:
        raise HTTPException(status_code=401, detail="User not found")
    
    if fresh_user.get("role") != "admin":
        # SECURITY: Log unauthorized admin access attempt
        await log_security_event(
            event_type=SecurityEventType.UNAUTHORIZED_ADMIN_ACCESS,
            severity=SecuritySeverity.HIGH,
            ip_address="",
            user_id=current_user.id,
            email=fresh_user.get("email"),
            details={
                "token_role": current_user.role,
                "database_role": fresh_user.get("role"),
                "endpoint": "admin_endpoint"
            },
            endpoint="admin"
        )
        raise HTTPException(status_code=403, detail="Admin access required")
    
    return current_user

async def get_current_admin(request: Request, auth_token: str = Cookie(None)):
    """
    Verify admin JWT token with enhanced security checks.
    
    Security checks:
    1. Token exists (cookie or Authorization header)
    2. Token signature valid
    3. Token not expired
    4. Token type is "admin"
    5. Admin exists in database
    6. Admin account is active
    """
    token = auth_token
    if not token:
        # Try from Authorization header
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
    
    if not token:
        raise HTTPException(
            status_code=401, 
            detail="Admin authentication required",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    try:
        # Decode and validate JWT with all security checks
        payload = jwt.decode(
            token, 
            SECRET_KEY, 
            algorithms=[ALGORITHM],
            options={
                "verify_signature": True,
                "verify_exp": True,
                "verify_nbf": True,
                "require": ["sub", "exp", "iat", "type"]  # Required claims
            }
        )
        
        # CRITICAL: Verify this is an admin token
        if payload.get("type") != "admin":
            raise HTTPException(
                status_code=403, 
                detail="Access denied. Admin privileges required."
            )
        
        admin_id = payload.get("sub")
        if not admin_id:
            raise HTTPException(
                status_code=401, 
                detail="Invalid token format",
                headers={"WWW-Authenticate": "Bearer"}
            )
        
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=401, 
            detail="Admin session expired. Please login again.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    except JWTError:
        # Catches all JWT errors including invalid tokens, signature errors, etc.
        raise HTTPException(
            status_code=401, 
            detail="Invalid admin token. Please login again.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    # Fetch admin from database
    admin = await db.admins.find_one({"id": admin_id}, {"_id": 0})
    if not admin:
        raise HTTPException(
            status_code=401, 
            detail="Admin account not found. Token invalid.",
            headers={"WWW-Authenticate": "Bearer"}
        )
    
    # Check if admin account is active
    if not admin.get('is_active', True):
        raise HTTPException(
            status_code=403, 
            detail="Admin account is disabled. Contact super admin."
        )
    
    return AdminUser(**admin)

# ============ SEED ADMIN ============

async def seed_admin():
    """
    Create or sync admin account from environment variables on every startup.
    SECURITY: Admin credentials MUST be set in .env file, never hardcoded.

    Behavior:
    - If admin doesn't exist → create with ADMIN_PASSWORD, force-change-password ON.
    - If admin exists → update its password hash to match ADMIN_PASSWORD, clear any
      lockouts, and turn must_change_password OFF. This makes the .env file the
      single source of truth for the admin password and lets the operator rotate
      the password by editing .env and redeploying. Critical for deployments where
      a different DB is provisioned per environment (so the password set during
      first deploy can be rotated later).
    """
    admin_email = os.environ.get('ADMIN_EMAIL')
    admin_password = os.environ.get('ADMIN_PASSWORD')

    # Skip if admin credentials not configured
    if not admin_email or not admin_password:
        logging.warning("ADMIN_EMAIL and ADMIN_PASSWORD not set in environment variables. Skipping admin creation.")
        return

    # Validate password strength
    if len(admin_password) < 8:
        logging.error("ADMIN_PASSWORD must be at least 8 characters long")
        return

    new_hash = pwd_context.hash(admin_password)

    # Check if admin already exists → sync password & clear lockouts
    admin_exists = await db.admins.find_one({"email": admin_email})
    if admin_exists:
        await db.admins.update_one(
            {"email": admin_email},
            {"$set": {
                "password": new_hash,
                "is_active": True,
                "must_change_password": False,
                "is_locked": False,
                "failed_login_attempts": 0,
                "locked_until": None,
            }}
        )
        logging.info(f"✅ Admin password synced from env: {admin_email}")
        return

    # Create admin account
    admin = {
        'id': 'admin-' + os.urandom(8).hex(),
        'email': admin_email,
        'password': new_hash,
        'name': 'System Admin',
        'role': 'admin',
        'is_active': True,
        'created_at': datetime.now(timezone.utc).isoformat(),
        'must_change_password': False
    }

    await db.admins.insert_one(admin)
    logging.info(f"✅ Admin account created: {admin_email}")

# ============ AUTH ROUTES ============

@api_router.post("/auth/register")
@limiter.limit("3/minute")  # Rate limit: 3 registration attempts per minute per IP
async def register(request: Request, response: Response, user_data: UserCreate):
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    
    # ===== EMAIL VALIDATION =====
    is_valid_email, email_error = validate_email_format(user_data.email)
    if not is_valid_email:
        # Log security event
        event_type = SecurityEventType.DISPOSABLE_EMAIL if "disposable" in email_error.lower() else SecurityEventType.INVALID_EMAIL
        await log_security_event(
            event_type=event_type,
            severity=SecuritySeverity.LOW,
            ip_address=client_ip,
            user_agent=user_agent,
            email=user_data.email,
            details={"error": email_error},
            endpoint="/auth/register"
        )
        raise HTTPException(status_code=400, detail=email_error)
    
    # ===== PASSWORD STRENGTH VALIDATION =====
    is_valid_password, password_error = validate_password_strength(
        user_data.password, 
        user_data.email
    )
    if not is_valid_password:
        # Log weak password attempt
        await log_security_event(
            event_type=SecurityEventType.WEAK_PASSWORD,
            severity=SecuritySeverity.LOW,
            ip_address=client_ip,
            user_agent=user_agent,
            email=user_data.email,
            details={"error": password_error},
            endpoint="/auth/register"
        )
        raise HTTPException(status_code=400, detail=password_error)
    
    # ===== CHECK EXISTING USER =====
    existing = await db.users.find_one({"email": user_data.email.lower()})
    if existing:
        # Log registration attempt for existing email
        await log_security_event(
            event_type=SecurityEventType.REGISTRATION_BLOCKED,
            severity=SecuritySeverity.MEDIUM,
            ip_address=client_ip,
            user_agent=user_agent,
            email=user_data.email,
            details={"reason": "email_already_exists"},
            endpoint="/auth/register"
        )
        # Show specific error message for better UX
        raise HTTPException(
            status_code=400, 
            detail="This email address is already registered. Please login or use a different email."
        )
    
    # ===== CREATE USER =====
    # Sanitize inputs
    user = User(
        email=user_data.email.lower().strip(),
        name=sanitize_input(user_data.name),
        phone=sanitize_input(user_data.phone) if user_data.phone else None
    )
    user_dict = user.model_dump()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    user_dict['password'] = get_password_hash(user_data.password)
    
    await db.users.insert_one(user_dict)
    
    # ===== AUTO-LOGIN: CREATE SESSION COOKIE =====
    # Create JWT token for the newly registered user
    access_token = create_access_token({"sub": user.id, "role": user.role})
    
    # Set the cookie in the response (same as login endpoint)
    response.set_cookie(
        key="auth_token",
        value=access_token,
        httponly=True,
        secure=COOKIE_SECURE,
        samesite=COOKIE_SAMESITE,
        max_age=ACCESS_TOKEN_EXPIRE_DAYS * 24 * 60 * 60,
        path="/"
    )
    
    return {"message": "Registration successful", "user_id": user.id}

@api_router.post("/auth/login")
@limiter.limit("10/minute")  # Rate limit: 10 login attempts per minute per IP
async def login(request: Request, credentials: UserLogin, response: Response):
    # Get client IP address and user agent
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    
    # ===== BRUTE FORCE PROTECTION =====
    # Check if account or IP is locked
    is_allowed, lockout_message = await check_login_attempts(
        credentials.email.lower(), 
        client_ip
    )
    if not is_allowed:
        # Log account lockout trigger
        await log_security_event(
            event_type=SecurityEventType.ACCOUNT_LOCKED,
            severity=SecuritySeverity.HIGH,
            ip_address=client_ip,
            user_agent=user_agent,
            email=credentials.email.lower(),
            details={"message": lockout_message},
            endpoint="/auth/login"
        )
        raise HTTPException(status_code=429, detail=lockout_message)
    
    # ===== AUTHENTICATION =====
    user_doc = await db.users.find_one({"email": credentials.email.lower()}, {"_id": 0})
    
    # Security: Always check password even if user doesn't exist (timing attack prevention)
    if not user_doc:
        # Perform a dummy hash operation to prevent timing attacks
        pwd_context.hash("dummy_password_to_prevent_timing_attack")
        # Record failed attempt
        await record_failed_login(credentials.email.lower(), client_ip)
        # Log failed login attempt
        await log_security_event(
            event_type=SecurityEventType.FAILED_LOGIN,
            severity=SecuritySeverity.MEDIUM,
            ip_address=client_ip,
            user_agent=user_agent,
            email=credentials.email.lower(),
            details={"reason": "invalid_credentials"},
            endpoint="/auth/login"
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not verify_password(credentials.password, user_doc['password']):
        # Record failed attempt
        await record_failed_login(credentials.email.lower(), client_ip)
        # Log failed login attempt
        await log_security_event(
            event_type=SecurityEventType.FAILED_LOGIN,
            severity=SecuritySeverity.MEDIUM,
            ip_address=client_ip,
            user_agent=user_agent,
            email=credentials.email.lower(),
            user_id=user_doc['id'],
            details={"reason": "invalid_password"},
            endpoint="/auth/login"
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user_doc.get('is_active', True):
        # Don't record as failed attempt for inactive accounts
        raise HTTPException(status_code=403, detail="Account is inactive. Please contact support.")
    
    # ===== SUCCESSFUL LOGIN =====
    # Clear any previous failed login attempts
    await clear_failed_login_attempts(credentials.email.lower(), client_ip)
    
    # Note: Successful user logins are NOT logged to reduce database storage
    # Only suspicious activities, failures, and admin logins are logged for security monitoring
    
    access_token = create_access_token({"sub": user_doc['id'], "role": user_doc['role']})
    user = User(**user_doc)
    
    # Set httpOnly, secure cookie
    # Security: Multi-layer cookie protection
    # httponly=True: Prevents JavaScript access, mitigating XSS attacks
    # secure=True (production): Ensures cookie only sent over HTTPS, preventing MITM
    # samesite='lax': Provides CSRF protection while allowing safe top-level navigation
    #                 Cookies are NOT sent with cross-site POST/PUT/DELETE requests
    # max_age: Token expiration aligned with subscription billing cycle
    response.set_cookie(
        key="auth_token",
        value=access_token,
        httponly=True,           # XSS Protection: No client-side JavaScript access
        secure=COOKIE_SECURE,    # MITM Protection: HTTPS only in production
        samesite=COOKIE_SAMESITE,  # CSRF Protection: 'lax' (recommended) or 'strict'
        max_age=ACCESS_TOKEN_EXPIRE_DAYS * 24 * 60 * 60,
        path="/"
    )
    
    return {
        "message": "Login successful",
        "user": user.model_dump()
    }

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("auth_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.get("/auth/me")
async def get_me(current_user: User = Depends(get_current_user)):
    """Get current user. Normalizes subscription_plan: 'none' -> 'Free Plan'"""
    user_dict = current_user.model_dump()
    # Normalize subscription_plan: 'none' or empty -> 'Free Plan'
    if user_dict.get('subscription_plan') == 'none' or not user_dict.get('subscription_plan'):
        user_dict['subscription_plan'] = 'Free Plan'
    return user_dict

@api_router.put("/auth/update-profile")
async def update_profile(
    request: Request,
    profile_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    SECURITY: Update user's own profile (name only).
    Prevents modification of subscription_plan, role, subscription_status, or other sensitive fields.
    """
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    
    # SECURITY: Whitelist allowed fields - prevent tampering with subscription/role data
    ALLOWED_FIELDS = {"name"}
    
    # Check for forbidden fields (subscription tampering attempts)
    forbidden_fields = set(profile_data.keys()) - ALLOWED_FIELDS
    sensitive_fields = {"subscription_plan", "role", "subscription_status", "plan_start_date", 
                        "plan_end_date", "is_active", "id", "email", "password", "created_at"}
    
    if any(field in profile_data for field in sensitive_fields):
        # SECURITY: Log subscription tampering attempt
        await log_security_event(
            event_type=SecurityEventType.SUBSCRIPTION_TAMPER_ATTEMPT,
            severity=SecuritySeverity.HIGH,
            ip_address=client_ip,
            user_agent=user_agent,
            user_id=current_user.id,
            email=current_user.email,
            details={
                "attempted_fields": list(profile_data.keys()),
                "forbidden_fields": list(forbidden_fields),
                "endpoint": "/auth/update-profile"
            },
            endpoint="/auth/update-profile"
        )
        raise HTTPException(
            status_code=403,
            detail="You cannot modify subscription or account security fields. Contact support for subscription changes."
        )
    
    update_fields = {}
    
    # Only allow updating name field
    if 'name' in profile_data and profile_data['name']:
        update_fields['name'] = sanitize_input(profile_data['name'])
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    # SECURITY: Always include user ID in filter to prevent cross-user updates
    result = await db.users.update_one(
        {"id": current_user.id},
        {"$set": update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Profile updated successfully"}

@api_router.put("/auth/change-password")
async def change_password(
    password_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Change user's password"""
    old_password = password_data.get('old_password')
    new_password = password_data.get('new_password')
    
    if not old_password or not new_password:
        raise HTTPException(status_code=400, detail="Both old and new passwords are required")
    
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    
    # Verify old password
    user = await db.users.find_one({"id": current_user.id})
    if not user or not pwd_context.verify(old_password, user['password']):
        raise HTTPException(status_code=400, detail="Old password is incorrect")
    
    # Update password
    hashed_password = pwd_context.hash(new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"password": hashed_password}}
    )
    
    return {"message": "Password changed successfully"}

# ============ WORKER ROUTES ============

@api_router.post("/workers", response_model=Worker)
async def create_worker(
    worker_data: WorkerCreate, 
    current_user: User = Depends(get_active_subscription_user)
):
    # Check worker limit before creating
    await check_worker_limit(current_user)
    
    date_parts = worker_data.date_of_joining.split('-')
    joining_date = datetime(int(date_parts[2]), int(date_parts[1]), int(date_parts[0]), tzinfo=timezone.utc)
    
    worker = Worker(
        contractor_id=current_user.id,
        name=sanitize_input(worker_data.name),
        phone_number=sanitize_input(worker_data.phone_number),
        address=sanitize_input(worker_data.address) if worker_data.address else None,
        wage_per_day=worker_data.wage_per_day,
        wage_from_employer=worker_data.wage_from_employer,
        date_of_joining=joining_date,
        notes=sanitize_input(worker_data.notes) if worker_data.notes else None,
        pending_settlement=worker_data.initial_pending_settlement or 0.0,  # Set initial pending wage
        advance_paid=worker_data.initial_advance_paid or 0.0,  # Set initial advance paid
        room_id=worker_data.room_id or None  # ✅ Optional room assignment from form
    )
    
    worker_dict = worker.model_dump()
    worker_dict['date_of_joining'] = worker_dict['date_of_joining'].isoformat()
    worker_dict['created_at'] = worker_dict['created_at'].isoformat()

    # ✅ If room_id is provided, validate it belongs to this contractor and isn't full.
    if worker_data.room_id:
        room = await db.rooms.find_one({"id": worker_data.room_id, "contractor_id": current_user.id})
        if not room:
            raise HTTPException(status_code=400, detail="Selected room not found")
        if room.get("max_occupants"):
            current_count = await db.workers.count_documents({
                "contractor_id": current_user.id,
                "room_id": worker_data.room_id,
                "status": "Active",
            })
            if current_count >= int(room["max_occupants"]):
                raise HTTPException(
                    status_code=400,
                    detail=f"Room '{room.get('name', 'Selected')}' is at full capacity ({room['max_occupants']} occupants)."
                )
    
    # SECURITY: Insert worker
    await db.workers.insert_one(worker_dict)
    
    # SECURITY: Post-insert verification to catch race conditions
    # Double-check limit wasn't exceeded due to concurrent requests
    try:
        await check_worker_limit(current_user, allow_exactly_at_limit=False)
    except HTTPException:
        # Limit exceeded after insert - rollback by deleting the worker
        await db.workers.delete_one({"id": worker_dict["id"]})
        raise HTTPException(
            status_code=403,
            detail="Worker limit reached. Another operation may have used the last available slot. Please try again."
        )
    
    activity = Activity(
        contractor_id=current_user.id,
        type="worker_added",
        description=f"Added worker: {worker.name}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return worker

@api_router.get("/workers", response_model=List[Worker])
async def get_workers(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for worker in workers:
        if isinstance(worker['date_of_joining'], str):
            worker['date_of_joining'] = datetime.fromisoformat(worker['date_of_joining'])
        if isinstance(worker['created_at'], str):
            worker['created_at'] = datetime.fromisoformat(worker['created_at'])
    return workers

@api_router.get("/workers/without-room")
async def list_workers_without_room(current_user: User = Depends(get_active_subscription_user)):
    """Active workers who are not yet assigned to any room — for the
    Add-Member / Create-Room worker picker. Declared before /workers/{worker_id}
    so FastAPI routes the literal path before the wildcard."""
    workers = await db.workers.find({
        "contractor_id": current_user.id,
        "status": "Active",
        "$or": [{"room_id": {"$exists": False}}, {"room_id": None}, {"room_id": ""}],
    }, {"_id": 0}).to_list(1000)
    return workers

@api_router.get("/workers/{worker_id}", response_model=Worker)
async def get_worker(worker_id: str, current_user: User = Depends(get_active_subscription_user)):
    worker = await db.workers.find_one({"id": worker_id, "contractor_id": current_user.id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    if isinstance(worker['date_of_joining'], str):
        worker['date_of_joining'] = datetime.fromisoformat(worker['date_of_joining'])
    if isinstance(worker['created_at'], str):
        worker['created_at'] = datetime.fromisoformat(worker['created_at'])
    return Worker(**worker)

@api_router.put("/workers/{worker_id}", response_model=Worker)
async def update_worker(
    worker_id: str, 
    worker_data: WorkerUpdate, 
    current_user: User = Depends(get_active_subscription_user)
):
    worker = await db.workers.find_one({"id": worker_id, "contractor_id": current_user.id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # SECURITY: Check if status is being changed from Inactive to Active
    # If so, verify the user hasn't exceeded their limit
    current_status = worker.get("status", "Active")
    new_status = worker_data.status if worker_data.status is not None else current_status
    
    if current_status != "Active" and new_status == "Active":
        # Reactivating an inactive worker - check limit first
        await check_worker_limit(current_user, allow_exactly_at_limit=False)
    
    update_data = {}
    for k, v in worker_data.model_dump().items():
        if v is not None:
            # SECURITY: Prevent modification of contractor_id or other sensitive fields
            if k in ["contractor_id", "id", "created_at"]:
                continue
            if isinstance(v, str):
                update_data[k] = sanitize_input(v)
            else:
                update_data[k] = v
    
    if update_data:
        # SECURITY: Ensure contractor_id is always in the filter to prevent cross-user updates
        await db.workers.update_one({"id": worker_id, "contractor_id": current_user.id}, {"$set": update_data})
    
    updated_worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    if isinstance(updated_worker['date_of_joining'], str):
        updated_worker['date_of_joining'] = datetime.fromisoformat(updated_worker['date_of_joining'])
    if isinstance(updated_worker['created_at'], str):
        updated_worker['created_at'] = datetime.fromisoformat(updated_worker['created_at'])
    return Worker(**updated_worker)

@api_router.delete("/workers/{worker_id}")
async def delete_worker(worker_id: str, current_user: User = Depends(get_active_subscription_user)):
    # ✅ Block deletion if any attendance / commission / settlement / charge history exists.
    # The user is offered the option to deactivate (set status=Inactive) instead.
    worker = await db.workers.find_one({"id": worker_id, "contractor_id": current_user.id})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")

    attendance_exists = await db.attendance.find_one(
        {"contractor_id": current_user.id,
         "$or": [{"worker_id": worker_id}, {"selected_workers": worker_id}]}
    ) or await db.worker_attendance.find_one({"contractor_id": current_user.id, "worker_id": worker_id})

    commission_exists = await db.commissions.find_one({"contractor_id": current_user.id, "worker_id": worker_id})
    settlement_exists = await db.wage_settlements.find_one({"contractor_id": current_user.id, "worker_id": worker_id}) if hasattr(db, "wage_settlements") else None
    has_pending = float(worker.get("pending_settlement", 0) or 0) != 0 or float(worker.get("advance_paid", 0) or 0) != 0

    if attendance_exists or commission_exists or settlement_exists or has_pending:
        raise HTTPException(
            status_code=400,
            detail=(
                "This worker has attendance, commission, settlement or pending balance history "
                "and cannot be deleted. You can deactivate the worker instead — their records will "
                "be preserved and they won't appear in active lists."
            )
        )

    result = await db.workers.delete_one({"id": worker_id, "contractor_id": current_user.id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Worker not found")
    return {"message": "Worker deleted successfully"}

@api_router.post("/workers/add-charge")
async def add_worker_charge(
    charge_data: WorkerCharge,
    current_user: User = Depends(get_active_subscription_user)
):
    worker = await db.workers.find_one({"id": charge_data.worker_id, "contractor_id": current_user.id})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    charge = {
        "amount": charge_data.amount,
        "reason": sanitize_input(charge_data.reason),
        "date": charge_data.date or datetime.now(timezone.utc).strftime("%d-%m-%Y")
    }
    
    await db.workers.update_one(
        {"id": charge_data.worker_id},
        {
            "$push": {"charges": charge},
            "$inc": {"pending_settlement": -charge_data.amount}
        }
    )
    
    return {"message": "Charge added successfully"}

# ============ ROOMS ROUTES ============

class RoomBase(BaseModel):
    name: str
    key_number: Optional[str] = None
    max_occupants: Optional[int] = None

class RoomCreate(RoomBase):
    member_ids: List[str] = []  # ✅ Optional worker IDs to assign at creation

class RoomUpdate(BaseModel):
    name: Optional[str] = None
    key_number: Optional[str] = None
    max_occupants: Optional[int] = None

class Room(RoomBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    member_count: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.get("/rooms")
async def get_rooms(current_user: User = Depends(get_active_subscription_user)):
    """Get all rooms for the current contractor"""
    rooms = await db.rooms.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(1000)
    
    # Calculate member count for each room
    for room in rooms:
        member_count = await db.workers.count_documents({
            "contractor_id": current_user.id,
            "room_id": room["id"],
            "status": "Active"
        })
        room["member_count"] = member_count
    
    return rooms

@api_router.get("/rooms/{room_id}")
async def get_room(room_id: str, current_user: User = Depends(get_active_subscription_user)):
    """Get a specific room with its members"""
    room = await db.rooms.find_one({"id": room_id, "contractor_id": current_user.id}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    
    # Get members
    members = await db.workers.find({
        "contractor_id": current_user.id,
        "room_id": room_id,
        "status": "Active"
    }, {"_id": 0}).to_list(1000)
    
    room["members"] = members
    room["member_count"] = len(members)
    
    return room

@api_router.post("/rooms")
async def create_room(room_data: RoomCreate, current_user: User = Depends(get_active_subscription_user)):
    """Create a new room. Optionally assign initial members in the same call.
    Workers already assigned to another room are silently skipped."""
    room_id = str(uuid.uuid4())
    room = {
        "id": room_id,
        "contractor_id": current_user.id,
        "name": sanitize_input(room_data.name),
        "key_number": sanitize_input(room_data.key_number) if room_data.key_number else None,
        "max_occupants": room_data.max_occupants,
        "member_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }

    await db.rooms.insert_one(room)

    # ✅ Assign initial members (skip those already in any room)
    if room_data.member_ids:
        candidate_ids = [mid for mid in room_data.member_ids if mid]
        if room_data.max_occupants:
            candidate_ids = candidate_ids[: int(room_data.max_occupants)]
        if candidate_ids:
            await db.workers.update_many(
                {
                    "contractor_id": current_user.id,
                    "id": {"$in": candidate_ids},
                    "status": "Active",
                    "$or": [{"room_id": {"$exists": False}}, {"room_id": None}, {"room_id": ""}],
                },
                {"$set": {"room_id": room_id}},
            )

    # Return the room with the live member count
    created_room = await db.rooms.find_one({"id": room_id}, {"_id": 0})
    created_room["member_count"] = await db.workers.count_documents(
        {"contractor_id": current_user.id, "room_id": room_id, "status": "Active"}
    )
    return created_room

@api_router.put("/rooms/{room_id}")
async def update_room(
    room_id: str,
    room_data: RoomUpdate,
    current_user: User = Depends(get_active_subscription_user)
):
    """Update a room"""
    room = await db.rooms.find_one({"id": room_id, "contractor_id": current_user.id}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    
    update_data = {}
    if room_data.name is not None:
        update_data["name"] = sanitize_input(room_data.name)
    if room_data.key_number is not None:
        update_data["key_number"] = sanitize_input(room_data.key_number) if room_data.key_number else None
    if room_data.max_occupants is not None:
        update_data["max_occupants"] = room_data.max_occupants
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.rooms.update_one(
            {"id": room_id, "contractor_id": current_user.id},
            {"$set": update_data}
        )
    
    updated_room = await db.rooms.find_one({"id": room_id}, {"_id": 0})
    
    # Calculate member count
    member_count = await db.workers.count_documents({
        "contractor_id": current_user.id,
        "room_id": room_id,
        "status": "Active"
    })
    updated_room["member_count"] = member_count
    
    return updated_room

@api_router.delete("/rooms/{room_id}")
async def delete_room(room_id: str, current_user: User = Depends(get_active_subscription_user)):
    """Delete a room and remove room assignment from all workers"""
    room = await db.rooms.find_one({"id": room_id, "contractor_id": current_user.id})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    
    # Remove room_id from all workers in this room
    await db.workers.update_many(
        {"contractor_id": current_user.id, "room_id": room_id},
        {"$unset": {"room_id": ""}}
    )
    
    # Delete the room
    await db.rooms.delete_one({"id": room_id, "contractor_id": current_user.id})
    
    return {"message": "Room deleted successfully"}

# ============ EMPLOYER ROUTES (Similar pattern with subscription check) ============
# Continuing with remaining endpoints...

# ============ EMPLOYER ROUTES ============

@api_router.post("/employers", response_model=Employer)
async def create_employer(employer_data: EmployerCreate, current_user: User = Depends(get_active_subscription_user)):
    # Check employer limit before creating
    await check_employer_limit(current_user)
    
    employer = Employer(
        contractor_id=current_user.id,
        name=sanitize_input(employer_data.name),
        address=sanitize_input(employer_data.address) if employer_data.address else None,
        phone_number=sanitize_input(employer_data.phone_number),
        work_location=sanitize_input(employer_data.work_location) if employer_data.work_location else None,
        notes=sanitize_input(employer_data.notes) if employer_data.notes else None,
        pending_payment=employer_data.initial_pending_payment or 0.0  # Set initial pending
    )
    
    employer_dict = employer.model_dump()
    employer_dict['created_at'] = employer_dict['created_at'].isoformat()
    
    # SECURITY: Insert employer
    await db.employers.insert_one(employer_dict)
    
    # SECURITY: Post-insert verification to catch race conditions
    # Double-check limit wasn't exceeded due to concurrent requests
    try:
        await check_employer_limit(current_user, allow_exactly_at_limit=False)
    except HTTPException:
        # Limit exceeded after insert - rollback by deleting the employer
        await db.employers.delete_one({"id": employer_dict["id"]})
        raise HTTPException(
            status_code=403,
            detail="Employer limit reached. Another operation may have used the last available slot. Please try again."
        )
    
    activity = Activity(
        contractor_id=current_user.id,
        type="employer_added",
        description=f"Added employer: {employer.name}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return employer

@api_router.get("/employers", response_model=List[Employer])
async def get_employers(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for employer in employers:
        if isinstance(employer['created_at'], str):
            employer['created_at'] = datetime.fromisoformat(employer['created_at'])
    return employers

@api_router.get("/employers/{employer_id}", response_model=Employer)
async def get_employer(employer_id: str, current_user: User = Depends(get_active_subscription_user)):
    employer = await db.employers.find_one({"id": employer_id, "contractor_id": current_user.id}, {"_id": 0})
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    if isinstance(employer['created_at'], str):
        employer['created_at'] = datetime.fromisoformat(employer['created_at'])
    return Employer(**employer)

@api_router.put("/employers/{employer_id}", response_model=Employer)
async def update_employer(employer_id: str, employer_data: EmployerUpdate, current_user: User = Depends(get_active_subscription_user)):
    employer = await db.employers.find_one({"id": employer_id, "contractor_id": current_user.id}, {"_id": 0})
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    
    # SECURITY: Check if status is being changed from Inactive to Active
    # If so, verify the user hasn't exceeded their limit
    current_status = employer.get("status", "Active")
    new_status = employer_data.status if employer_data.status is not None else current_status
    
    if current_status != "Active" and new_status == "Active":
        # Reactivating an inactive employer - check limit first
        await check_employer_limit(current_user, allow_exactly_at_limit=False)
    
    update_data = {}
    for k, v in employer_data.model_dump().items():
        if v is not None:
            # SECURITY: Prevent modification of contractor_id or other sensitive fields
            if k in ["contractor_id", "id", "created_at"]:
                continue
            if isinstance(v, str):
                update_data[k] = sanitize_input(v)
            else:
                update_data[k] = v
    
    if update_data:
        # SECURITY: Ensure contractor_id is always in the filter to prevent cross-user updates
        await db.employers.update_one({"id": employer_id, "contractor_id": current_user.id}, {"$set": update_data})
    
    updated_employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
    if isinstance(updated_employer['created_at'], str):
        updated_employer['created_at'] = datetime.fromisoformat(updated_employer['created_at'])
    return Employer(**updated_employer)

@api_router.delete("/employers/{employer_id}")
async def delete_employer(employer_id: str, current_user: User = Depends(get_active_subscription_user)):
    # ✅ Block deletion if any attendance / commission / payment history exists.
    employer = await db.employers.find_one({"id": employer_id, "contractor_id": current_user.id})
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")

    attendance_exists = await db.attendance.find_one({"contractor_id": current_user.id, "employer_id": employer_id})
    commission_exists = await db.commissions.find_one({"contractor_id": current_user.id, "employer_id": employer_id})
    has_pending = float(employer.get("pending_payment", 0) or 0) != 0 or float(employer.get("advance_received", 0) or 0) != 0

    if attendance_exists or commission_exists or has_pending:
        raise HTTPException(
            status_code=400,
            detail=(
                "This employer has attendance, commission or pending payment history and "
                "cannot be deleted. You can deactivate the employer instead — their records will "
                "be preserved and they won't appear in active lists."
            )
        )

    result = await db.employers.delete_one({"id": employer_id, "contractor_id": current_user.id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employer not found")
    return {"message": "Employer deleted successfully"}

# ============ ATTENDANCE ROUTES ============

@api_router.post("/attendance/employer", response_model=EmployerAttendance)
async def create_employer_attendance(attendance_data: EmployerAttendanceCreate, current_user: User = Depends(get_active_subscription_user)):
    # Check total active workers
    active_workers_count = await db.workers.count_documents({"contractor_id": current_user.id, "status": "Active"})
    
    # PROBLEM 1 FIX: Calculate truly assigned workers (those with employer_id set)
    # Count workers already assigned to employers for this date
    assigned_workers = await db.worker_attendance.distinct("worker_id", {
        "contractor_id": current_user.id,
        "date": attendance_data.date,
        "employer_id": {"$ne": "", "$exists": True},
        "status": {"$in": ["Present", "Late"]}
    })
    already_assigned_count = len(assigned_workers)
    
    # If specific workers are selected, check if they're already assigned
    if attendance_data.selected_workers:
        for worker_id in attendance_data.selected_workers:
            if worker_id in assigned_workers:
                worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
                raise HTTPException(
                    status_code=400,
                    detail=f"Worker '{worker['name'] if worker else worker_id}' is already assigned to another employer for this date"
                )
    
    # Calculate available unassigned workers
    available_workers = active_workers_count - already_assigned_count
    
    if attendance_data.workers_count > available_workers:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot exceed available workers. Total active: {active_workers_count}, Already assigned: {already_assigned_count}, Available: {available_workers}"
        )
    
    # Calculate total_amount: Use individual worker rates when specific workers selected
    if attendance_data.selected_workers and len(attendance_data.selected_workers) > 0:
        # Sum individual worker wage_from_employer rates
        total_from_workers = 0.0
        for worker_id in attendance_data.selected_workers:
            worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
            if worker:
                total_from_workers += float(worker.get("wage_from_employer", 500.0))
        total_amount = total_from_workers + attendance_data.additional_charges
    else:
        # Use payment_per_worker when no specific workers selected
        total_amount = (attendance_data.workers_count * attendance_data.payment_per_worker) + attendance_data.additional_charges
    
    attendance = EmployerAttendance(
        contractor_id=current_user.id,
        employer_id=attendance_data.employer_id,
        date=attendance_data.date,
        workers_count=attendance_data.workers_count,
        selected_workers=attendance_data.selected_workers,
        payment_per_worker=attendance_data.payment_per_worker,
        additional_charges=attendance_data.additional_charges,
        charge_description=sanitize_input(attendance_data.charge_description) if attendance_data.charge_description else None,
        total_amount=total_amount
    )
    
    attendance_dict = attendance.model_dump()
    attendance_dict['created_at'] = attendance_dict['created_at'].isoformat()
    await db.employer_attendance.insert_one(attendance_dict)
    
    # Update employer pending payment
    await db.employers.update_one(
        {"id": attendance_data.employer_id},
        {"$inc": {"pending_payment": total_amount}}
    )
    
    # If workers selected, auto-mark their attendance AND CALCULATE COMMISSIONS
    if attendance_data.selected_workers:
        for worker_id in attendance_data.selected_workers:
            worker = await db.workers.find_one({"id": worker_id})
            if worker:
                # PROBLEM 2 FIX: Check if worker attendance already exists (marked earlier without employer)
                existing = await db.worker_attendance.find_one({
                    "contractor_id": current_user.id,
                    "worker_id": worker_id,
                    "date": attendance_data.date
                })
                
                if existing:
                    # Update existing worker attendance with employer_id
                    await db.worker_attendance.update_one(
                        {
                            "contractor_id": current_user.id,
                            "worker_id": worker_id,
                            "date": attendance_data.date
                        },
                        {"$set": {"employer_id": attendance_data.employer_id}}
                    )
                    
                    # ✅ Calculate commission: Use payment_per_worker if provided, otherwise use worker's wage_from_employer
                    worker_wage = float(worker.get('wage_per_day', 0))
                    if attendance_data.payment_per_worker and attendance_data.payment_per_worker > 0:
                        # Use payment_per_worker when provided
                        payment_from_employer = attendance_data.payment_per_worker
                    else:
                        # Use worker's wage_from_employer (set during worker profile creation)
                        payment_from_employer = float(worker.get('wage_from_employer', 500.0))
                    
                    commission_amount = max(payment_from_employer - worker_wage, 0.0)
                    
                    # Check if commission already exists
                    existing_commission = await db.commissions.find_one({
                        "contractor_id": current_user.id,
                        "worker_id": worker_id,
                        "employer_id": attendance_data.employer_id,
                        "date": attendance_data.date
                    })
                    
                    if not existing_commission:
                        commission = Commission(
                            contractor_id=current_user.id,
                            date=attendance_data.date,
                            employer_id=attendance_data.employer_id,
                            worker_id=worker_id,
                            payment_from_employer=payment_from_employer,
                            wage_to_worker=worker_wage,
                            commission_amount=commission_amount,
                            attendance_id=attendance.id
                        )
                        
                        commission_dict = commission.model_dump()
                        commission_dict['created_at'] = commission_dict['created_at'].isoformat()
                        await db.commissions.insert_one(commission_dict)
                else:
                    # Create new worker attendance
                    worker_wage = float(worker.get('wage_per_day', 0))
                    
                    worker_att = WorkerAttendance(
                        contractor_id=current_user.id,
                        worker_id=worker_id,
                        employer_id=attendance_data.employer_id,
                        date=attendance_data.date,
                        status="Present",
                        wage_earned=worker_wage
                    )
                    worker_att_dict = worker_att.model_dump()
                    worker_att_dict['created_at'] = worker_att_dict['created_at'].isoformat()
                    await db.worker_attendance.insert_one(worker_att_dict)
                    
                    # Update worker pending settlement
                    await db.workers.update_one(
                        {"id": worker_id},
                        {"$inc": {"pending_settlement": worker_wage}}
                    )
                    
                    # ✅ Calculate commission: Use payment_per_worker if provided, otherwise use worker's wage_from_employer
                    if attendance_data.payment_per_worker and attendance_data.payment_per_worker > 0:
                        # Use payment_per_worker when provided
                        payment_from_employer = attendance_data.payment_per_worker
                    else:
                        # Use worker's wage_from_employer (set during worker profile creation)
                        payment_from_employer = float(worker.get('wage_from_employer', 500.0))
                    
                    commission_amount = max(payment_from_employer - worker_wage, 0.0)
                    
                    commission = Commission(
                        contractor_id=current_user.id,
                        date=attendance_data.date,
                        employer_id=attendance_data.employer_id,
                        worker_id=worker_id,
                        payment_from_employer=payment_from_employer,
                        wage_to_worker=worker_wage,
                        commission_amount=commission_amount,
                        attendance_id=attendance.id
                    )
                    
                    commission_dict = commission.model_dump()
                    commission_dict['created_at'] = commission_dict['created_at'].isoformat()
                    await db.commissions.insert_one(commission_dict)
    else:
        # ✅ When workers are NOT selected, calculate commissions using default wage from preferences
        # Get contractor preferences to get default_worker_wage
        preference = await db.contractor_preferences.find_one({"contractor_id": current_user.id}, {"_id": 0})
        
        if preference:
            default_wage = float(preference.get("default_worker_wage", 450.0))
        else:
            # If no preference exists, use default value
            default_wage = 450.0
        
        # Determine payment_from_employer
        if attendance_data.payment_per_worker and attendance_data.payment_per_worker > 0:
            payment_from_employer = attendance_data.payment_per_worker
        else:
            # Use default employer rate from preferences
            if preference:
                payment_from_employer = float(preference.get("default_employer_rate", 500.0))
            else:
                payment_from_employer = 500.0
        
        # Calculate commission per worker: payment_from_employer - default_wage
        commission_per_worker = max(payment_from_employer - default_wage, 0.0)
        total_commission = commission_per_worker * attendance_data.workers_count
        
        # Create summary commission record (using special marker "SUMMARY" for worker_id when workers not selected)
        # Check if summary commission already exists for this employer/date
        existing_summary_commission = await db.commissions.find_one({
            "contractor_id": current_user.id,
            "employer_id": attendance_data.employer_id,
            "date": attendance_data.date,
            "worker_id": "SUMMARY"  # Special marker for summary record when workers not selected
        })
        
        if not existing_summary_commission and total_commission >= 0:
            commission = Commission(
                contractor_id=current_user.id,
                date=attendance_data.date,
                employer_id=attendance_data.employer_id,
                worker_id="SUMMARY",  # Special marker for summary record when workers not selected
                payment_from_employer=payment_from_employer,
                wage_to_worker=default_wage,  # Standard/default wage used for all workers
                commission_amount=total_commission,  # Total commission = (payment - default_wage) × workers_count
                attendance_id=attendance.id,
                workers_count=attendance_data.workers_count  # Store actual number of workers
            )
            
            commission_dict = commission.model_dump()
            commission_dict['created_at'] = commission_dict['created_at'].isoformat()
            await db.commissions.insert_one(commission_dict)
    
    # Log activity
    employer = await db.employers.find_one({"id": attendance_data.employer_id}, {"_id": 0})
    activity = Activity(
        contractor_id=current_user.id,
        type="workers_sent",
        description=f"Sent {attendance_data.workers_count} workers to {employer['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return attendance

@api_router.get("/attendance/employer")
async def get_employer_attendance(
    date: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    query = {"contractor_id": current_user.id}
    if date:
        query["date"] = date
    
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    attendance_list = await db.employer_attendance.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for att in attendance_list:
        if isinstance(att['created_at'], str):
            att['created_at'] = datetime.fromisoformat(att['created_at'])
    return attendance_list

# Put this in server.py replacing existing @api_router.post("/attendance/worker") implementation
from typing import List
from uuid import uuid4
from datetime import datetime, timezone

@api_router.post("/attendance/worker")
async def create_worker_attendance(
    attendance_data: List[WorkerAttendanceCreate],
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Create worker attendance records (NO commission calculation to prevent doubling).
    Commissions are only calculated when employer attendance is marked with specific workers selected.

    - Worker attendance saved to: worker_attendance
    - Commission is NOT calculated here to prevent doubling when attendance is marked in both modes
    """
    results = []

    for att in attendance_data:
        # sanitize inputs if you have a sanitize_input() helper
        # att.worker_id, att.employer_id, att.date, att.status expected

        # --- 1) Prevent duplicate worker attendance (same contractor, worker, date) ---
        existing = await db.worker_attendance.find_one({
            "contractor_id": current_user.id,
            "worker_id": att.worker_id,
            "date": att.date
        })
        if existing:
            # Skip duplicates (you can also collect info if needed)
            continue

        # --- 2) Load worker document to get worker base wage ---
        worker_doc = await db.workers.find_one({"id": att.worker_id}, {"_id": 0})
        if not worker_doc:
            # worker not found -> skip (or optionally return error)
            continue

        # worker's base wage (what you pay the worker)
        wage_to_worker = float(worker_doc.get("wage_per_day", 0) or 0)

        # --- 3) Compute wage_earned based on status ---
        wage_earned = 0.0
        status = (att.status or "").strip()
        if status.lower() == "present":
            wage_earned = wage_to_worker
        elif status.lower() == "late":
            wage_earned = wage_to_worker * 0.75
        else:
            wage_earned = 0.0

        # --- 4) Prepare and insert worker attendance doc (NO commission calculation) ---
        attendance_id = str(uuid4())
        now_iso = datetime.now(timezone.utc).isoformat()

        attendance_doc = {
            "id": attendance_id,
            "contractor_id": current_user.id,
            "employer_id": att.employer_id,
            "worker_id": att.worker_id,
            "date": att.date,
            "status": att.status,
            "mode": "worker",
            "wage_amount": wage_to_worker,             # store the worker wage here (per your schema)
            "wage_earned": wage_earned,                # actual amount earned depending on status
            "additional_charges": att.additional_charges if getattr(att, "additional_charges", None) else [],
            "created_by": current_user.id,
            "created_at": now_iso,
            "updated_at": now_iso
        }

        await db.worker_attendance.insert_one(attendance_doc)

        # --- 5) Update worker pending settlement (atomic increment) ---
        if wage_earned and wage_earned > 0:
            await db.workers.update_one(
                {"id": att.worker_id},
                {"$inc": {"pending_settlement": wage_earned}}
            )

        # ✅ NOTE: Commission is NOT calculated here to prevent doubling
        # Commissions are only calculated when employer attendance is marked with specific workers selected

        results.append(attendance_doc)

    return {
        "message": f"Recorded {len(results)} worker attendance entries",
        "count": len(results),
        "attendance": results
    }


@api_router.get("/attendance/worker")
async def get_worker_attendance(
    date: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    query = {"contractor_id": current_user.id}
    if date:
        query["date"] = date
    
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    attendance_list = await db.worker_attendance.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for att in attendance_list:
        if isinstance(att['created_at'], str):
            att['created_at'] = datetime.fromisoformat(att['created_at'])
    return attendance_list

# ============ PAYMENT ROUTES ============

@api_router.post("/payments/collect", response_model=PaymentCollection)
async def collect_payment(payment_data: PaymentCollectionCreate, current_user: User = Depends(get_active_subscription_user)):
    # ✅ CORRECTED LOGIC: Advance is already deducted from pending_payment when received
    # So we just collect the actual cash payment here
    employer = await db.employers.find_one({"id": payment_data.employer_id})
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    
    current_pending = float(employer.get('pending_payment', 0))
    
    # Validate payment amount - allow some flexibility for adjustments
    if payment_data.amount > current_pending * 1.1 and current_pending > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Payment amount (₹{payment_data.amount:.2f}) exceeds pending payment (₹{current_pending:.2f}) by too much. Please verify the amount."
        )
    
    # Use provided payment_date or default to now
    if payment_data.payment_date:
        try:
            payment_date = datetime.fromisoformat(payment_data.payment_date.replace('Z', '+00:00'))
        except:
            payment_date = datetime.now(timezone.utc)
    else:
        payment_date = datetime.now(timezone.utc)
    
    # Create payment record (advance_deducted kept for backward compatibility but should be 0)
    payment = PaymentCollection(
        contractor_id=current_user.id,
        employer_id=payment_data.employer_id,
        amount=payment_data.amount,
        advance_deducted=0,  # No longer used - advance already deducted when received
        payment_mode=payment_data.payment_mode,
        remarks=sanitize_input(payment_data.remarks) if payment_data.remarks else None,
        date=payment_date
    )
    
    payment_dict = payment.model_dump()
    payment_dict['date'] = payment_dict['date'].isoformat()
    await db.payment_collections.insert_one(payment_dict)
    
    # Update employer's pending payment
    await db.employers.update_one(
        {"id": payment_data.employer_id},
        {"$inc": {"pending_payment": -payment_data.amount}}
    )
    
    employer = await db.employers.find_one({"id": payment_data.employer_id}, {"_id": 0})
    
    # Create activity log
    activity = Activity(
        contractor_id=current_user.id,
        type="payment_collected",
        description=f"Collected ₹{payment_data.amount} from {employer['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return payment

@api_router.get("/payments/collections")
async def get_payment_collections(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    payments = await db.payment_collections.find({"contractor_id": current_user.id}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for payment in payments:
        if isinstance(payment['date'], str):
            payment['date'] = datetime.fromisoformat(payment['date'])
    return payments

@api_router.post("/payments/settle-wage", response_model=WageSettlement)
async def settle_wage(settlement_data: WageSettlementCreate, current_user: User = Depends(get_active_subscription_user)):
    # Get worker to find current pending
    worker = await db.workers.find_one({"id": settlement_data.worker_id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Use worker's current pending_settlement as the total (this is accumulated from all attendance)
    current_pending = worker.get('pending_settlement', 0)
    
    # ✅ SAFETY CHECK: Verify settlement amount doesn't exceed pending by too much
    # Allow some flexibility (can settle up to 10% more than pending for adjustments)
    if settlement_data.amount_paid > current_pending * 1.1 and current_pending > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Settlement amount (₹{settlement_data.amount_paid:.2f}) exceeds pending wage (₹{current_pending:.2f}) by too much. Please verify the amount."
        )
    
    # Get ALL unsettled attendance (we settle everything pending, not just a date range)
    attendances = await db.worker_attendance.find({
        "contractor_id": current_user.id,
        "worker_id": settlement_data.worker_id
    }, {"_id": 0}).to_list(10000)
    
    # Calculate total from attendance
    days_worked = len([a for a in attendances if a['status'] in ['Present', 'Late']])
    total_wage_from_attendance = sum(a['wage_earned'] for a in attendances)
    
    # If extra charge is provided, create it
    if settlement_data.extra_charge_amount and settlement_data.extra_charge_amount > 0:
        extra_charge = {
            "id": str(uuid.uuid4()),
            "contractor_id": current_user.id,
            "worker_id": settlement_data.worker_id,
            "amount": settlement_data.extra_charge_amount,
            "reason": sanitize_input(settlement_data.extra_charge_reason) if settlement_data.extra_charge_reason else "Deduction during settlement",
            "date": datetime.now(timezone.utc).strftime("%d-%m-%Y"),
            "is_settled": False,
            "settled_at": None
        }
        await db.extra_charges.insert_one(extra_charge)
    
    settlement = WageSettlement(
        contractor_id=current_user.id,
        worker_id=settlement_data.worker_id,
        start_date=settlement_data.start_date,
        end_date=settlement_data.end_date,
        days_worked=days_worked,
        total_wage=current_pending,  # Use current pending as total wage
        advance_deducted=settlement_data.advance_deducted,
        charges_deducted=settlement_data.charges_deducted,
        amount_paid=settlement_data.amount_paid,
        settlement_type=settlement_data.settlement_type
    )
    
    settlement_dict = settlement.model_dump()
    settlement_dict['settlement_date'] = settlement_dict['settlement_date'].isoformat()
    await db.wage_settlements.insert_one(settlement_dict)
    
    # ✅ Mark all unsettled extra charges for this worker as settled
    if settlement_data.charges_deducted > 0:
        await db.extra_charges.update_many(
            {
                "contractor_id": current_user.id,
                "worker_id": settlement_data.worker_id,
                "is_settled": {"$ne": True}  # Only update unsettled charges
            },
            {
                "$set": {
                    "is_settled": True,
                    "settled_at": settlement.id
                }
            }
        )
    
    # FIX: Deduct only the amount being settled (for partial payments)
    # Calculate the total deduction (amount paid + advances used + charges)
    total_deduction = settlement_data.amount_paid + settlement_data.advance_deducted + settlement_data.charges_deducted
    
    await db.workers.update_one(
        {"id": settlement_data.worker_id},
        {
            "$inc": {
                "pending_settlement": -total_deduction,  # Deduct what was actually settled
                "advance_paid": -settlement_data.advance_deducted  # Reduce advance by amount used
            }
        }
    )
    
    activity = Activity(
        contractor_id=current_user.id,
        type="wage_settled",
        description=f"Settled ₹{settlement_data.amount_paid} wage for {worker['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return settlement

@api_router.get("/payments/settlements")
async def get_wage_settlements(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    settlements = await db.wage_settlements.find({"contractor_id": current_user.id}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for settlement in settlements:
        if isinstance(settlement['settlement_date'], str):
            settlement['settlement_date'] = datetime.fromisoformat(settlement['settlement_date'])
    return settlements

@api_router.delete("/payments/collection/{payment_id}")
async def reverse_payment_collection(
    payment_id: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Reverse a payment collection - adds the amount back to employer's pending payment
    """
    # Find the payment record
    payment = await db.payment_collections.find_one({
        "id": payment_id,
        "contractor_id": current_user.id
    })
    
    if not payment:
        raise HTTPException(status_code=404, detail="Payment record not found")
    
    # Reverse the payment - add amount back to employer's pending payment
    await db.employers.update_one(
        {"id": payment['employer_id']},
        {"$inc": {"pending_payment": payment['amount']}}
    )
    
    # Delete the payment record
    await db.payment_collections.delete_one({"id": payment_id})
    
    # Log activity
    employer = await db.employers.find_one({"id": payment['employer_id']}, {"_id": 0})
    activity = Activity(
        contractor_id=current_user.id,
        type="payment_reversed",
        description=f"Reversed payment collection of ₹{payment['amount']} from {employer['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return {
        "message": "Payment reversed successfully",
        "amount": payment['amount'],
        "employer_id": payment['employer_id']
    }

@api_router.delete("/payments/settlement/{settlement_id}")
async def reverse_wage_settlement(
    settlement_id: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Reverse a wage settlement - adds the amount back to worker's pending settlement
    """
    # Find the settlement record
    settlement = await db.wage_settlements.find_one({
        "id": settlement_id,
        "contractor_id": current_user.id
    })
    
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement record not found")
    
    # Reverse the settlement - add back to worker's pending settlement and advance
    total_deduction = settlement['amount_paid'] + settlement['advance_deducted'] + settlement['charges_deducted']
    
    await db.workers.update_one(
        {"id": settlement['worker_id']},
        {
            "$inc": {
                "pending_settlement": total_deduction,  # Add back what was settled
                "advance_paid": settlement['advance_deducted']  # Add back advance that was used
            }
        }
    )
    
    # ✅ Unmark extra charges that were part of this settlement
    if settlement['charges_deducted'] > 0:
        await db.extra_charges.update_many(
            {
                "contractor_id": current_user.id,
                "worker_id": settlement['worker_id'],
                "settled_at": settlement_id
            },
            {
                "$set": {
                    "is_settled": False,
                    "settled_at": None
                }
            }
        )
    
    # Delete the settlement record
    await db.wage_settlements.delete_one({"id": settlement_id})
    
    # Log activity
    worker = await db.workers.find_one({"id": settlement['worker_id']}, {"_id": 0})
    activity = Activity(
        contractor_id=current_user.id,
        type="settlement_reversed",
        description=f"Reversed wage settlement of ₹{settlement['amount_paid']} for {worker['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return {
        "message": "Settlement reversed successfully",
        "amount": settlement['amount_paid'],
        "worker_id": settlement['worker_id']
    }

# ============ ADVANCE ROUTES ============

@api_router.post("/advances", response_model=Advance)
async def create_advance(advance_data: AdvanceCreate, current_user: User = Depends(get_active_subscription_user)):
    # Use provided date or default to today
    if advance_data.date:
        try:
            # Convert YYYY-MM-DD to datetime
            advance_date = datetime.strptime(advance_data.date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        advance_date = datetime.now(timezone.utc)
    
    advance = Advance(
        contractor_id=current_user.id,
        worker_id=advance_data.worker_id,
        amount=advance_data.amount,
        purpose=sanitize_input(advance_data.purpose) if advance_data.purpose else None,
        date=advance_date
    )
    
    advance_dict = advance.model_dump()
    advance_dict['date'] = advance_dict['date'].isoformat()
    await db.advances.insert_one(advance_dict)
    
    await db.workers.update_one(
        {"id": advance_data.worker_id},
        {"$inc": {"advance_paid": advance_data.amount}}
    )
    
    worker = await db.workers.find_one({"id": advance_data.worker_id}, {"_id": 0})
    activity = Activity(
        contractor_id=current_user.id,
        type="advance_given",
        description=f"Advanced ₹{advance_data.amount} to {worker['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return advance

@api_router.get("/advances")
async def get_advances(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    advances = await db.advances.find({"contractor_id": current_user.id}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for advance in advances:
        if isinstance(advance['date'], str):
            advance['date'] = datetime.fromisoformat(advance['date'])
    return advances

@api_router.delete("/advances/{advance_id}")
async def delete_advance(advance_id: str, current_user: User = Depends(get_active_subscription_user)):
    """
    Delete an advance payment.
    Prevents deletion if the advance has been reduced in any settlement.
    """
    # Find the advance record
    advance = await db.advances.find_one({
        "id": advance_id,
        "contractor_id": current_user.id
    })
    
    if not advance:
        raise HTTPException(status_code=404, detail="Advance record not found")
    
    # Get worker to check current advance balance
    worker = await db.workers.find_one({"id": advance['worker_id']}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Get all settlements for this worker to check if advance has been used
    settlements = await db.wage_settlements.find({
        "contractor_id": current_user.id,
        "worker_id": advance['worker_id']
    }, {"_id": 0}).to_list(1000)
    
    # Calculate total advance deducted in settlements
    total_advance_deducted = sum(s.get('advance_deducted', 0) for s in settlements)
    
    # Calculate total advances given
    all_advances = await db.advances.find({
        "contractor_id": current_user.id,
        "worker_id": advance['worker_id']
    }, {"_id": 0}).to_list(1000)
    total_advances_given = sum(a['amount'] for a in all_advances)
    
    # If we delete this advance, will it cause issues?
    # Total advances after deletion
    total_after_deletion = total_advances_given - advance['amount']
    
    # If total deducted is more than what would remain, prevent deletion
    if total_advance_deducted > total_after_deletion:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete this advance. ₹{total_advance_deducted} has already been deducted in settlements, but only ₹{total_after_deletion} would remain if deleted."
        )
    
    # Safe to delete - reverse the advance
    await db.workers.update_one(
        {"id": advance['worker_id']},
        {"$inc": {"advance_paid": -advance['amount']}}
    )
    
    # Delete the advance record
    await db.advances.delete_one({"id": advance_id})
    
    # Log activity
    activity = Activity(
        contractor_id=current_user.id,
        type="advance_deleted",
        description=f"Deleted advance of ₹{advance['amount']} for {worker['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return {
        "message": "Advance deleted successfully",
        "amount": advance['amount'],
        "worker_id": advance['worker_id']
    }

# ============ EMPLOYER ADVANCE ROUTES ============

@api_router.post("/employer-advances")
async def create_employer_advance(
    advance_data: dict,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Record advance payment received from an employer.
    This will be deducted from future payment collections.
    """
    employer_id = advance_data.get('employer_id')
    amount = float(advance_data.get('amount'))
    purpose = advance_data.get('purpose', '')
    date_str = advance_data.get('date')  # YYYY-MM-DD format
    
    if not employer_id or not amount:
        raise HTTPException(status_code=400, detail="Employer ID and amount are required")
    
    # Verify employer exists
    employer = await db.employers.find_one({
        "id": employer_id,
        "contractor_id": current_user.id
    }, {"_id": 0})
    
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    
    # Use provided date or default to today
    if date_str:
        try:
            advance_date = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    else:
        advance_date = datetime.now(timezone.utc)
    
    # Create employer advance record
    employer_advance = {
        "id": str(uuid.uuid4()),
        "contractor_id": current_user.id,
        "employer_id": employer_id,
        "amount": amount,
        "purpose": sanitize_input(purpose) if purpose else None,
        "date": advance_date.isoformat()
    }
    
    await db.employer_advances.insert_one(employer_advance)
    
    # ✅ FIX: When employer gives advance, it should reduce their pending payment
    # Advance is prepayment for future work, so it directly reduces what they owe
    await db.employers.update_one(
        {"id": employer_id},
        {"$inc": {
            "advance_received": amount,  # Track total advance received
            "pending_payment": -amount   # Reduce pending payment by advance amount
        }}
    )
    
    # Log activity
    activity = Activity(
        contractor_id=current_user.id,
        type="employer_advance_received",
        description=f"Received advance of ₹{amount} from {employer['name']} (reduced pending payment)"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    # Return the advance with proper format
    return {
        "id": employer_advance["id"],
        "contractor_id": employer_advance["contractor_id"],
        "employer_id": employer_advance["employer_id"],
        "amount": employer_advance["amount"],
        "purpose": employer_advance["purpose"],
        "date": employer_advance["date"]
    }

@api_router.get("/employer-advances")
async def get_employer_advances(current_user: User = Depends(get_active_subscription_user)):
    """Get all employer advances"""
    advances = await db.employer_advances.find({
        "contractor_id": current_user.id
    }, {"_id": 0}).to_list(1000)
    
    for advance in advances:
        if isinstance(advance.get('date'), str):
            try:
                advance['date'] = datetime.fromisoformat(advance['date'])
            except:
                advance['date'] = datetime.now(timezone.utc)
    
    return advances

@api_router.delete("/employer-advances/{advance_id}")
async def delete_employer_advance(
    advance_id: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Delete an employer advance payment.
    Prevents deletion if the advance has been deducted in any payment collection.
    """
    # Find the advance record
    advance = await db.employer_advances.find_one({
        "id": advance_id,
        "contractor_id": current_user.id
    })
    
    if not advance:
        raise HTTPException(status_code=404, detail="Employer advance record not found")
    
    # Get employer to check current advance balance
    employer = await db.employers.find_one({"id": advance['employer_id']}, {"_id": 0})
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    
    # Get all payment collections for this employer to check if advance has been used
    collections = await db.payment_collections.find({
        "contractor_id": current_user.id,
        "employer_id": advance['employer_id']
    }, {"_id": 0}).to_list(1000)
    
    # Calculate total advance deducted in collections (if we add this feature)
    total_advance_deducted = sum(c.get('advance_deducted', 0) for c in collections)
    
    # Calculate total advances received
    all_advances = await db.employer_advances.find({
        "contractor_id": current_user.id,
        "employer_id": advance['employer_id']
    }, {"_id": 0}).to_list(1000)
    total_advances_received = sum(a['amount'] for a in all_advances)
    
    # Total advances after deletion
    total_after_deletion = total_advances_received - advance['amount']
    
    # If total deducted is more than what would remain, prevent deletion
    if total_advance_deducted > total_after_deletion:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete this advance. ₹{total_advance_deducted} has already been deducted in payment collections, but only ₹{total_after_deletion} would remain if deleted."
        )
    
    # Safe to delete - reverse the advance and restore pending payment
    await db.employers.update_one(
        {"id": advance['employer_id']},
        {"$inc": {
            "advance_received": -advance['amount'],
            "pending_payment": advance['amount']  # Restore pending payment
        }}
    )
    
    # Delete the advance record
    await db.employer_advances.delete_one({"id": advance_id})
    
    # Log activity
    activity = Activity(
        contractor_id=current_user.id,
        type="employer_advance_deleted",
        description=f"Deleted employer advance of ₹{advance['amount']} from {employer['name']} (restored pending payment)"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return {
        "message": "Employer advance deleted successfully",
        "amount": advance['amount'],
        "employer_id": advance['employer_id']
    }

# ============ EXTRA CHARGES ROUTES ============

@api_router.post("/extra-charges", response_model=ExtraCharge)
async def create_extra_charge(charge_data: WorkerCharge, current_user: User = Depends(get_active_subscription_user)):
    charge_dict = {
        "id": str(uuid.uuid4()),
        "contractor_id": current_user.id,
        "worker_id": charge_data.worker_id,
        "amount": charge_data.amount,
        "reason": sanitize_input(charge_data.reason),
        "date": charge_data.date if charge_data.date else datetime.now(timezone.utc).strftime("%d-%m-%Y"),
        "is_settled": False,
        "settled_at": None
    }
    
    await db.extra_charges.insert_one(charge_dict)
    
    worker = await db.workers.find_one({"id": charge_data.worker_id}, {"_id": 0})
    if worker:
        activity = Activity(
            contractor_id=current_user.id,
            type="extra_charge_added",
            description=f"Added ₹{charge_data.amount} extra charge for {worker['name']}: {charge_data.reason}"
        )
        activity_dict = activity.model_dump()
        activity_dict['date'] = activity_dict['date'].isoformat()
        await db.activities.insert_one(activity_dict)
    
    return ExtraCharge(**charge_dict)

@api_router.get("/extra-charges")
async def get_extra_charges(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_active_subscription_user)
):
    # Limit maximum page size to prevent memory issues
    limit = min(limit, 500)
    charges = await db.extra_charges.find({"contractor_id": current_user.id}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return charges

@api_router.delete("/extra-charges/{charge_id}")
async def delete_extra_charge(charge_id: str, current_user: User = Depends(get_active_subscription_user)):
    charge = await db.extra_charges.find_one({"id": charge_id, "contractor_id": current_user.id}, {"_id": 0})
    if not charge:
        raise HTTPException(status_code=404, detail="Extra charge not found")
    
    # ✅ SECURITY: Prevent deletion if charge has been deducted in a settlement
    if charge.get("is_settled", False):
        raise HTTPException(
            status_code=400,
            detail="Cannot delete extra charge that has already been deducted in a wage settlement. Please reverse the settlement first."
        )
    
    await db.extra_charges.delete_one({"id": charge_id})
    
    # Log activity
    worker = await db.workers.find_one({"id": charge['worker_id']}, {"_id": 0})
    if worker:
        activity = Activity(
            contractor_id=current_user.id,
            type="extra_charge_deleted",
            description=f"Deleted extra charge of ₹{charge['amount']} for {worker['name']}: {charge.get('reason', 'N/A')}"
        )
        activity_dict = activity.model_dump()
        activity_dict['date'] = activity_dict['date'].isoformat()
        await db.activities.insert_one(activity_dict)
    
    return {"message": "Extra charge deleted successfully"}

@api_router.post("/extra-charges/split-by-room")
async def split_extra_charge_by_room(
    split_data: dict,
    current_user: User = Depends(get_active_subscription_user)
):
    """Split an extra charge equally among all workers in a room"""
    room_id = split_data.get('room_id')
    total_amount = split_data.get('total_amount')
    reason = split_data.get('reason')
    date = split_data.get('date', datetime.now(timezone.utc).strftime("%d-%m-%Y"))
    
    if not room_id or not total_amount or not reason:
        raise HTTPException(status_code=400, detail="Missing required fields: room_id, total_amount, reason")
    
    # Verify room exists and belongs to contractor
    room = await db.rooms.find_one({"id": room_id, "contractor_id": current_user.id}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Room not found")
    
    # Get all active workers in this room
    workers_in_room = await db.workers.find({
        "contractor_id": current_user.id,
        "room_id": room_id,
        "status": "Active"
    }, {"_id": 0}).to_list(1000)
    
    if not workers_in_room:
        raise HTTPException(status_code=400, detail="No active workers found in this room")
    
    # Calculate amount per worker
    num_workers = len(workers_in_room)
    amount_per_worker = round(total_amount / num_workers, 2)
    
    # Create extra charges for each worker
    created_charges = []
    for worker in workers_in_room:
        charge_dict = {
            "id": str(uuid.uuid4()),
            "contractor_id": current_user.id,
            "worker_id": worker['id'],
            "amount": amount_per_worker,
            "reason": sanitize_input(f"{reason} (Split among {num_workers} workers in {room['name']})"),
            "date": date,
            "is_settled": False,
            "settled_at": None
        }
        
        await db.extra_charges.insert_one(charge_dict)
        created_charges.append({
            "worker_id": worker['id'],
            "worker_name": worker['name'],
            "amount": amount_per_worker
        })
    
    # Log activity
    activity = Activity(
        contractor_id=current_user.id,
        type="extra_charge_split",
        description=f"Split ₹{total_amount} charge ({reason}) among {num_workers} workers in room {room['name']}"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)
    
    return {
        "message": f"Successfully split ₹{total_amount} among {num_workers} workers",
        "room_name": room['name'],
        "total_amount": total_amount,
        "amount_per_worker": amount_per_worker,
        "num_workers": num_workers,
        "charges_created": created_charges
    }

# ============ WORKER SETTLEMENT CALCULATION ============

@api_router.get("/payments/worker-settlement-details/{worker_id}")
async def get_worker_settlement_details(
    worker_id: str,
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Calculate settlement details for a worker including:
    - Days worked and total wage earned
    - Current advance balance
    - Extra charges in the period
    """
    # Get worker
    worker = await db.workers.find_one({"id": worker_id, "contractor_id": current_user.id}, {"_id": 0})
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Get attendance records for the period
    attendances = await db.worker_attendance.find({
        "contractor_id": current_user.id,
        "worker_id": worker_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(1000)
    
    days_worked = len([a for a in attendances if a['status'] in ['Present', 'Late']])
    total_wage = sum(a['wage_earned'] for a in attendances)
    
    # Get unsettled extra charges for the worker
    extra_charges = await db.extra_charges.find({
        "contractor_id": current_user.id,
        "worker_id": worker_id,
        "is_settled": {"$ne": True}  # Only get unsettled charges
    }, {"_id": 0}).to_list(1000)
    
    total_extra_charges = sum(charge['amount'] for charge in extra_charges)
    
    return {
        "worker_id": worker_id,
        "worker_name": worker['name'],
        "start_date": start_date,
        "end_date": end_date,
        "days_worked": days_worked,
        "total_wage_earned": total_wage,
        "advance_balance": worker.get('advance_paid', 0),
        "extra_charges": extra_charges,
        "total_extra_charges": total_extra_charges,
        "net_payable": total_wage - worker.get('advance_paid', 0) - total_extra_charges
    }

# ============ PAYMENT SUMMARY ROUTES ============

@api_router.get("/payments/employer-summaries")
async def get_employer_payment_summaries(
    search: Optional[str] = None,
    filter_status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Get all employers with their payment summaries
    - Total amount pending (filtered by date if provided)
    - Total amount collected (filtered by date if provided)
    - Payment status
    """
    query = {"contractor_id": current_user.id, "status": "Active"}
    
    employers = await db.employers.find(query, {"_id": 0}).to_list(1000)
    
    # Filter by search
    if search:
        employers = [e for e in employers if search.lower() in e['name'].lower()]
    
    summaries = []
    for employer in employers:
        if start_date and end_date:
            # Date-filtered mode: calculate amounts for the specific date range
            # Get employer attendance records within date range
            new_attendance = await db.employer_attendance.find({
                "contractor_id": current_user.id,
                "employer_id": employer['id']
            }, {"_id": 0}).to_list(10000)
            
            old_attendance = await db.attendance.find({
                "contractor_id": current_user.id,
                "employer_id": employer['id'],
                "mode": "employer"
            }, {"_id": 0}).to_list(10000)
            
            all_attendance = new_attendance + old_attendance
            
            # Filter by date range
            filtered_attendance = []
            for att in all_attendance:
                try:
                    att_date = datetime.strptime(att['date'], "%d-%m-%Y")
                    start = datetime.strptime(start_date, "%d-%m-%Y")
                    end = datetime.strptime(end_date, "%d-%m-%Y")
                    if start <= att_date <= end:
                        filtered_attendance.append(att)
                except:
                    continue
            
            # Calculate total work amount in date range
            # Use payment_per_worker if available, otherwise use wage_amount
            total_work_amount = 0
            for att in filtered_attendance:
                if att.get('payment_per_worker') and att.get('payment_per_worker', 0) > 0:
                    # Use payment_per_worker * workers_count + additional_charges
                    workers_count = att.get('workers_count', 0)
                    additional_charges = sum(charge.get('amount', 0) for charge in att.get('additional_charges', []))
                    total_work_amount += (att.get('payment_per_worker', 0) * workers_count) + additional_charges
                else:
                    # Fallback to wage_amount or total_amount
                    total_work_amount += att.get('total_amount', att.get('wage_amount', 0))
            
            # Get payments within date range
            all_payments = await db.payment_collections.find({
                "contractor_id": current_user.id,
                "employer_id": employer['id']
            }, {"_id": 0}).to_list(1000)
            
            filtered_payments = []
            for payment in all_payments:
                try:
                    payment_date = payment['date']
                    if isinstance(payment_date, str):
                        payment_date = datetime.fromisoformat(payment_date)
                    # Normalize to remove time component
                    payment_date = payment_date.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
                    start = datetime.strptime(start_date, "%d-%m-%Y").replace(hour=0, minute=0, second=0)
                    end = datetime.strptime(end_date, "%d-%m-%Y").replace(hour=23, minute=59, second=59)
                    if start <= payment_date <= end:
                        filtered_payments.append(payment)
                except Exception as e:
                    print(f"Error filtering payment date: {e}, payment: {payment}")
                    continue
            
            total_collected = sum(p['amount'] for p in filtered_payments)
            pending_payment = total_work_amount - total_collected
        else:
            # No date filter: use current pending/collected amounts
            payments = await db.payment_collections.find({
                "contractor_id": current_user.id,
                "employer_id": employer['id']
            }, {"_id": 0}).to_list(1000)
            
            total_collected = sum(p['amount'] for p in payments)
            pending_payment = employer.get('pending_payment', 0)
        
        # Determine status
        status = "Paid" if pending_payment == 0 else "Pending"
        
        summaries.append({
            "employer_id": employer['id'],
            "employer_name": employer['name'],
            "total_pending": pending_payment,
            "total_collected": total_collected,
            "advance_received": employer.get('advance_received', 0),
            "status": status,
            "phone_number": employer.get('phone_number', '')
        })
    
    # Filter by status
    if filter_status and filter_status != "all":
        summaries = [s for s in summaries if s['status'].lower() == filter_status.lower()]
    
    # Sort alphabetically by name
    summaries.sort(key=lambda x: x['employer_name'].lower())
    
    return summaries

@api_router.get("/payments/employer-history/{employer_id}")
async def get_employer_payment_history(
    employer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Get payment history for a specific employer
    Shows all payments with running balance
    """
    query = {
        "contractor_id": current_user.id,
        "employer_id": employer_id
    }
    
    payments = await db.payment_collections.find(query, {"_id": 0}).to_list(1000)
    
    # Convert dates and sort by date
    for payment in payments:
        if isinstance(payment['date'], str):
            try:
                payment['date'] = datetime.fromisoformat(payment['date'])
            except:
                # Handle edge cases
                payment['date'] = datetime.now()
    
    payments.sort(key=lambda x: x['date'])
    
    # Filter by date range if provided
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%d-%m-%Y").replace(hour=0, minute=0, second=0)
            end_dt = datetime.strptime(end_date, "%d-%m-%Y").replace(hour=23, minute=59, second=59)
            filtered_payments = []
            for p in payments:
                # Normalize payment date to remove timezone info if present
                pay_date = p['date'].replace(tzinfo=None) if p['date'].tzinfo else p['date']
                if start_dt <= pay_date <= end_dt:
                    filtered_payments.append(p)
            payments = filtered_payments
        except Exception as e:
            print(f"Error filtering by date: {e}")
            # If date filtering fails, return all payments
            pass
    
    # Get employer initial pending
    employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
    
    # Calculate running balance
    history = []
    for payment in payments:
        history.append({
            "id": payment['id'],
            "date": payment['date'].isoformat(),
            "amount_paid": payment['amount'],
            "advance_deducted": payment.get('advance_deducted', 0),
            "total_settled": payment['amount'] + payment.get('advance_deducted', 0),
            "payment_mode": payment['payment_mode'],
            "remarks": payment.get('remarks', ''),
        })
    
    return {
        "employer_name": employer['name'],
        "current_pending": employer.get('pending_payment', 0),
        "current_advance": employer.get('advance_received', 0),
        "history": history
    }

@api_router.get("/payments/worker-summaries")
async def get_worker_wage_summaries(
    search: Optional[str] = None,
    filter_status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Get all workers with their wage summaries
    - Total pending wage (after deductions, filtered by date if provided)
    - Total settled (filtered by date if provided)
    - Settlement status
    """
    query = {"contractor_id": current_user.id, "status": "Active"}
    
    workers = await db.workers.find(query, {"_id": 0}).to_list(1000)
    
    # Filter by search
    if search:
        workers = [w for w in workers if search.lower() in w['name'].lower()]
    
    summaries = []
    for worker in workers:
        if start_date and end_date:
            # Date-filtered mode: calculate amounts for the specific date range
            # Get worker attendance records within date range
            new_attendance = await db.worker_attendance.find({
                "contractor_id": current_user.id,
                "worker_id": worker['id']
            }, {"_id": 0}).to_list(10000)
            
            old_attendance = await db.attendance.find({
                "contractor_id": current_user.id,
                "worker_id": worker['id'],
                "mode": "worker"
            }, {"_id": 0}).to_list(10000)
            
            all_attendance = new_attendance + old_attendance
            
            # Filter by date range
            filtered_attendance = []
            for att in all_attendance:
                try:
                    att_date = datetime.strptime(att['date'], "%d-%m-%Y")
                    start = datetime.strptime(start_date, "%d-%m-%Y")
                    end = datetime.strptime(end_date, "%d-%m-%Y")
                    if start <= att_date <= end:
                        filtered_attendance.append(att)
                except:
                    continue
            
            # Calculate total wage earned in date range
            pending_settlement = sum(att.get('wage_earned', att.get('wage_amount', 0)) for att in filtered_attendance)
            
            # Get settlements within date range
            all_settlements = await db.wage_settlements.find({
                "contractor_id": current_user.id,
                "worker_id": worker['id']
            }, {"_id": 0}).to_list(1000)
            
            filtered_settlements = []
            for settlement in all_settlements:
                try:
                    settlement_date = settlement['settlement_date']
                    if isinstance(settlement_date, str):
                        settlement_date = datetime.fromisoformat(settlement_date)
                    # Normalize to remove time component
                    settlement_date = settlement_date.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
                    start = datetime.strptime(start_date, "%d-%m-%Y").replace(hour=0, minute=0, second=0)
                    end = datetime.strptime(end_date, "%d-%m-%Y").replace(hour=23, minute=59, second=59)
                    if start <= settlement_date <= end:
                        filtered_settlements.append(settlement)
                except Exception as e:
                    print(f"Error filtering settlement date: {e}, settlement: {settlement}")
                    continue
            
            total_settled = sum(s['amount_paid'] for s in filtered_settlements)
            
            # For date-filtered view, advance and extra charges are not date-specific
            advance_balance = worker.get('advance_paid', 0)
            extra_charges_list = await db.extra_charges.find({
                "contractor_id": current_user.id,
                "worker_id": worker['id']
            }, {"_id": 0}).to_list(1000)
            total_extra_charges = sum(c['amount'] for c in extra_charges_list)
        else:
            # No date filter: use current pending/settled amounts
            pending_settlement = worker.get('pending_settlement', 0)
            advance_balance = worker.get('advance_paid', 0)
            
            # Get extra charges
            extra_charges = await db.extra_charges.find({
                "contractor_id": current_user.id,
                "worker_id": worker['id']
            }, {"_id": 0}).to_list(1000)
            total_extra_charges = sum(c['amount'] for c in extra_charges)
            
            # Get all settlements
            settlements = await db.wage_settlements.find({
                "contractor_id": current_user.id,
                "worker_id": worker['id']
            }, {"_id": 0}).to_list(1000)
            
            total_settled = sum(s['amount_paid'] for s in settlements)
        
        # Calculate net pending (after deductions)
        net_pending = pending_settlement - advance_balance - total_extra_charges
        
        # Determine status
        status = "Settled" if net_pending <= 0 else "Pending"
        
        summaries.append({
            "worker_id": worker['id'],
            "worker_name": worker['name'],
            "pending_wage": pending_settlement,
            "advance_balance": advance_balance,
            "extra_charges": total_extra_charges,
            "net_pending": max(0, net_pending),
            "total_settled": total_settled,
            "status": status,
            "phone_number": worker.get('phone_number', '')
        })
    
    # Filter by status
    if filter_status and filter_status != "all":
        summaries = [s for s in summaries if s['status'].lower() == filter_status.lower()]
    
    # Sort alphabetically by name
    summaries.sort(key=lambda x: x['worker_name'].lower())
    
    return summaries

@api_router.get("/payments/worker-history/{worker_id}")
async def get_worker_settlement_history(
    worker_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Get settlement history for a specific worker
    Shows all settlements with running balance
    """
    query = {
        "contractor_id": current_user.id,
        "worker_id": worker_id
    }
    
    settlements = await db.wage_settlements.find(query, {"_id": 0}).to_list(1000)
    
    # Convert dates and sort
    for settlement in settlements:
        if isinstance(settlement['settlement_date'], str):
            try:
                settlement['settlement_date'] = datetime.fromisoformat(settlement['settlement_date'])
            except:
                settlement['settlement_date'] = datetime.now()
    
    settlements.sort(key=lambda x: x['settlement_date'])
    
    # Filter by date range if provided
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%d-%m-%Y").replace(hour=0, minute=0, second=0)
            end_dt = datetime.strptime(end_date, "%d-%m-%Y").replace(hour=23, minute=59, second=59)
            filtered_settlements = []
            for s in settlements:
                # Normalize settlement date to remove timezone info if present
                settle_date = s['settlement_date'].replace(tzinfo=None) if s['settlement_date'].tzinfo else s['settlement_date']
                if start_dt <= settle_date <= end_dt:
                    filtered_settlements.append(s)
            settlements = filtered_settlements
        except Exception as e:
            print(f"Error filtering by date: {e}")
            # If date filtering fails, return all settlements
            pass
    
    # Get worker
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    
    # Calculate running balance through history
    # Start with all settlements to calculate proper running balance
    all_settlements = await db.wage_settlements.find({
        "contractor_id": current_user.id,
        "worker_id": worker_id
    }, {"_id": 0}).to_list(1000)
    
    for s in all_settlements:
        if isinstance(s['settlement_date'], str):
            s['settlement_date'] = datetime.fromisoformat(s['settlement_date'])
    all_settlements.sort(key=lambda x: x['settlement_date'])
    
    # Build history with running balance
    history = []
    for settlement in settlements:
        # Find index in all settlements to calculate balance after this settlement
        idx = next((i for i, s in enumerate(all_settlements) if s['id'] == settlement['id']), -1)
        
        # Calculate pending after this settlement
        # Sum all settlements up to and including this one
        settled_so_far = sum(s['amount_paid'] for s in all_settlements[:idx+1])
        pending_after = worker.get('pending_settlement', 0)  # This is current pending
        
        history.append({
            "id": settlement['id'],
            "date": settlement['settlement_date'].isoformat(),
            "amount_settled": settlement['amount_paid'],
            "advance_deducted": settlement['advance_deducted'],
            "charges_deducted": settlement['charges_deducted'],
            "settlement_type": settlement['settlement_type'],
            "balance_after": pending_after  # Current pending (will be shown in UI)
        })
    
    return {
        "worker_name": worker['name'],
        "current_pending": worker.get('pending_settlement', 0),
        "current_advance": worker.get('advance_paid', 0),
        "history": history
    }

@api_router.get("/payments/export-employer-history/{employer_id}")
async def export_employer_history(
    employer_id: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """Export employer payment history as beautiful HTML statement"""
    history_data = await get_employer_payment_history(employer_id, None, None, current_user)
    
    # Get employer details
    employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
    
    # Calculate totals
    total_collected = sum(p['amount_paid'] for p in history_data['history'])
    total_advance_deducted = sum(p.get('advance_deducted', 0) for p in history_data['history'])
    total_settled = sum(p.get('total_settled', p['amount_paid']) for p in history_data['history'])
    total_payments = len(history_data['history'])
    
    # Generate HTML
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Payment Statement - {history_data['employer_name']}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; background-color: #f9fafb; position: relative; }}
            .watermark {{ position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-45deg); font-size: 120px; font-weight: bold; color: rgba(0, 0, 0, 0.03); z-index: 0; white-space: nowrap; pointer-events: none; }}
            .container {{ max-width: 1000px; margin: 0 auto; background: white; padding: 40px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); position: relative; z-index: 1; }}
            .logo-container {{ text-align: center; margin-bottom: 20px; padding-bottom: 15px; border-bottom: 2px solid #e5e7eb; }}
            .header {{ text-align: center; margin-bottom: 30px; border-bottom: 3px solid #1e3a8a; padding-bottom: 20px; }}
            .header h1 {{ margin: 0; color: #1e40af; font-size: 32px; }}
            .header p {{ margin: 5px 0; color: #6b7280; }}
            .info-section {{ display: flex; justify-content: space-between; margin-bottom: 30px; background: #eff6ff; padding: 20px; border-radius: 8px; }}
            .info-box {{ width: 48%; }}
            .info-box h3 {{ margin: 0 0 10px 0; color: #1e40af; border-bottom: 2px solid #2563eb; padding-bottom: 5px; }}
            .info-box p {{ margin: 8px 0; color: #374151; }}
            .info-label {{ font-weight: 600; color: #6b7280; font-size: 14px; }}
            .info-value {{ font-weight: 700; color: #111827; font-size: 16px; }}
            
            .summary-cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 30px; }}
            .summary-card {{ background: linear-gradient(135deg, #2563eb 0%, #1e3a8a 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
            .summary-card h4 {{ margin: 0 0 5px 0; font-size: 14px; opacity: 0.9; }}
            .summary-card p {{ margin: 0; font-size: 24px; font-weight: bold; }}
            
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
            th {{ background: linear-gradient(135deg, #1e3a8a 0%, #1e40af 100%); color: white; padding: 14px; text-align: left; font-size: 14px; text-transform: uppercase; letter-spacing: 0.5px; }}
            td {{ padding: 12px; border-bottom: 1px solid #e5e7eb; color: #374151; }}
            tr:hover {{ background-color: #eff6ff; }}
            
            .payment-mode {{ padding: 4px 12px; border-radius: 12px; font-size: 12px; font-weight: 600; display: inline-block; }}
            .mode-cash {{ background-color: #dcfce7; color: #15803d; }}
            .mode-upi {{ background-color: #dbeafe; color: #1e40af; }}
            .mode-bank {{ background-color: #fef3c7; color: #92400e; }}
            .mode-cheque {{ background-color: #f3e8ff; color: #6b21a8; }}
            
            .total-section {{ margin-top: 30px; background: #eff6ff; padding: 20px; border-radius: 8px; border: 2px solid #93c5fd; }}
            .total-row {{ display: flex; justify-content: space-between; font-size: 16px; margin: 8px 0; padding: 8px 0; }}
            .total-row.grand-total {{ font-size: 28px; font-weight: bold; color: #1e3a8a; border-top: 3px solid #2563eb; padding-top: 15px; margin-top: 15px; }}
            
            .status-box {{ background: #fef2f2; border-left: 4px solid #ef4444; padding: 15px; margin-top: 20px; }}
            .status-box.settled {{ background: #f0fdf4; border-left-color: #10b981; }}
            
            .footer {{ margin-top: 50px; text-align: center; color: #6b7280; border-top: 2px solid #e5e7eb; padding-top: 20px; }}
            
            @media print {{
                body {{ margin: 0; background: white; }}
                .container {{ box-shadow: none; }}
                .no-print {{ display: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="watermark">GUESTWORKER</div>
        <div class="container">
            <div class="logo-container">
                <div style="font-size: 28px; background: linear-gradient(to right, #4f46e5, #9333ea); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; font-weight: 800; letter-spacing: 1px;">GuestWorker</div>
            </div>
            <div class="header">
                <h1>PAYMENT STATEMENT</h1>
                <p style="font-size: 18px; color: #1e40af; font-weight: 600;">{history_data['employer_name']}</p>
                <p><strong>Contractor:</strong> {current_user.name}</p>
                <p><strong>Generated On:</strong> {datetime.now().strftime("%d-%m-%Y %H:%M")}</p>
            </div>
            
            <div class="info-section">
                <div class="info-box">
                    <h3>Employer Information</h3>
                    <p><span class="info-label">Name:</span> <span class="info-value">{history_data['employer_name']}</span></p>
                    {"<p><span class='info-label'>Company:</span> <span class='info-value'>" + employer.get('company', 'N/A') + "</span></p>" if employer.get('company') else ""}
                    {"<p><span class='info-label'>Phone:</span> <span class='info-value'>" + employer.get('phone_number', 'N/A') + "</span></p>" if employer.get('phone_number') else ""}
                </div>
                <div class="info-box">
                    <h3>Statement Summary</h3>
                    <p><span class="info-label">Total Payments Made:</span> <span class="info-value">{total_payments}</span></p>
                    <p><span class="info-label">Total Amount Collected:</span> <span class="info-value">₹{total_collected:,.2f}</span></p>
                    <p><span class="info-label">Current Pending:</span> <parameter name="info-value" style="color: #2563eb;">₹{history_data['current_pending']:,.2f}</span></p>
                </div>
            </div>
            
            <div class="summary-cards">
                <div class="summary-card">
                    <h4>Total Payments</h4>
                    <p>{total_payments}</p>
                </div>
                <div class="summary-card" style="background: linear-gradient(135deg, #10b981 0%, #059669 100%);">
                    <h4>Cash Collected</h4>
                    <p>₹{total_collected:,.2f}</p>
                </div>
                <div class="summary-card" style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);">
                    <h4>Advance Deducted</h4>
                    <p>₹{total_advance_deducted:,.2f}</p>
                </div>
                <div class="summary-card" style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);">
                    <h4>Pending Balance</h4>
                    <p>₹{history_data['current_pending']:,.2f}</p>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Cash Collected</th>
                        <th>Advance Deducted</th>
                        <th>Total Settled</th>
                        <th>Mode</th>
                        <th>Remarks</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for payment in history_data['history']:
        payment_date = datetime.fromisoformat(payment['date']).strftime("%d-%m-%Y")
        mode_class = f"mode-{payment['payment_mode'].lower()}"
        advance_ded = payment.get('advance_deducted', 0)
        total_set = payment.get('total_settled', payment['amount_paid'])
        
        html_content += f"""
                    <tr>
                        <td style="font-weight: 600;">{payment_date}</td>
                        <td style="font-weight: 700; color: #059669; font-size: 16px;">₹{payment['amount_paid']:,.2f}</td>
                        <td style="font-weight: 600; color: #d97706;">{f"₹{advance_ded:,.2f}" if advance_ded > 0 else "-"}</td>
                        <td style="font-weight: 700; color: #1e40af; font-size: 16px;">₹{total_set:,.2f}</td>
                        <td><span class="payment-mode {mode_class}">{payment['payment_mode']}</span></td>
                        <td><i style="color: #6b7280;">{payment['remarks'] or '-'}</i></td>
                    </tr>
        """
    
    status_class = 'settled' if history_data['current_pending'] <= 0 else ''
    status_message = '✓ Account Fully Settled' if history_data['current_pending'] <= 0 else f'⚠ Outstanding Balance: ₹{history_data["current_pending"]:,.2f}'
    
    html_content += f"""
                </tbody>
            </table>
            
            <div class="total-section">
                <div class="total-row">
                    <span>Total Number of Payments:</span>
                    <strong>{total_payments} transactions</strong>
                </div>
                <div class="total-row grand-total">
                    <span>TOTAL AMOUNT COLLECTED:</span>
                    <span>₹{total_collected:,.2f}</span>
                </div>
            </div>
            
            <div class="status-box {status_class}">
                <p style="margin: 0; font-size: 16px; font-weight: 600;">{status_message}</p>
            </div>
            
            <div class="footer">
                <p style="font-size: 16px; font-weight: 600; color: #f97316;">Thank you for your business!</p>
                <p><small>This is a computer-generated statement. Generated on {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}</small></p>
            </div>
            
            <div class="no-print" style="text-align: center; margin-top: 30px;">
                <button onclick="window.print()" style="padding: 12px 40px; font-size: 16px; cursor: pointer; background: linear-gradient(135deg, #1e3a8a 0%, #1e40af 100%); color: white; border: none; border-radius: 8px; font-weight: 600; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    🖨️ Print Statement
                </button>
            </div>
        </div>
    </body>
    </html>
    """
    
    return HTMLResponse(content=html_content)

@api_router.get("/payments/export-worker-history/{worker_id}")
async def export_worker_history(
    worker_id: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """Export worker settlement history as beautiful HTML statement"""
    history_data = await get_worker_settlement_history(worker_id, None, None, current_user)
    
    # Get worker details
    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
    
    # Calculate totals
    total_settled = sum(s['amount_settled'] for s in history_data['history'])
    total_advance_deducted = sum(s['advance_deducted'] for s in history_data['history'])
    total_charges_deducted = sum(s['charges_deducted'] for s in history_data['history'])
    total_settlements = len(history_data['history'])
    
    # Generate HTML
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Settlement Statement - {history_data['worker_name']}</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; background-color: #f9fafb; position: relative; }}
            .watermark {{ position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-45deg); font-size: 120px; font-weight: bold; color: rgba(0, 0, 0, 0.03); z-index: 0; white-space: nowrap; pointer-events: none; }}
            .container {{ max-width: 1000px; margin: 0 auto; background: white; padding: 40px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); position: relative; z-index: 1; }}
            .logo-container {{ text-align: center; margin-bottom: 20px; padding-bottom: 15px; border-bottom: 2px solid #e5e7eb; }}
            .header {{ text-align: center; margin-bottom: 30px; border-bottom: 3px solid #10b981; padding-bottom: 20px; }}
            .header h1 {{ margin: 0; color: #059669; font-size: 32px; }}
            .header p {{ margin: 5px 0; color: #6b7280; }}
            .info-section {{ display: flex; justify-content: space-between; margin-bottom: 30px; background: #f0fdf4; padding: 20px; border-radius: 8px; }}
            .info-box {{ width: 48%; }}
            .info-box h3 {{ margin: 0 0 10px 0; color: #059669; border-bottom: 2px solid #10b981; padding-bottom: 5px; }}
            .info-box p {{ margin: 8px 0; color: #374151; }}
            .info-label {{ font-weight: 600; color: #6b7280; font-size: 14px; }}
            .info-value {{ font-weight: 700; color: #111827; font-size: 16px; }}
            
            .summary-cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 30px; }}
            .summary-card {{ background: linear-gradient(135deg, #10b981 0%, #14b8a6 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
            .summary-card h4 {{ margin: 0 0 5px 0; font-size: 13px; opacity: 0.9; }}
            .summary-card p {{ margin: 0; font-size: 22px; font-weight: bold; }}
            
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
            th {{ background: linear-gradient(135deg, #047857 0%, #14b8a6 100%); color: white; padding: 14px; text-align: left; font-size: 13px; text-transform: uppercase; letter-spacing: 0.5px; }}
            td {{ padding: 12px; border-bottom: 1px solid #e5e7eb; color: #374151; }}
            tr:hover {{ background-color: #f0fdf4; }}
            
            .settlement-type {{ padding: 4px 12px; border-radius: 12px; font-size: 12px; font-weight: 600; display: inline-block; }}
            .type-full {{ background-color: #dcfce7; color: #15803d; }}
            .type-partial {{ background-color: #fef3c7; color: #92400e; }}
            
            .deduction {{ color: #dc2626; font-weight: 600; }}
            
            .total-section {{ margin-top: 30px; background: #f0fdf4; padding: 20px; border-radius: 8px; border: 2px solid #bbf7d0; }}
            .total-row {{ display: flex; justify-content: space-between; font-size: 16px; margin: 8px 0; padding: 8px 0; }}
            .total-row.grand-total {{ font-size: 28px; font-weight: bold; color: #059669; border-top: 3px solid #10b981; padding-top: 15px; margin-top: 15px; }}
            
            .status-box {{ background: #fef2f2; border-left: 4px solid #ef4444; padding: 15px; margin-top: 20px; }}
            .status-box.settled {{ background: #f0fdf4; border-left-color: #10b981; }}
            
            .footer {{ margin-top: 50px; text-align: center; color: #6b7280; border-top: 2px solid #e5e7eb; padding-top: 20px; }}
            
            @media print {{
                body {{ margin: 0; background: white; }}
                .container {{ box-shadow: none; }}
                .no-print {{ display: none; }}
            }}
        </style>
    </head>
    <body>
        <div class="watermark">GUESTWORKER</div>
        <div class="container">
            <div class="logo-container">
                <div style="font-size: 28px; background: linear-gradient(to right, #4f46e5, #9333ea); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; font-weight: 800; letter-spacing: 1px;">GuestWorker</div>
            </div>
            <div class="header">
                <h1>SETTLEMENT STATEMENT</h1>
                <p style="font-size: 18px; color: #10b981; font-weight: 600;">{history_data['worker_name']}</p>
                <p><strong>Contractor:</strong> {current_user.name}</p>
                <p><strong>Generated On:</strong> {datetime.now().strftime("%d-%m-%Y %H:%M")}</p>
            </div>
            
            <div class="info-section">
                <div class="info-box">
                    <h3>Worker Information</h3>
                    <p><span class="info-label">Name:</span> <span class="info-value">{history_data['worker_name']}</span></p>
                    {"<p><span class='info-label'>Phone:</span> <span class='info-value'>" + worker.get('phone_number', 'N/A') + "</span></p>" if worker.get('phone_number') else ""}
                    {"<p><span class='info-label'>Daily Wage:</span> <span class='info-value'>₹" + str(worker.get('wage_per_day', 0)) + "</span></p>" if worker.get('wage_per_day') else ""}
                </div>
                <div class="info-box">
                    <h3>Settlement Summary</h3>
                    <p><span class="info-label">Total Settlements:</span> <span class="info-value">{total_settlements}</span></p>
                    <p><span class="info-label">Total Amount Paid:</span> <span class="info-value">₹{total_settled:,.2f}</span></p>
                    <p><span class="info-label">Current Pending:</span> <span class="info-value" style="color: #059669;">₹{history_data['current_pending']:,.2f}</span></p>
                </div>
            </div>
            
            <div class="summary-cards">
                <div class="summary-card">
                    <h4>Total Settlements</h4>
                    <p>{total_settlements}</p>
                </div>
                <div class="summary-card" style="background: linear-gradient(135deg, #3b82f6 0%, #1e40af 100%);">
                    <h4>Total Paid</h4>
                    <p>₹{total_settled:,.2f}</p>
                </div>
                <div class="summary-card" style="background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);">
                    <h4>Advance Used</h4>
                    <p>₹{total_advance_deducted:,.2f}</p>
                </div>
                <div class="summary-card" style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);">
                    <h4>Charges Deducted</h4>
                    <p>₹{total_charges_deducted:,.2f}</p>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Amount Paid</th>
                        <th>Advance Deducted</th>
                        <th>Charges Deducted</th>
                        <th>Settlement Type</th>
                    </tr>
                </thead>
                <tbody>
    """
    
    for settlement in history_data['history']:
        settlement_date = datetime.fromisoformat(settlement['date']).strftime("%d-%m-%Y")
        type_class = f"type-{settlement['settlement_type'].lower()}"
        
        html_content += f"""
                    <tr>
                        <td style="font-weight: 600;">{settlement_date}</td>
                        <td style="font-weight: 700; color: #059669; font-size: 16px;">₹{settlement['amount_settled']:,.2f}</td>
                        <td class="deduction">₹{settlement['advance_deducted']:,.2f}</td>
                        <td class="deduction">₹{settlement['charges_deducted']:,.2f}</td>
                        <td><span class="settlement-type {type_class}">{settlement['settlement_type']}</span></td>
                    </tr>
        """
    
    status_class = 'settled' if history_data['current_pending'] <= 0 else ''
    status_message = '✓ All Wages Settled' if history_data['current_pending'] <= 0 else f'⚠ Pending Settlement: ₹{history_data["current_pending"]:,.2f}'
    
    html_content += f"""
                </tbody>
            </table>
            
            <div class="total-section">
                <div class="total-row">
                    <span>Total Number of Settlements:</span>
                    <strong>{total_settlements} transactions</strong>
                </div>
                <div class="total-row">
                    <span>Total Advance Deducted:</span>
                    <strong>₹{total_advance_deducted:,.2f}</strong>
                </div>
                <div class="total-row">
                    <span>Total Charges Deducted:</span>
                    <strong>₹{total_charges_deducted:,.2f}</strong>
                </div>
                <div class="total-row grand-total">
                    <span>TOTAL AMOUNT PAID:</span>
                    <span>₹{total_settled:,.2f}</span>
                </div>
            </div>
            
            <div class="status-box {status_class}">
                <p style="margin: 0; font-size: 16px; font-weight: 600;">{status_message}</p>
                {"<p style='margin: 10px 0 0 0; font-size: 14px;'>Current Advance Balance: ₹" + str(history_data.get('current_advance', 0)) + "</p>" if history_data.get('current_advance', 0) > 0 else ""}
            </div>
            
            <div class="footer">
                <p style="font-size: 16px; font-weight: 600; color: #10b981;">Thank you for your dedicated service!</p>
                <p><small>This is a computer-generated statement. Generated on {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}</small></p>
            </div>
            
            <div class="no-print" style="text-align: center; margin-top: 30px;">
                <button onclick="window.print()" style="padding: 12px 40px; font-size: 16px; cursor: pointer; background: linear-gradient(135deg, #10b981 0%, #059669 100%); color: white; border: none; border-radius: 8px; font-weight: 600; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    🖨️ Print Statement
                </button>
            </div>
        </div>
    </body>
    </html>
    """
    
    return HTMLResponse(content=html_content)

@api_router.get("/employers/{employer_id}/work-history")
async def get_employer_work_history(
    employer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get employer's work history (attendance records) with optional date filtering"""
    
    try:
        # Build query
        query = {
        "contractor_id": current_user.id,
        "employer_id": employer_id
        }
        
        print(f"🔍 Fetching work history for employer {employer_id}, contractor {current_user.id}")
        print(f"📅 Date range: {start_date} to {end_date}")
        
        # Get ALL employer attendance records from BOTH collections
        new_attendances = await db.employer_attendance.find(query, {"_id": 0}).to_list(10000)
        
        # Also check old attendance collection
        old_query = {
            "contractor_id": current_user.id,
            "employer_id": employer_id,
            "mode": "employer"
        }
        old_attendances = await db.attendance.find(old_query, {"_id": 0}).to_list(10000)
        
        # Combine both lists
        all_attendances = new_attendances + old_attendances
        
        print(f"📊 Found {len(new_attendances)} in new collection, {len(old_attendances)} in old collection")
        
        # Filter by date range if provided
        if start_date or end_date:
            filtered_attendances = []
            for att in all_attendances:
                try:
                    att_date = datetime.strptime(att['date'], "%d-%m-%Y")
                    if start_date:
                        start = datetime.strptime(start_date, "%d-%m-%Y")
                        if att_date < start:
                            continue
                    if end_date:
                        end = datetime.strptime(end_date, "%d-%m-%Y")
                        if att_date > end:
                            continue
                    filtered_attendances.append(att)
                except Exception as e:
                    print(f"Error parsing date {att.get('date')}: {e}")
                    continue
            all_attendances = filtered_attendances
        
        employer = await db.employers.find_one({"id": employer_id}, {"_id": 0})
        
        if not all_attendances:
            print(f"⚠️ No attendance records found")
            return {
                "employer_name": employer['name'] if employer else "Unknown",
                "employer_company": employer.get('company', '') if employer else '',
                "history": [],
                "total_amount": 0,
                "total_days": 0
            }
        
        # Sort by date (oldest first for cumulative calculation)
        all_attendances.sort(key=lambda x: datetime.strptime(x['date'], "%d-%m-%Y"))
        
        # Build history with more details
        history = []
        cumulative = 0
        for record in all_attendances:
            cumulative += record.get('total_amount', record.get('wage_amount', 0))
            
            # Get worker names if selected_workers exist
            worker_names = []
            if record.get('selected_workers'):
                for worker_id in record['selected_workers']:
                    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0, "name": 1})
                    if worker:
                        worker_names.append(worker['name'])
            
            # Extract additional charges from array format or direct field
            additional_charges_amount = 0
            additional_charges_reason = ''
            
            # Check if additional_charges is an array (new format)
            if isinstance(record.get('additional_charges'), list) and len(record['additional_charges']) > 0:
                additional_charges_amount = record['additional_charges'][0].get('amount', 0)
                additional_charges_reason = record['additional_charges'][0].get('description', '')
            # Check if it's a direct number (old format)
            elif isinstance(record.get('additional_charges'), (int, float)):
                additional_charges_amount = record.get('additional_charges', 0)
                additional_charges_reason = record.get('charge_description', '')
            
            # Also check for alternative field names
            if not additional_charges_reason:
                additional_charges_reason = record.get('additional_charges_reason', '') or record.get('charge_description', '')
            
            history.append({
                "id": record.get('id'),
                "date": record['date'],
                "workers_count": record.get('workers_count', 0),
                "amount": record.get('total_amount', record.get('wage_amount', 0)),
                "additional_charges": additional_charges_amount,
                "additional_charges_reason": additional_charges_reason,
                "remarks": record.get('remarks', ''),
                "selected_workers": worker_names,
                "cumulative_total": cumulative
            })
        
        print(f"✅ Returning {len(history)} records")
        
        return {
            "employer_name": employer['name'] if employer else "Unknown",
            "employer_company": employer.get('company', '') if employer else '',
            "history": history,
            "total_amount": cumulative,
            "total_days": len(all_attendances)
        }
    except Exception as e:
        print(f"❌ Error in get_employer_work_history: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/employers/{employer_id}/work-history/invoice")
async def generate_employer_invoice(
    employer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Generate HTML invoice for employer work history"""
    
    try:
        # Reuse the work history endpoint logic
        work_history = await get_employer_work_history(employer_id, start_date, end_date, current_user)
        
        # Generate HTML invoice
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Invoice - {work_history['employer_name']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; position: relative; }}
                .watermark {{ position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-45deg); font-size: 120px; font-weight: bold; color: rgba(0, 0, 0, 0.03); z-index: -1; white-space: nowrap; pointer-events: none; }}
                .logo-container {{ text-align: center; margin-bottom: 20px; }}
                .logo {{ width: 80px; height: 80px; margin: 0 auto 10px; }}
                .header {{ text-align: center; margin-bottom: 30px; border-bottom: 3px solid #047857; padding-bottom: 20px; }}
                .header h1 {{ margin: 0; color: #047857; }}
                .header p {{ margin: 5px 0; color: #666; }}
                .info-section {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
                .info-box {{ width: 48%; }}
                .info-box h3 {{ margin: 0 0 10px 0; color: #333; border-bottom: 2px solid #ddd; padding-bottom: 5px; }}
                .info-box p {{ margin: 5px 0; }}
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
                th {{ background-color: #333; color: white; padding: 12px; text-align: left; }}
                td {{ padding: 10px; border-bottom: 1px solid #ddd; }}
                tr:hover {{ background-color: #f5f5f5; }}
                .total-section {{ margin-top: 30px; text-align: right; }}
                .total-row {{ font-size: 18px; margin: 10px 0; }}
                .total-row.grand-total {{ font-size: 24px; font-weight: bold; color: #333; border-top: 3px solid #333; padding-top: 10px; }}
                .footer {{ margin-top: 50px; text-align: center; color: #666; border-top: 2px solid #ddd; padding-top: 20px; }}
                @media print {{
                    body {{ margin: 20px; }}
                    .no-print {{ display: none; }}
                }}
            </style>
        </head>
        <body>
            <div class="watermark">GUESTWORKER</div>
            <div class="logo-container">
                <div style="font-size: 28px; background: linear-gradient(to right, #4f46e5, #9333ea); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; font-weight: 800; letter-spacing: 1px;">GuestWorker</div>
            </div>
            <div class="header">
                <h1>WORK INVOICE</h1>
                <p><strong>Contractor:</strong> {current_user.name}</p>
                <p><strong>Date:</strong> {datetime.now().strftime("%d-%m-%Y")}</p>
            </div>
            
            <div class="info-section">
                <div class="info-box">
                    <h3>Bill To:</h3>
                    <p><strong>{work_history['employer_name']}</strong></p>
                    {"<p>" + work_history['employer_company'] + "</p>" if work_history.get('employer_company') else ""}
                </div>
                <div class="info-box">
                    <h3>Invoice Details:</h3>
                    <p><strong>Period:</strong> {start_date or "All time"} to {end_date or "Present"}</p>
                    <p><strong>Total Days:</strong> {work_history['total_days']}</p>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Workers</th>
                        <th>Worker Names</th>
                        <th style="text-align: right;">Amount</th>
                        <th style="text-align: right;">Extra Charges</th>
                        <th style="text-align: right;">Total</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for record in work_history['history']:
            worker_names = ", ".join(record['selected_workers']) if record['selected_workers'] else "Not specified"
            extra_charges_display = ""
            if record['additional_charges'] > 0:
                extra_charges_display = f"₹{record['additional_charges']:,.2f}"
                if record['additional_charges_reason']:
                    extra_charges_display += f"<br><small style='color: #666;'>{record['additional_charges_reason']}</small>"
            else:
                extra_charges_display = "-"
            
            html_content += f"""
                    <tr>
                        <td>{record['date']}</td>
                        <td style="text-align: center;">{record['workers_count']}</td>
                        <td><small>{worker_names}</small></td>
                        <td style="text-align: right;">₹{record['amount']:,.2f}</td>
                        <td style="text-align: right;">{extra_charges_display}</td>
                        <td style="text-align: right;"><strong>₹{record['cumulative_total']:,.2f}</strong></td>
                    </tr>
            """
        
        html_content += f"""
                </tbody>
            </table>
            
            <div class="total-section">
                <div class="total-row">
                    <strong>Number of Working Days:</strong> {work_history['total_days']}
                </div>
                <div class="total-row grand-total">
                    <strong>GRAND TOTAL:</strong> ₹{work_history['total_amount']:,.2f}
                </div>
            </div>
            
            <div class="footer">
                <p>Thank you for your business!</p>
                <p><small>Generated on {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}</small></p>
            </div>
            
            <div class="no-print" style="text-align: center; margin-top: 30px;">
                <button onclick="window.print()" style="padding: 10px 30px; font-size: 16px; cursor: pointer; background-color: #333; color: white; border: none; border-radius: 5px;">
                    Print Invoice
                </button>
            </div>
        </body>
        </html>
        """
        
        return HTMLResponse(content=html_content)
        
    except Exception as e:
        print(f"❌ Error generating invoice: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/workers/{worker_id}/work-history")
async def get_worker_work_history(
    worker_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get worker's attendance history (all records) with optional date filtering"""
    
    try:
        print(f"🔍 Fetching work history for worker {worker_id}, contractor {current_user.id}")
        print(f"📅 Date range: {start_date} to {end_date}")
        
        # Get ALL worker attendance records from BOTH collections
        new_attendances = await db.worker_attendance.find({
        "contractor_id": current_user.id,
        "worker_id": worker_id
    }, {"_id": 0}).to_list(10000)
    
        # Also check old attendance collection
        old_attendances = await db.attendance.find({
            "contractor_id": current_user.id,
            "worker_id": worker_id,
            "mode": "worker"
        }, {"_id": 0}).to_list(10000)
        
        # Combine both lists
        all_attendances = new_attendances + old_attendances
        
        print(f"📊 Found {len(new_attendances)} in new collection, {len(old_attendances)} in old collection")
        
        # Filter by date range if provided
        if start_date or end_date:
            filtered_attendances = []
            for att in all_attendances:
                try:
                    att_date = datetime.strptime(att['date'], "%d-%m-%Y")
                    if start_date:
                        start = datetime.strptime(start_date, "%d-%m-%Y")
                        if att_date < start:
                            continue
                    if end_date:
                        end = datetime.strptime(end_date, "%d-%m-%Y")
                        if att_date > end:
                            continue
                    filtered_attendances.append(att)
                except Exception as e:
                    print(f"Error parsing date {att.get('date')}: {e}")
                    continue
            all_attendances = filtered_attendances
        
        worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
        
        if not all_attendances:
            print(f"⚠️ No attendance records found")
            return {
                "worker_name": worker['name'] if worker else "Unknown",
                "history": [],
                "total_days_worked": 0,
                "total_wage_earned": 0,
                "present_days": 0,
                "absent_days": 0
            }
        
        # Sort by date (oldest first for chronological order)
        all_attendances.sort(key=lambda x: datetime.strptime(x['date'], "%d-%m-%Y"))
        
        # Build history with more details
        history = []
        cumulative_wage = 0
        for att in all_attendances:
            wage_earned = att.get('wage_earned', att.get('wage_amount', 0))
            cumulative_wage += wage_earned
            
            # Get employer name if assigned
            employer_name = None
            if att.get('employer_id'):
                employer = await db.employers.find_one({"id": att['employer_id']}, {"_id": 0, "name": 1})
                if employer:
                    employer_name = employer['name']
            
            history.append({
                "id": att.get('id'),
                "date": att['date'],
                "status": att.get('status', 'Unknown'),
                "wage_earned": wage_earned,
                "employer_name": employer_name,
                "cumulative_wage": cumulative_wage
            })
        
        present_count = len([h for h in history if h['status'] == 'Present'])
        absent_count = len([h for h in history if h['status'] == 'Absent'])
        
        print(f"✅ Returning {len(history)} records")
        
        return {
            "worker_name": worker['name'] if worker else "Unknown",
            "history": history,
            "total_days_worked": present_count,
            "total_wage_earned": cumulative_wage,
            "present_days": present_count,
            "absent_days": absent_count
        }
    except Exception as e:
        print(f"❌ Error in get_worker_work_history: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/subscription/invoice/{transaction_id}")
async def generate_subscription_invoice(
    transaction_id: str,
    current_user: User = Depends(get_current_user)
):
    """Generate HTML invoice for subscription payment (Razorpay or Activation Key)"""
    
    try:
        # Get transaction/payment record
        transaction = await db.payment_orders.find_one({
            "id": transaction_id,
            "contractor_id": current_user.id
        })
        
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        
        # Get user details
        user = await db.users.find_one({"id": current_user.id}, {"_id": 0})
        
        # Determine payment method and details
        payment_method = transaction.get('payment_method', 'razorpay')
        amount = transaction.get('amount', 0)
        plan_name = transaction.get('plan_name', 'Subscription Plan')
        created_at = transaction.get('created_at')
        
        # Parse date
        if isinstance(created_at, str):
            try:
                invoice_date = datetime.fromisoformat(created_at)
            except:
                invoice_date = datetime.now(timezone.utc)
        else:
            invoice_date = created_at if created_at else datetime.now(timezone.utc)
        
        # Format invoice number
        invoice_number = f"INV-{transaction_id[:8].upper()}"
        
        # Determine if this is a paid transaction
        if payment_method == 'activation_key':
            is_paid = transaction.get('is_paid', False)
            activation_key = transaction.get('activation_key', 'N/A')
            payment_status = "Paid" if is_paid else "Complimentary"
            payment_method_label = f"Activation Key ({activation_key})"
        elif payment_method == 'razorpay_subscription':
            is_paid = True
            is_renewal = transaction.get('is_renewal', False)
            payment_status = "Paid"
            payment_method_label = "Razorpay Subscription" + (" (Auto-Renewal)" if is_renewal else "")
            razorpay_payment_id = transaction.get('razorpay_payment_id', 'N/A')
        else:
            is_paid = True
            payment_status = "Paid"
            payment_method_label = "Razorpay"
            razorpay_payment_id = transaction.get('razorpay_payment_id', 'N/A')
        
        # Duration
        duration_days = transaction.get('duration_days', 30)
        
        # Generate HTML invoice
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Invoice - {invoice_number}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; position: relative; }}
                .watermark {{ position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-45deg); font-size: 120px; font-weight: bold; color: rgba(0, 0, 0, 0.03); z-index: -1; white-space: nowrap; pointer-events: none; }}
                .header {{ text-align: center; margin-bottom: 30px; border-bottom: 2px solid #333; padding-bottom: 20px; }}
                .header h1 {{ margin: 0; color: #333; font-size: 32px; }}
                .company-info {{ text-align: center; margin-bottom: 30px; }}
                .company-name {{ font-size: 28px; background: linear-gradient(to right, #4f46e5, #9333ea); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; font-weight: 800; letter-spacing: 1px; }}
                .invoice-details {{ display: flex; justify-content: space-between; margin-bottom: 30px; }}
                .info-box {{ flex: 1; padding: 15px; background: #f9f9f9; border-radius: 5px; margin: 0 10px; }}
                .info-box h3 {{ margin-top: 0; color: #555; font-size: 14px; text-transform: uppercase; }}
                .info-box p {{ margin: 5px 0; color: #333; }}
                table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
                th {{ background-color: #333; color: white; font-weight: bold; }}
                .total-row {{ font-weight: bold; background-color: #f0f0f0; }}
                .amount {{ text-align: right; }}
                .footer {{ margin-top: 50px; text-align: center; color: #777; font-size: 12px; border-top: 1px solid #ddd; padding-top: 20px; }}
                .status-badge {{ display: inline-block; padding: 5px 15px; border-radius: 20px; font-size: 12px; font-weight: bold; }}
                .status-paid {{ background-color: #10b981; color: white; }}
                .status-complimentary {{ background-color: #6366f1; color: white; }}
                @media print {{
                    .no-print {{ display: none; }}
                    body {{ margin: 20px; }}
                }}
            </style>
        </head>
        <body>
            <div class="watermark">GuestWorker</div>
            
            <div class="company-info">
                <div class="company-name">GuestWorker</div>
            </div>
            
            <div class="header">
                <h1>SUBSCRIPTION INVOICE</h1>
                <p><strong>Invoice Number:</strong> {invoice_number}</p>
                <p><strong>Date:</strong> {invoice_date.strftime("%d-%m-%Y")}</p>
                <p><span class="status-badge status-{payment_status.lower().replace(' ', '-')}">{payment_status}</span></p>
            </div>
            
            <div class="invoice-details">
                <div class="info-box">
                    <h3>Billed To:</h3>
                    <p><strong>{user.get('name', 'N/A')}</strong></p>
                    <p>{user.get('email', 'N/A')}</p>
                    {"<p>" + user.get('phone', '') + "</p>" if user.get('phone') else ""}
                </div>
                <div class="info-box">
                    <h3>Payment Details:</h3>
                    <p><strong>Method:</strong> {payment_method_label}</p>
                    {f"<p><strong>Payment ID:</strong> {razorpay_payment_id}</p>" if payment_method in ['razorpay', 'razorpay_subscription'] else ""}
                    <p><strong>Status:</strong> {payment_status}</p>
                </div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Description</th>
                        <th>Duration</th>
                        <th class="amount">Amount (₹)</th>
                    </tr>
                </thead>
                <tbody>
                    <tr>
                        <td><strong>{plan_name}</strong><br><small>Subscription Plan</small></td>
                        <td>{duration_days} days</td>
                        <td class="amount">₹{amount:.2f}</td>
                    </tr>
                    <tr class="total-row">
                        <td colspan="2" style="text-align: right;"><strong>Total Amount</strong></td>
                        <td class="amount"><strong>₹{amount:.2f}</strong></td>
                    </tr>
                </tbody>
            </table>
            
            <div class="footer">
                <p><strong>Thank you for your subscription!</strong></p>
                <p>This is a computer-generated invoice and does not require a signature.</p>
                <p>For any queries, please contact support@guestworker.app</p>
                <p style="margin-top: 20px; font-size: 10px;">GuestWorker - Contractor Management System</p>
            </div>
            
            <div class="no-print" style="text-align: center; margin-top: 30px;">
                <button onclick="window.print()" style="padding: 10px 30px; font-size: 16px; cursor: pointer; background-color: #333; color: white; border: none; border-radius: 5px;">
                    Print Invoice
                </button>
            </div>
        </body>
        </html>
        """
        
        return HTMLResponse(content=html_content)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error generating subscription invoice: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating invoice: {str(e)}")

@api_router.get("/workers/{worker_id}/work-history/invoice")
async def generate_worker_invoice(
    worker_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Generate HTML invoice for worker work history"""
    
    try:
        # Reuse the work history endpoint logic
        work_history = await get_worker_work_history(worker_id, start_date, end_date, current_user)
        
        # Get worker details
        worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
        
        # Generate HTML invoice
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Wage Statement - {work_history['worker_name']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; background-color: #f9fafb; position: relative; }}
                .watermark {{ position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%) rotate(-45deg); font-size: 120px; font-weight: bold; color: rgba(0, 0, 0, 0.03); z-index: 0; white-space: nowrap; pointer-events: none; }}
                .container {{ max-width: 900px; margin: 0 auto; background: white; padding: 40px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); position: relative; z-index: 1; }}
                .logo-container {{ text-align: center; margin-bottom: 20px; }}
                .logo {{ width: 80px; height: 80px; margin: 0 auto 10px; }}
                .header {{ text-align: center; margin-bottom: 30px; border-bottom: 3px solid #047857; padding-bottom: 20px; }}
                .header h1 {{ margin: 0; color: #059669; font-size: 32px; }}
                .header p {{ margin: 5px 0; color: #6b7280; }}
                .info-section {{ display: flex; justify-content: space-between; margin-bottom: 30px; background: #f0fdf4; padding: 20px; border-radius: 8px; }}
                .info-box {{ width: 48%; }}
                .info-box h3 {{ margin: 0 0 10px 0; color: #047857; border-bottom: 2px solid #10b981; padding-bottom: 5px; }}
                .info-box p {{ margin: 8px 0; color: #374151; }}
                .info-label {{ font-weight: 600; color: #6b7280; font-size: 14px; }}
                .info-value {{ font-weight: 700; color: #111827; font-size: 16px; }}
                
                .summary-cards {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; margin-bottom: 30px; }}
                .summary-card {{ background: linear-gradient(135deg, #047857 0%, #14b8a6 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
                .summary-card h4 {{ margin: 0 0 5px 0; font-size: 14px; opacity: 0.9; }}
                .summary-card p {{ margin: 0; font-size: 24px; font-weight: bold; }}
                
                table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                th {{ background: linear-gradient(135deg, #3b82f6 0%, #1e40af 100%); color: white; padding: 14px; text-align: left; font-size: 14px; text-transform: uppercase; letter-spacing: 0.5px; }}
                td {{ padding: 12px; border-bottom: 1px solid #e5e7eb; color: #374151; }}
                tr:hover {{ background-color: #f3f4f6; }}
                tr.present {{ background-color: #f0fdf4; }}
                tr.absent {{ background-color: #fef2f2; }}
                
                .status-badge {{ padding: 4px 12px; border-radius: 12px; font-size: 12px; font-weight: 600; display: inline-block; }}
                .status-present {{ background-color: #dcfce7; color: #15803d; }}
                .status-absent {{ background-color: #fee2e2; color: #991b1b; }}
                
                .total-section {{ margin-top: 30px; background: #f9fafb; padding: 20px; border-radius: 8px; border: 2px solid #e5e7eb; }}
                .total-row {{ display: flex; justify-content: space-between; font-size: 16px; margin: 8px 0; padding: 8px 0; }}
                .total-row.grand-total {{ font-size: 28px; font-weight: bold; color: #1e40af; border-top: 3px solid #3b82f6; padding-top: 15px; margin-top: 15px; }}
                
                .footer {{ margin-top: 50px; text-align: center; color: #6b7280; border-top: 2px solid #e5e7eb; padding-top: 20px; }}
                .footer-note {{ background: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin-top: 20px; text-align: left; }}
                
                @media print {{
                    body {{ margin: 0; background: white; }}
                    .container {{ box-shadow: none; }}
                    .no-print {{ display: none; }}
                }}
            </style>
        </head>
        <body>
            <div class="watermark">GUESTWORKER</div>
            <div class="container">
                <div class="logo-container">
                    <div style="font-size: 28px; background: linear-gradient(to right, #4f46e5, #9333ea); -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text; font-weight: 800; letter-spacing: 1px;">GuestWorker</div>
                </div>
                <div class="header">
                    <h1>WAGE STATEMENT</h1>
                    <p style="font-size: 18px; color: #3b82f6; font-weight: 600;">{work_history['worker_name']}</p>
                    <p><strong>Contractor:</strong> {current_user.name}</p>
                    <p><strong>Generated On:</strong> {datetime.now().strftime("%d-%m-%Y %H:%M")}</p>
                </div>
                
                <div class="info-section">
                    <div class="info-box">
                        <h3>Worker Information</h3>
                        <p><span class="info-label">Name:</span> <span class="info-value">{work_history['worker_name']}</span></p>
                        <p><span class="info-label">Phone:</span> <span class="info-value">{worker.get('phone_number', 'N/A')}</span></p>
                        {"<p><span class='info-label'>Address:</span> <span class='info-value'>" + worker.get('address', 'N/A') + "</span></p>" if worker.get('address') else ""}
                    </div>
                    <div class="info-box">
                        <h3>Statement Period</h3>
                        <p><span class="info-label">From:</span> <span class="info-value">{start_date or "Beginning"}</span></p>
                        <p><span class="info-label">To:</span> <span class="info-value">{end_date or "Present"}</span></p>
                        <p><span class="info-label">Daily Rate:</span> <span class="info-value">₹{worker.get('wage_per_day', 0):,.2f}</span></p>
                    </div>
                </div>
                
                <div class="summary-cards">
                    <div class="summary-card" style="background: linear-gradient(135deg, #10b981 0%, #059669 100%);">
                        <h4>Present Days</h4>
                        <p>{work_history['present_days']}</p>
                    </div>
                    <div class="summary-card" style="background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);">
                        <h4>Absent Days</h4>
                        <p>{work_history['absent_days']}</p>
                    </div>
                    <div class="summary-card">
                        <h4>Total Earned</h4>
                        <p>₹{work_history['total_wage_earned']:,.2f}</p>
                    </div>
                </div>
                
                <table>
                    <thead>
                        <tr>
                            <th>Date</th>
                            <th>Status</th>
                            <th>Employer</th>
                            <th style="text-align: right;">Daily Wage</th>
                            <th style="text-align: right;">Cumulative Total</th>
                        </tr>
                    </thead>
                    <tbody>
        """
        
        for record in work_history['history']:
            status_class = record['status'].lower()
            status_badge_class = f"status-{status_class}"
            employer_display = record['employer_name'] if record.get('employer_name') else '<i style="color: #9ca3af;">Not assigned</i>'
            wage_display = f"₹{record['wage_earned']:,.2f}" if record['status'] == 'Present' else "-"
            
            html_content += f"""
                    <tr class="{status_class}">
                        <td style="font-weight: 600;">{record['date']}</td>
                        <td><span class="status-badge {status_badge_class}">{record['status']}</span></td>
                        <td>{employer_display}</td>
                        <td style="text-align: right; font-weight: 600; color: #059669;">{wage_display}</td>
                        <td style="text-align: right; font-weight: 700; color: #1e40af;">₹{record['cumulative_wage']:,.2f}</td>
                    </tr>
            """
        
        html_content += f"""
                    </tbody>
                </table>
                
                <div class="total-section">
                    <div class="total-row">
                        <span>Total Working Days (Present):</span>
                        <strong>{work_history['present_days']} days</strong>
                    </div>
                    <div class="total-row">
                        <span>Total Absent Days:</span>
                        <strong>{work_history['absent_days']} days</strong>
                    </div>
                    <div class="total-row grand-total">
                        <span>TOTAL WAGE EARNED:</span>
                        <span>₹{work_history['total_wage_earned']:,.2f}</span>
                    </div>
                </div>
                
                <div class="footer-note">
                    <p style="margin: 0; font-size: 14px; color: #92400e;"><strong>📌 Note:</strong> This statement shows the total wages earned during the specified period. For payment and settlement details, please refer to the payment records section.</p>
                </div>
                
                <div class="footer">
                    <p style="font-size: 16px; font-weight: 600; color: #3b82f6;">Thank you for your dedicated work!</p>
                    <p><small>This is a computer-generated document. Generated on {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}</small></p>
                </div>
                
                <div class="no-print" style="text-align: center; margin-top: 30px;">
                    <button onclick="window.print()" style="padding: 12px 40px; font-size: 16px; cursor: pointer; background: linear-gradient(135deg, #3b82f6 0%, #1e40af 100%); color: white; border: none; border-radius: 8px; font-weight: 600; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                        🖨️ Print Statement
                    </button>
                </div>
            </div>
        </body>
        </html>
        """
        
        return HTMLResponse(content=html_content)
        
    except Exception as e:
        print(f"❌ Error generating worker invoice: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating invoice: {str(e)}")

# ============ ACCOUNT DELETION REQUEST ============

class AccountDeletionRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    contractor_name: str
    contractor_email: str
    reason: Optional[str] = None
    requested_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    status: str = "pending"  # pending, approved, rejected, auto_deleted
    approved_at: Optional[datetime] = None
    scheduled_deletion_date: Optional[datetime] = None  # 30 days after approval

class DeletedUserArchive(BaseModel):
    """Archive record for deleted users - keeps minimal info for admin reference"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    contractor_name: str
    contractor_email: str
    phone: Optional[str] = None
    subscription_plan: Optional[str] = None
    subscription_status: Optional[str] = None
    deletion_reason: Optional[str] = None
    deletion_type: str  # "user_requested", "admin_approved", "auto_inactive"
    requested_at: Optional[datetime] = None
    deleted_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    deleted_by: Optional[str] = None  # admin email or "system"
    
    # Summary statistics (not detailed data)
    total_workers: int = 0
    total_employers: int = 0
    total_attendance_records: int = 0
    total_payments: int = 0
    account_created_at: Optional[datetime] = None
    last_login_at: Optional[datetime] = None

class ProfileUpdate(BaseModel):
    name: str

class PasswordChange(BaseModel):
    old_password: str
    new_password: str

@api_router.put("/account/profile")
async def update_profile(
    profile_data: ProfileUpdate,
    current_user: User = Depends(get_current_user)
):
    """Update user profile (name)"""
    
    # Sanitize input
    name = sanitize_input(profile_data.name.strip())
    
    if not name or len(name) < 2:
        raise HTTPException(status_code=400, detail="Name must be at least 2 characters long")
    
    if len(name) > 100:
        raise HTTPException(status_code=400, detail="Name is too long (max 100 characters)")
    
    # Update user name
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"name": name}}
    )
    
    return {"message": "Profile updated successfully", "name": name}

@api_router.post("/account/change-password")
async def change_password(
    password_data: PasswordChange,
    current_user: User = Depends(get_current_user)
):
    """Change user password"""
    
    # Verify old password
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not pwd_context.verify(password_data.old_password, user_doc['password']):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    # Validate new password strength
    is_valid, error_message = validate_password_strength(
        password_data.new_password, 
        current_user.email
    )
    if not is_valid:
        raise HTTPException(status_code=400, detail=error_message)
    
    # Hash and update password
    hashed_password = pwd_context.hash(password_data.new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"password": hashed_password}}
    )
    
    return {"message": "Password changed successfully"}

# ========================= Notifications - User =========================
@api_router.get("/notifications")
async def get_notifications(limit: int = 20, current_user: User = Depends(get_current_user)):
    # Auto-generate attendance reminder once per day on first login if conditions are met
    try:
        # Check if user has active workers or employers
        active_workers_count = await db.workers.count_documents({
            "contractor_id": current_user.id,
            "status": "Active"
        })
        active_employers_count = await db.employers.count_documents({
            "contractor_id": current_user.id,
            "status": "Active"
        })
        
        # Only send reminder if user has active workers or employers
        if active_workers_count > 0 or active_employers_count > 0:
            today_str = datetime.now(timezone.utc).astimezone().date().isoformat()
            # Consider any attendance where status is not 'Absent'
            has_att = await db.attendance.count_documents({
                "contractor_id": current_user.id,
                "date": today_str,
                "status": {"$ne": "Absent"}
            })
            
            if has_att == 0:
                now = datetime.now(timezone.utc)
                # Check if reminder was already sent today
                user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0, "last_attendance_alert_at": 1})
                last_sent = user_doc.get("last_attendance_alert_at") if user_doc else None
                
                if isinstance(last_sent, str):
                    try:
                        last_sent = datetime.fromisoformat(last_sent)
                    except Exception:
                        last_sent = None
                
                # Only send once per day (24 hours)
                should_send = not last_sent or (now - last_sent) >= timedelta(hours=24)
                
                if should_send:
                    # Check if reminder already exists for today (not read)
                    existing_reminder = await db.notifications.find_one({
                        "user_id": current_user.id,
                        "title": "Reminder: Mark Today's Attendance",
                        "type": "alert",
                        "read": False
                    })
                    
                    if not existing_reminder:
                        await db.notifications.insert_one({
                            "id": str(uuid.uuid4()),
                            "user_id": current_user.id,
                            "title": "Reminder: Mark Today's Attendance",
                            "message": "You have not marked worker attendance for today. Please update attendance to keep records accurate.",
                            "type": "alert",
                            "link": "/attendance",
                            "action_url": "/attendance",
                            "action_label": "Mark Attendance",
                            "read": False,
                            "created_at": now
                        })
                        await db.users.update_one({"id": current_user.id}, {"$set": {"last_attendance_alert_at": now}})
    except Exception:
        pass

    notifications = await db.notifications.find({"user_id": current_user.id}).sort("created_at", -1).limit(limit).to_list(length=limit)
    for n in notifications:
        n.pop("_id", None)
    return {"items": notifications}

@api_router.post("/notifications/read")
async def mark_notification_read(notification_id: str, current_user: User = Depends(get_current_user)):
    await db.notifications.update_one({"id": notification_id, "user_id": current_user.id}, {"$set": {"read": True}})
    return {"message": "Marked as read"}

@api_router.post("/notifications/read-all")
async def mark_all_notifications_read(current_user: User = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": current_user.id, "read": False}, {"$set": {"read": True}})
    return {"message": "All marked as read"}

@api_router.post("/notifications/mark-attendance-reminder-read")
async def mark_attendance_reminder_read(current_user: User = Depends(get_current_user)):
    """Mark attendance reminder notifications as read when notification panel is opened"""
    await db.notifications.update_many({
        "user_id": current_user.id,
        "title": "Reminder: Mark Today's Attendance",
        "type": "alert",
        "read": False
    }, {"$set": {"read": True}})
    return {"message": "Attendance reminders marked as read"}

@api_router.delete("/notifications/{notification_id}")
async def delete_notification(notification_id: str, current_user: User = Depends(get_current_user)):
    await db.notifications.delete_one({"id": notification_id, "user_id": current_user.id})
    return {"message": "Deleted"}

# (moved admin broadcast endpoint below get_current_admin definition)

class DeletionRequest(BaseModel):
    reason: Optional[str] = None

@api_router.post("/account/request-deletion")
async def request_account_deletion(
    request_data: DeletionRequest,
    current_user: User = Depends(get_current_user)
):
    """Request account deletion - sends request to admin (available to all users)"""
    
    # Check if already requested
    existing = await db.deletion_requests.find_one({
        "contractor_id": current_user.id,
        "status": "pending"
    })
    
    if existing:
        raise HTTPException(status_code=400, detail="Deletion request already pending")
    
    deletion_request = AccountDeletionRequest(
        contractor_id=current_user.id,
        contractor_name=current_user.name,
        contractor_email=current_user.email,
        reason=sanitize_input(request_data.reason) if request_data.reason else None
    )
    
    request_dict = deletion_request.model_dump()
    request_dict['requested_at'] = request_dict['requested_at'].isoformat()
    await db.deletion_requests.insert_one(request_dict)
    
    return {"message": "Account deletion request submitted successfully"}

# ============ COMPREHENSIVE REPORTS ROUTES ============

@api_router.get("/reports/summary")
async def get_reports_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get comprehensive summary of all data"""
    
    # Workers
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
    active_workers = len([w for w in workers if w['status'] == 'Active'])
    deleted_workers = len([w for w in workers if w['status'] == 'Inactive'])
    
    # Employers
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
    active_employers = len([e for e in employers if e['status'] == 'Active'])
    deleted_employers = len([e for e in employers if e['status'] == 'Inactive'])
    
    # Attendance
    attendance_query = {"contractor_id": current_user.id}
    if start_date and end_date:
        attendance_query["date"] = {"$gte": start_date, "$lte": end_date}
    
    worker_attendance = await db.worker_attendance.find(attendance_query, {"_id": 0}).to_list(10000)
    total_days_worked = len([a for a in worker_attendance if a['status'] in ['Present', 'Late']])
    total_wages_generated = sum(a.get('wage_earned', 0) for a in worker_attendance)
    
    # Payments
    payment_query = {"contractor_id": current_user.id}
    payments = await db.payment_collections.find(payment_query, {"_id": 0}).to_list(10000)
    
    # Filter by date if needed
    if start_date and end_date:
        payments = [p for p in payments if start_date <= p.get('date', '')[:10] <= end_date]
    
    total_payments_collected = sum(p['amount'] for p in payments)
    
    # Settlements
    settlement_query = {"contractor_id": current_user.id}
    settlements = await db.wage_settlements.find(settlement_query, {"_id": 0}).to_list(10000)
    
    # Filter by date if needed
    if start_date and end_date:
        settlements = [s for s in settlements if start_date <= s.get('settlement_date', '')[:10] <= end_date]
    
    total_wages_settled = sum(s['amount_paid'] for s in settlements)
    
    # Advances
    advance_query = {"contractor_id": current_user.id}
    advances = await db.advances.find(advance_query, {"_id": 0}).to_list(10000)
    total_advances = sum(a['amount'] for a in advances)
    
    # Extra Charges
    charges_query = {"contractor_id": current_user.id}
    charges = await db.extra_charges.find(charges_query, {"_id": 0}).to_list(10000)
    total_extra_charges = sum(c['amount'] for c in charges)
    
    # Commission calculation from database
    commissions_query = {"contractor_id": current_user.id}
    if start_date and end_date:
        commissions_query["date"] = {"$gte": start_date, "$lte": end_date}
    
    commissions = await db.commissions.find(commissions_query, {"_id": 0}).to_list(10000)
    commission = sum(c['commission_amount'] for c in commissions)
    
    return {
        "workers": {
            "total": len(workers),
            "active": active_workers,
            "deleted": deleted_workers
        },
        "employers": {
            "total": len(employers),
            "active": active_employers,
            "deleted": deleted_employers
        },
        "attendance": {
            "total_days": total_days_worked,
            "total_records": len(worker_attendance)
        },
        "financial": {
            "total_wages_generated": total_wages_generated,
            "total_payments_collected": total_payments_collected,
            "total_wages_settled": total_wages_settled,
            "total_advances": total_advances,
            "total_extra_charges": total_extra_charges,
            "commission_earned": commission,  # Frontend expects commission_earned
            "total_commission": commission,  # Keep both for compatibility
            "pending_settlements": total_wages_generated - total_wages_settled
        }
    }

@api_router.get("/reports/workers-detailed")
async def get_workers_detailed_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Detailed worker report with all related data"""
    query = {"contractor_id": current_user.id}
    if status:
        query["status"] = status
    
    workers = await db.workers.find(query, {"_id": 0}).to_list(10000)
    
    detailed_workers = []
    for worker in workers:
        # Get attendance
        att_query = {"contractor_id": current_user.id, "worker_id": worker['id']}
        if start_date and end_date:
            att_query["date"] = {"$gte": start_date, "$lte": end_date}
        
        attendances = await db.worker_attendance.find(att_query, {"_id": 0}).to_list(10000)
        days_worked = len([a for a in attendances if a['status'] in ['Present', 'Late']])
        total_wage_earned = sum(a.get('wage_earned', 0) for a in attendances)
        
        # Get settlements
        settlements = await db.wage_settlements.find({
            "contractor_id": current_user.id,
            "worker_id": worker['id']
        }, {"_id": 0}).to_list(10000)
        total_settled = sum(s['amount_paid'] for s in settlements)
        
        # Get advances
        advances = await db.advances.find({
            "contractor_id": current_user.id,
            "worker_id": worker['id']
        }, {"_id": 0}).to_list(10000)
        total_advances = sum(a['amount'] for a in advances)
        
        detailed_workers.append({
            **worker,
            "days_worked": days_worked,
            "total_wage_earned": total_wage_earned,
            "total_settled": total_settled,
            "total_advances": total_advances,
            "pending": worker.get('pending_settlement', 0)
        })
    
    return detailed_workers

@api_router.get("/reports/employers-detailed")
async def get_employers_detailed_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Detailed employer report with all related data"""
    query = {"contractor_id": current_user.id}
    if status:
        query["status"] = status
    
    employers = await db.employers.find(query, {"_id": 0}).to_list(10000)
    
    detailed_employers = []
    for employer in employers:
        # Get attendance records
        att_query = {"contractor_id": current_user.id, "employer_id": employer['id']}
        if start_date and end_date:
            att_query["date"] = {"$gte": start_date, "$lte": end_date}
        
        attendances = await db.employer_attendance.find(att_query, {"_id": 0}).to_list(10000)
        total_work_days = len(attendances)
        total_amount_owed = sum(a['total_amount'] for a in attendances)
        
        # Get payments
        payments = await db.payment_collections.find({
            "contractor_id": current_user.id,
            "employer_id": employer['id']
        }, {"_id": 0}).to_list(10000)
        total_collected = sum(p['amount'] for p in payments)
        
        detailed_employers.append({
            **employer,
            "total_work_days": total_work_days,
            "total_amount_owed": total_amount_owed,
            "total_collected": total_collected,
            "pending": employer.get('pending_payment', 0)
        })
    
    return detailed_employers

@api_router.get("/reports/attendance-detailed")
async def get_attendance_detailed_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Detailed attendance report"""
    query = {"contractor_id": current_user.id}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    worker_attendance = await db.worker_attendance.find(query, {"_id": 0}).to_list(10000)
    employer_attendance = await db.employer_attendance.find(query, {"_id": 0}).to_list(10000)
    
    # Sort by date
    worker_attendance.sort(key=lambda x: x['date'], reverse=True)
    employer_attendance.sort(key=lambda x: x['date'], reverse=True)
    
    return {
        "worker_attendance": worker_attendance,
        "employer_attendance": employer_attendance
    }

@api_router.get("/reports/financial-detailed")
async def get_financial_detailed_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Detailed financial report"""
    query = {"contractor_id": current_user.id}
    
    # Payments collected
    payments = await db.payment_collections.find(query, {"_id": 0}).to_list(10000)
    payments.sort(key=lambda x: x['date'], reverse=True)
    
    # Wage settlements
    settlements = await db.wage_settlements.find(query, {"_id": 0}).to_list(10000)
    for s in settlements:
        if isinstance(s['settlement_date'], str):
            s['settlement_date'] = datetime.fromisoformat(s['settlement_date'])
    settlements.sort(key=lambda x: x['settlement_date'], reverse=True)
    
    # Advances
    advances = await db.advances.find(query, {"_id": 0}).to_list(10000)
    advances.sort(key=lambda x: x['date'], reverse=True)
    
    # Extra charges
    charges = await db.extra_charges.find(query, {"_id": 0}).to_list(10000)
    charges.sort(key=lambda x: x['date'], reverse=True)
    
    return {
        "payments_collected": payments,
        "wage_settlements": settlements,
        "advances": advances,
        "extra_charges": charges
    }

@api_router.get("/reports/export-all")
async def export_all_data(
    current_user: User = Depends(get_active_subscription_user)
):
    """Export all data as CSV"""
    import zipfile
    from io import BytesIO
    
    # Create in-memory zip file
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # Workers
        workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
        workers_csv = io.StringIO()
        if workers:
            writer = csv.DictWriter(workers_csv, fieldnames=workers[0].keys())
            writer.writeheader()
            writer.writerows(workers)
            zip_file.writestr('workers.csv', workers_csv.getvalue())
        
        # Employers
        employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
        employers_csv = io.StringIO()
        if employers:
            writer = csv.DictWriter(employers_csv, fieldnames=employers[0].keys())
            writer.writeheader()
            writer.writerows(employers)
            zip_file.writestr('employers.csv', employers_csv.getvalue())
        
        # Attendance
        attendance = await db.worker_attendance.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
        attendance_csv = io.StringIO()
        if attendance:
            writer = csv.DictWriter(attendance_csv, fieldnames=attendance[0].keys())
            writer.writeheader()
            writer.writerows(attendance)
            zip_file.writestr('attendance.csv', attendance_csv.getvalue())
        
        # Payments
        payments = await db.payment_collections.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
        payments_csv = io.StringIO()
        if payments:
            writer = csv.DictWriter(payments_csv, fieldnames=payments[0].keys())
            writer.writeheader()
            writer.writerows(payments)
            zip_file.writestr('payments.csv', payments_csv.getvalue())
        
        # Settlements
        settlements = await db.wage_settlements.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
        settlements_csv = io.StringIO()
        if settlements:
            # Convert datetime to string
            for s in settlements:
                if isinstance(s.get('settlement_date'), datetime):
                    s['settlement_date'] = s['settlement_date'].isoformat()
            writer = csv.DictWriter(settlements_csv, fieldnames=settlements[0].keys())
            writer.writeheader()
            writer.writerows(settlements)
            zip_file.writestr('settlements.csv', settlements_csv.getvalue())
        
        # Advances
        advances = await db.advances.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
        advances_csv = io.StringIO()
        if advances:
            writer = csv.DictWriter(advances_csv, fieldnames=advances[0].keys())
            writer.writeheader()
            writer.writerows(advances)
            zip_file.writestr('advances.csv', advances_csv.getvalue())
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=all_data_export.zip"}
    )

# ============ REPORTS & EXPORT ============

@api_router.get("/reports/commissions")
async def get_commissions_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get commissions data for viewing"""
    query = {"contractor_id": current_user.id}
    
    # Get all commissions and filter in Python (since dates are DD-MM-YYYY strings)
    commissions = await db.commissions.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date and end_date:
        # Convert DD-MM-YYYY to comparable format
        def date_to_comparable(date_str):
            try:
                day, month, year = date_str.split('-')
                return int(year) * 10000 + int(month) * 100 + int(day)
            except:
                return 0
        
        start_comparable = date_to_comparable(start_date)
        end_comparable = date_to_comparable(end_date)
        
        filtered_commissions = []
        for comm in commissions:
            comm_comparable = date_to_comparable(comm.get('date', ''))
            if start_comparable <= comm_comparable <= end_comparable:
                filtered_commissions.append(comm)
        commissions = filtered_commissions
    
    # Return with IDs so frontend can map names, including workers_count
    result = []
    for commission in commissions:
        result.append({
            "id": commission.get('id'),
            "date": commission['date'],
            "worker_id": commission.get('worker_id'),
            "employer_id": commission.get('employer_id'),
            "payment_from_employer": commission.get('payment_from_employer', 0),
            "wage_to_worker": commission.get('wage_to_worker', 0),
            "commission_amount": round(commission.get('commission_amount', 0), 2),
            "workers_count": commission.get('workers_count')  # Include workers_count for SUMMARY records
        })
    
    return result

@api_router.get("/reports/attendance")
async def get_attendance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get attendance data for viewing"""
    query = {"contractor_id": current_user.id}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    attendance = await db.worker_attendance.find(query, {"_id": 0}).to_list(10000)
    
    # Enrich with worker names
    result = []
    for att in attendance:
        worker = await db.workers.find_one({"id": att['worker_id']}, {"_id": 0})
        employer = await db.employers.find_one({"id": att['employer_id']}, {"_id": 0})
        result.append({
            "date": att['date'],
            "worker_name": worker['name'] if worker else "Unknown",
            "employer_name": employer['name'] if employer else "Unknown",
            "status": att['status'],
            "wage_earned": att.get('wage_earned', 0)
        })
    
    return result

@api_router.get("/reports/payments")
async def get_payments_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get payment collection data"""
    query = {"contractor_id": current_user.id}
    payments = await db.payment_collections.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if needed
    if start_date and end_date:
        payments = [p for p in payments if start_date <= p.get('date', '')[:10] <= end_date]
    
    return payments

@api_router.get("/reports/wages")
async def get_wages_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get wage settlement data"""
    query = {"contractor_id": current_user.id}
    settlements = await db.wage_settlements.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if needed
    if start_date and end_date:
        settlements = [s for s in settlements if start_date <= s.get('date', '')[:10] <= end_date]
    
    return settlements

@api_router.get("/reports/advances")
async def get_advances_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get advance data"""
    # Worker advances
    worker_advances = await db.advances.find(
        {"contractor_id": current_user.id},
        {"_id": 0}
    ).to_list(10000)
    
    # Employer advances
    employer_advances = await db.employer_advances.find(
        {"contractor_id": current_user.id},
        {"_id": 0}
    ).to_list(10000)
    
    # Filter by date if needed
    if start_date and end_date:
        worker_advances = [a for a in worker_advances if start_date <= a.get('date', '')[:10] <= end_date]
        employer_advances = [a for a in employer_advances if start_date <= a.get('date', '')[:10] <= end_date]
    
    return {
        "worker_advances": worker_advances,
        "employer_advances": employer_advances
    }

# ============ DASHBOARD ROUTES ============

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(
    filter: str = "today",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Comprehensive dashboard statistics with accurate worker-present logic
    """
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%d-%m-%Y")

    # ---- Date filter logic ----
    date_filter = {}
    if filter == "today":
        date_filter = {"date": today_str}
    elif filter == "week":
        week_ago = (now - timedelta(days=7)).strftime("%d-%m-%Y")
        date_filter = {"date": {"$gte": week_ago, "$lte": today_str}}
    elif filter == "month":
        month_ago = (now - timedelta(days=30)).strftime("%d-%m-%Y")
        date_filter = {"date": {"$gte": month_ago, "$lte": today_str}}
    elif filter == "year":
        year_ago = (now - timedelta(days=365)).strftime("%d-%m-%Y")
        date_filter = {"date": {"$gte": year_ago, "$lte": today_str}}
    elif filter == "custom" and start_date and end_date:
        date_filter = {"date": {"$gte": start_date, "$lte": end_date}}

    # ---- Static counts ----
    active_workers_count = await db.workers.count_documents({
        "contractor_id": current_user.id,
        "status": "Active"
    })
    
    active_employers_count = await db.employers.count_documents({
        "contractor_id": current_user.id,
        "status": "Active"
    })

    # ---- Worker attendance stats ----
    worker_attendance_query = {"contractor_id": current_user.id, "mode": "worker", **date_filter}
    worker_attendance = await db.attendance.find(worker_attendance_query, {"_id": 0}).to_list(10000)

    present_ids = {a['worker_id'] for a in worker_attendance if a.get('status') in ['Present', 'Late']}
    absent_ids = {a['worker_id'] for a in worker_attendance if a.get('status') == 'Absent'}

    workers_present = len(present_ids)
    workers_absent = len(absent_ids)

    # ---- Employers active today (those who have any attendance record) ----
    employers_today = len(set(a['employer_id'] for a in worker_attendance if a.get('employer_id')))

    # ---- Commissions ----
    commissions_query = {"contractor_id": current_user.id, **date_filter}
    commissions = await db.commissions.find(commissions_query, {"_id": 0}).to_list(10000)
    total_commission = sum(c.get('commission_amount', 0) for c in commissions)

    # ---- Payments collected ----
    payments_query = {"contractor_id": current_user.id}
    payments = await db.payment_collections.find(payments_query, {"_id": 0}).to_list(10000)
    total_payments_collected = sum(p.get('amount', 0) for p in payments)

    # ---- Wages settled ----
    settlements_query = {"contractor_id": current_user.id}
    settlements = await db.wage_settlements.find(settlements_query, {"_id": 0}).to_list(10000)
    total_wages_settled = sum(s.get('amount_paid', 0) for s in settlements)

    # ---- Pending amounts ----
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)

    total_pending_wages = sum(w.get('pending_settlement', 0) for w in workers)
    total_pending_payments = sum(e.get('pending_payment', 0) for e in employers)

    # ---- Days worked ----
    total_days_worked = len(set(a['date'] for a in worker_attendance))

    return {
        "active_workers": active_workers_count,
        "active_employers": active_employers_count,
        "workers_present": workers_present,
        "workers_present_ratio": f"{workers_present}/{active_workers_count}",
        "workers_absent": workers_absent,
        "employers_today": employers_today,
        "total_commission": round(total_commission, 2),
        "payments_collected": round(total_payments_collected, 2),
        "wages_settled": round(total_wages_settled, 2),
        "pending_wages": round(total_pending_wages, 2),
        "pending_payments": round(total_pending_payments, 2),
        "days_worked": total_days_worked,
        "filter": filter,
        "date_range": f"{date_filter.get('date', {}).get('$gte', today_str)} to {date_filter.get('date', {}).get('$lte', today_str)}"
                      if isinstance(date_filter.get("date"), dict) else today_str
    }


@api_router.get("/dashboard/activities")
async def get_activities(limit: int = 5, current_user: User = Depends(get_active_subscription_user)):
    today = datetime.now(timezone.utc)
    today_start = datetime.combine(today.date(), datetime.min.time()).replace(tzinfo=timezone.utc)
    
    activities = await db.activities.find({
        "contractor_id": current_user.id,
        "date": {"$gte": today_start.isoformat()}
    }, {"_id": 0}).sort("date", -1).to_list(limit)
    
    for activity in activities:
        if isinstance(activity['date'], str):
            activity['date'] = datetime.fromisoformat(activity['date'])
    
    return activities

# ============ REPORTS & EXPORT ============

# Individual report data endpoints (for viewing in frontend)
@api_router.get("/reports/attendance")
async def get_attendance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get attendance data for viewing"""
    query = {"contractor_id": current_user.id}
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    attendance = await db.worker_attendance.find(query, {"_id": 0}).to_list(10000)
    
    # Enrich with worker names
    result = []
    for att in attendance:
        worker = await db.workers.find_one({"id": att['worker_id']}, {"_id": 0})
        employer = await db.employers.find_one({"id": att['employer_id']}, {"_id": 0})
        result.append({
            "date": att['date'],
            "worker_name": worker['name'] if worker else "Unknown",
            "employer_name": employer['name'] if employer else "Unknown",
            "status": att['status'],
            "wage_earned": att.get('wage_earned', 0)
        })
    
    return result

@api_router.get("/reports/payments")
async def get_payments_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get payments data for viewing"""
    query = {"contractor_id": current_user.id}
    
    payments = await db.payment_collections.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, "%d-%m-%Y").replace(tzinfo=timezone.utc)
        end_dt = datetime.strptime(end_date, "%d-%m-%Y").replace(tzinfo=timezone.utc)
        payments = [
            p for p in payments 
            if start_dt <= datetime.fromisoformat(p['date']) <= end_dt
        ]
    
    # Enrich with employer names
    result = []
    for payment in payments:
        employer = await db.employers.find_one({"id": payment['employer_id']}, {"_id": 0})
        result.append({
            "date": payment['date'] if isinstance(payment['date'], str) else payment['date'].isoformat(),
            "employer_name": employer['name'] if employer else "Unknown",
            "amount": payment['amount'],
            "payment_mode": payment['payment_mode'],
            "remarks": payment.get('remarks', '')
        })
    
    return result

@api_router.get("/reports/wages")
async def get_wages_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get wages settlement data for viewing"""
    query = {"contractor_id": current_user.id}
    
    settlements = await db.wage_settlements.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, "%d-%m-%Y").replace(tzinfo=timezone.utc)
        end_dt = datetime.strptime(end_date, "%d-%m-%Y").replace(tzinfo=timezone.utc)
        settlements = [
            s for s in settlements 
            if start_dt <= datetime.fromisoformat(s['settlement_date']) <= end_dt
        ]
    
    # Enrich with worker names
    result = []
    for settlement in settlements:
        worker = await db.workers.find_one({"id": settlement['worker_id']}, {"_id": 0})
        result.append({
            "date": settlement['settlement_date'] if isinstance(settlement['settlement_date'], str) else settlement['settlement_date'].isoformat(),
            "worker_name": worker['name'] if worker else "Unknown",
            "amount_paid": settlement['amount_paid'],
            "advance_deducted": settlement['advance_deducted'],
            "charges_deducted": settlement['charges_deducted'],
            "settlement_type": settlement['settlement_type']
        })
    
    return result

@api_router.get("/reports/advances")
async def get_advances_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get advances data for viewing"""
    query = {"contractor_id": current_user.id}
    
    advances = await db.advances.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, "%d-%m-%Y").replace(tzinfo=timezone.utc)
        end_dt = datetime.strptime(end_date, "%d-%m-%Y").replace(tzinfo=timezone.utc)
        advances = [
            a for a in advances 
            if start_dt <= datetime.fromisoformat(a['date']) <= end_dt
        ]
    
    # Enrich with worker names
    result = []
    for advance in advances:
        worker = await db.workers.find_one({"id": advance['worker_id']}, {"_id": 0})
        result.append({
            "date": advance['date'] if isinstance(advance['date'], str) else advance['date'].isoformat(),
            "worker_name": worker['name'] if worker else "Unknown",
            "amount": advance['amount'],
            "purpose": advance.get('purpose', '')
        })
    
    return result

@api_router.get("/reports/commissions")
async def get_commissions_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get commission earnings from database with proper IDs"""
    query = {"contractor_id": current_user.id}
    
    # Get all commissions and filter in Python (since dates are DD-MM-YYYY strings)
    commissions = await db.commissions.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date and end_date:
        # Convert DD-MM-YYYY to comparable format
        def date_to_comparable(date_str):
            try:
                day, month, year = date_str.split('-')
                return int(year) * 10000 + int(month) * 100 + int(day)
            except:
                return 0
        
        start_comparable = date_to_comparable(start_date)
        end_comparable = date_to_comparable(end_date)
        
        filtered_commissions = []
        for comm in commissions:
            comm_comparable = date_to_comparable(comm.get('date', ''))
            if start_comparable <= comm_comparable <= end_comparable:
                filtered_commissions.append(comm)
        commissions = filtered_commissions
    
    # Return with IDs so frontend can map names
    result = []
    for commission in commissions:
        result.append({
            "id": commission.get('id'),
            "date": commission['date'],
            "worker_id": commission['worker_id'],
            "employer_id": commission['employer_id'],
            "payment_from_employer": commission.get('payment_from_employer', 0),
            "wage_to_worker": commission.get('wage_to_worker', 0),
            "commission_amount": round(commission.get('commission_amount', 0), 2),
            "workers_count": commission.get('workers_count')  # Include workers_count for SUMMARY records
        })
    
    return result

# ============ ATTENDANCE ANALYTICS REPORTS ============

def _dedupe_attendance_by_worker_date(records):
    """
    Deduplicate attendance records by (worker_id, date).
    
    A single physical day of attendance can be persisted to multiple sources:
      1. db.worker_attendance (new system)
      2. db.attendance with mode='worker' (bulk endpoint)
      3. db.attendance with mode='employer' and worker_id inside selected_workers
         (which we expand into per-worker rows for analytics)
    
    Without dedup, 1 day of presence would be counted as 2-3 days. We keep only the
    first occurrence per (worker_id, date), preferring records with an explicit
    non-Present status (so 'Absent' / 'Late' takes precedence over derived 'Present').
    """
    seen = {}
    deduped = []
    for r in records:
        worker_id = r.get('worker_id')
        date = r.get('date')
        if not worker_id or not date:
            continue
        key = (worker_id, date)
        if key in seen:
            existing_idx = seen[key]
            existing = deduped[existing_idx]
            existing_status = (existing.get('status') or '').strip().lower()
            new_status = (r.get('status') or '').strip().lower()
            # Prefer explicit non-Present statuses (Absent, Late) over default Present
            if existing_status == 'present' and new_status in ('absent', 'late'):
                deduped[existing_idx] = r
            continue
        seen[key] = len(deduped)
        deduped.append(r)
    return deduped

@api_router.get("/reports/attendance-analytics")
async def get_attendance_analytics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Get comprehensive attendance analytics including:
    - Overall stats (percentage, present/absent counts)
    - Day-of-week trends
    - Smart predictions for worker absences
    """
    # Fetch from both new and old attendance collections
    query = {"contractor_id": current_user.id}
    
    # Get from new worker_attendance collection
    new_attendance = await db.worker_attendance.find(query, {"_id": 0}).to_list(10000)
    
    # Get from old attendance collection (all modes - employer records contain worker data)
    old_query = {"contractor_id": current_user.id}
    old_attendance = await db.attendance.find(old_query, {"_id": 0}).to_list(10000)
    
    # Transform old attendance to match new schema
    attendance_records = []
    
    # Add new attendance records
    attendance_records.extend(new_attendance)
    
    # Transform and add old attendance records
    for old_rec in old_attendance:
        if old_rec.get('selected_workers'):
            # This is an employer attendance with selected workers
            for worker_id in old_rec.get('selected_workers', []):
                attendance_records.append({
                    'worker_id': worker_id,
                    'employer_id': old_rec.get('employer_id', ''),
                    'date': old_rec.get('date', ''),
                    'status': 'Present',  # Old records don't have status, assume Present
                    'wage_earned': old_rec.get('wage_amount', 0) / max(len(old_rec.get('selected_workers', [])), 1)
                })
        elif old_rec.get('worker_id'):
            # Direct worker attendance record
            attendance_records.append({
                'worker_id': old_rec.get('worker_id'),
                'employer_id': old_rec.get('employer_id', ''),
                'date': old_rec.get('date', ''),
                'status': old_rec.get('status', 'Present'),
                'wage_earned': old_rec.get('wage_amount', 0)
            })
    
    if not attendance_records:
        return {
            "overall_stats": {},
            "day_of_week_trends": {},
            "predictions": [],
            "message": "No attendance data found for the selected period"
        }
    
    # Deduplicate: same worker on same day may exist in multiple collections
    attendance_records = _dedupe_attendance_by_worker_date(attendance_records)
    
    # Filter by date range at endpoint level if provided
    filtered_records = attendance_records
    if start_date and end_date:
        from attendance_analytics import convert_dd_mm_yyyy_to_iso
        filtered_records = []
        for r in attendance_records:
            record_date = r.get('date', '')
            if record_date:
                # Convert to ISO format for proper comparison
                iso_date = convert_dd_mm_yyyy_to_iso(record_date)
                if start_date <= iso_date <= end_date:
                    filtered_records.append(r)
    
    # Pass filtered records to analytics functions
    overall_stats = get_overall_attendance_stats(filtered_records, start_date, end_date)
    day_trends = analyze_day_of_week_trends(filtered_records)
    
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(1000)
    worker_map = {w['id']: w['name'] for w in workers}
    
    all_predictions = []
    for worker_id, worker_name in worker_map.items():
        worker_attendance = [r for r in filtered_records if r.get('worker_id') == worker_id]
        if len(worker_attendance) >= 5:
            predictions = predict_worker_absences(worker_attendance, worker_name)
            all_predictions.extend(predictions)
    
    all_predictions.sort(key=lambda x: (x['confidence'] == 'High', x['absence_rate']), reverse=True)
    
    return {
        "overall_stats": overall_stats,
        "day_of_week_trends": day_trends,
        "predictions": all_predictions[:10],
        "total_predictions": len(all_predictions)
    }

@api_router.get("/reports/worker-attendance-analysis/{worker_id}")
async def get_worker_attendance_analysis(
    worker_id: str,
    period: str = "all",
    current_user: User = Depends(get_active_subscription_user)
):
    """Get detailed attendance analysis for a specific worker"""
    worker = await db.workers.find_one(
        {"id": worker_id, "contractor_id": current_user.id},
        {"_id": 0}
    )
    
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    
    # Fetch from both collections
    new_query = {"contractor_id": current_user.id, "worker_id": worker_id}
    new_attendance = await db.worker_attendance.find(new_query, {"_id": 0}).to_list(10000)
    
    # Get from old attendance collection (all modes)
    old_query = {"contractor_id": current_user.id}
    old_attendance = await db.attendance.find(old_query, {"_id": 0}).to_list(10000)
    
    # Transform old attendance records
    attendance_records = list(new_attendance)
    for old_rec in old_attendance:
        if old_rec.get('selected_workers') and worker_id in old_rec.get('selected_workers', []):
            attendance_records.append({
                'worker_id': worker_id,
                'employer_id': old_rec.get('employer_id', ''),
                'date': old_rec.get('date', ''),
                'status': 'Present',
                'wage_earned': old_rec.get('wage_amount', 0) / max(len(old_rec.get('selected_workers', [])), 1)
            })
        elif old_rec.get('worker_id') == worker_id:
            attendance_records.append({
                'worker_id': worker_id,
                'employer_id': old_rec.get('employer_id', ''),
                'date': old_rec.get('date', ''),
                'status': old_rec.get('status', 'Present'),
                'wage_earned': old_rec.get('wage_amount', 0)
            })
    
    # Deduplicate: same worker on same day may exist in multiple collections
    attendance_records = _dedupe_attendance_by_worker_date(attendance_records)
    
    if not attendance_records:
        return {
            "worker_info": worker,
            "message": "No attendance data found for this worker"
        }
    
    total_days = len(attendance_records)
    present_days = sum(1 for r in attendance_records if r.get('status') in ["Present", "Late"])
    absent_days = sum(1 for r in attendance_records if r.get('status') == "Absent")
    attendance_percentage = round((present_days / total_days) * 100, 2) if total_days > 0 else 0
    
    absent_dates = [r.get('date') for r in attendance_records if r.get('status') == "Absent"]
    patterns = detect_absence_patterns(attendance_records)
    worker_day_trends = analyze_day_of_week_trends(attendance_records)
    predictions = predict_worker_absences(attendance_records, worker.get('name', 'Unknown'))
    
    monthly_stats = {}
    for record in attendance_records:
        try:
            from attendance_analytics import convert_dd_mm_yyyy_to_iso
            iso_date = convert_dd_mm_yyyy_to_iso(record.get('date', ''))
            date_obj = datetime.fromisoformat(iso_date)
            month_key = date_obj.strftime('%Y-%m')
            if month_key not in monthly_stats:
                monthly_stats[month_key] = {"present": 0, "absent": 0, "total": 0}
            
            monthly_stats[month_key]["total"] += 1
            if record.get('status') in ["Present", "Late"]:
                monthly_stats[month_key]["present"] += 1
            elif record.get('status') == "Absent":
                monthly_stats[month_key]["absent"] += 1
        except:
            continue
    
    for month, stats in monthly_stats.items():
        stats["attendance_percentage"] = round((stats["present"] / stats["total"]) * 100, 2) if stats["total"] > 0 else 0
    
    return {
        "worker_info": {
            "id": worker.get('id'),
            "name": worker.get('name'),
            "phone": worker.get('phone'),
            "status": worker.get('status')
        },
        "overall_stats": {
            "total_days": total_days,
            "present_days": present_days,
            "absent_days": absent_days,
            "attendance_percentage": attendance_percentage,
            "period": period
        },
        "absent_dates": absent_dates,
        "monthly_breakdown": monthly_stats,
        "patterns": patterns,
        "day_of_week_trends": worker_day_trends,
        "predictions": predictions
    }

@api_router.get("/reports/attendance-leaderboard")
async def get_attendance_leaderboard(
    period: str = "monthly",
    limit: int = 50,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get attendance leaderboard showing top-performing workers"""
    # Fetch from both collections
    query = {"contractor_id": current_user.id}
    new_attendance = await db.worker_attendance.find(query, {"_id": 0}).to_list(10000)
    
    # Get all old attendance (employer mode records contain worker data in selected_workers)
    old_query = {"contractor_id": current_user.id}
    old_attendance = await db.attendance.find(old_query, {"_id": 0}).to_list(10000)
    
    # Transform and combine
    attendance_records = list(new_attendance)
    for old_rec in old_attendance:
        if old_rec.get('selected_workers'):
            for worker_id in old_rec.get('selected_workers', []):
                attendance_records.append({
                    'worker_id': worker_id,
                    'date': old_rec.get('date', ''),
                    'status': 'Present'
                })
        elif old_rec.get('worker_id'):
            attendance_records.append({
                'worker_id': old_rec.get('worker_id'),
                'date': old_rec.get('date', ''),
                'status': old_rec.get('status', 'Present')
            })
    
    # Deduplicate: same worker on same day may exist in multiple collections
    attendance_records = _dedupe_attendance_by_worker_date(attendance_records)
    
    # Filter by period BEFORE calculating leaderboard
    from attendance_analytics import convert_dd_mm_yyyy_to_iso
    now = datetime.now(timezone.utc)
    
    if period == "monthly" or period == "yearly":
        filtered_records = []
        for record in attendance_records:
            try:
                iso_date = convert_dd_mm_yyyy_to_iso(record.get('date', ''))
                record_date = datetime.fromisoformat(iso_date)
                
                if period == "monthly":
                    if record_date.year == now.year and record_date.month == now.month:
                        filtered_records.append(record)
                elif period == "yearly":
                    if record_date.year == now.year:
                        filtered_records.append(record)
            except:
                continue
        attendance_records = filtered_records
    
    workers = await db.workers.find(
        {"contractor_id": current_user.id, "status": "Active"},
        {"_id": 0}
    ).to_list(1000)
    
    if not attendance_records or not workers:
        return {
            "leaderboard": [],
            "period": period,
            "message": "No data available for leaderboard"
        }
    
    leaderboard = calculate_worker_leaderboard(attendance_records, workers, period)
    leaderboard = leaderboard[:limit]
    
    period_label = ""
    if period == "monthly":
        period_label = now.strftime("%B %Y")
    elif period == "yearly":
        period_label = str(now.year)
    else:
        period_label = "All Time"
    
    return {
        "leaderboard": leaderboard,
        "period": period,
        "period_label": period_label,
        "total_workers": len(leaderboard)
    }

@api_router.get("/reports/attendance-summary")
async def get_attendance_summary(
    current_user: User = Depends(get_active_subscription_user)
):
    """Get quick attendance summary for dashboard"""
    now = datetime.now(timezone.utc)
    today = now.date().isoformat()
    start_of_week = (now - timedelta(days=now.weekday())).date().isoformat()
    start_of_month = now.replace(day=1).date().isoformat()
    
    today_records = await db.worker_attendance.find({
        "contractor_id": current_user.id,
        "date": today
    }, {"_id": 0}).to_list(1000)
    
    today_present = sum(1 for r in today_records if r.get('status') in ["Present", "Late"])
    today_absent = sum(1 for r in today_records if r.get('status') == "Absent")
    
    week_records = await db.worker_attendance.find({
        "contractor_id": current_user.id,
        "date": {"$gte": start_of_week}
    }, {"_id": 0}).to_list(10000)
    
    week_present = sum(1 for r in week_records if r.get('status') in ["Present", "Late"])
    week_absent = sum(1 for r in week_records if r.get('status') == "Absent")
    week_total = len(week_records)
    week_percentage = round((week_present / week_total) * 100, 2) if week_total > 0 else 0
    
    month_records = await db.worker_attendance.find({
        "contractor_id": current_user.id,
        "date": {"$gte": start_of_month}
    }, {"_id": 0}).to_list(10000)
    
    month_present = sum(1 for r in month_records if r.get('status') in ["Present", "Late"])
    month_absent = sum(1 for r in month_records if r.get('status') == "Absent")
    month_total = len(month_records)
    month_percentage = round((month_present / month_total) * 100, 2) if month_total > 0 else 0
    
    return {
        "today": {
            "present": today_present,
            "absent": today_absent,
            "total": len(today_records)
        },
        "this_week": {
            "present": week_present,
            "absent": week_absent,
            "total": week_total,
            "attendance_percentage": week_percentage
        },
        "this_month": {
            "present": month_present,
            "absent": month_absent,
            "total": month_total,
            "attendance_percentage": month_percentage
        }
    }

# Export endpoints for Excel downloads
@api_router.get("/reports/export-attendance")
async def export_attendance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Export attendance report as CSV"""
    data = await get_attendance_report(start_date, end_date, current_user)
    
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=attendance_report.csv"}
    )

@api_router.get("/reports/export-payments")
async def export_payments_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Export payments report as CSV"""
    data = await get_payments_report(start_date, end_date, current_user)
    
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=payments_report.csv"}
    )

@api_router.get("/reports/export-wages")
async def export_wages_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Export wages report as CSV"""
    data = await get_wages_report(start_date, end_date, current_user)
    
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=wages_report.csv"}
    )

@api_router.get("/reports/export-advances")
async def export_advances_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Export advances report as CSV"""
    data = await get_advances_report(start_date, end_date, current_user)
    
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=advances_report.csv"}
    )

@api_router.get("/reports/export-commissions")
async def export_commissions_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Export commissions report as CSV"""
    data = await get_commissions_report(start_date, end_date, current_user)
    
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=commissions_report.csv"}
    )

@api_router.get("/reports/export-workers")
async def export_workers(current_user: User = Depends(get_active_subscription_user)):
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(1000)
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=['name', 'phone_number', 'wage_per_day', 'wage_from_employer', 'status', 'pending_settlement', 'advance_paid'])
    writer.writeheader()
    
    for worker in workers:
        writer.writerow({
            'name': worker['name'],
            'phone_number': worker['phone_number'],
            'wage_per_day': worker['wage_per_day'],
            'wage_from_employer': worker.get('wage_from_employer', 500),
            'status': worker['status'],
            'pending_settlement': worker['pending_settlement'],
            'advance_paid': worker.get('advance_paid', worker.get('advance_balance', 0))
        })
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=workers.csv"}
    )

@api_router.get("/reports/export-employers")
async def export_employers(current_user: User = Depends(get_active_subscription_user)):
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(1000)
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=['name', 'phone_number', 'work_location', 'status', 'pending_payment'])
    writer.writeheader()
    
    for employer in employers:
        writer.writerow({
            'name': employer['name'],
            'phone_number': employer['phone_number'],
            'work_location': employer.get('work_location', ''),
            'status': employer['status'],
            'pending_payment': employer['pending_payment']
        })
    
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=employers.csv"}
    )

@api_router.get("/reports/summary")
async def get_reports_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    query = {"contractor_id": current_user.id}
    if start_date and end_date:
        query["date" if "date" in query else "settlement_date"] = {"$gte": start_date, "$lte": end_date}
    
    # Wage settlements
    settlements = await db.wage_settlements.find(query if not start_date else {"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
    total_wages_settled = sum(s['amount_paid'] for s in settlements)
    total_advances_deducted = sum(s['advance_deducted'] for s in settlements)
    total_charges_deducted = sum(s['charges_deducted'] for s in settlements)
    
    # Payment collections
    payments = await db.payment_collections.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
    total_payments_collected = sum(p['amount'] for p in payments)
    
    # Advances given
    advances = await db.advances.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(10000)
    total_advances_given = sum(a['amount'] for a in advances)
    
    # Worker charges
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(1000)
    total_worker_charges = sum(
        sum(charge['amount'] for charge in worker.get('charges', []))
        for worker in workers
    )
    
    return {
        "total_wages_settled": round(total_wages_settled, 2),
        "total_payments_collected": round(total_payments_collected, 2),
        "total_advances_given": round(total_advances_given, 2),
        "total_advances_deducted": round(total_advances_deducted, 2),
        "total_charges_collected": round(total_worker_charges, 2),
        "total_charges_deducted": round(total_charges_deducted, 2),
        "settlements_count": len(settlements),
        "payments_count": len(payments)
    }

# ============ NEW COMPREHENSIVE REPORT ENDPOINTS ============

@api_router.get("/reports/workers-complete")
async def get_workers_complete_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Complete worker report with all details including wages, advances, attendance
    """
    workers = await db.workers.find(
        {"contractor_id": current_user.id, "status": "Active"}, 
        {"_id": 0}
    ).to_list(1000)
    
    result = []
    for worker in workers:
        # Get ALL attendance records for this worker
        attendance_query = {
            "contractor_id": current_user.id,
            "worker_id": worker['id']
        }
        
        all_attendance = await db.worker_attendance.find(attendance_query, {"_id": 0}).to_list(10000)
        
        # Filter by date if provided (using proper date parsing)
        if start_date and end_date:
            def date_to_comparable(date_str):
                try:
                    day, month, year = date_str.split('-')
                    return int(year) * 10000 + int(month) * 100 + int(day)
                except:
                    return 0
            
            start_comparable = date_to_comparable(start_date)
            end_comparable = date_to_comparable(end_date)
            
            worker_attendance = [
                att for att in all_attendance 
                if start_comparable <= date_to_comparable(att.get('date', '')) <= end_comparable
            ]
        else:
            worker_attendance = all_attendance
        
        # Calculate attendance stats
        present_days = len([a for a in worker_attendance if a.get('status') == 'Present'])
        absent_days = len([a for a in worker_attendance if a.get('status') == 'Absent'])
        total_earnings = sum(a.get('wage_earned', 0) for a in worker_attendance)
        
        # Get unique employers
        unique_employers = len(set(a.get('employer_id') for a in worker_attendance if a.get('employer_id')))
        
        # Calculate attendance rate
        total_days = present_days + absent_days
        attendance_rate = (present_days / total_days * 100) if total_days > 0 else 0
        
        result.append({
            "id": worker['id'],
            "name": worker['name'],
            "phone_number": worker.get('phone_number', ''),
            "wage_per_day": worker.get('wage_per_day', 0),
            "wage_from_employer": worker.get('wage_from_employer', 0),
            "pending_settlement": worker.get('pending_settlement', 0),
            "advance_paid": worker.get('advance_paid', 0),
            "status": worker.get('status', 'Active'),
            "present_days": present_days,
            "absent_days": absent_days,
            "total_days": total_days,
            "total_earnings": total_earnings,
            "avg_earnings_per_day": total_earnings / present_days if present_days > 0 else 0,
            "unique_employers": unique_employers,
            "attendance_rate": round(attendance_rate, 1)
        })
    
    return result

@api_router.get("/reports/employers-complete")
async def get_employers_complete_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Complete employer report with all details including payments, advances, activity
    """
    employers = await db.employers.find(
        {"contractor_id": current_user.id, "status": "Active"}, 
        {"_id": 0}
    ).to_list(1000)
    
    result = []
    for employer in employers:
        # Get ALL employer attendance records
        attendance_query = {
            "contractor_id": current_user.id,
            "employer_id": employer['id']
        }
        
        all_attendance = await db.employer_attendance.find(attendance_query, {"_id": 0}).to_list(10000)
        
        # Filter by date if provided (using proper date parsing)
        if start_date and end_date:
            def date_to_comparable(date_str):
                try:
                    day, month, year = date_str.split('-')
                    return int(year) * 10000 + int(month) * 100 + int(day)
                except:
                    return 0
            
            start_comparable = date_to_comparable(start_date)
            end_comparable = date_to_comparable(end_date)
            
            employer_attendance = [
                att for att in all_attendance 
                if start_comparable <= date_to_comparable(att.get('date', '')) <= end_comparable
            ]
        else:
            employer_attendance = all_attendance
        
        # Calculate activity stats
        total_days_with_workers = len(employer_attendance)
        total_worker_days = sum(att.get('workers_count', 0) for att in employer_attendance)
        total_cost = sum(att.get('total_amount', 0) for att in employer_attendance)
        
        # Get unique workers
        unique_workers = set()
        for att in employer_attendance:
            if att.get('selected_workers'):
                unique_workers.update(att['selected_workers'])
        
        avg_workers_per_day = total_worker_days / total_days_with_workers if total_days_with_workers > 0 else 0
        avg_cost_per_day = total_cost / total_days_with_workers if total_days_with_workers > 0 else 0
        
        # Get payments
        payment_query = {
            "contractor_id": current_user.id,
            "employer_id": employer['id']
        }
        payments = await db.payment_collections.find(payment_query, {"_id": 0}).to_list(1000)
        total_collected = sum(p['amount'] for p in payments)
        
        result.append({
            "id": employer['id'],
            "name": employer['name'],
            "phone_number": employer.get('phone_number', ''),
            "pending_payment": employer.get('pending_payment', 0),
            "advance_received": employer.get('advance_received', 0),
            "status": employer.get('status', 'Active'),
            "total_days_with_workers": total_days_with_workers,
            "total_worker_days": total_worker_days,
            "unique_workers": len(unique_workers),
            "total_cost": total_cost,
            "total_collected": total_collected,
            "avg_workers_per_day": round(avg_workers_per_day, 1),
            "avg_cost_per_day": round(avg_cost_per_day, 2)
        })
    
    return result

@api_router.get("/reports/attendance-complete")
async def get_attendance_complete_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Complete attendance report combining worker and employer attendance
    """
    print(f"🔍 Attendance Report Request - User: {current_user.id}, Dates: {start_date} to {end_date}")
    
    # Get ALL attendance first (without date filter in query)
    query = {"contractor_id": current_user.id}
    
    # Get worker attendance
    all_worker_attendance = await db.worker_attendance.find(query, {"_id": 0}).to_list(10000)
    print(f"📊 Found {len(all_worker_attendance)} worker attendance records")
    
    # Get employer attendance
    all_employer_attendance = await db.employer_attendance.find(query, {"_id": 0}).to_list(10000)
    print(f"📊 Found {len(all_employer_attendance)} employer attendance records")
    
    # Filter by date if provided (using proper date parsing)
    if start_date and end_date:
        def date_to_comparable(date_str):
            try:
                day, month, year = date_str.split('-')
                return int(year) * 10000 + int(month) * 100 + int(day)
            except:
                return 0
        
        start_comparable = date_to_comparable(start_date)
        end_comparable = date_to_comparable(end_date)
        
        worker_attendance = [
            att for att in all_worker_attendance 
            if start_comparable <= date_to_comparable(att.get('date', '')) <= end_comparable
        ]
        employer_attendance = [
            att for att in all_employer_attendance 
            if start_comparable <= date_to_comparable(att.get('date', '')) <= end_comparable
        ]
        print(f"📅 After date filtering: {len(worker_attendance)} worker, {len(employer_attendance)} employer records")
    else:
        worker_attendance = all_worker_attendance
        employer_attendance = all_employer_attendance
    
    # Get workers and employers for name mapping
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    
    worker_map = {w['id']: w['name'] for w in workers}
    employer_map = {e['id']: e['name'] for e in employers}
    
    result = []
    
    # Process worker attendance
    for att in worker_attendance:
        result.append({
            "date": att['date'],
            "worker_id": att['worker_id'],
            "worker_name": worker_map.get(att['worker_id'], 'Unknown'),
            "employer_id": att.get('employer_id', ''),
            "employer_name": employer_map.get(att.get('employer_id', ''), 'Unknown'),
            "status": att['status'],
            "wage_earned": att.get('wage_earned', 0),
            "type": "worker"
        })
    
    # Process employer attendance
    for att in employer_attendance:
        # If specific workers selected, create entry for each
        if att.get('selected_workers') and len(att['selected_workers']) > 0:
            for worker_id in att['selected_workers']:
                result.append({
                    "date": att['date'],
                    "worker_id": worker_id,
                    "worker_name": worker_map.get(worker_id, 'Unknown'),
                    "employer_id": att['employer_id'],
                    "employer_name": employer_map.get(att['employer_id'], 'Unknown'),
                    "status": "Present",
                    "wage_earned": att.get('payment_per_worker', 0),
                    "type": "employer"
                })
        else:
            # Generic entry - use payment_per_worker if available
            wage_earned = 0
            if att.get('payment_per_worker') and att.get('payment_per_worker', 0) > 0:
                wage_earned = att.get('payment_per_worker', 0) * att.get('workers_count', 0)
            else:
                wage_earned = att.get('total_amount', att.get('wage_amount', 0))
            
            result.append({
                "date": att['date'],
                "worker_id": None,
                "worker_name": f"{att.get('workers_count', 0)} workers",
                "employer_id": att['employer_id'],
                "employer_name": employer_map.get(att['employer_id'], 'Unknown'),
                "status": "Present",
                "wage_earned": wage_earned,
                "workers_count": att.get('workers_count', 0),
                "type": "employer"
            })
    
    # Sort by date
    result.sort(key=lambda x: x['date'], reverse=True)
    
    return result

@api_router.get("/reports/commissions-complete")
async def get_commissions_complete_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Complete commission report with worker and employer names
    """
    query = {"contractor_id": current_user.id}
    
    commissions = await db.commissions.find(query, {"_id": 0}).to_list(10000)
    
    # Filter by date if provided
    if start_date and end_date:
        def date_to_comparable(date_str):
            try:
                day, month, year = date_str.split('-')
                return int(year) * 10000 + int(month) * 100 + int(day)
            except:
                return 0
        
        start_comparable = date_to_comparable(start_date)
        end_comparable = date_to_comparable(end_date)
        
        commissions = [
            c for c in commissions 
            if start_comparable <= date_to_comparable(c.get('date', '')) <= end_comparable
        ]
    
    # Get workers and employers for name mapping
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    
    worker_map = {w['id']: w['name'] for w in workers}
    employer_map = {e['id']: e['name'] for e in employers}
    
    result = []
    for comm in commissions:
        result.append({
            "id": comm.get('id'),
            "date": comm['date'],
            "worker_id": comm['worker_id'],
            "worker_name": worker_map.get(comm['worker_id'], 'Unknown'),
            "employer_id": comm['employer_id'],
            "employer_name": employer_map.get(comm['employer_id'], 'Unknown'),
            "payment_from_employer": comm.get('payment_from_employer', 0),
            "wage_to_worker": comm.get('wage_to_worker', 0),
            "commission_amount": round(comm.get('commission_amount', 0), 2)
        })
    
    return result

@api_router.get("/reports/advances-complete")
async def get_advances_complete_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Complete advances report for both workers and employers
    """
    # Get worker advances
    worker_advances = await db.advances.find(
        {"contractor_id": current_user.id}, 
        {"_id": 0}
    ).to_list(10000)
    
    # Get employer advances
    employer_advances = await db.employer_advances.find(
        {"contractor_id": current_user.id}, 
        {"_id": 0}
    ).to_list(10000)
    
    # Get workers and employers for name mapping
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    
    worker_map = {w['id']: w['name'] for w in workers}
    employer_map = {e['id']: e['name'] for e in employers}
    
    result = {
        "worker_advances": [],
        "employer_advances": [],
        "summary": {
            "total_worker_advances": 0,
            "total_employer_advances": 0,
            "net_advance_position": 0
        }
    }
    
    # Process worker advances
    for adv in worker_advances:
        adv_date = adv['date']
        if isinstance(adv_date, datetime):
            adv_date_str = adv_date.strftime("%d-%m-%Y")
        else:
            adv_date_str = str(adv_date)
        
        result["worker_advances"].append({
            "id": adv.get('id'),
            "date": adv_date_str,
            "worker_id": adv['worker_id'],
            "worker_name": worker_map.get(adv['worker_id'], 'Unknown'),
            "amount": adv['amount'],
            "purpose": adv.get('purpose', '')
        })
    
    # Process employer advances
    for adv in employer_advances:
        adv_date = adv['date']
        if isinstance(adv_date, datetime):
            adv_date_str = adv_date.strftime("%d-%m-%Y")
        else:
            adv_date_str = str(adv_date)
        
        result["employer_advances"].append({
            "id": adv.get('id'),
            "date": adv_date_str,
            "employer_id": adv['employer_id'],
            "employer_name": employer_map.get(adv['employer_id'], 'Unknown'),
            "amount": adv['amount'],
            "purpose": adv.get('purpose', '')
        })
    
    # Calculate totals
    result["summary"]["total_worker_advances"] = sum(a['amount'] for a in worker_advances)
    result["summary"]["total_employer_advances"] = sum(a['amount'] for a in employer_advances)
    result["summary"]["net_advance_position"] = result["summary"]["total_employer_advances"] - result["summary"]["total_worker_advances"]
    
    return result

@api_router.get("/reports/business-overview-complete")
async def get_business_overview_complete(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Complete business overview with all financial metrics
    """
    # Get all workers and employers
    workers = await db.workers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(1000)
    employers = await db.employers.find({"contractor_id": current_user.id}, {"_id": 0}).to_list(1000)
    
    # Get commissions
    commissions_query = {"contractor_id": current_user.id}
    commissions = await db.commissions.find(commissions_query, {"_id": 0}).to_list(10000)
    
    # Filter commissions by date
    if start_date and end_date:
        def date_to_comparable(date_str):
            try:
                day, month, year = date_str.split('-')
                return int(year) * 10000 + int(month) * 100 + int(day)
            except:
                return 0
        
        start_comparable = date_to_comparable(start_date)
        end_comparable = date_to_comparable(end_date)
        
        commissions = [
            c for c in commissions 
            if start_comparable <= date_to_comparable(c.get('date', '')) <= end_comparable
        ]
    
    # Calculate totals
    total_workers = len(workers)
    active_workers = len([w for w in workers if w.get('status') == 'Active'])
    total_employers = len(employers)
    active_employers = len([e for e in employers if e.get('status') == 'Active'])
    
    total_pending_wages = sum(w.get('pending_settlement', 0) for w in workers)
    total_pending_payments = sum(e.get('pending_payment', 0) for e in employers)
    total_worker_advances = sum(w.get('advance_paid', 0) for w in workers)
    total_employer_advances = sum(e.get('advance_received', 0) for e in employers)
    total_commissions = sum(c.get('commission_amount', 0) for c in commissions)
    
    return {
        "workers": {
            "total": total_workers,
            "active": active_workers,
            "pending_wages": total_pending_wages,
            "advances_given": total_worker_advances
        },
        "employers": {
            "total": total_employers,
            "active": active_employers,
            "pending_payments": total_pending_payments,
            "advances_received": total_employer_advances
        },
        "financial": {
            "total_commissions": round(total_commissions, 2),
            "net_receivable": round(total_pending_payments - total_pending_wages, 2),
            "net_advance_position": round(total_employer_advances - total_worker_advances, 2)
        }
    }

# ============ ADMIN ROUTES ============

@api_router.get("/admin/users", dependencies=[Depends(get_current_admin)])
async def get_all_users():
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    for user in users:
        if isinstance(user['created_at'], str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

@api_router.put("/admin/users/{user_id}/activate", dependencies=[Depends(get_current_admin)])
async def activate_user(user_id: str):
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "subscription_plan": "contractor",
            "subscription_status": "active",
            "plan_start_date": datetime.now(timezone.utc).isoformat(),
            "plan_end_date": (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
        }}
    )
    return {"message": "User activated successfully"}

@api_router.put("/admin/users/{user_id}/deactivate", dependencies=[Depends(get_current_admin)])
async def deactivate_user(user_id: str):
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"subscription_status": "inactive", "is_active": False}}
    )
    return {"message": "User deactivated successfully"}

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ---------------- Default Subscription Plans ----------------
DEFAULT_SUBSCRIPTION_PLANS = [
    {
        "name": "Contractor Plus",
        "price": 499,
        "duration_days": 30,
        "max_workers": 50,
        "max_employers": 25,
        "description": "Perfect for small to medium contractors",
        "features": [
            "Track up to 50 workers",
            "Manage up to 25 employers",
            "Dual attendance tracking (worker & employer view)",
            "Payment & wage management",
            "Advance tracking & settlements",
            "Commission auto-calculation",
            "Export to CSV / Excel / PDF",
            "Dashboard analytics",
            "Mobile responsive interface",
            "7-day free trial available",
            "Email support",
        ],
        "is_active": True,
        "trial_eligible": True,
    },
    {
        "name": "Contractor Pro",
        "price": 999,
        "duration_days": 30,
        "max_workers": 250,
        "max_employers": 100,
        "description": "Ideal for growing contractor businesses",
        "features": [
            "Everything in Contractor Plus",
            "Track up to 250 workers",
            "Manage up to 100 employers",
            "Advanced reporting & insights",
            "Priority support (faster response)",
            "Bulk attendance import",
            "Custom date-range work history",
            "Branded PDF / Excel exports",
            "Multiple payment method tracking",
            "7-day free trial available",
        ],
        "is_active": True,
        "trial_eligible": True,
    },
    {
        "name": "Enterprise",
        "price": 1999,
        "duration_days": 30,
        "max_workers": None,   # Unlimited
        "max_employers": None,  # Unlimited
        "description": "For large-scale contractor operations",
        "features": [
            "Everything in Contractor Pro",
            "Unlimited workers",
            "Unlimited employers",
            "24/7 priority support",
            "Dedicated account manager",
            "Custom integrations (on request)",
            "Audit-grade work history exports",
            "GST-ready invoice templates",
            "Onboarding & training assistance",
            "SLA-backed uptime",
        ],
        "is_active": True,
        "trial_eligible": False,
        "coming_soon": True,
    },
]


async def seed_default_plans():
    """
    Idempotent default-plan seeder, run on startup.

    For each canonical plan:
      • Creates the plan if missing.
      • If the plan exists but its `features` list is shorter/empty
        compared to the canonical list, fills in the canonical features
        and the description (only when those fields are blank/sparse).
      • Never overwrites admin-set price, duration, or limits.
    """
    try:
        for tpl in DEFAULT_SUBSCRIPTION_PLANS:
            now_iso = datetime.now(timezone.utc).isoformat()
            existing = await db.subscription_plans.find_one({"name": tpl["name"]}, {"_id": 0})
            if not existing:
                plan_doc = {
                    "id": str(uuid.uuid4()),
                    "name": tpl["name"],
                    "price": tpl["price"],
                    "duration_days": tpl["duration_days"],
                    "max_workers": tpl["max_workers"],
                    "max_employers": tpl["max_employers"],
                    "features": list(tpl["features"]),
                    "description": tpl["description"],
                    "is_active": tpl["is_active"],
                    "trial_eligible": tpl.get("trial_eligible", False),
                    "coming_soon": tpl.get("coming_soon", False),
                    "created_by": "system",
                    "created_at": now_iso,
                    "updated_at": now_iso,
                }
                await db.subscription_plans.insert_one(plan_doc)
                logger.info(f"[seed_default_plans] Created default plan: {tpl['name']}")
                continue

            # Plan exists — only enrich missing/incomplete fields. Don't disturb pricing/limits.
            updates: Dict[str, Any] = {}
            existing_features = existing.get("features") or []
            if len(existing_features) < len(tpl["features"]):
                updates["features"] = list(tpl["features"])
            if not (existing.get("description") or "").strip():
                updates["description"] = tpl["description"]
            if "trial_eligible" not in existing:
                updates["trial_eligible"] = tpl.get("trial_eligible", False)
            if "coming_soon" not in existing:
                updates["coming_soon"] = tpl.get("coming_soon", False)
            if updates:
                updates["updated_at"] = now_iso
                await db.subscription_plans.update_one(
                    {"id": existing["id"]},
                    {"$set": updates},
                )
                logger.info(
                    f"[seed_default_plans] Enriched plan '{tpl['name']}' with: {list(updates.keys())}"
                )
    except Exception as exc:
        logger.warning(f"[seed_default_plans] Skipped due to error: {exc}")


@app.on_event("startup")
async def startup_db():
    await seed_admin()
    await seed_default_plans()

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# ============ NEW UNIFIED ATTENDANCE SYSTEM ============

class AttendanceEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # DD-MM-YYYY
    contractor_id: str
    employer_id: Optional[str] = None  # For employer-based or worker assignment
    worker_id: Optional[str] = None  # For worker-based
    workers_count: Optional[int] = None  # For employer-based summary
    selected_workers: List[str] = []  # ✅ Specific worker IDs assigned to employer
    status: Optional[str] = None  # Present/Absent/Late for worker-based
    wage_amount: float = 0.0
    payment_per_worker: Optional[float] = None  # ✅ Payment per worker for employer-based attendance (when workers not selected)
    additional_charges: List[Dict[str, Any]] = []
    extra_payment_per_worker: float = 0.0  # ✅ Per-worker bonus (e.g. skill premium) - paid by employer, received by worker
    extra_payment_reason: Optional[str] = None  # ✅ Reason for extra per-worker payment
    additional_charges_as_commission: bool = False  # ✅ When True, additional_charges count toward commission
    remarks: Optional[str] = None  # ✅ Optional remarks field
    mode: str  # "employer" or "worker"
    created_by: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployerAttendanceEntry(BaseModel):
    employer_id: str
    workers_count: int
    selected_workers: List[str] = []  # Optional worker IDs
    payment_per_worker: float = 500.0
    additional_charges: float = 0.0
    charge_description: Optional[str] = None
    extra_payment_per_worker: float = 0.0  # ✅ Per-worker bonus (e.g. skill premium) - paid by employer AND received by worker
    extra_payment_reason: Optional[str] = None  # ✅ Reason for extra per-worker payment
    additional_charges_as_commission: bool = False  # ✅ Toggle: count additional_charges as commission
    remarks: Optional[str] = None  # ✅ Optional remarks

class WorkerAttendanceEntry(BaseModel):
    worker_id: str
    status: str  # Present, Absent, Late
    employer_id: Optional[str] = None

class AttendanceBulkCreate(BaseModel):
    date: str  # DD-MM-YYYY
    mode: str  # "employer" or "worker"
    employer_entries: List[EmployerAttendanceEntry] = []
    worker_entries: List[WorkerAttendanceEntry] = []

class ContractorPreference(BaseModel):
    contractor_id: str
    attendance_mode: str = "Hybrid"  # EmployerBased, WorkerBased, Hybrid
    default_worker_wage: float = 450.0
    default_employer_rate: float = 500.0

class ContractorPreferenceUpdate(BaseModel):
    attendance_mode: Optional[str] = None
    default_worker_wage: Optional[float] = None
    default_employer_rate: Optional[float] = None

# ============ CONTRACTOR PREFERENCES ROUTES ============

@api_router.get("/preferences", response_model=ContractorPreference)
async def get_contractor_preferences(current_user: User = Depends(get_active_subscription_user)):
    """Get contractor preferences, creating default if not exists"""
    preference = await db.contractor_preferences.find_one({"contractor_id": current_user.id}, {"_id": 0})
    
    if not preference:
        # Create default preferences
        default_pref = ContractorPreference(
            contractor_id=current_user.id,
            attendance_mode="Hybrid",
            default_worker_wage=450.0,
            default_employer_rate=500.0
        )
        pref_dict = default_pref.model_dump()
        await db.contractor_preferences.insert_one(pref_dict)
        return default_pref
    
    return ContractorPreference(**preference)

@api_router.put("/preferences", response_model=ContractorPreference)
async def update_contractor_preferences(
    preferences: ContractorPreferenceUpdate,
    current_user: User = Depends(get_active_subscription_user)
):
    """Update contractor preferences"""
    update_data = {}
    for k, v in preferences.model_dump().items():
        if v is not None:
            update_data[k] = v
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    # Check if preferences exist
    existing = await db.contractor_preferences.find_one({"contractor_id": current_user.id})
    
    if existing:
        await db.contractor_preferences.update_one(
            {"contractor_id": current_user.id},
            {"$set": update_data}
        )
    else:
        # Create new preferences with defaults
        default_pref = ContractorPreference(contractor_id=current_user.id)
        pref_dict = default_pref.model_dump()
        pref_dict.update(update_data)
        await db.contractor_preferences.insert_one(pref_dict)
    
    # Return updated preferences
    updated = await db.contractor_preferences.find_one({"contractor_id": current_user.id}, {"_id": 0})
    return ContractorPreference(**updated)

@api_router.get("/attendance/available-workers")
async def get_available_workers(
    date: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Returns available unassigned workers count and list.
    Available = Total Active Workers - Workers assigned in employer-mode - Workers marked absent in worker-mode
    """
    try:
        # Convert ISO → "27-10-2025"
        parsed_date = datetime.fromisoformat(date.replace("Z", "")).strftime("%d-%m-%Y")
    except Exception:
        parsed_date = date

    # ✅ 1. Get total active workers count
    total_active_workers = await db.workers.count_documents({
        "contractor_id": current_user.id,
        "status": "Active"
    })

    # ✅ 2. Calculate total workers assigned in employer-based attendance
    # This includes both named workers (selected_workers) and unnamed workers (workers_count)
    employer_attendance_records = await db.attendance.find({
        "contractor_id": current_user.id,
        "date": parsed_date,
        "mode": "employer"
    }, {"_id": 0, "workers_count": 1, "selected_workers": 1}).to_list(1000)

    # Count total assigned workers from employer-based records
    total_assigned_in_employer_mode = sum(
        record.get("workers_count", 0) for record in employer_attendance_records
    )

    # ✅ 3. Get workers with specific assignments (to exclude from available list)
    assigned_worker_ids = []
    for record in employer_attendance_records:
        if record.get("selected_workers"):
            assigned_worker_ids.extend(record["selected_workers"])

    # ✅ 4. Get workers marked absent (in worker-mode)
    absent_worker_ids = await db.attendance.distinct("worker_id", {
        "contractor_id": current_user.id,
        "date": parsed_date,
        "status": "Absent"
    })

    # Count absent workers (filter out any None/empty values)
    total_absent_workers = len([wid for wid in absent_worker_ids if wid])

    # ✅ 5. Merge excluded IDs (for the worker list)
    excluded_ids = [wid for wid in (assigned_worker_ids + absent_worker_ids) if wid]

    # ✅ 6. Calculate available count
    # Available = Total - Assigned in Employer Mode - Marked Absent in Worker Mode
    available_count = total_active_workers - total_assigned_in_employer_mode - total_absent_workers

    # ✅ 7. Fetch workers not specifically assigned or absent (for dropdown)
    available_workers = await db.workers.find(
        {
            "contractor_id": current_user.id,
            "status": "Active",
            "id": {"$nin": excluded_ids},
        },
        {"_id": 0}
    ).to_list(1000)

    return {
        "available_workers": available_workers,
        "available_count": max(0, available_count),  # Ensure non-negative
        "total_active_workers": total_active_workers,
        "total_assigned_in_employer_mode": total_assigned_in_employer_mode,
        "total_absent_workers": total_absent_workers
    }

@api_router.post("/attendance/bulk")
async def create_bulk_attendance(
    attendance_data: AttendanceBulkCreate,
    current_user: User = Depends(get_active_subscription_user)
):
    """Unified endpoint for both employer and worker attendance, with commission logic"""
    results = {
        "employer_entries": [],
        "worker_entries": [],
        "warnings": []
    }

    date = attendance_data.date

    # Validate total workers
    active_workers_count = await db.workers.count_documents({
        "contractor_id": current_user.id,
        "status": "Active"
    })

    # Process Employer-Based Entries
    if attendance_data.mode in ["employer", "Hybrid"] and attendance_data.employer_entries:
        total_workers_sent = sum(entry.workers_count for entry in attendance_data.employer_entries)

        if total_workers_sent > active_workers_count:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot send {total_workers_sent} workers. Only {active_workers_count} active workers available."
            )

        for employer_entry in attendance_data.employer_entries:
                # Prevent over-assignment and duplicate workers
            assigned_ids = await get_assigned_worker_ids(current_user.id, date)

            # ✅ SELF (Own Work): when the contractor sends workers to his own jobs,
            # there is no employer to bill and no commission to earn. Worker still gets wage.
            is_self_work = (employer_entry.employer_id or "").upper() == "SELF"

            if employer_entry.selected_workers:
                overlap = [wid for wid in employer_entry.selected_workers if wid in assigned_ids]
                if overlap:
                    raise HTTPException(
                    status_code=400,
                    detail=f"Some selected workers are already assigned today: {overlap}"
                    )

            existing = await db.attendance.find_one({
                "contractor_id": current_user.id,
                "employer_id": employer_entry.employer_id,
                "date": date,
                "mode": "employer"
            })
            
            # Capture employer attendance ID for commission records
            employer_attendance_id = None

            # ✅ NEW: If workers_count == total active workers, auto-select ALL workers
            auto_selected_all = False
            if employer_entry.workers_count == active_workers_count and (not employer_entry.selected_workers or len(employer_entry.selected_workers) == 0):
                # Get all active workers
                all_workers = await db.workers.find({
                    "contractor_id": current_user.id,
                    "status": "Active"
                }, {"_id": 0, "id": 1}).to_list(1000)
                
                employer_entry.selected_workers = [w["id"] for w in all_workers]
                auto_selected_all = True
            
            # ✅ PRIORITY 1: If specific workers are selected, use their individual wage_from_employer rates
            if employer_entry.selected_workers and len(employer_entry.selected_workers) > 0:
                # Sum individual worker rates from employer (defaults / payment_per_worker NOT used here)
                total_from_workers = 0.0
                for worker_id in employer_entry.selected_workers:
                    worker = await db.workers.find_one({"id": worker_id}, {"_id": 0})
                    if worker:
                        total_from_workers += float(worker.get("wage_from_employer", 0) or 0)
                # ✅ Extra payment per worker (skill bonus) is paid by employer for each selected worker
                extra_total = float(employer_entry.extra_payment_per_worker or 0.0) * len(employer_entry.selected_workers)
                total_amount = total_from_workers + extra_total + employer_entry.additional_charges
            # ✅ PRIORITY 2: If payment_per_worker is provided and no specific workers selected, use it
            elif employer_entry.payment_per_worker and employer_entry.payment_per_worker > 0:
                # Use manual payment_per_worker * count
                extra_total = float(employer_entry.extra_payment_per_worker or 0.0) * employer_entry.workers_count
                total_amount = (employer_entry.workers_count * employer_entry.payment_per_worker) + extra_total + employer_entry.additional_charges
            else:
                # Fallback: Use manual payment_per_worker * count (even if 0)
                extra_total = float(employer_entry.extra_payment_per_worker or 0.0) * employer_entry.workers_count
                total_amount = (employer_entry.workers_count * (employer_entry.payment_per_worker or 0)) + extra_total + employer_entry.additional_charges

            if existing:
                # Update existing employer record
                old_amount = existing.get('wage_amount', 0)
                diff = total_amount - old_amount
                employer_attendance_id = existing['id']  # Capture attendance ID for commission records

                # ✅ Delete old commissions for this employer+date (will be recalculated below)
                await db.commissions.delete_many({
                    "contractor_id": current_user.id,
                    "employer_id": employer_entry.employer_id,
                    "date": date
                })

                # ✅ Clear employer_id from worker attendance records (will be reassigned below)
                await db.attendance.update_many(
                    {
                        "contractor_id": current_user.id,
                        "employer_id": employer_entry.employer_id,
                        "date": date,
                        "mode": "worker"
                    },
                    {"$set": {"employer_id": ""}}
                )

                await db.attendance.update_one(
                    {"id": existing['id']},
                    {"$set": {
                        "workers_count": employer_entry.workers_count,
                        "selected_workers": employer_entry.selected_workers,  # ✅ Save selected workers
                        "payment_per_worker": employer_entry.payment_per_worker,  # ✅ Save payment_per_worker
                        "wage_amount": total_amount,
                        "additional_charges": [{
                            "amount": employer_entry.additional_charges,
                            "description": employer_entry.charge_description or ""
                        }] if employer_entry.additional_charges > 0 else [],
                        "extra_payment_per_worker": float(employer_entry.extra_payment_per_worker or 0.0),  # ✅ Save extra per-worker bonus
                        "extra_payment_reason": employer_entry.extra_payment_reason or "",  # ✅ Save reason
                        "additional_charges_as_commission": bool(employer_entry.additional_charges_as_commission),  # ✅ Toggle
                        "remarks": employer_entry.remarks or "",  # ✅ Save remarks
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )

                await db.employers.update_one(
                    {"id": employer_entry.employer_id},
                    {"$inc": {"pending_payment": diff}}
                ) if not is_self_work else None
            else:
                # Create new employer record
                entry = AttendanceEntry(
                    date=date,
                    contractor_id=current_user.id,
                    employer_id=employer_entry.employer_id,
                    workers_count=employer_entry.workers_count,
                    selected_workers=employer_entry.selected_workers,  # ✅ Save selected workers
                    payment_per_worker=employer_entry.payment_per_worker,  # ✅ Save payment_per_worker
                    wage_amount=total_amount,
                    additional_charges=[{
                        "amount": employer_entry.additional_charges,
                        "description": employer_entry.charge_description or ""
                    }] if employer_entry.additional_charges > 0 else [],
                    extra_payment_per_worker=float(employer_entry.extra_payment_per_worker or 0.0),  # ✅ Save extra per-worker bonus
                    extra_payment_reason=employer_entry.extra_payment_reason or None,  # ✅ Save reason
                    additional_charges_as_commission=bool(employer_entry.additional_charges_as_commission),  # ✅ Toggle
                    remarks=employer_entry.remarks or "",  # ✅ Save remarks
                    mode="employer",
                    created_by=current_user.id
                )

                entry_dict = entry.model_dump()
                entry_dict['created_at'] = entry_dict['created_at'].isoformat()
                entry_dict['updated_at'] = entry_dict['updated_at'].isoformat()
                await db.attendance.insert_one(entry_dict)
                employer_attendance_id = entry_dict['id']  # Capture attendance ID for commission records

                await db.employers.update_one(
                    {"id": employer_entry.employer_id},
                    {"$inc": {"pending_payment": total_amount}}
                ) if not is_self_work else None

            # === NEW === Commission + Worker Attendance Auto Marking ===
            # ✅ Use individual worker rates when available
            if employer_entry.selected_workers:
                for worker_id in employer_entry.selected_workers:
                    # Check if worker attendance already exists
                    existing_worker_att = await db.attendance.find_one({
                        "contractor_id": current_user.id,
                        "worker_id": worker_id,
                        "date": date,
                        "mode": "worker"
                    })
                    
                    worker = await db.workers.find_one({"id": worker_id})
                    if not worker:
                        continue

                    # ✅ When specific workers are selected, ALWAYS use the worker's own
                    # wage_from_employer rate. Ignore form's `payment_per_worker` (which is
                    # often auto-filled from preferences default and would otherwise leak in
                    # as a wrong rate). User rule: "if a specific worker is selected, the
                    # default set wage is not used".
                    base_wage = float(worker.get("wage_per_day", 0))
                    base_payment_from_employer = float(worker.get("wage_from_employer", 0) or 0)

                    # ✅ Extra per-worker bonus (e.g. skill premium): paid by employer AND received by worker
                    extra = float(employer_entry.extra_payment_per_worker or 0.0)
                    payment_from_employer = base_payment_from_employer + extra
                    wage_to_worker = base_wage + extra

                    # Worker actually receives wage_per_day + extra (added to settlement)
                    wage_amount = wage_to_worker
                    # Commission = (employer payment + extra) - (worker wage + extra) = base diff (extra is pass-through)
                    commission_amount = max(payment_from_employer - wage_to_worker, 0.0)
                    
                    # If worker already marked present but employer_id is empty, update it
                    if existing_worker_att:
                        existing_employer_id = existing_worker_att.get("employer_id", "")
                        
                        # If already assigned to an employer, skip
                        if existing_employer_id and existing_employer_id != "":
                            results["warnings"].append(f"Skipped worker {worker_id} (already assigned to employer)")
                            continue
                        
                        # Update employer_id for worker already marked present
                        await db.attendance.update_one(
                            {"id": existing_worker_att['id']},
                            {"$set": {
                                "employer_id": employer_entry.employer_id,
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }}
                        )
                        
                        # Insert/update commission record (skip entirely for SELF / own-work)
                        if is_self_work:
                            results["warnings"].append(f"Assigned worker {worker_id} to own work")
                            continue
                        existing_commission = await db.commissions.find_one({
                            "contractor_id": current_user.id,
                            "worker_id": worker_id,
                            "date": date
                        })
                        
                        if existing_commission:
                            await db.commissions.update_one(
                                {"_id": existing_commission['_id']},
                                {"$set": {
                                    "employer_id": employer_entry.employer_id,
                                    "payment_from_employer": payment_from_employer,  # ✅ Includes extra_payment_per_worker
                                    "wage_to_worker": wage_to_worker,  # ✅ Includes extra_payment_per_worker
                                    "commission_amount": commission_amount,
                                    "extra_payment_per_worker": extra,
                                    "extra_payment_reason": employer_entry.extra_payment_reason or None
                                }}
                            )
                        else:
                            commission_doc = {
                                "contractor_id": current_user.id,
                                "date": date,
                                "employer_id": employer_entry.employer_id,
                                "worker_id": worker_id,
                                "payment_from_employer": payment_from_employer,  # ✅ base + extra
                                "wage_to_worker": wage_to_worker,  # ✅ base + extra
                                "commission_amount": commission_amount,
                                "extra_payment_per_worker": extra,
                                "extra_payment_reason": employer_entry.extra_payment_reason or None,
                                "attendance_id": employer_attendance_id if employer_attendance_id else str(uuid.uuid4()),
                                "created_at": datetime.now(timezone.utc).isoformat(),
                            }
                            await db.commissions.insert_one(commission_doc)
                        
                        results["warnings"].append(f"Assigned worker {worker_id} to employer")
                        continue

                    # Insert new worker attendance
                    worker_att = AttendanceEntry(
                        date=date,
                        contractor_id=current_user.id,
                        worker_id=worker_id,
                        employer_id=employer_entry.employer_id,
                        status="Present",
                        wage_amount=wage_amount,
                        mode="worker",
                        created_by=current_user.id
                    )

                    att_dict = worker_att.model_dump()
                    att_dict['created_at'] = att_dict['created_at'].isoformat()
                    att_dict['updated_at'] = att_dict['updated_at'].isoformat()
                    await db.attendance.insert_one(att_dict)

                    await db.workers.update_one(
                        {"id": worker_id},
                        {"$inc": {"pending_settlement": wage_amount}}
                    )

                    # ✅ Insert commission record - Calculate using worker's wage_from_employer or payment_per_worker
                    if is_self_work:
                        # SELF / own-work: no commission row
                        continue
                    commission_doc = {
                        "contractor_id": current_user.id,
                        "date": date,
                        "employer_id": employer_entry.employer_id,
                        "worker_id": worker_id,
                        "payment_from_employer": payment_from_employer,  # ✅ base + extra
                        "wage_to_worker": wage_to_worker,  # ✅ base + extra
                        "commission_amount": commission_amount,
                        "extra_payment_per_worker": extra,
                        "extra_payment_reason": employer_entry.extra_payment_reason or None,
                        "attendance_id": employer_attendance_id if employer_attendance_id else str(uuid.uuid4()),
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                    await db.commissions.insert_one(commission_doc)

                # ✅ ADDITIONAL CHARGES → recorded as commission ONLY when the toggle is on
                # When `additional_charges_as_commission` is True, additional_charges count toward
                # contractor commission. When False (default), they only inflate the employer's bill.
                if (not is_self_work
                        and employer_entry.additional_charges_as_commission
                        and employer_entry.additional_charges
                        and employer_entry.additional_charges > 0):
                    existing_addl = await db.commissions.find_one({
                        "contractor_id": current_user.id,
                        "employer_id": employer_entry.employer_id,
                        "date": date,
                        "worker_id": "ADDITIONAL_CHARGES"
                    })
                    addl_doc = {
                        "contractor_id": current_user.id,
                        "date": date,
                        "employer_id": employer_entry.employer_id,
                        "worker_id": "ADDITIONAL_CHARGES",  # marker — not a real worker
                        "payment_from_employer": float(employer_entry.additional_charges),
                        "wage_to_worker": 0.0,
                        "commission_amount": float(employer_entry.additional_charges),
                        "charge_description": employer_entry.charge_description or "",
                        "attendance_id": employer_attendance_id if employer_attendance_id else str(uuid.uuid4()),
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                    if existing_addl:
                        await db.commissions.update_one({"_id": existing_addl["_id"]}, {"$set": addl_doc})
                    else:
                        await db.commissions.insert_one(addl_doc)
            else:
                # ✅ When workers are NOT selected:
                # - payment_from_employer = entered payment_per_worker (or default_employer_rate)
                # - wage_to_worker = default_worker_wage from preferences (only sensible value
                #   when no specific worker is known)
                # - commission per worker = payment_from_employer - default_worker_wage
                if is_self_work:
                    # SELF / own-work: no commission row when no specific worker
                    results["employer_entries"].append(employer_entry.model_dump())
                    continue
                preference = await db.contractor_preferences.find_one({"contractor_id": current_user.id}, {"_id": 0})
                if preference:
                    default_wage = float(preference.get("default_worker_wage", 450.0))
                    default_employer_rate = float(preference.get("default_employer_rate", 500.0))
                else:
                    default_wage = 450.0
                    default_employer_rate = 500.0

                # Extra per-worker bonus: paid by employer AND received by worker → pass-through, cancels in commission
                extra = float(employer_entry.extra_payment_per_worker or 0.0)

                if employer_entry.payment_per_worker and employer_entry.payment_per_worker > 0:
                    base_payment_from_employer = float(employer_entry.payment_per_worker)
                else:
                    base_payment_from_employer = default_employer_rate

                payment_from_employer = base_payment_from_employer + extra
                wage_to_worker = default_wage + extra
                commission_per_worker = max(payment_from_employer - wage_to_worker, 0.0)

                total_commission = commission_per_worker * employer_entry.workers_count

                # Create summary commission record (using special marker "SUMMARY" for worker_id)
                # Check if summary commission already exists for this employer/date
                existing_summary_commission = await db.commissions.find_one({
                    "contractor_id": current_user.id,
                    "employer_id": employer_entry.employer_id,
                    "date": date,
                    "worker_id": "SUMMARY"  # Special marker for summary record when workers not selected
                })
                
                if not existing_summary_commission and total_commission >= 0:
                    commission_doc = {
                        "contractor_id": current_user.id,
                        "date": date,
                        "employer_id": employer_entry.employer_id,
                        "worker_id": "SUMMARY",  # Special marker for summary record when workers not selected
                        "payment_from_employer": payment_from_employer,
                        "wage_to_worker": wage_to_worker,
                        "commission_amount": total_commission,
                        "extra_payment_per_worker": extra,
                        "extra_payment_reason": employer_entry.extra_payment_reason or None,
                        "attendance_id": employer_attendance_id if employer_attendance_id else str(uuid.uuid4()),
                        "workers_count": employer_entry.workers_count,
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                    await db.commissions.insert_one(commission_doc)

                # ✅ ADDITIONAL CHARGES → recorded as commission ONLY when the toggle is on
                if (employer_entry.additional_charges_as_commission
                        and employer_entry.additional_charges
                        and employer_entry.additional_charges > 0):
                    existing_addl = await db.commissions.find_one({
                        "contractor_id": current_user.id,
                        "employer_id": employer_entry.employer_id,
                        "date": date,
                        "worker_id": "ADDITIONAL_CHARGES"
                    })
                    addl_doc = {
                        "contractor_id": current_user.id,
                        "date": date,
                        "employer_id": employer_entry.employer_id,
                        "worker_id": "ADDITIONAL_CHARGES",
                        "payment_from_employer": float(employer_entry.additional_charges),
                        "wage_to_worker": 0.0,
                        "commission_amount": float(employer_entry.additional_charges),
                        "charge_description": employer_entry.charge_description or "",
                        "attendance_id": employer_attendance_id if employer_attendance_id else str(uuid.uuid4()),
                        "created_at": datetime.now(timezone.utc).isoformat(),
                    }
                    if existing_addl:
                        await db.commissions.update_one({"_id": existing_addl["_id"]}, {"$set": addl_doc})
                    else:
                        await db.commissions.insert_one(addl_doc)

            results["employer_entries"].append(employer_entry.model_dump())

    # Process Worker-Based Entries (manual marking mode)
    if attendance_data.mode in ["worker", "Hybrid"] and attendance_data.worker_entries:
        # ✅ VALIDATION: Present workers in worker-mode must be >= workers assigned in employer-mode
        # Get total workers assigned in employer-mode for this date
        employer_records = await db.attendance.find({
            "contractor_id": current_user.id,
            "date": date,
            "mode": "employer"
        }, {"_id": 0, "workers_count": 1}).to_list(1000)
        
        total_assigned_in_employer_mode = sum(
            record.get("workers_count", 0) for record in employer_records
        )
        
        # Count present workers from current submission + existing worker-mode records
        present_workers_count = sum(
            1 for entry in attendance_data.worker_entries if entry.status == "Present"
        )
        
        # Add existing worker-mode present records not being updated
        existing_worker_ids_in_submission = {entry.worker_id for entry in attendance_data.worker_entries}
        existing_present_count = await db.attendance.count_documents({
            "contractor_id": current_user.id,
            "date": date,
            "mode": "worker",
            "status": "Present",
            "worker_id": {"$nin": list(existing_worker_ids_in_submission)}
        })
        
        total_present_workers = present_workers_count + existing_present_count
        
        if total_present_workers < total_assigned_in_employer_mode:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot save worker attendance. You have assigned {total_assigned_in_employer_mode} workers to employers, but only marking {total_present_workers} workers as present. Please ensure at least {total_assigned_in_employer_mode} workers are marked present in worker-based attendance."
            )
        
        for worker_entry in attendance_data.worker_entries:
            existing = await db.attendance.find_one({
                "contractor_id": current_user.id,
                "worker_id": worker_entry.worker_id,
                "date": date,
                "mode": "worker"
            })

            worker = await db.workers.find_one({"id": worker_entry.worker_id})
            if not worker:
                continue

            base_wage = float(worker.get("wage_per_day", 0))
            if worker_entry.status == "Present":
                wage_amount = base_wage
            elif worker_entry.status == "Late":
                wage_amount = base_wage * 0.75
            else:
                wage_amount = 0.0

            # ✅ NOTE: Commission is NOT calculated here to prevent doubling
            # Commissions are only calculated when employer attendance is marked with specific workers selected

            if existing:
                old_wage = existing.get('wage_amount', 0)
                diff = wage_amount - old_wage

                await db.attendance.update_one(
                    {"id": existing['id']},
                    {"$set": {
                        "status": worker_entry.status,
                        "employer_id": worker_entry.employer_id or existing.get("employer_id", ""),
                        "wage_amount": wage_amount,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )

                await db.workers.update_one(
                    {"id": worker_entry.worker_id},
                    {"$inc": {"pending_settlement": diff}}
                )

                # ✅ If status changed to Absent, delete any existing commission (commission should only exist if worker was assigned via employer mode)
                if worker_entry.status == "Absent":
                    await db.commissions.delete_many({
                        "contractor_id": current_user.id,
                        "worker_id": worker_entry.worker_id,
                        "date": date
                    })

            else:
                # Create worker attendance
                entry = AttendanceEntry(
                    date=date,
                    contractor_id=current_user.id,
                    worker_id=worker_entry.worker_id,
                    employer_id=worker_entry.employer_id,
                    status=worker_entry.status,
                    wage_amount=wage_amount,
                    mode="worker",
                    created_by=current_user.id
                )

                entry_dict = entry.model_dump()
                entry_dict['created_at'] = entry_dict['created_at'].isoformat()
                entry_dict['updated_at'] = entry_dict['updated_at'].isoformat()
                await db.attendance.insert_one(entry_dict)

                if wage_amount > 0:
                    await db.workers.update_one(
                        {"id": worker_entry.worker_id},
                        {"$inc": {"pending_settlement": wage_amount}}
                    )

                # ✅ NOTE: Commission is NOT created here to prevent doubling
                # Commissions are only calculated when employer attendance is marked with specific workers selected

            results["worker_entries"].append(worker_entry.model_dump())

    # Log activity
    activity = Activity(
        contractor_id=current_user.id,
        type="attendance_recorded",
        description=f"Recorded attendance for {date} ({attendance_data.mode} mode)"
    )
    activity_dict = activity.model_dump()
    activity_dict['date'] = activity_dict['date'].isoformat()
    await db.activities.insert_one(activity_dict)

    return results

@api_router.get("/attendance/fetch")
async def fetch_attendance(
    date: str,
    mode: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Fetch attendance for a specific date
    FIX: Query new collections (employer_attendance, worker_attendance)
    FIX: Return proper data structure with selected_workers
    FIX: Don't add additional_charges to wage_amount
    """
    result = []
    
    # Fetch employer attendance from new collection
    if mode is None or mode == "employer":
        employer_records = await db.employer_attendance.find({
            "contractor_id": current_user.id,
            "date": date
        }, {"_id": 0}).to_list(1000)
        
        for record in employer_records:
            # Convert datetime fields
            if isinstance(record.get('created_at'), str):
                record['created_at'] = datetime.fromisoformat(record['created_at'])
            
            # Calculate wage_amount WITHOUT additional charges
            # wage_amount = workers_count × payment_per_worker (NOT including additional_charges)
            wage_amount = record['workers_count'] * record.get('payment_per_worker', 0)
            
            # Format for frontend
            formatted_record = {
                "id": record['id'],
                "contractor_id": record['contractor_id'],
                "employer_id": record['employer_id'],
                "date": record['date'],
                "mode": "employer",
                "workers_count": record['workers_count'],
                "payment_per_worker": record.get('payment_per_worker', 0),  # FIX: Add this field
                "wage_amount": wage_amount,  # Pure wage without additional charges
                "selected_workers": record.get('selected_workers', []),
                "additional_charges": [{
                    "amount": record.get('additional_charges', 0),
                    "description": record.get('charge_description', '')
                }] if record.get('additional_charges', 0) > 0 else [],
                "remarks": record.get('charge_description', ''),
                "created_at": record.get('created_at')
            }
            result.append(formatted_record)
    
    # Fetch worker attendance from new collection
    if mode is None or mode == "worker":
        worker_records = await db.worker_attendance.find({
            "contractor_id": current_user.id,
            "date": date
        }, {"_id": 0}).to_list(1000)
        
        for record in worker_records:
            # Convert datetime fields
            if isinstance(record.get('created_at'), str):
                record['created_at'] = datetime.fromisoformat(record['created_at'])
            
            # Format for frontend
            formatted_record = {
                "id": record['id'],
                "contractor_id": record['contractor_id'],
                "worker_id": record['worker_id'],
                "employer_id": record.get('employer_id', ''),
                "date": record['date'],
                "mode": "worker",
                "status": record['status'],
                "wage_amount": record.get('wage_earned', 0),
                "created_at": record.get('created_at')
            }
            result.append(formatted_record)
    
    # Fallback: Also check old attendance collection for backward compatibility
    old_records = await db.attendance.find({
        "contractor_id": current_user.id,
        "date": date
    }, {"_id": 0}).to_list(1000)
    
    for record in old_records:
        if isinstance(record.get('created_at'), str):
            record['created_at'] = datetime.fromisoformat(record['created_at'])
        if isinstance(record.get('updated_at'), str):
            record['updated_at'] = datetime.fromisoformat(record['updated_at'])
        
        # Only add if not already in result (avoid duplicates)
        if not any(r['id'] == record['id'] for r in result):
            result.append(record)
    
    return result

@api_router.get("/attendance/reconcile")
async def reconcile_attendance(
    date: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """Check for mismatches between employer and worker attendance"""
    
    # Get employer-based attendance
    employer_records = await db.attendance.find({
        "contractor_id": current_user.id,
        "date": date,
        "mode": "employer"
    }, {"_id": 0}).to_list(1000)
    
    # Get worker-based attendance
    worker_records = await db.attendance.find({
        "contractor_id": current_user.id,
        "date": date,
        "mode": "worker",
        "status": {"$in": ["Present", "Late"]}
    }, {"_id": 0}).to_list(1000)
    
    # Calculate totals
    total_employer_workers = sum(r.get('workers_count', 0) for r in employer_records)
    total_present_workers = len(worker_records)
    
    mismatches = []
    
    # Check per employer
    for emp_record in employer_records:
        employer_id = emp_record['employer_id']
        employer = await db.employers.find_one({"id": employer_id})
        
        emp_workers_count = emp_record['workers_count']
        actual_workers = len([w for w in worker_records if w.get('employer_id') == employer_id])
        
        if emp_workers_count != actual_workers:
            mismatches.append({
                "employer_id": employer_id,
                "employer_name": employer['name'] if employer else "Unknown",
                "recorded_count": emp_workers_count,
                "actual_count": actual_workers,
                "difference": emp_workers_count - actual_workers
            })
    
    return {
        "date": date,
        "total_employer_based": total_employer_workers,
        "total_worker_based": total_present_workers,
        "overall_difference": total_employer_workers - total_present_workers,
        "mismatches": mismatches,
        "status": "matched" if len(mismatches) == 0 else "mismatch"
    }

@api_router.delete("/attendance/{attendance_id}")
async def delete_attendance(
    attendance_id: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """
    Delete an attendance record
    FIX: When employer attendance is deleted, delete all worker attendance 
    records that have that employer_id for that date (make them "not marked")
    SAFETY: Prevent deletion if it would cause negative balances (payments already collected/settled)
    """
    # Try to find in old attendance collection first (for backward compatibility)
    record = await db.attendance.find_one({
        "id": attendance_id,
        "contractor_id": current_user.id
    })
    
    # If not found in old collection, check new employer_attendance collection
    if not record:
        record = await db.employer_attendance.find_one({
            "id": attendance_id,
            "contractor_id": current_user.id
        })
        if record:
            record['mode'] = 'employer'  # Set mode for processing
    
    # If still not found, check worker_attendance collection
    if not record:
        record = await db.worker_attendance.find_one({
            "id": attendance_id,
            "contractor_id": current_user.id
        })
        if record:
            record['mode'] = 'worker'  # Set mode for processing
    
    if not record:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    # Handle EMPLOYER attendance deletion
    # FIX: Prioritize mode field - only treat as employer if mode is explicitly "employer"
    # Also require valid employer_id to prevent errors when deleting worker attendance
    mode = record.get('mode')
    is_employer_attendance = (
        mode == "employer" and 
        'workers_count' in record and
        record.get('employer_id') and 
        record.get('employer_id') != ''
    )
    
    if is_employer_attendance:
        employer_id = record['employer_id']
        date = record['date']
        total_amount = record.get('total_amount', record.get('wage_amount', 0))

        # ✅ SELF / own-work attendance: no employer to refund, no commission to delete (none exist).
        is_self_work = (employer_id or "").upper() == "SELF"
        if is_self_work:
            employer = None
        else:
            # ✅ SAFETY CHECK: Verify deletion won't cause negative balance
            employer = await db.employers.find_one({"id": employer_id})
            if not employer:
                raise HTTPException(status_code=404, detail="Employer not found")

            current_pending = float(employer.get('pending_payment', 0))
            new_pending = current_pending - total_amount

            if new_pending < -0.01:  # Allow small floating point errors
                # Calculate how much has been paid
                amount_paid = total_amount + abs(new_pending)
                raise HTTPException(
                    status_code=400,
                    detail=f"Cannot delete this attendance record. ₹{amount_paid:.2f} has already been collected from this employer. Please reverse the payment first or contact support."
                )
        
        # Check all associated worker records for negative balance issues
        worker_records = await db.worker_attendance.find({
            "contractor_id": current_user.id,
            "employer_id": employer_id,
            "date": date
        }, {"_id": 0}).to_list(1000)
        
        old_worker_records = await db.attendance.find({
            "contractor_id": current_user.id,
            "employer_id": employer_id,
            "date": date,
            "mode": "worker"
        }, {"_id": 0}).to_list(1000)
        
        all_worker_records = worker_records + old_worker_records
        
        # Check each worker's balance before deletion
        for worker_record in all_worker_records:
            worker_id = worker_record['worker_id']
            wage_earned = worker_record.get('wage_earned', worker_record.get('wage_amount', 0))
            
            worker = await db.workers.find_one({"id": worker_id})
            if worker:
                current_worker_pending = float(worker.get('pending_settlement', 0))
                new_worker_pending = current_worker_pending - wage_earned
                
                if new_worker_pending < -0.01:
                    worker_name = worker.get('name', worker_id)
                    amount_settled = wage_earned + abs(new_worker_pending)
                    raise HTTPException(
                        status_code=400,
                        detail=f"Cannot delete this attendance. ₹{amount_settled:.2f} has already been settled with worker '{worker_name}'. Please reverse the settlement first."
                    )
        
        # Reverse employer pending payment (skip for SELF/own-work)
        if not is_self_work:
            await db.employers.update_one(
                {"id": employer_id},
                {"$inc": {"pending_payment": -total_amount}}
            )
        
        # FIX: Delete ALL worker attendance records that have this employer_id for this date
        # This includes workers in selected_workers list AND any others assigned to this employer
        
        # Find all worker attendance records with this employer_id and date
        worker_records = await db.worker_attendance.find({
            "contractor_id": current_user.id,
            "employer_id": employer_id,
            "date": date
        }, {"_id": 0}).to_list(1000)
        
        # Also check old collection
        old_worker_records = await db.attendance.find({
            "contractor_id": current_user.id,
            "employer_id": employer_id,
            "date": date,
            "mode": "worker"
        }, {"_id": 0}).to_list(1000)
        
        # Combine both lists
        all_worker_records = worker_records + old_worker_records
        
        # Delete each worker attendance record
        for worker_record in all_worker_records:
            worker_id = worker_record['worker_id']
            
            # Reverse worker pending settlement
            wage_earned = worker_record.get('wage_earned', worker_record.get('wage_amount', 0))
            await db.workers.update_one(
                {"id": worker_id},
                {"$inc": {"pending_settlement": -wage_earned}}
            )
            
            # Delete worker attendance record (completely remove it)
            deleted_from_new = await db.worker_attendance.delete_one({"id": worker_record['id']})
            if deleted_from_new.deleted_count == 0:
                # If not in new collection, delete from old
                await db.attendance.delete_one({"id": worker_record['id']})

        # ✅ Delete ALL commission rows for this contractor/employer/date.
        # Catches per-worker rows, SUMMARY rows (when no specific workers were selected),
        # and ADDITIONAL_CHARGES rows (when the toggle was on).
        commissions_deleted = await db.commissions.delete_many({
            "contractor_id": current_user.id,
            "employer_id": employer_id,
            "date": date
        })

        # Delete the employer attendance record from appropriate collection
        deleted = await db.employer_attendance.delete_one({"id": attendance_id})
        if deleted.deleted_count == 0:
            await db.attendance.delete_one({"id": attendance_id})
        
        # Return count of deleted worker records
        return {
            "message": "Employer attendance deleted successfully",
            "deleted_worker_records": len(all_worker_records),
            "workers_now_unassigned": len(all_worker_records),
            "commissions_deleted": commissions_deleted.deleted_count
        }
        
    # Handle WORKER attendance deletion
    # FIX: Check if it's worker mode OR has worker_id (prioritize worker over employer)
    if mode == "worker" or 'worker_id' in record:
        worker_id = record['worker_id']
        wage_earned = record.get('wage_earned', record.get('wage_amount', 0))
        
        # ✅ SAFETY CHECK: Verify deletion won't cause negative balance
        worker = await db.workers.find_one({"id": worker_id})
        if not worker:
            raise HTTPException(status_code=404, detail="Worker not found")
        
        current_pending = float(worker.get('pending_settlement', 0))
        new_pending = current_pending - wage_earned
        
        if new_pending < -0.01:  # Allow small floating point errors
            worker_name = worker.get('name', worker_id)
            amount_settled = wage_earned + abs(new_pending)
            raise HTTPException(
                status_code=400,
                detail=f"Cannot delete this attendance record. ₹{amount_settled:.2f} has already been settled with '{worker_name}'. Please reverse the settlement first."
            )
        
        # Reverse worker pending settlement
        await db.workers.update_one(
            {"id": worker_id},
            {"$inc": {"pending_settlement": -wage_earned}}
        )
        
        # Delete related commission record
        if record.get('employer_id'):
            await db.commissions.delete_many({
                "contractor_id": current_user.id,
                "worker_id": record['worker_id'],
                "employer_id": record['employer_id'],
                "date": record['date']
            })
        
        # Delete worker attendance record from appropriate collection
        deleted = await db.worker_attendance.delete_one({"id": attendance_id})
        if deleted.deleted_count == 0:
            await db.attendance.delete_one({"id": attendance_id})
        
        return {"message": "Worker attendance record deleted successfully"}
    
    return {"message": "Attendance record deleted successfully"}

@api_router.get("/attendance/yesterday")
async def get_yesterday_attendance(
    current_user: User = Depends(get_active_subscription_user)
):
    """Get yesterday's attendance for quick copy"""
    yesterday = (datetime.now(timezone.utc) - timedelta(days=1)).strftime("%d-%m-%Y")
    
    records = await db.attendance.find({
        "contractor_id": current_user.id,
        "date": yesterday
    }, {"_id": 0}).to_list(1000)
    
    return records


# ============ ADMIN ROUTES ============

class AdminLogin(BaseModel):
    email: EmailStr
    password: str

class AdminUser(BaseModel):
    id: str
    email: str
    name: str
    role: str
    is_active: bool
    created_at: str

def create_admin_token(data: dict):
    """
    Create JWT token for admin with shorter expiration for security.
    Admin tokens are valid for 7 days and must include "admin" type claim.
    """
    to_encode = data.copy()
    now = datetime.now(timezone.utc)
    expire = now + timedelta(days=7)  # Admin token valid for 7 days
    
    # Add security claims
    to_encode.update({
        "exp": expire,      # Expiration time
        "iat": now,         # Issued at time
        "nbf": now,         # Not before time
        "type": "admin"     # CRITICAL: Mark as admin token
    })
    
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

@api_router.post("/admin/login")
@limiter.limit("5/minute")  # Rate limit: 5 admin login attempts per minute per IP (stricter than user login)
async def admin_login(request: Request, admin_data: AdminLogin, response: Response):
    """Admin login endpoint with enhanced brute force protection"""
    # Get client IP address and user agent
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    
    # ===== BRUTE FORCE PROTECTION =====
    # Check if account or IP is locked (admin accounts have stricter protection)
    is_allowed, lockout_message = await check_login_attempts(
        f"admin_{admin_data.email.lower()}",  # Prefix with 'admin_' to separate from user attempts
        client_ip
    )
    if not is_allowed:
        # Log admin account lockout (CRITICAL severity)
        await log_security_event(
            event_type=SecurityEventType.ACCOUNT_LOCKED,
            severity=SecuritySeverity.CRITICAL,
            ip_address=client_ip,
            user_agent=user_agent,
            email=f"admin_{admin_data.email.lower()}",
            details={"message": lockout_message, "account_type": "admin"},
            endpoint="/admin/login"
        )
        raise HTTPException(status_code=429, detail=lockout_message)
    
    # ===== AUTHENTICATION =====
    admin = await db.admins.find_one({"email": admin_data.email.lower()}, {"_id": 0})
    
    # Security: Always check password even if admin doesn't exist (timing attack prevention)
    if not admin:
        # Perform a dummy hash operation to prevent timing attacks
        pwd_context.hash("dummy_password_to_prevent_timing_attack")
        # Record failed attempt
        await record_failed_login(f"admin_{admin_data.email.lower()}", client_ip)
        # Log failed admin login (HIGH severity)
        await log_security_event(
            event_type=SecurityEventType.ADMIN_LOGIN_FAILED,
            severity=SecuritySeverity.HIGH,
            ip_address=client_ip,
            user_agent=user_agent,
            email=admin_data.email.lower(),
            details={"reason": "invalid_credentials", "account_type": "admin"},
            endpoint="/admin/login"
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not pwd_context.verify(admin_data.password, admin['password']):
        # Record failed attempt
        await record_failed_login(f"admin_{admin_data.email.lower()}", client_ip)
        # Log failed admin login (HIGH severity)
        await log_security_event(
            event_type=SecurityEventType.ADMIN_LOGIN_FAILED,
            severity=SecuritySeverity.HIGH,
            ip_address=client_ip,
            user_agent=user_agent,
            email=admin_data.email.lower(),
            user_id=admin['id'],
            details={"reason": "invalid_password", "account_type": "admin"},
            endpoint="/admin/login"
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not admin.get('is_active'):
        # Don't record as failed attempt for disabled accounts
        raise HTTPException(status_code=403, detail="Admin account is disabled. Please contact super admin.")
    
    # ===== SUCCESSFUL LOGIN =====
    # Clear any previous failed login attempts
    await clear_failed_login_attempts(f"admin_{admin_data.email.lower()}", client_ip)
    
    # Log successful admin login
    await log_security_event(
        event_type=SecurityEventType.ADMIN_LOGIN_SUCCESS,
        severity=SecuritySeverity.LOW,
        ip_address=client_ip,
        user_agent=user_agent,
        email=admin_data.email.lower(),
        user_id=admin['id'],
        details={"login_method": "email_password", "account_type": "admin"},
        endpoint="/admin/login"
    )
    
    # Update last_login timestamp for admin
    await db.admins.update_one(
        {"id": admin['id']},
        {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Create JWT token
    access_token = create_admin_token({"sub": admin['id']})
    
    # Security: Multi-layer cookie protection for admin sessions
    # Admin tokens have shorter lifespan (7 days) for enhanced security
    # httponly=True: Prevents JavaScript access, mitigating XSS attacks
    # secure=True (production): Ensures cookie only sent over HTTPS, preventing MITM
    # samesite='lax'/'strict': CSRF protection - blocks cross-site request forgery
    response.set_cookie(
        key="auth_token",
        value=access_token,
        httponly=True,           # XSS Protection: No client-side JavaScript access
        secure=COOKIE_SECURE,    # MITM Protection: HTTPS only in production
        samesite=COOKIE_SAMESITE,  # CSRF Protection: Blocks cross-site state-changing requests
        max_age=604800,          # 7 days (shorter than user tokens for security)
        path="/"
    )
    
    return {
        "message": "Login successful",
        "admin": {
            "id": admin['id'],
            "email": admin['email'],
            "name": admin['name'],
            "role": admin['role']
        }
    }

@api_router.post("/admin/logout")
async def admin_logout(response: Response):
    """Admin logout"""
    response.delete_cookie("auth_token", path="/")
    return {"message": "Logged out successfully"}

@api_router.get("/admin/me")
async def get_admin_profile(admin: AdminUser = Depends(get_current_admin)):
    """Get current admin profile"""
    return admin

# ========================= Trial Settings (Admin) =========================
@api_router.get("/admin/trial-settings", dependencies=[Depends(get_current_admin)])
async def get_trial_settings():
    doc = await db.settings.find_one({"key": "trial_settings"}, {"_id": 0})
    if not doc:
        default = {"key": "trial_settings", "duration_days": 14}
        await db.settings.update_one({"key": "trial_settings"}, {"$set": default}, upsert=True)
        return {"duration_days": 14}
    return {"duration_days": int(doc.get("duration_days", 14))}

@api_router.post("/admin/trial-settings", dependencies=[Depends(get_current_admin)])
async def update_trial_settings(settings: TrialSettings):
    if settings.duration_days < 0 or settings.duration_days > 60:
        raise HTTPException(status_code=400, detail="Duration must be between 0 and 60 days")
    await db.settings.update_one({"key": "trial_settings"}, {"$set": {"key": "trial_settings", "duration_days": settings.duration_days}}, upsert=True)
    return {"message": "Updated", "duration_days": settings.duration_days}

# ========================= Notifications - Admin Broadcast =========================
@api_router.post("/admin/notifications/send", dependencies=[Depends(get_current_admin)])
async def admin_send_notifications(payload: AdminBroadcastRequest):
    broadcast_id = str(uuid.uuid4())
    # Determine recipient filter
    query: Dict[str, Any] = {}
    if payload.filter == "free":
        query = {"subscription_plan": {"$in": ["none", "free"]}}
    elif payload.filter == "trial":
        query = {"subscription_status": "trial"}
    elif payload.filter == "paid":
        query = {"subscription_status": "active"}
    else:
        query = {}

    users = await db.users.find(query, {"_id": 0, "id": 1}).to_list(100000)
    user_ids = [u["id"] for u in users]

    if not user_ids:
        return {"sent": 0}

    now = datetime.now(timezone.utc)
    
    # Determine CTA button URL and label based on type
    action_url = None
    action_label = None
    
    if payload.cta_type:
        cta_mappings = {
            "pricing": {"url": "/pricing", "label": "View Pricing"},
            "subscription": {"url": "/manage-subscription", "label": "Manage Subscription"},
            "help": {"url": "/help", "label": "Get Help"},
            "attendance": {"url": "/attendance", "label": "Mark Attendance"},
            "custom": {"url": payload.cta_url, "label": payload.cta_label}
        }
        
        if payload.cta_type in cta_mappings:
            mapping = cta_mappings[payload.cta_type]
            action_url = mapping["url"]
            action_label = mapping["label"]
    
    docs = [
        {
            "id": str(uuid.uuid4()),
            "user_id": uid,
            "title": payload.title,
            "message": payload.message,
            "type": payload.type,
            "broadcast_id": broadcast_id,
            "action_url": action_url,
            "action_label": action_label,
            "read": False,
            "created_at": now,
        }
        for uid in user_ids
    ]

    if docs:
        await db.notifications.insert_many(docs)

    # Save admin broadcast meta
    await db.admin_notifications.insert_one({
        "broadcast_id": broadcast_id,
        "title": payload.title,
        "message": payload.message,
        "type": payload.type,
        "filter": payload.filter,
        "sent": len(docs),
        "created_at": now,
    })

    return {"sent": len(docs), "broadcast_id": broadcast_id}

@api_router.get("/admin/notifications/sent", dependencies=[Depends(get_current_admin)])
async def admin_list_sent_notifications(limit: int = 50):
    items = await db.admin_notifications.find({}, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(length=limit)
    return {"items": items}

@api_router.delete("/admin/notifications/{broadcast_id}", dependencies=[Depends(get_current_admin)])
async def admin_delete_broadcast(broadcast_id: str):
    # Delete broadcast meta and all user notifications with this broadcast_id
    await db.admin_notifications.delete_one({"broadcast_id": broadcast_id})
    result = await db.notifications.delete_many({"broadcast_id": broadcast_id})
    return {"deleted": result.deleted_count}

# ========================= Promotions/Offers - Admin =========================
@api_router.post("/admin/promotions/create", dependencies=[Depends(get_current_admin)])
async def create_promotion(promotion: PromotionCreate):
    """Create a new promotion/offer"""
    promotion_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    # Get eligible users based on category
    user_query: Dict[str, Any] = {}
    
    if promotion.target_category == "all":
        user_query = {}
    elif promotion.target_category == "free_users":
        user_query = {"subscription_plan": {"$in": ["none", "free"]}}
    elif promotion.target_category == "trial_users":
        user_query = {"subscription_status": "trial"}
    elif promotion.target_category == "paid_users":
        user_query = {"subscription_status": "active"}
    elif promotion.target_category == "cancelled_users":
        user_query = {"subscription_status": "cancelled"}
    elif promotion.target_category == "expired_trial":
        user_query = {
            "subscription_status": {"$in": ["none", "free"]},
            "trial_activated_at": {"$exists": True}
        }
    elif promotion.target_category == "low_activity":
        seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)
        user_query = {"last_login": {"$lt": seven_days_ago.isoformat()}}
    elif promotion.target_category == "custom_users":
        user_query = {"id": {"$in": promotion.custom_user_ids or []}}
    
    eligible_users = await db.users.find(user_query, {"_id": 0, "id": 1}).to_list(100000)
    user_ids = [u["id"] for u in eligible_users]
    
    # Create promotion document
    promotion_doc = {
        "id": promotion_id,
        "name": promotion.name,
        "description": promotion.description,
        "discount_type": promotion.discount_type,
        "discount_value": promotion.discount_value,
        "target_category": promotion.target_category,
        "custom_user_ids": promotion.custom_user_ids or [],
        "plan_targets": promotion.plan_targets or [],
        "valid_from": promotion.valid_from.isoformat(),
        "valid_until": promotion.valid_until.isoformat(),
        "max_uses": promotion.max_uses,
        "current_uses": 0,
        "eligible_user_ids": user_ids,
        "active": promotion.active,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.promotions.insert_one(promotion_doc)
    
    # Send notifications if enabled
    if promotion.send_notification and user_ids:
        notification_title = promotion.notification_title or f"Special Offer: {promotion.name}"
        notification_message = promotion.notification_message or promotion.description
        
        notification_docs = [
            {
                "id": str(uuid.uuid4()),
                "user_id": uid,
                "title": notification_title,
                "message": notification_message,
                "type": "promo",
                "action_url": "/pricing",
                "action_label": "View Offer",
                "promotion_id": promotion_id,
                "read": False,
                "created_at": now
            }
            for uid in user_ids
        ]
        
        if notification_docs:
            await db.notifications.insert_many(notification_docs)
    
    return {
        "promotion_id": promotion_id,
        "eligible_users": len(user_ids),
        "notifications_sent": len(user_ids) if promotion.send_notification else 0
    }

@api_router.get("/admin/promotions/list", dependencies=[Depends(get_current_admin)])
async def list_promotions(active_only: bool = False):
    """List all promotions"""
    query = {"active": True} if active_only else {}
    promotions = await db.promotions.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"promotions": promotions}

@api_router.get("/admin/promotions/users", dependencies=[Depends(get_current_admin)])
async def get_all_users_for_promotion():
    """Get all users with basic info for custom user selection"""
    users = await db.users.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1, "subscription_plan": 1, "subscription_status": 1}).to_list(10000)
    return {"users": users}

@api_router.put("/admin/promotions/{promotion_id}", dependencies=[Depends(get_current_admin)])
async def update_promotion(promotion_id: str, update: PromotionUpdate):
    """Update an existing promotion"""
    update_data = {k: v for k, v in update.dict(exclude_unset=True).items() if v is not None}
    
    if update_data:
        if "valid_until" in update_data and isinstance(update_data["valid_until"], datetime):
            update_data["valid_until"] = update_data["valid_until"].isoformat()
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.promotions.update_one({"id": promotion_id}, {"$set": update_data})
    
    return {"message": "Promotion updated successfully"}

@api_router.delete("/admin/promotions/{promotion_id}", dependencies=[Depends(get_current_admin)])
async def delete_promotion(promotion_id: str):
    """Delete a promotion"""
    await db.promotions.delete_one({"id": promotion_id})
    await db.notifications.delete_many({"promotion_id": promotion_id})
    return {"message": "Promotion deleted successfully"}

# ========================= Site-Wide Offers - Admin =========================
@api_router.post("/admin/site-offers/create", dependencies=[Depends(get_current_admin)])
async def create_site_wide_offer(offer: SiteWideOfferCreate):
    """Create a new site-wide offer (visible to all visitors)"""
    offer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    offer_dict = {
        "id": offer_id,
        "name": offer.name,
        "description": offer.description,
        "offer_reason": offer.offer_reason,
        "discount_type": offer.discount_type,
        "discount_value": offer.discount_value,
        "plan_targets": offer.plan_targets,
        "valid_from": offer.valid_from.isoformat(),
        "valid_until": offer.valid_until.isoformat(),
        "badge_text": offer.badge_text,
        "badge_color": offer.badge_color,
        "show_countdown": offer.show_countdown,
        "active": offer.active,
        "created_at": now.isoformat(),
        "created_by": "admin"
    }
    
    await db.site_wide_offers.insert_one(offer_dict)
    
    return {
        "message": "Site-wide offer created successfully",
        "offer_id": offer_id
    }

@api_router.get("/admin/site-offers/list", dependencies=[Depends(get_current_admin)])
async def list_site_wide_offers(active_only: bool = False):
    """List all site-wide offers"""
    query = {"active": True} if active_only else {}
    offers = await db.site_wide_offers.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"offers": offers}

@api_router.put("/admin/site-offers/{offer_id}", dependencies=[Depends(get_current_admin)])
async def update_site_wide_offer(offer_id: str, update: SiteWideOfferUpdate):
    """Update an existing site-wide offer"""
    update_data = {k: v for k, v in update.dict(exclude_unset=True).items() if v is not None}
    
    if "valid_until" in update_data:
        update_data["valid_until"] = update_data["valid_until"].isoformat()
    
    if update_data:
        await db.site_wide_offers.update_one({"id": offer_id}, {"$set": update_data})
    
    return {"message": "Site-wide offer updated successfully"}

@api_router.delete("/admin/site-offers/{offer_id}", dependencies=[Depends(get_current_admin)])
async def delete_site_wide_offer(offer_id: str):
    """Delete a site-wide offer"""
    await db.site_wide_offers.delete_one({"id": offer_id})
    return {"message": "Site-wide offer deleted successfully"}

# ========================= Site-Wide Offers - Public =========================
@api_router.get("/site-offers/active")
async def get_active_site_offers():
    """Get active site-wide offers (public endpoint - no auth required)"""
    now = datetime.now(timezone.utc)
    
    offers = await db.site_wide_offers.find({
        "active": True,
        "valid_from": {"$lte": now.isoformat()},
        "valid_until": {"$gte": now.isoformat()}
    }, {"_id": 0}).sort("created_at", -1).to_list(10)
    
    return {"offers": offers}

@api_router.get("/site-offers/calculate-price")
async def calculate_site_offer_price(plan: str):
    """Calculate discounted price for a plan based on active site-wide offers (public endpoint)"""
    now = datetime.now(timezone.utc)
    
    # Get active site-wide offers for this plan
    offers = await db.site_wide_offers.find({
        "active": True,
        "valid_from": {"$lte": now.isoformat()},
        "valid_until": {"$gte": now.isoformat()},
        "$or": [
            {"plan_targets": []},  # Applies to all plans
            {"plan_targets": plan}  # Applies to specific plan
        ]
    }, {"_id": 0}).sort("discount_value", -1).to_list(1)
    
    if not offers:
        return {"has_offer": False, "original_price": None, "discounted_price": None}
    
    offer = offers[0]
    
    # Get original price from database
    plan_doc = await db.subscription_plans.find_one({"name": plan, "is_active": True})
    if not plan_doc:
        return {"has_offer": False, "original_price": None, "discounted_price": None}
    
    original_price = plan_doc.get("price", 0)
    
    # Calculate discounted price
    if offer["discount_type"] == "percentage":
        discount_amount = (original_price * offer["discount_value"]) / 100
        discounted_price = original_price - discount_amount
    else:  # fixed_amount
        discounted_price = max(0, original_price - offer["discount_value"])
    
    return {
        "has_offer": True,
        "original_price": original_price,
        "discounted_price": round(discounted_price, 2),
        "discount_value": offer["discount_value"],
        "discount_type": offer["discount_type"],
        "offer_name": offer["name"],
        "offer_reason": offer["offer_reason"],
        "badge_text": offer.get("badge_text"),
        "badge_color": offer.get("badge_color", "red"),
        "show_countdown": offer.get("show_countdown", False),
        "valid_until": offer["valid_until"]
    }

# ========================= Promotions/Offers - User =========================
@api_router.get("/promotions/my-offers")
async def get_my_promotions(current_user: User = Depends(get_current_user)):
    """Get active promotions for current user"""
    now = datetime.now(timezone.utc)
    
    promotions = await db.promotions.find({
        "active": True,
        "eligible_user_ids": current_user.id,
        "valid_from": {"$lte": now.isoformat()},
        "valid_until": {"$gte": now.isoformat()}
    }, {"_id": 0}).to_list(100)
    
    # Filter out promotions that have reached max uses
    valid_promotions = [
        p for p in promotions 
        if p.get("max_uses") is None or p.get("current_uses", 0) < p.get("max_uses")
    ]
    
    return {"promotions": valid_promotions}

@api_router.get("/promotions/calculate-price")
async def calculate_promotional_pricing(plan: str, current_user: User = Depends(get_current_user)):
    """Calculate promotional price for a plan"""
    now = datetime.now(timezone.utc)
    
    # Get active promotions for this user and plan
    promotions = await db.promotions.find({
        "active": True,
        "eligible_user_ids": current_user.id,
        "valid_from": {"$lte": now.isoformat()},
        "valid_until": {"$gte": now.isoformat()},
        "$or": [
            {"plan_targets": []},
            {"plan_targets": plan}
        ]
    }, {"_id": 0}).sort("discount_value", -1).to_list(1)
    
    if not promotions:
        return {"has_promotion": False, "original_price": None, "promotional_price": None}
    
    promotion = promotions[0]
    
    # Get original price from database
    plan_doc = await db.subscription_plans.find_one({"name": plan, "is_active": True})
    if not plan_doc:
        return {"has_promotion": False, "original_price": None, "promotional_price": None}
    
    original_price = plan_doc.get("price", 0)
    
    # Calculate promotional price
    if promotion["discount_type"] == "percentage":
        discount_amount = original_price * (promotion["discount_value"] / 100)
        promotional_price = round(original_price - discount_amount, 2)
    elif promotion["discount_type"] == "fixed_amount":
        promotional_price = max(0, round(original_price - promotion["discount_value"], 2))
    elif promotion["discount_type"] == "custom_price":
        promotional_price = round(promotion["discount_value"], 2)
    else:
        promotional_price = original_price
    
    return {
        "has_promotion": True,
        "promotion_id": promotion["id"],
        "promotion_name": promotion["name"],
        "original_price": original_price,
        "promotional_price": promotional_price,
        "discount_type": promotion["discount_type"],
        "discount_value": promotion["discount_value"],
        "valid_until": promotion["valid_until"]
    }

@api_router.get("/admin/dashboard/stats")
async def get_admin_dashboard_stats(admin: AdminUser = Depends(get_current_admin)):
    """Get dashboard statistics for admin"""
    
    # Total contractors
    total_contractors = await db.users.count_documents({})
    active_contractors = await db.users.count_documents({"subscription_status": "active"})
    
    # Active contractors on Contractor Plus plan
    contractor_plan_users = await db.users.count_documents({
        "subscription_status": "active",
        "subscription_plan": "Contractor Plus"
    })
    
    # Platform Revenue = Active Contractor Plan users × ₹799
    # Note: Currently all users are test users, so we show 0 revenue
    # When real payments are integrated via Razorpay, this will be calculated from actual payments
    platform_revenue = 0  # Will be updated when Razorpay integration is complete
    
    # Total workers and employers across all contractors
    total_workers = await db.workers.count_documents({"status": "Active"})
    total_employers = await db.employers.count_documents({"status": "Active"})
    
    # Total payments collected (sum of all employer payments)
    payments_cursor = db.payments.find({"type": "employer_payment"})
    total_payments = 0
    async for payment in payments_cursor:
        total_payments += payment.get('amount', 0)
    
    # Total wages paid to workers
    settlements_cursor = db.settlements.find({})
    total_wages_paid = 0
    async for settlement in settlements_cursor:
        total_wages_paid += settlement.get('amount_paid', 0)
    
    # Count activation keys
    total_keys = await db.activation_keys.count_documents({})
    active_keys = await db.activation_keys.count_documents({"is_active": True})
    
    # Recent contractors (last 10)
    recent_contractors = await db.users.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "subscription_status": 1, "subscription_end_date": 1}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "total_contractors": total_contractors,
        "active_contractors": active_contractors,
        "contractor_plan_users": contractor_plan_users,
        "total_workers": total_workers,
        "total_employers": total_employers,
        "total_payments_collected": round(total_payments, 2),
        "platform_revenue": platform_revenue,
        "total_wages_paid": round(total_wages_paid, 2),
        "total_activation_keys": total_keys,
        "active_activation_keys": active_keys,
        "recent_contractors": recent_contractors
    }

@api_router.get("/admin/security-logs")
async def get_security_logs(
    admin: AdminUser = Depends(get_current_admin),
    event_type: Optional[str] = None,
    severity: Optional[str] = None,
    ip_address: Optional[str] = None,
    email: Optional[str] = None,
    time_range: Optional[str] = "24h",  # 1h, 24h, 7d, 30d, all
    resolved: Optional[bool] = None,
    limit: int = 100,
    skip: int = 0
):
    """Get security logs with filtering options"""
    
    # Build query
    query = {}
    
    if event_type:
        query["event_type"] = event_type
    
    if severity:
        query["severity"] = severity
    
    if ip_address:
        query["ip_address"] = ip_address
    
    if email:
        query["email"] = {"$regex": email, "$options": "i"}
    
    if resolved is not None:
        query["resolved"] = resolved
    
    # Time range filtering
    now = datetime.now(timezone.utc)
    time_ranges = {
        "1h": now - timedelta(hours=1),
        "24h": now - timedelta(hours=24),
        "7d": now - timedelta(days=7),
        "30d": now - timedelta(days=30),
    }
    
    if time_range and time_range != "all":
        start_time = time_ranges.get(time_range, time_ranges["24h"])
        query["timestamp"] = {"$gte": start_time}
    
    # Get total count
    total_count = await db.security_logs.count_documents(query)
    
    # Get logs
    logs_cursor = db.security_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit)
    logs = await logs_cursor.to_list(limit)
    
    # Get statistics
    stats = {
        "total_events": total_count,
        "by_severity": {
            "critical": await db.security_logs.count_documents({**query, "severity": SecuritySeverity.CRITICAL}),
            "high": await db.security_logs.count_documents({**query, "severity": SecuritySeverity.HIGH}),
            "medium": await db.security_logs.count_documents({**query, "severity": SecuritySeverity.MEDIUM}),
            "low": await db.security_logs.count_documents({**query, "severity": SecuritySeverity.LOW}),
        },
        "by_type": {
            "failed_login": await db.security_logs.count_documents({**query, "event_type": SecurityEventType.FAILED_LOGIN}),
            "admin_login_failed": await db.security_logs.count_documents({**query, "event_type": SecurityEventType.ADMIN_LOGIN_FAILED}),
            "account_locked": await db.security_logs.count_documents({**query, "event_type": SecurityEventType.ACCOUNT_LOCKED}),
            "disposable_email": await db.security_logs.count_documents({**query, "event_type": SecurityEventType.DISPOSABLE_EMAIL}),
            "weak_password": await db.security_logs.count_documents({**query, "event_type": SecurityEventType.WEAK_PASSWORD}),
        },
        "top_ips": []
    }
    
    # Get top 10 IPs with most events
    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$ip_address", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    top_ips = await db.security_logs.aggregate(pipeline).to_list(10)
    stats["top_ips"] = [{"ip": item["_id"], "count": item["count"]} for item in top_ips]
    
    return {
        "logs": logs,
        "stats": stats,
        "pagination": {
            "total": total_count,
            "skip": skip,
            "limit": limit,
            "has_more": total_count > (skip + limit)
        }
    }

@api_router.put("/admin/security-logs/{log_id}/acknowledge")
async def acknowledge_security_log(
    log_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Mark a security log as acknowledged (reviewed but not necessarily resolved)"""
    
    if not validate_uuid(log_id):
        raise HTTPException(status_code=400, detail="Invalid log ID format")
    
    update = {
        "acknowledged": True,
        "acknowledged_at": datetime.now(timezone.utc),
        "acknowledged_by": admin.email
    }
    
    result = await db.security_logs.update_one(
        {"id": log_id},
        {"$set": update}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Security log not found")
    
    return {"message": "Security log acknowledged"}

@api_router.put("/admin/security-logs/{log_id}/resolve")
async def resolve_security_log(
    log_id: str,
    notes: Optional[str] = None,
    admin: AdminUser = Depends(get_current_admin)
):
    """Mark a security log as resolved"""
    
    if not validate_uuid(log_id):
        raise HTTPException(status_code=400, detail="Invalid log ID format")
    
    update = {
        "resolved": True,
        "resolved_at": datetime.now(timezone.utc),
        "resolved_by": admin.email,
        "acknowledged": True,
        "acknowledged_at": datetime.now(timezone.utc),
        "acknowledged_by": admin.email
    }
    
    if notes:
        update["notes"] = sanitize_input(notes)
    
    result = await db.security_logs.update_one(
        {"id": log_id},
        {"$set": update}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Security log not found")
    
    return {"message": "Security log marked as resolved"}

@api_router.post("/admin/security-logs/bulk-acknowledge")
async def bulk_acknowledge_security_logs(
    log_ids: list[str],
    admin: AdminUser = Depends(get_current_admin)
):
    """Acknowledge multiple security logs at once"""
    
    if not log_ids or len(log_ids) == 0:
        raise HTTPException(status_code=400, detail="No log IDs provided")
    
    if len(log_ids) > 100:
        raise HTTPException(status_code=400, detail="Cannot acknowledge more than 100 logs at once")
    
    # Validate all IDs
    for log_id in log_ids:
        if not validate_uuid(log_id):
            raise HTTPException(status_code=400, detail=f"Invalid log ID format: {log_id}")
    
    update = {
        "acknowledged": True,
        "acknowledged_at": datetime.now(timezone.utc),
        "acknowledged_by": admin.email
    }
    
    result = await db.security_logs.update_many(
        {"id": {"$in": log_ids}},
        {"$set": update}
    )
    
    return {
        "message": f"Acknowledged {result.modified_count} security logs",
        "count": result.modified_count
    }

@api_router.get("/admin/security-logs/dashboard")
async def get_security_dashboard(admin: AdminUser = Depends(get_current_admin)):
    """Get security dashboard overview for admin panel - shows new alerts since last login"""
    
    now = datetime.now(timezone.utc)
    last_24h = now - timedelta(hours=24)
    last_7d = now - timedelta(days=7)
    
    # Get admin's last login time from database
    admin_doc = await db.admins.find_one({"id": admin.id}, {"last_login": 1})
    last_login_str = admin_doc.get('last_login') if admin_doc else None
    
    # Parse last_login timestamp (fallback to 24h ago if not set)
    if last_login_str:
        try:
            last_login = datetime.fromisoformat(last_login_str.replace('Z', '+00:00'))
        except:
            last_login = last_24h
    else:
        last_login = last_24h
    
    security_logs = db.security_logs
    
    # NEW: Threat indicators since last login (for alert banner)
    new_threats_since_login = await security_logs.count_documents({
        "severity": {"$in": [SecuritySeverity.HIGH, SecuritySeverity.CRITICAL]},
        "timestamp": {"$gte": last_login}
    })
    
    # Threat indicators (last 24h - for stats)
    threats_24h = await security_logs.count_documents({
        "severity": {"$in": [SecuritySeverity.HIGH, SecuritySeverity.CRITICAL]},
        "timestamp": {"$gte": last_24h}
    })
    
    threats_7d = await security_logs.count_documents({
        "severity": {"$in": [SecuritySeverity.HIGH, SecuritySeverity.CRITICAL]},
        "timestamp": {"$gte": last_7d}
    })
    
    # Failed login attempts (last 24h)
    failed_logins_24h = await security_logs.count_documents({
        "event_type": {"$in": [SecurityEventType.FAILED_LOGIN, SecurityEventType.ADMIN_LOGIN_FAILED]},
        "timestamp": {"$gte": last_24h}
    })
    
    # Active lockouts
    active_lockouts = await db.failed_login_attempts.count_documents({
        "lockout_until": {"$gt": now}
    })
    
    # Disposable email attempts (last 7d)
    disposable_emails = await security_logs.count_documents({
        "event_type": SecurityEventType.DISPOSABLE_EMAIL,
        "timestamp": {"$gte": last_7d}
    })
    
    # Weak password attempts (last 7d)
    weak_passwords = await security_logs.count_documents({
        "event_type": SecurityEventType.WEAK_PASSWORD,
        "timestamp": {"$gte": last_7d}
    })
    
    # Unresolved critical/high events
    unresolved_critical = await security_logs.count_documents({
        "severity": {"$in": [SecuritySeverity.HIGH, SecuritySeverity.CRITICAL]},
        "resolved": False
    })
    
    # Recent critical events (last 10)
    recent_critical = await security_logs.find(
        {"severity": {"$in": [SecuritySeverity.HIGH, SecuritySeverity.CRITICAL]}},
        {"_id": 0}
    ).sort("timestamp", -1).limit(10).to_list(10)
    
    return {
        "new_threats_since_login": new_threats_since_login,  # NEW: For alert banner
        "last_login": last_login_str,  # NEW: Admin's last login timestamp
        "threats_24h": threats_24h,
        "threats_7d": threats_7d,
        "failed_logins_24h": failed_logins_24h,
        "active_lockouts": active_lockouts,
        "disposable_emails": disposable_emails,
        "weak_passwords": weak_passwords,
        "unresolved_critical": unresolved_critical,
        "recent_critical_events": recent_critical
    }

@api_router.get("/admin/contractors")
async def get_all_contractors(
    admin: AdminUser = Depends(get_current_admin),
    search: Optional[str] = None,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get all contractors with search and filter. Normalizes subscription_plan: 'none' -> 'Free Plan'"""
    query = {}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone_number": {"$regex": search, "$options": "i"}}
        ]
    
    if status:
        query["subscription_status"] = status
    
    contractors = await db.users.find(
        query,
        {"_id": 0, "password": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    total = await db.users.count_documents(query)
    
    # Get additional stats for each contractor and normalize plan names
    for contractor in contractors:
        contractor_id = contractor['id']
        contractor['worker_count'] = await db.workers.count_documents({"contractor_id": contractor_id})
        contractor['employer_count'] = await db.employers.count_documents({"contractor_id": contractor_id})
        # Normalize subscription_plan: 'none' or empty -> 'Free Plan'
        if contractor.get('subscription_plan') == 'none' or not contractor.get('subscription_plan'):
            contractor['subscription_plan'] = 'Free Plan'
    
    return {
        "contractors": contractors,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@api_router.get("/admin/contractors/{contractor_id}")
async def get_contractor_details(
    contractor_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Get detailed information about a specific contractor. Normalizes subscription_plan: 'none' -> 'Free Plan'"""
    contractor = await db.users.find_one({"id": contractor_id}, {"_id": 0, "password": 0})
    
    if not contractor:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    # Get workers
    workers = await db.workers.find(
        {"contractor_id": contractor_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get employers
    employers = await db.employers.find(
        {"contractor_id": contractor_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get payments
    payments = await db.payments.find(
        {"contractor_id": contractor_id},
        {"_id": 0}
    ).sort("date", -1).limit(100).to_list(100)
    
    # Get attendance records count
    attendance_count = await db.attendance.count_documents({"contractor_id": contractor_id})
    
    # Normalize subscription_plan: 'none' or empty -> 'Free Plan'
    if contractor.get('subscription_plan') == 'none' or not contractor.get('subscription_plan'):
        contractor['subscription_plan'] = 'Free Plan'
    
    return {
        "contractor": contractor,
        "workers": workers,
        "employers": employers,
        "payments": payments,
        "attendance_count": attendance_count
    }

@api_router.put("/admin/contractors/{contractor_id}/status")
async def update_contractor_status(
    contractor_id: str,
    status: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Activate or suspend a contractor"""
    new_status = status.get('subscription_status')
    if new_status not in ['active', 'suspended', 'expired', 'inactive']:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.users.update_one(
        {"id": contractor_id},
        {"$set": {"subscription_status": new_status}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    return {"message": f"Contractor status updated to {new_status}"}

@api_router.delete("/admin/contractors/{contractor_id}")
async def delete_contractor(
    contractor_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete a contractor and all their data"""
    # Delete user
    user_result = await db.users.delete_one({"id": contractor_id})
    
    if user_result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    # Delete associated data
    await db.workers.delete_many({"contractor_id": contractor_id})
    await db.employers.delete_many({"contractor_id": contractor_id})
    await db.attendance.delete_many({"contractor_id": contractor_id})
    await db.payments.delete_many({"contractor_id": contractor_id})
    await db.settlements.delete_many({"contractor_id": contractor_id})
    
    return {"message": "Contractor and all associated data deleted successfully"}

@api_router.put("/admin/contractors/{contractor_id}/profile")
async def update_contractor_profile(
    contractor_id: str,
    profile_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Update contractor profile (name, email, phone)"""
    update_fields = {}
    
    if 'name' in profile_data:
        update_fields['name'] = profile_data['name']
    if 'email' in profile_data:
        # Check if email already exists for another user
        existing = await db.users.find_one({"email": profile_data['email'], "id": {"$ne": contractor_id}})
        if existing:
            raise HTTPException(status_code=400, detail="Email already in use")
        update_fields['email'] = profile_data['email']
    if 'phone_number' in profile_data:
        update_fields['phone_number'] = profile_data['phone_number']
    
    if not update_fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    result = await db.users.update_one(
        {"id": contractor_id},
        {"$set": update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    return {"message": "Contractor profile updated successfully"}

@api_router.post("/admin/contractors/{contractor_id}/reset-password")
async def reset_contractor_password(
    contractor_id: str,
    password_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Reset contractor password"""
    new_password = password_data.get('new_password')
    if not new_password or len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    hashed_password = pwd_context.hash(new_password)
    
    result = await db.users.update_one(
        {"id": contractor_id},
        {"$set": {"password": hashed_password}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    return {"message": "Password reset successfully"}

@api_router.put("/admin/contractors/{contractor_id}/plan")
async def update_contractor_plan(
    contractor_id: str,
    plan_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Change contractor subscription plan"""
    plan = plan_data.get('plan')
    extend_days = plan_data.get('extend_days', 0)
    
    # Validate plan name against allowed plans
    valid_plans = ['Contractor Plus', 'Contractor Pro', 'Enterprise']
    if plan not in valid_plans:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid plan. Allowed plans: {', '.join(valid_plans)}"
        )
    
    update_data = {"subscription_plan": plan}
    
    # If extending days, update end date
    if extend_days > 0:
        contractor = await db.users.find_one({"id": contractor_id}, {"_id": 0})
        if contractor:
            # Get the current end date
            current_end_str = contractor.get('plan_end_date') or contractor.get('subscription_end_date')
            if current_end_str:
                try:
                    current_end = datetime.fromisoformat(current_end_str)
                    # If already expired, extend from now
                    if current_end < datetime.now(timezone.utc):
                        new_end = datetime.now(timezone.utc) + timedelta(days=extend_days)
                    else:
                        # Extend from current end date
                        new_end = current_end + timedelta(days=extend_days)
                except:
                    new_end = datetime.now(timezone.utc) + timedelta(days=extend_days)
            else:
                new_end = datetime.now(timezone.utc) + timedelta(days=extend_days)
            
            update_data["subscription_end_date"] = new_end.isoformat()
            update_data["plan_end_date"] = new_end.isoformat()
            update_data["subscription_status"] = "active"
    
    result = await db.users.update_one(
        {"id": contractor_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    return {"message": "Subscription plan updated successfully"}

@api_router.put("/admin/contractors/{contractor_id}/validity")
async def update_contractor_validity(
    contractor_id: str,
    validity_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Change contractor subscription validity (set new end date or add days)"""
    action = validity_data.get('action')  # 'set_date' or 'add_days'
    
    contractor = await db.users.find_one({"id": contractor_id}, {"_id": 0})
    if not contractor:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    if action == 'set_date':
        # Set specific end date
        end_date_str = validity_data.get('end_date')
        if not end_date_str:
            raise HTTPException(status_code=400, detail="end_date is required")
        try:
            new_end = datetime.fromisoformat(end_date_str)
        except:
            raise HTTPException(status_code=400, detail="Invalid date format")
        
    elif action == 'add_days':
        # Add days to current validity
        days = validity_data.get('days', 0)
        if days <= 0:
            raise HTTPException(status_code=400, detail="Days must be positive")
        
        current_end_str = contractor.get('plan_end_date') or contractor.get('subscription_end_date')
        if current_end_str:
            try:
                current_end = datetime.fromisoformat(current_end_str)
                # If expired, add from now
                if current_end < datetime.now(timezone.utc):
                    new_end = datetime.now(timezone.utc) + timedelta(days=days)
                else:
                    new_end = current_end + timedelta(days=days)
            except:
                new_end = datetime.now(timezone.utc) + timedelta(days=days)
        else:
            new_end = datetime.now(timezone.utc) + timedelta(days=days)
    
    elif action == 'set_days':
        # Set exact number of days from now
        days = validity_data.get('days', 30)
        new_end = datetime.now(timezone.utc) + timedelta(days=days)
    
    else:
        raise HTTPException(status_code=400, detail="Invalid action. Use 'set_date', 'add_days', or 'set_days'")
    
    # Update the subscription
    await db.users.update_one(
        {"id": contractor_id},
        {
            "$set": {
                "subscription_end_date": new_end.isoformat(),
                "plan_end_date": new_end.isoformat(),
                "subscription_status": "active"
            }
        }
    )
    
    return {
        "message": "Subscription validity updated successfully",
        "new_end_date": new_end.isoformat()
    }

@api_router.get("/admin/contractors/{contractor_id}/financial-summary")
async def get_contractor_financial_summary(
    contractor_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Get financial summary for a contractor"""
    
    # Total wages processed
    attendance_cursor = db.attendance.find({"contractor_id": contractor_id})
    total_wages = 0
    commission_earned = 0
    async for record in attendance_cursor:
        wage = record.get('wage_amount', 0)
        total_wages += wage
        if record.get('mode') == 'employer':
            commission_earned += wage * 0.05
    
    # Payments collected
    payments_cursor = db.payments.find({"contractor_id": contractor_id, "type": "employer_payment"})
    payments_collected = 0
    async for payment in payments_cursor:
        payments_collected += payment.get('amount', 0)
    
    # Outstanding balances
    outstanding_cursor = db.employers.find({"contractor_id": contractor_id})
    outstanding = 0
    async for employer in outstanding_cursor:
        outstanding += employer.get('pending_payment', 0)
    
    # Settlements
    settlements_cursor = db.settlements.find({"contractor_id": contractor_id})
    total_settlements = 0
    async for settlement in settlements_cursor:
        total_settlements += settlement.get('amount_paid', 0)
    
    return {
        "total_wages_processed": round(total_wages, 2),
        "commission_earned": round(commission_earned, 2),
        "payments_collected": round(payments_collected, 2),
        "outstanding_balances": round(outstanding, 2),
        "total_settlements": round(total_settlements, 2)
    }

@api_router.get("/admin/contractors/{contractor_id}/activity-log")
async def get_contractor_activity_log(
    contractor_id: str,
    admin: AdminUser = Depends(get_current_admin),
    limit: int = 50
):
    """Get activity log for a contractor"""
    
    activities = []
    
    # Recent attendance records
    attendance = await db.attendance.find(
        {"contractor_id": contractor_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    for record in attendance:
        activities.append({
            "type": "attendance",
            "action": f"{record.get('mode')} attendance marked",
            "date": record.get('date'),
            "timestamp": record.get('created_at', ''),
            "details": f"Wage: ₹{record.get('wage_amount', 0)}"
        })
    
    # Recent payments
    payments = await db.payments.find(
        {"contractor_id": contractor_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    for payment in payments:
        activities.append({
            "type": "payment",
            "action": f"{payment.get('type')} recorded",
            "date": payment.get('date'),
            "timestamp": payment.get('created_at', ''),
            "details": f"Amount: ₹{payment.get('amount', 0)}"
        })
    
    # Sort all activities by timestamp
    activities.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
    
    return activities[:limit]

@api_router.get("/admin/contractors/{contractor_id}/workers-stats")
async def get_contractor_workers_stats(
    contractor_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Get worker statistics for a contractor"""
    
    total_workers = await db.workers.count_documents({"contractor_id": contractor_id})
    active_workers = await db.workers.count_documents({"contractor_id": contractor_id, "status": "Active"})
    inactive_workers = await db.workers.count_documents({"contractor_id": contractor_id, "status": "Inactive"})
    
    # Last 7 days attendance
    seven_days_ago = (datetime.now(timezone.utc) - timedelta(days=7))
    recent_attendance = await db.attendance.count_documents({
        "contractor_id": contractor_id,
        "created_at": {"$gte": seven_days_ago.isoformat()}
    })
    
    return {
        "total_workers": total_workers,
        "active_workers": active_workers,
        "inactive_workers": inactive_workers,
        "attendance_last_7_days": recent_attendance
    }

@api_router.post("/admin/bulk-actions")
async def bulk_actions(
    action_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Perform bulk actions on contractors"""
    action = action_data.get('action')
    contractor_ids = action_data.get('contractor_ids', [])
    
    if not contractor_ids:
        raise HTTPException(status_code=400, detail="No contractors selected")
    
    if action == "activate":
        result = await db.users.update_many(
            {"id": {"$in": contractor_ids}},
            {"$set": {"subscription_status": "active"}}
        )
        return {"message": f"Activated {result.modified_count} contractors"}
    
    elif action == "suspend":
        result = await db.users.update_many(
            {"id": {"$in": contractor_ids}},
            {"$set": {"subscription_status": "suspended"}}
        )
        return {"message": f"Suspended {result.modified_count} contractors"}
    
    elif action == "delete":
        # Delete contractors and their data
        for contractor_id in contractor_ids:
            await db.users.delete_one({"id": contractor_id})
            await db.workers.delete_many({"contractor_id": contractor_id})
            await db.employers.delete_many({"contractor_id": contractor_id})
            await db.attendance.delete_many({"contractor_id": contractor_id})
            await db.payments.delete_many({"contractor_id": contractor_id})
            await db.settlements.delete_many({"contractor_id": contractor_id})
        
        return {"message": f"Deleted {len(contractor_ids)} contractors"}
    
    else:
        raise HTTPException(status_code=400, detail="Invalid action")

@api_router.get("/admin/contractors/{contractor_id}/export")
async def export_contractor_data(
    contractor_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Export contractor data as CSV"""
    contractor = await db.users.find_one({"id": contractor_id}, {"_id": 0, "password": 0})
    
    if not contractor:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    # Get all related data
    workers = await db.workers.find({"contractor_id": contractor_id}, {"_id": 0}).to_list(1000)
    employers = await db.employers.find({"contractor_id": contractor_id}, {"_id": 0}).to_list(1000)
    attendance = await db.attendance.find({"contractor_id": contractor_id}, {"_id": 0}).to_list(10000)
    
    # Create CSV in memory
    output = io.StringIO()
    
    # Write contractor info
    output.write("CONTRACTOR INFORMATION\n")
    output.write(f"Name,{contractor.get('name', '')}\n")
    output.write(f"Email,{contractor.get('email', '')}\n")
    output.write(f"Phone,{contractor.get('phone_number', '')}\n")
    output.write(f"Business Name,{contractor.get('business_name', '')}\n")
    output.write(f"Plan,{contractor.get('subscription_plan', '')}\n")
    output.write(f"Status,{contractor.get('subscription_status', '')}\n")
    output.write("\n\n")
    
    # Write workers
    output.write("WORKERS\n")
    if workers:
        writer = csv.DictWriter(output, fieldnames=workers[0].keys())
        writer.writeheader()
        writer.writerows(workers)
    output.write("\n\n")
    
    # Write employers
    output.write("EMPLOYERS\n")
    if employers:
        writer = csv.DictWriter(output, fieldnames=employers[0].keys())
        writer.writeheader()
        writer.writerows(employers)
    output.write("\n\n")
    
    # Write attendance summary
    output.write("ATTENDANCE RECORDS\n")
    if attendance:
        writer = csv.DictWriter(output, fieldnames=attendance[0].keys())
        writer.writeheader()
        writer.writerows(attendance)
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=contractor_{contractor_id}_data.csv"}
    )

# ============ ACTIVATION KEY MANAGEMENT ============

import secrets
import string

def generate_activation_key():
    """Generate a secure random activation key"""
    # Format: XXXX-XXXX-XXXX-XXXX (16 characters, uppercase alphanumeric)
    chars = string.ascii_uppercase + string.digits
    key = '-'.join([''.join(secrets.choice(chars) for _ in range(4)) for _ in range(4)])
    return key

@api_router.post("/admin/activation-keys/generate")
async def generate_activation_key_endpoint(
    key_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Generate a new activation key"""
    plan = key_data.get('plan')
    max_uses = key_data.get('max_uses', 1)
    duration_days = key_data.get('duration_days', 30)
    notes = key_data.get('notes', '')
    is_paid = key_data.get('is_paid', False)  # Whether this key was sold (for invoice generation)
    price = key_data.get('price', 0)  # Price if paid key
    
    # SECURITY: Validate plan name against allowed plans
    valid_plans = ['Contractor Plus', 'Contractor Pro', 'Enterprise']
    if plan not in valid_plans:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid plan. Allowed plans: {', '.join(valid_plans)}"
        )
    
    if max_uses < 1:
        raise HTTPException(status_code=400, detail="Max uses must be at least 1")
    
    # Generate unique key
    key = generate_activation_key()
    
    # Check if key already exists (very unlikely but possible)
    existing = await db.activation_keys.find_one({"key": key})
    while existing:
        key = generate_activation_key()
        existing = await db.activation_keys.find_one({"key": key})
    
    activation_key = {
        "id": str(uuid.uuid4()),
        "key": key,
        "plan": plan,
        "max_uses": max_uses,
        "current_uses": 0,
        "duration_days": duration_days,
        "is_active": True,
        "is_paid": is_paid,  # NEW: Track if key was sold
        "price": price if is_paid else 0,  # NEW: Price if paid
        "notes": notes,
        "created_by": admin.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "used_by": []  # List of user IDs who used this key
    }
    
    await db.activation_keys.insert_one(activation_key)
    
    return {
        "message": "Activation key generated successfully",
        "key": key,
        "details": {
            "plan": plan,
            "max_uses": max_uses,
            "duration_days": duration_days
        }
    }

@api_router.get("/admin/activation-keys")
async def get_activation_keys(
    admin: AdminUser = Depends(get_current_admin),
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all activation keys with user information"""
    query = {}
    
    if status == "active":
        query["is_active"] = True
    elif status == "inactive":
        query["is_active"] = False
    elif status == "expired":
        query["$expr"] = {"$gte": ["$current_uses", "$max_uses"]}
    
    keys = await db.activation_keys.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # For each key, fetch the users who used it
    for key in keys:
        key_value = key.get("key")
        
        # Find users who activated with this key
        users = await db.users.find(
            {"activation_key": key_value},
            {"_id": 0, "id": 1, "name": 1, "email": 1, "plan_start_date": 1}
        ).to_list(100)
        
        # Format user data
        used_by = []
        for user in users:
            used_by.append({
                "user_id": user.get("id"),
                "user_name": user.get("name"),
                "user_email": user.get("email"),
                "used_at": user.get("plan_start_date")
            })
        
        key["used_by"] = used_by
    
    total = await db.activation_keys.count_documents(query)
    
    return {
        "keys": keys,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@api_router.put("/admin/activation-keys/{key_id}/toggle")
async def toggle_activation_key(
    key_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Toggle activation key active status"""
    key = await db.activation_keys.find_one({"id": key_id})
    
    if not key:
        raise HTTPException(status_code=404, detail="Activation key not found")
    
    new_status = not key.get('is_active', True)
    
    await db.activation_keys.update_one(
        {"id": key_id},
        {"$set": {"is_active": new_status}}
    )
    
    return {"message": f"Activation key {'activated' if new_status else 'deactivated'} successfully"}

@api_router.delete("/admin/activation-keys/{key_id}")
async def delete_activation_key(
    key_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete an activation key"""
    result = await db.activation_keys.delete_one({"id": key_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Activation key not found")
    
    return {"message": "Activation key deleted successfully"}

# ============ EXTENSION KEYS MANAGEMENT (ADMIN) ============

@api_router.post("/admin/extension-keys/generate")
async def generate_extension_key(
    key_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Generate a new validity extension key for a specific plan"""
    
    plan = key_data.get("plan")  # Plan name (optional - if not provided, works for any plan)
    duration_days = key_data.get("duration_days", 30)
    max_uses = key_data.get("max_uses", 1)
    notes = key_data.get("notes", "")
    expiry_date = key_data.get("expiry_date")  # Optional expiry date
    
    # Validate plan if provided
    if plan:
        valid_plans = ['Contractor Plus', 'Contractor Pro', 'Enterprise']
        if plan not in valid_plans:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid plan. Allowed plans: {', '.join(valid_plans)}"
            )
    
    # Generate unique extension key
    extension_key = f"EXT-{str(uuid.uuid4())[:8].upper()}-{str(uuid.uuid4())[:8].upper()}"
    
    key_doc = {
        "id": str(uuid.uuid4()),
        "key": extension_key,
        "plan": plan,  # Plan name (optional)
        "duration_days": duration_days,
        "max_uses": max_uses,
        "current_uses": 0,
        "is_active": True,
        "used": False,
        "notes": notes,
        "expiry_date": expiry_date,
        "created_by": admin.email,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.extension_keys.insert_one(key_doc)
    
    return {
        "message": "Extension key generated successfully",
        "key": extension_key,
        "details": {
            "duration_days": duration_days,
            "max_uses": max_uses
        }
    }

@api_router.get("/admin/extension-keys")
async def get_extension_keys(
    admin: AdminUser = Depends(get_current_admin),
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all extension keys with usage information"""
    query = {}
    
    if status == "active":
        query["is_active"] = True
        query["used"] = False
    elif status == "inactive":
        query["is_active"] = False
    elif status == "used":
        query["used"] = True
    
    keys = await db.extension_keys.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # For each key, get usage information
    for key in keys:
        key_value = key.get("key")
        
        # Find users who used this extension key
        usages = await db.extension_key_usage.find(
            {"extension_key": key_value},
            {"_id": 0, "user_id": 1, "user_email": 1, "days_extended": 1, "applied_at": 1}
        ).to_list(100)
        
        key["used_by"] = usages
    
    total = await db.extension_keys.count_documents(query)
    
    return {
        "keys": keys,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@api_router.put("/admin/extension-keys/{key_id}/toggle")
async def toggle_extension_key(
    key_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Toggle extension key active status"""
    key = await db.extension_keys.find_one({"id": key_id})
    
    if not key:
        raise HTTPException(status_code=404, detail="Extension key not found")
    
    new_status = not key.get('is_active', True)
    
    await db.extension_keys.update_one(
        {"id": key_id},
        {"$set": {"is_active": new_status}}
    )
    
    return {"message": f"Extension key {'activated' if new_status else 'deactivated'} successfully"}

@api_router.delete("/admin/extension-keys/{key_id}")
async def delete_extension_key(
    key_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete an extension key"""
    result = await db.extension_keys.delete_one({"id": key_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Extension key not found")
    
    return {"message": "Extension key deleted successfully"}

# ============ ACCOUNT DELETION REQUESTS MANAGEMENT (ADMIN) ============

@api_router.get("/admin/deletion-requests")
async def get_deletion_requests(
    admin: AdminUser = Depends(get_current_admin),
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all account deletion requests"""
    try:
        query = {}
        
        if status:
            query["status"] = status
        
        requests = await db.deletion_requests.find(
            query,
            {"_id": 0}
        ).sort("requested_at", -1).skip(skip).limit(limit).to_list(limit)
        
        total = await db.deletion_requests.count_documents(query)
        
        return {
            "requests": requests,
            "total": total,
            "skip": skip,
            "limit": limit
        }
    except Exception as e:
        print(f"Error fetching deletion requests: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to fetch deletion requests: {str(e)}")

@api_router.put("/admin/deletion-requests/{request_id}/approve")
async def approve_deletion_request(
    request_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Approve account deletion request - schedules deletion for 30 days later"""
    try:
        # Find the deletion request
        deletion_req = await db.deletion_requests.find_one({"id": request_id})
        
        if not deletion_req:
            raise HTTPException(status_code=404, detail="Deletion request not found")
        
        if deletion_req.get("status") != "pending":
            raise HTTPException(status_code=400, detail="Request already processed")
        
        # Schedule deletion for 30 days from now
        now = datetime.now(timezone.utc)
        scheduled_deletion = now + timedelta(days=30)
        
        # Update deletion request status to approved with scheduled date
        await db.deletion_requests.update_one(
            {"id": request_id},
            {"$set": {
                "status": "approved",
                "approved_at": now.isoformat(),
                "scheduled_deletion_date": scheduled_deletion.isoformat(),
                "processed_by": admin.email
            }}
        )
        
        return {
            "message": f"Account deletion approved. Data will be permanently deleted on {scheduled_deletion.strftime('%Y-%m-%d')}",
            "scheduled_deletion_date": scheduled_deletion.isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error approving deletion request: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to process deletion: {str(e)}")

@api_router.put("/admin/deletion-requests/{request_id}/reject")
async def reject_deletion_request(
    request_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Reject account deletion request"""
    try:
        deletion_req = await db.deletion_requests.find_one({"id": request_id})
        
        if not deletion_req:
            raise HTTPException(status_code=404, detail="Deletion request not found")
        
        if deletion_req.get("status") != "pending":
            raise HTTPException(status_code=400, detail="Request already processed")
        
        # Update deletion request status
        await db.deletion_requests.update_one(
            {"id": request_id},
            {"$set": {
                "status": "rejected",
                "processed_at": datetime.now(timezone.utc).isoformat(),
                "processed_by": admin.email
            }}
        )
        
        return {"message": "Deletion request rejected"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error rejecting deletion request: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to reject request: {str(e)}")

# ============ PROMO/DISCOUNT CODES MANAGEMENT (ADMIN) ============

@api_router.post("/admin/promo-codes/generate")
async def generate_promo_code(
    promo_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Generate a new promo/discount code for a specific plan"""
    
    code = promo_data.get("code", "").strip().upper()
    plan = promo_data.get("plan")  # Plan name (optional - if not provided, works for any plan)
    discount_type = promo_data.get("discount_type", "percentage")  # percentage or fixed
    discount_value = promo_data.get("discount_value", 0)
    max_uses = promo_data.get("max_uses", 0)  # 0 = unlimited
    description = promo_data.get("description", "")
    expiry_date = promo_data.get("expiry_date")  # Optional
    
    # Validate plan if provided
    if plan:
        valid_plans = ['Contractor Plus', 'Contractor Pro', 'Enterprise']
        if plan not in valid_plans:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid plan. Allowed plans: {', '.join(valid_plans)}"
            )
    
    if not code:
        # Auto-generate code if not provided
        code = f"PROMO{str(uuid.uuid4())[:8].upper()}"
    
    # Check if code already exists
    existing = await db.promo_codes.find_one({"code": code})
    if existing:
        raise HTTPException(status_code=400, detail="Promo code already exists")
    
    promo_doc = {
        "id": str(uuid.uuid4()),
        "code": code,
        "plan": plan,  # Plan name (optional)
        "discount_type": discount_type,
        "discount_value": discount_value,
        "max_uses": max_uses,
        "times_used": 0,
        "used_by": [],
        "is_active": True,
        "description": description,
        "expiry_date": expiry_date,
        "created_by": admin.email,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.promo_codes.insert_one(promo_doc)
    
    return {
        "message": "Promo code generated successfully",
        "code": code,
        "details": {
            "discount_type": discount_type,
            "discount_value": discount_value,
            "max_uses": max_uses
        }
    }

@api_router.get("/admin/promo-codes")
async def get_promo_codes(
    admin: AdminUser = Depends(get_current_admin),
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all promo codes"""
    try:
        query = {}
        
        if status == "active":
            query["is_active"] = True
        elif status == "inactive":
            query["is_active"] = False
        elif status == "expired":
            # Will be checked on the fly with expiry_date
            pass
        
        promos = await db.promo_codes.find(
            query,
            {"_id": 0}
        ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
        
        # Check expiry status for each promo
        now = datetime.now(timezone.utc)
        for promo in promos:
            try:
                if promo.get("expiry_date"):
                    # Handle both string and datetime formats
                    expiry_date = promo["expiry_date"]
                    if isinstance(expiry_date, str):
                        expiry = datetime.fromisoformat(expiry_date.replace('Z', '+00:00'))
                    else:
                        expiry = expiry_date
                    promo["is_expired"] = now > expiry
                else:
                    promo["is_expired"] = False
            except Exception as e:
                print(f"Error parsing expiry date for promo {promo.get('code')}: {e}")
                promo["is_expired"] = False
            
            # Check if max uses reached
            max_uses = promo.get("max_uses", 0)
            times_used = promo.get("times_used", 0)
            promo["is_exhausted"] = max_uses > 0 and times_used >= max_uses
        
        total = await db.promo_codes.count_documents(query)
        
        return {
            "promo_codes": promos,
            "total": total,
            "skip": skip,
            "limit": limit
        }
    except Exception as e:
        print(f"Error fetching promo codes: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to fetch promo codes: {str(e)}")

@api_router.put("/admin/promo-codes/{promo_id}/toggle")
async def toggle_promo_code(
    promo_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Toggle promo code active status"""
    promo = await db.promo_codes.find_one({"id": promo_id})
    
    if not promo:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    new_status = not promo.get('is_active', True)
    
    await db.promo_codes.update_one(
        {"id": promo_id},
        {"$set": {"is_active": new_status}}
    )
    
    return {"message": f"Promo code {'activated' if new_status else 'deactivated'} successfully"}

@api_router.put("/admin/promo-codes/{promo_id}")
async def update_promo_code(
    promo_id: str,
    promo_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Update a promo code"""
    promo = await db.promo_codes.find_one({"id": promo_id})
    
    if not promo:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    # Allowed fields to update
    update_fields = {}
    if "discount_value" in promo_data:
        update_fields["discount_value"] = promo_data["discount_value"]
    if "max_uses" in promo_data:
        update_fields["max_uses"] = promo_data["max_uses"]
    if "description" in promo_data:
        update_fields["description"] = promo_data["description"]
    if "expiry_date" in promo_data:
        update_fields["expiry_date"] = promo_data["expiry_date"]
    
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.promo_codes.update_one(
            {"id": promo_id},
            {"$set": update_fields}
        )
    
    return {"message": "Promo code updated successfully"}

@api_router.delete("/admin/promo-codes/{promo_id}")
async def delete_promo_code(
    promo_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete a promo code"""
    result = await db.promo_codes.delete_one({"id": promo_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    return {"message": "Promo code deleted successfully"}

# ============ SUBSCRIPTION PLANS MANAGEMENT (ADMIN) ============

@api_router.post("/admin/plans")
async def create_plan(
    plan_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Create a new subscription plan"""
    
    name = plan_data.get("name", "").strip()
    price = plan_data.get("price", 0)
    duration_days = plan_data.get("duration_days", 30)
    features = plan_data.get("features", [])
    description = plan_data.get("description", "")
    is_active = plan_data.get("is_active", True)
    
    if not name:
        raise HTTPException(status_code=400, detail="Plan name is required")
    
    # Check if plan with same name exists
    existing = await db.subscription_plans.find_one({"name": name})
    if existing:
        raise HTTPException(status_code=400, detail="Plan with this name already exists")
    
    # Get plan limits (max_workers, max_employers)
    max_workers = plan_data.get("max_workers")  # None means unlimited
    max_employers = plan_data.get("max_employers")  # None means unlimited
    
    plan_doc = {
        "id": str(uuid.uuid4()),
        "name": name,
        "price": price,
        "duration_days": duration_days,
        "features": features,
        "description": description,
        "is_active": is_active,
        "max_workers": max_workers,  # None = unlimited
        "max_employers": max_employers,  # None = unlimited
        "created_by": admin.email,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.subscription_plans.insert_one(plan_doc)
    
    return {
        "message": "Plan created successfully",
        "plan": {
            "id": plan_doc["id"],
            "name": name,
            "price": price,
            "duration_days": duration_days
        }
    }

@api_router.get("/admin/plans")
async def get_all_plans_admin(
    admin: AdminUser = Depends(get_current_admin),
    include_inactive: bool = False
):
    """Get all subscription plans (admin view)"""
    
    query = {} if include_inactive else {"is_active": True}
    
    plans = await db.subscription_plans.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    # Get subscriber counts for each plan
    for plan in plans:
        # Count active subscribers
        subscriber_count = await db.users.count_documents({
            "plan_type": plan["name"],
            "subscription_status": "active"
        })
        plan["subscriber_count"] = subscriber_count
    
    return {"plans": plans}

@api_router.get("/plans")
async def get_active_plans():
    """Get all active subscription plans (public endpoint for pricing page)"""
    
    plans = await db.subscription_plans.find(
        {"is_active": True},
        {"_id": 0}
    ).sort("price", 1).to_list(100)
    
    return {"plans": plans}

@api_router.put("/admin/plans/{plan_id}")
async def update_plan(
    plan_id: str,
    plan_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Update a subscription plan"""
    
    plan = await db.subscription_plans.find_one({"id": plan_id})
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Build update document
    update_fields = {}
    
    if "name" in plan_data:
        # Check if new name conflicts with existing plan
        existing = await db.subscription_plans.find_one({
            "name": plan_data["name"],
            "id": {"$ne": plan_id}
        })
        if existing:
            raise HTTPException(status_code=400, detail="Plan with this name already exists")
        update_fields["name"] = plan_data["name"]
    
    if "price" in plan_data:
        update_fields["price"] = plan_data["price"]
    if "duration_days" in plan_data:
        update_fields["duration_days"] = plan_data["duration_days"]
    if "features" in plan_data:
        update_fields["features"] = plan_data["features"]
    if "description" in plan_data:
        update_fields["description"] = plan_data["description"]
    if "is_active" in plan_data:
        update_fields["is_active"] = plan_data["is_active"]
    if "max_workers" in plan_data:
        update_fields["max_workers"] = plan_data["max_workers"]  # None = unlimited
    if "max_employers" in plan_data:
        update_fields["max_employers"] = plan_data["max_employers"]  # None = unlimited
    if "razorpay_plan_id" in plan_data:
        update_fields["razorpay_plan_id"] = plan_data["razorpay_plan_id"]
    
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        update_fields["updated_by"] = admin.email
        
        await db.subscription_plans.update_one(
            {"id": plan_id},
            {"$set": update_fields}
        )
    
    return {"message": "Plan updated successfully"}

@api_router.delete("/admin/plans/{plan_id}")
async def delete_plan(
    plan_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete a subscription plan"""
    
    plan = await db.subscription_plans.find_one({"id": plan_id})
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Check if any users are using this plan
    users_count = await db.users.count_documents({
        "plan_type": plan["name"],
        "subscription_status": "active"
    })
    
    if users_count > 0:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete plan: {users_count} active subscribers are using this plan"
        )
    
    result = await db.subscription_plans.delete_one({"id": plan_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    return {"message": "Plan deleted successfully"}

@api_router.post("/admin/plans/{plan_id}/create-razorpay-plan")
async def create_razorpay_plan_for_plan(
    plan_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """
    Create a Razorpay subscription plan for this plan
    This creates a monthly recurring plan in Razorpay with the same price
    """
    # Get the plan
    plan = await db.subscription_plans.find_one({"id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Check if already has Razorpay plan ID
    if plan.get("razorpay_plan_id"):
        raise HTTPException(
            status_code=400, 
            detail=f"Plan already has Razorpay Plan ID: {plan['razorpay_plan_id']}"
        )
    
    try:
        # Create Razorpay plan
        # Price in paise (multiply by 100)
        price_in_paise = int(plan.get("price", 0) * 100)
        
        razorpay_plan = razorpay_client.plan.create({
            "period": "monthly",
            "interval": 1,
            "item": {
                "name": plan.get("name", "Subscription Plan"),
                "description": plan.get("description", f"{plan.get('name')} - Monthly Subscription"),
                "amount": price_in_paise,
                "currency": "INR"
            },
            "notes": {
                "plan_id": plan_id,
                "created_from": "admin_panel",
                "created_by": admin.email
            }
        })
        
        razorpay_plan_id = razorpay_plan["id"]
        
        # Update the plan with Razorpay plan ID
        await db.subscription_plans.update_one(
            {"id": plan_id},
            {"$set": {
                "razorpay_plan_id": razorpay_plan_id,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": admin.email
            }}
        )
        
        return {
            "message": "Razorpay plan created successfully",
            "razorpay_plan_id": razorpay_plan_id,
            "plan_name": plan.get("name"),
            "price": plan.get("price")
        }
        
    except Exception as e:
        print(f"Error creating Razorpay plan: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create Razorpay plan: {str(e)}"
        )

# ============ CONTRACTOR ACTIVATION KEY REDEMPTION ============

@api_router.post("/auth/redeem-key")
async def redeem_activation_key(
    key_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Redeem an activation key to activate subscription"""
    activation_code = key_data.get('key', '').strip().upper()
    
    if not activation_code:
        raise HTTPException(status_code=400, detail="Activation key is required")
    
    # Find the activation key
    key = await db.activation_keys.find_one({"key": activation_code}, {"_id": 0})
    
    if not key:
        raise HTTPException(status_code=404, detail="Invalid activation key")
    
    # Check if key is active
    if not key.get('is_active'):
        raise HTTPException(status_code=400, detail="This activation key has been deactivated")
    
    # Check if key has reached max uses
    if key.get('current_uses', 0) >= key.get('max_uses', 1):
        raise HTTPException(status_code=400, detail="This activation key has reached its usage limit")
    
    # Check if user already used this key
    if current_user.id in key.get('used_by', []):
        raise HTTPException(status_code=400, detail="You have already used this activation key")
    
    # Get user from database to access plan_end_date
    user_from_db = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_from_db:
        raise HTTPException(status_code=404, detail="User not found")
    
    # SECURITY: Check if user already has an active subscription
    current_status = user_from_db.get('subscription_status', 'inactive')
    if current_status == 'active':
        raise HTTPException(
            status_code=400,
            detail="You already have an active subscription. Please wait for it to expire before activating a new plan."
        )
    
    # Calculate new subscription end date
    # If user is expired, suspended, or inactive - always start fresh from now
    current_end = user_from_db.get('subscription_end_date') or user_from_db.get('plan_end_date')
    
    if current_status in ['expired', 'suspended', 'inactive']:
        # Start fresh from now for expired/suspended/inactive users
        new_end = datetime.now(timezone.utc) + timedelta(days=key.get('duration_days', 30))
    elif current_end:
        try:
            if isinstance(current_end, str):
                current_end_dt = datetime.fromisoformat(current_end)
            else:
                current_end_dt = current_end
            # If subscription is still active and not expired, extend from current end date
            if current_end_dt > datetime.now(timezone.utc):
                new_end = current_end_dt + timedelta(days=key.get('duration_days', 30))
            else:
                # If expired, start from now
                new_end = datetime.now(timezone.utc) + timedelta(days=key.get('duration_days', 30))
        except:
            new_end = datetime.now(timezone.utc) + timedelta(days=key.get('duration_days', 30))
    else:
        new_end = datetime.now(timezone.utc) + timedelta(days=key.get('duration_days', 30))
    
    new_start = datetime.now(timezone.utc)
    
    # Update user subscription
    await db.users.update_one(
        {"id": current_user.id},
        {
            "$set": {
                "subscription_plan": key.get('plan'),
                "subscription_status": "active",
                "subscription_end_date": new_end.isoformat(),
                "plan_end_date": new_end.isoformat(),
                "plan_start_date": new_start.isoformat(),
                "payment_method": "activation_key",
                "activation_key": activation_code  # Store key for transaction history
            }
        }
    )
    
    # Update key usage
    await db.activation_keys.update_one(
        {"key": activation_code},
        {
            "$inc": {"current_uses": 1},
            "$push": {
                "used_by": current_user.id
            },
            "$set": {
                "last_used_at": datetime.now(timezone.utc).isoformat(),
                "last_used_by": {
                    "user_id": current_user.id,
                    "user_email": current_user.email,
                    "user_name": current_user.name
                }
            }
        }
    )
    
    # Record transaction in payment_orders for admin visibility and invoice generation
    is_paid = key.get('is_paid', False)
    
    # If this is a paid key, get the actual plan price from the database
    price = 0
    if is_paid:
        plan_doc = await db.subscription_plans.find_one(
            {"name": key.get('plan')},
            {"_id": 0, "price": 1}
        )
        if plan_doc:
            price = plan_doc.get('price', 0)
        else:
            # Fallback to stored price in key if plan not found
            price = key.get('price', 0)
    
    transaction_record = {
        "id": str(uuid.uuid4()),
        "contractor_id": current_user.id,
        "plan_name": key.get('plan'),
        "amount": price,  # Actual plan price if paid, otherwise 0
        "status": "success",
        "payment_method": "activation_key",
        "activation_key": activation_code,
        "is_paid": is_paid,  # Track if this was a paid key
        "duration_days": key.get('duration_days', 30),
        "razorpay_order_id": None,
        "razorpay_payment_id": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.payment_orders.insert_one(transaction_record)
    
    return {
        "message": "Activation key redeemed successfully!",
        "plan": key.get('plan'),
        "duration_days": key.get('duration_days'),
        "expires_at": new_end.isoformat()
    }
# ============ MESSAGES SYSTEM ============

class MessageReplyItem(BaseModel):
    """Individual reply in a conversation"""
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sender: str  # "user" or "admin"
    sender_name: str
    sender_email: str
    message: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Message(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_email: str
    user_name: str
    type: str  # "account_deletion", "support", "feedback", "bug_report"
    subject: str
    message: str  # Initial message
    replies: List[Dict] = []  # Conversation thread
    status: str = "unread"  # "unread", "read", "replied", "resolved"
    priority: str = "normal"  # "low", "normal", "high", "urgent"
    
    # Admin tracking
    admin_reply: Optional[str] = None  # Deprecated, kept for backward compatibility
    replied_by: Optional[str] = None
    replied_at: Optional[datetime] = None
    resolved_at: Optional[datetime] = None
    opened_by_admin: bool = False  # Track if admin opened this conversation
    opened_by_admin_at: Optional[datetime] = None  # When admin first opened
    admin_last_seen_at: Optional[datetime] = None  # Last time admin viewed
    has_new_user_reply: bool = False  # Flag for admin when user replies after admin response
    
    # User read tracking (visible to admin only)
    opened_by_user: bool = False
    opened_by_user_at: Optional[datetime] = None
    user_last_seen_at: Optional[datetime] = None
    
    # Metadata
    has_unread_admin_reply: bool = False  # For user to know if admin replied
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MessageCreate(BaseModel):
    type: str
    subject: str
    message: str
    priority: Optional[str] = "normal"

class MessageReply(BaseModel):
    reply: str

class UserReply(BaseModel):
    message_id: str
    reply: str

@api_router.post("/messages/send")
async def send_message(
    message_data: MessageCreate,
    current_user: User = Depends(get_current_user)
):
    """User sends a message to admin"""
    # Validate message type (account_deletion removed - use dedicated endpoint instead)
    allowed_types = ["support", "feedback", "bug_report"]
    if message_data.type not in allowed_types:
        raise HTTPException(status_code=400, detail="Invalid message type")
    
    # For all message types, proceed normally
    # Validate priority
    allowed_priorities = ["low", "normal", "high", "urgent"]
    priority = message_data.priority or "normal"
    if priority not in allowed_priorities:
        raise HTTPException(status_code=400, detail="Invalid priority level")
    
    # Validate input lengths
    if len(message_data.subject) > 200:
        raise HTTPException(status_code=400, detail="Subject too long (max 200 characters)")
    if len(message_data.message) > 2000:
        raise HTTPException(status_code=400, detail="Message too long (max 2000 characters)")
    
    message = Message(
        user_id=current_user.id,
        user_email=current_user.email,
        user_name=sanitize_input(current_user.name),  # Sanitize user name
        type=message_data.type,
        subject=sanitize_input(message_data.subject),
        message=sanitize_input(message_data.message),
        priority=priority
    )
    
    message_dict = message.model_dump()
    message_dict['created_at'] = message_dict['created_at'].isoformat()
    message_dict['updated_at'] = message_dict['updated_at'].isoformat()
    
    await db.messages.insert_one(message_dict)
    
    return {
        "message": "Message sent successfully",
        "message_id": message.id
    }

@api_router.get("/admin/messages")
async def get_all_messages(
    admin: AdminUser = Depends(get_current_admin),
    status: Optional[str] = None,
    type: Optional[str] = None,
    priority: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get all messages with filters"""
    query = {}
    
    if status:
        query["status"] = status
    if type:
        query["type"] = type
    if priority:
        query["priority"] = priority
    if search:
        query["$or"] = [
            {"subject": {"$regex": search, "$options": "i"}},
            {"message": {"$regex": search, "$options": "i"}},
            {"user_email": {"$regex": search, "$options": "i"}},
            {"user_name": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.messages.count_documents(query)
    messages = await db.messages.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Convert datetime strings
    for msg in messages:
        if isinstance(msg.get('created_at'), str):
            msg['created_at'] = datetime.fromisoformat(msg['created_at'])
        if isinstance(msg.get('updated_at'), str):
            msg['updated_at'] = datetime.fromisoformat(msg['updated_at'])
        if msg.get('replied_at') and isinstance(msg['replied_at'], str):
            msg['replied_at'] = datetime.fromisoformat(msg['replied_at'])
        if msg.get('resolved_at') and isinstance(msg['resolved_at'], str):
            msg['resolved_at'] = datetime.fromisoformat(msg['resolved_at'])
    
    # Get unread count
    unread_count = await db.messages.count_documents({"status": "unread"})
    
    return {
        "messages": messages,
        "total": total,
        "unread_count": unread_count,
        "skip": skip,
        "limit": limit
    }

@api_router.get("/admin/messages/unread-count")
async def get_admin_unread_count(
    admin: AdminUser = Depends(get_current_admin)
):
    """Get count of messages requiring admin attention"""
    # Count messages that are:
    # 1. New (not opened by admin yet), OR
    # 2. Have new user replies since admin last checked
    unread_count = await db.messages.count_documents({
        "$or": [
            {"opened_by_admin": False},
            {"has_new_user_reply": True}
        ]
    })
    
    return {"unread_count": unread_count}

@api_router.get("/admin/messages/{message_id}")
async def get_message_details(
    message_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Get detailed message information"""
    # Validate message_id format to prevent injection
    if not validate_uuid(message_id):
        raise HTTPException(status_code=400, detail="Invalid message ID format")
    
    message = await db.messages.find_one({"id": message_id}, {"_id": 0})
    
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    # Mark as read if unread
    if message.get('status') == 'unread':
        await db.messages.update_one(
            {"id": message_id},
            {"$set": {
                "status": "read",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        message['status'] = 'read'
    
    # Convert datetime strings
    if isinstance(message.get('created_at'), str):
        message['created_at'] = datetime.fromisoformat(message['created_at'])
    if isinstance(message.get('updated_at'), str):
        message['updated_at'] = datetime.fromisoformat(message['updated_at'])
    if message.get('replied_at') and isinstance(message['replied_at'], str):
        message['replied_at'] = datetime.fromisoformat(message['replied_at'])
    
    return message

@api_router.post("/admin/messages/{message_id}/reply")
async def reply_to_message(
    message_id: str,
    reply_data: MessageReply,
    admin: AdminUser = Depends(get_current_admin)
):
    """Reply to a message - adds to conversation thread"""
    # Validate message_id format to prevent injection
    if not validate_uuid(message_id):
        raise HTTPException(status_code=400, detail="Invalid message ID format")
    
    # Validate reply length
    if not reply_data.reply or len(reply_data.reply.strip()) == 0:
        raise HTTPException(status_code=400, detail="Reply cannot be empty")
    if len(reply_data.reply) > 2000:
        raise HTTPException(status_code=400, detail="Reply too long (max 2000 characters)")
    
    message = await db.messages.find_one({"id": message_id})
    
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    now = datetime.now(timezone.utc)
    
    # Create reply item with sanitized data
    reply_item = {
        "id": str(uuid.uuid4()),
        "sender": "admin",
        "sender_name": sanitize_input(admin.name),  # Sanitize admin name
        "sender_email": admin.email,
        "message": sanitize_input(reply_data.reply),
        "created_at": now.isoformat()
    }
    
    # Add reply to conversation thread
    current_replies = message.get('replies', [])
    current_replies.append(reply_item)
    
    await db.messages.update_one(
        {"id": message_id},
        {"$set": {
            "replies": current_replies,
            "admin_reply": sanitize_input(reply_data.reply),  # Backward compatibility
            "replied_by": admin.email,
            "replied_at": now.isoformat(),
            "status": "replied",
            "has_unread_admin_reply": True,  # Mark as unread for user
            "has_new_user_reply": False,  # Admin has responded to user reply
            "admin_last_seen_at": now.isoformat(),
            "updated_at": now.isoformat()
        }}
    )
    
    # Create notification for user about admin reply
    notification = Notification(
        user_id=message['user_id'],
        title="New Reply from Support",
        message=f"Admin has replied to your message: \"{message['subject'][:50]}...\"",
        type="info",
        link=f"/help?tab=messages&message_id={message_id}",
        metadata={"message_id": message_id, "type": "admin_reply"}
    )
    notification_dict = notification.model_dump()
    notification_dict['created_at'] = notification_dict['created_at'].isoformat()
    await db.notifications.insert_one(notification_dict)
    
    return {"message": "Reply sent successfully"}

@api_router.put("/admin/messages/{message_id}/status")
async def update_message_status(
    message_id: str,
    status_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Update message status or priority"""
    # Validate message_id format to prevent injection
    if not validate_uuid(message_id):
        raise HTTPException(status_code=400, detail="Invalid message ID format")
    
    update_fields = {
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Handle status update
    new_status = status_data.get('status')
    if new_status:
        if new_status not in ['unread', 'read', 'replied', 'resolved']:
            raise HTTPException(status_code=400, detail="Invalid status")
        
        update_fields["status"] = new_status
        
        if new_status == 'resolved':
            update_fields['resolved_at'] = datetime.now(timezone.utc).isoformat()
    
    # Handle priority update
    new_priority = status_data.get('priority')
    if new_priority:
        if new_priority not in ['low', 'normal', 'high', 'urgent']:
            raise HTTPException(status_code=400, detail="Invalid priority")
        
        update_fields["priority"] = new_priority
    
    result = await db.messages.update_one(
        {"id": message_id},
        {"$set": update_fields}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Message not found")
    
    if new_status:
        return {"message": f"Message marked as {new_status}"}
    elif new_priority:
        return {"message": f"Priority updated to {new_priority}"}
    else:
        return {"message": "Message updated"}

@api_router.delete("/admin/messages/{message_id}")
async def delete_message(
    message_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete a message"""
    # Validate message_id format to prevent injection
    if not validate_uuid(message_id):
        raise HTTPException(status_code=400, detail="Invalid message ID format")
    
    result = await db.messages.delete_one({"id": message_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Message not found")
    
    return {"message": "Message deleted successfully"}

@api_router.put("/admin/messages/{message_id}/mark-opened")
async def mark_message_opened_by_admin(
    message_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Mark message as opened by admin"""
    # Validate message_id format to prevent injection
    if not validate_uuid(message_id):
        raise HTTPException(status_code=400, detail="Invalid message ID format")
    
    now = datetime.now(timezone.utc)
    
    message = await db.messages.find_one({"id": message_id})
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    update_fields = {
        "opened_by_admin": True,
        "admin_last_seen_at": now.isoformat(),
        "has_new_user_reply": False,  # Clear notification when admin opens
        "updated_at": now.isoformat()
    }
    
    # Set first opened time if not already set
    if not message.get('opened_by_admin'):
        update_fields["opened_by_admin_at"] = now.isoformat()
    
    await db.messages.update_one(
        {"id": message_id},
        {"$set": update_fields}
    )
    
    return {"message": "Message marked as opened"}

# ============ USER MESSAGE ENDPOINTS ============

@api_router.get("/messages/my-messages")
async def get_user_messages(
    current_user: User = Depends(get_current_user)
):
    """Get all messages for current user with conversation threads"""
    messages = await db.messages.find(
        {"user_id": current_user.id},
        {"_id": 0}
    ).sort("updated_at", -1).to_list(1000)
    
    # Format datetime fields and remove sensitive data
    for message in messages:
        if isinstance(message.get('created_at'), str):
            message['created_at'] = datetime.fromisoformat(message['created_at'])
        if message.get('updated_at') and isinstance(message['updated_at'], str):
            message['updated_at'] = datetime.fromisoformat(message['updated_at'])
        if message.get('replied_at') and isinstance(message['replied_at'], str):
            message['replied_at'] = datetime.fromisoformat(message['replied_at'])
        
        # CRITICAL: Remove ALL admin tracking fields that users should never see
        message.pop('opened_by_user', None)
        message.pop('opened_by_user_at', None)
        message.pop('user_last_seen_at', None)
        message.pop('opened_by_admin', None)
        message.pop('opened_by_admin_at', None)
        message.pop('admin_last_seen_at', None)
        message.pop('has_new_user_reply', None)
        message.pop('replied_by', None)  # Don't expose admin email who replied
        
        # Remove admin email addresses from replies for privacy
        if message.get('replies'):
            for reply in message['replies']:
                if reply.get('sender') == 'admin':
                    reply.pop('sender_email', None)  # Remove admin email
                    reply['sender_name'] = 'Admin Support'  # Standardize admin name
    
    return {"messages": messages}

@api_router.post("/messages/{message_id}/reply")
async def user_reply_to_message(
    message_id: str,
    reply_data: MessageReply,
    current_user: User = Depends(get_current_user)
):
    """User replies to an admin message"""
    # Validate message_id format to prevent injection
    if not validate_uuid(message_id):
        raise HTTPException(status_code=400, detail="Invalid message ID format")
    
    # Validate reply length
    if not reply_data.reply or len(reply_data.reply.strip()) == 0:
        raise HTTPException(status_code=400, detail="Reply cannot be empty")
    if len(reply_data.reply) > 2000:
        raise HTTPException(status_code=400, detail="Reply too long (max 2000 characters)")
    
    message = await db.messages.find_one({"id": message_id, "user_id": current_user.id})
    
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    # Prevent replies to resolved messages
    if message.get('status') == 'resolved':
        raise HTTPException(status_code=400, detail="Cannot reply to a resolved message. Please create a new message.")
    
    now = datetime.now(timezone.utc)
    
    # Create reply item with sanitized data
    reply_item = {
        "id": str(uuid.uuid4()),
        "sender": "user",
        "sender_name": sanitize_input(current_user.name),  # Sanitize user name
        "sender_email": current_user.email,
        "message": sanitize_input(reply_data.reply),
        "created_at": now.isoformat()
    }
    
    # Add reply to conversation thread
    current_replies = message.get('replies', [])
    current_replies.append(reply_item)
    
    await db.messages.update_one(
        {"id": message_id},
        {"$set": {
            "replies": current_replies,
            "status": "replied",  # Keep as replied
            "has_unread_admin_reply": False,  # User has seen and replied
            "has_new_user_reply": True,  # Notify admin of new user reply
            "user_last_seen_at": now.isoformat(),
            "updated_at": now.isoformat()
        }}
    )
    
    return {"message": "Reply sent successfully"}

@api_router.put("/messages/{message_id}/mark-opened")
async def mark_message_opened(
    message_id: str,
    current_user: User = Depends(get_current_user)
):
    """Mark message as opened by user (for admin tracking)"""
    # Validate message_id format to prevent injection
    if not validate_uuid(message_id):
        raise HTTPException(status_code=400, detail="Invalid message ID format")
    
    message = await db.messages.find_one({"id": message_id, "user_id": current_user.id})
    
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    now = datetime.now(timezone.utc)
    
    update_fields = {
        "user_last_seen_at": now.isoformat(),
        "has_unread_admin_reply": False
    }
    
    # Only set opened_by_user_at once (first time opened)
    if not message.get('opened_by_user'):
        update_fields['opened_by_user'] = True
        update_fields['opened_by_user_at'] = now.isoformat()
    
    await db.messages.update_one(
        {"id": message_id},
        {"$set": update_fields}
    )
    
    return {"message": "Message marked as opened"}

@api_router.get("/messages/unread-count")
async def get_user_unread_count(
    current_user: User = Depends(get_current_user)
):
    """Get count of unread admin replies for current user"""
    unread_count = await db.messages.count_documents({
        "user_id": current_user.id,
        "has_unread_admin_reply": True
    })
    
    return {"unread_count": unread_count}

@api_router.get("/admin/messages/user-profile/{user_id}")
async def get_user_profile_for_admin(
    user_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Get user profile information for admin (quick view)"""
    # Validate user_id format to prevent injection
    if not validate_uuid(user_id):
        raise HTTPException(status_code=400, detail="Invalid user ID format")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get statistics
    total_workers = await db.workers.count_documents({"contractor_id": user_id})
    total_employers = await db.employers.count_documents({"contractor_id": user_id})
    total_messages = await db.messages.count_documents({"user_id": user_id})
    
    # Get recent activity
    recent_attendance = await db.worker_attendance.find(
        {"contractor_id": user_id}
    ).sort("date", -1).limit(5).to_list(5)
    
    return {
        "user": user,
        "stats": {
            "total_workers": total_workers,
            "total_employers": total_employers,
            "total_messages": total_messages,
            "subscription_status": user.get('subscription_status', 'inactive'),
            "subscription_plan": user.get('subscription_plan', 'free')
        },
        "recent_activity": recent_attendance
    }

# ============ CONTACT FORM (UNLOGGED USERS) ============

class ContactForm(BaseModel):
    name: str
    email: str
    subject: str
    message: str

@api_router.post("/contact")
@limiter.limit("5/minute")  # Rate limit: 5 contact form submissions per minute per IP
async def submit_contact_form(request: Request, contact_data: ContactForm):
    """Submit contact form (public endpoint - no auth required)"""
    # Validate input
    if not contact_data.name or not contact_data.name.strip():
        raise HTTPException(status_code=400, detail="Name is required")
    if not contact_data.email or not contact_data.email.strip():
        raise HTTPException(status_code=400, detail="Email is required")
    
    # Use proper email validation
    is_valid_email, email_error = validate_email_format(contact_data.email)
    if not is_valid_email:
        raise HTTPException(status_code=400, detail=email_error)
    if not contact_data.subject or not contact_data.subject.strip():
        raise HTTPException(status_code=400, detail="Subject is required")
    if not contact_data.message or not contact_data.message.strip():
        raise HTTPException(status_code=400, detail="Message is required")
    
    # Validate lengths
    if len(contact_data.name) > 100:
        raise HTTPException(status_code=400, detail="Name too long (max 100 characters)")
    if len(contact_data.email) > 100:
        raise HTTPException(status_code=400, detail="Email too long (max 100 characters)")
    if len(contact_data.subject) > 200:
        raise HTTPException(status_code=400, detail="Subject too long (max 200 characters)")
    if len(contact_data.message) > 2000:
        raise HTTPException(status_code=400, detail="Message too long (max 2000 characters)")
    if len(contact_data.message) < 10:
        raise HTTPException(status_code=400, detail="Message must be at least 10 characters long")
    
    # Sanitize inputs
    contact_doc = {
        "id": str(uuid.uuid4()),
        "name": sanitize_input(contact_data.name.strip()),
        "email": sanitize_input(contact_data.email.strip().lower()),
        "subject": sanitize_input(contact_data.subject.strip()),
        "message": sanitize_input(contact_data.message.strip()),
        "status": "unread",  # unread, read, replied, resolved
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "is_from_logged_user": False  # Flag to distinguish from user messages
    }
    
    await db.contact_messages.insert_one(contact_doc)
    
    return {
        "message": "Contact form submitted successfully",
        "contact_id": contact_doc["id"]
    }

@api_router.get("/admin/contact-messages")
async def get_contact_messages(
    admin: AdminUser = Depends(get_current_admin),
    status: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get all contact messages from unlogged users (admin only)"""
    query = {}
    
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"subject": {"$regex": search, "$options": "i"}},
            {"message": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"name": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.contact_messages.count_documents(query)
    messages = await db.contact_messages.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Convert datetime strings to datetime objects for display
    for msg in messages:
        if isinstance(msg.get('created_at'), str):
            msg['created_at'] = datetime.fromisoformat(msg['created_at'])
        if isinstance(msg.get('updated_at'), str):
            msg['updated_at'] = datetime.fromisoformat(msg['updated_at'])
    
    # Get unread count
    unread_count = await db.contact_messages.count_documents({"status": "unread"})
    
    return {
        "messages": messages,
        "total": total,
        "unread_count": unread_count,
        "skip": skip,
        "limit": limit
    }

@api_router.put("/admin/contact-messages/{contact_id}/status")
async def update_contact_message_status(
    contact_id: str,
    status_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Update status of a contact message"""
    # Validate contact_id format
    if not validate_uuid(contact_id):
        raise HTTPException(status_code=400, detail="Invalid contact ID format")
    
    allowed_statuses = ["unread", "read", "replied", "resolved"]
    new_status = status_data.get("status")
    
    if new_status not in allowed_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Allowed: {allowed_statuses}")
    
    contact = await db.contact_messages.find_one({"id": contact_id})
    
    if not contact:
        raise HTTPException(status_code=404, detail="Contact message not found")
    
    update_fields = {
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Track when status changed to resolved
    if new_status == "resolved":
        update_fields["resolved_at"] = datetime.now(timezone.utc).isoformat()
        update_fields["resolved_by"] = admin.email
    
    if new_status == "read" and contact.get("status") == "unread":
        update_fields["read_at"] = datetime.now(timezone.utc).isoformat()
        update_fields["read_by"] = admin.email
    
    await db.contact_messages.update_one(
        {"id": contact_id},
        {"$set": update_fields}
    )
    
    return {"message": f"Contact message status updated to {new_status}"}

@api_router.delete("/admin/contact-messages/{contact_id}")
async def delete_contact_message(
    contact_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete a contact message"""
    # Validate contact_id format
    if not validate_uuid(contact_id):
        raise HTTPException(status_code=400, detail="Invalid contact ID format")
    
    result = await db.contact_messages.delete_one({"id": contact_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contact message not found")
    
    return {"message": "Contact message deleted successfully"}

@api_router.get("/admin/messages/stats/summary")
async def get_messages_stats(
    admin: AdminUser = Depends(get_current_admin)
):
    """Get message statistics"""
    total_messages = await db.messages.count_documents({})
    unread_messages = await db.messages.count_documents({"status": "unread"})
    read_messages = await db.messages.count_documents({"status": "read"})
    replied_messages = await db.messages.count_documents({"status": "replied"})
    resolved_messages = await db.messages.count_documents({"status": "resolved"})
    
    # Count by type
    account_deletion_count = await db.messages.count_documents({"type": "account_deletion"})
    support_count = await db.messages.count_documents({"type": "support"})
    feedback_count = await db.messages.count_documents({"type": "feedback"})
    bug_report_count = await db.messages.count_documents({"type": "bug_report"})
    
    # Recent messages (last 7 days)
    seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)
    recent_messages = await db.messages.count_documents({
        "created_at": {"$gte": seven_days_ago.isoformat()}
    })
    
    return {
        "total_messages": total_messages,
        "unread_messages": unread_messages,
        "read_messages": read_messages,
        "replied_messages": replied_messages,
        "resolved_messages": resolved_messages,
        "by_type": {
            "account_deletion": account_deletion_count,
            "support": support_count,
            "feedback": feedback_count,
            "bug_report": bug_report_count
        },
        "recent_messages_7_days": recent_messages
    }


# ============ PAYMENT GATEWAY SETTINGS ============

import razorpay
from cryptography.fernet import Fernet
import base64
import hashlib

# Generate encryption key from JWT secret
def get_encryption_key():
    key = hashlib.sha256(JWT_SECRET_KEY.encode()).digest()
    return base64.urlsafe_b64encode(key)

def encrypt_data(data: str) -> str:
    """Encrypt sensitive data"""
    f = Fernet(get_encryption_key())
    return f.encrypt(data.encode()).decode()

def decrypt_data(encrypted_data: str) -> str:
    """Decrypt sensitive data"""
    try:
        f = Fernet(get_encryption_key())
        return f.decrypt(encrypted_data.encode()).decode()
    except:
        return ""

class GatewaySettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    razorpay_key_id: str
    razorpay_key_secret: str
    is_active: bool = True
    test_mode: bool = True
    webhook_secret: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentOrder(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    razorpay_order_id: str
    amount: float
    currency: str = "INR"
    status: str = "created"  # created, paid, failed
    payment_id: Optional[str] = None
    signature: Optional[str] = None
    plan_type: str  # "contractor_plan", "enterprise_plan"
    duration_days: int = 30
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.post("/admin/gateway/settings")
async def save_gateway_settings(
    settings_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Save or update gateway settings"""
    
    # Encrypt sensitive data
    encrypted_key = encrypt_data(settings_data.get('razorpay_key_id', ''))
    encrypted_secret = encrypt_data(settings_data.get('razorpay_key_secret', ''))
    encrypted_webhook = encrypt_data(settings_data.get('webhook_secret', '')) if settings_data.get('webhook_secret') else None
    
    # Check if settings already exist
    existing = await db.gateway_settings.find_one({})
    
    if existing:
        # Update existing settings
        await db.gateway_settings.update_one(
            {"id": existing['id']},
            {"$set": {
                "razorpay_key_id": encrypted_key,
                "razorpay_key_secret": encrypted_secret,
                "webhook_secret": encrypted_webhook,
                "test_mode": settings_data.get('test_mode', True),
                "is_active": settings_data.get('is_active', True),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        # Create new settings
        settings = GatewaySettings(
            razorpay_key_id=encrypted_key,
            razorpay_key_secret=encrypted_secret,
            webhook_secret=encrypted_webhook,
            test_mode=settings_data.get('test_mode', True),
            is_active=settings_data.get('is_active', True)
        )
        
        settings_dict = settings.model_dump()
        settings_dict['created_at'] = settings_dict['created_at'].isoformat()
        settings_dict['updated_at'] = settings_dict['updated_at'].isoformat()
        
        await db.gateway_settings.insert_one(settings_dict)
    
    return {"message": "Gateway settings saved successfully"}

@api_router.get("/admin/gateway/settings")
async def get_gateway_settings(
    admin: AdminUser = Depends(get_current_admin)
):
    """Get gateway settings (decrypted for admin view)"""
    settings = await db.gateway_settings.find_one({})
    
    if not settings:
        return {
            "configured": False,
            "settings": None
        }
    
    # Decrypt for admin view
    return {
        "configured": True,
        "settings": {
            "razorpay_key_id": decrypt_data(settings['razorpay_key_id']),
            "razorpay_key_secret": decrypt_data(settings['razorpay_key_secret']),
            "webhook_secret": decrypt_data(settings['webhook_secret']) if settings.get('webhook_secret') else "",
            "test_mode": settings.get('test_mode', True),
            "is_active": settings.get('is_active', True)
        }
    }

@api_router.get("/public/landing-stats")
async def get_landing_stats():
    """Get public statistics for landing page (no auth required)"""
    
    # Total contractors
    total_contractors = await db.users.count_documents({})
    
    # Total active workers
    total_workers = await db.workers.count_documents({"status": "Active"})
    
    # Total active employers
    total_employers = await db.employers.count_documents({"status": "Active"})
    
    # Total payments processed - calculate maximum from all sources
    # 1. Sum from employer_payments collection
    pipeline_employer = [
        {
            "$group": {
                "_id": None,
                "total": {"$sum": "$amount_paid"}
            }
        }
    ]
    employer_result = await db.employer_payments.aggregate(pipeline_employer).to_list(length=1)
    employer_payments = employer_result[0]["total"] if employer_result else 0
    
    # 2. Sum from employer_attendance collection (total_amount field)
    pipeline_attendance = [
        {
            "$group": {
                "_id": None,
                "total": {"$sum": "$total_amount"}
            }
        }
    ]
    attendance_result = await db.employer_attendance.aggregate(pipeline_attendance).to_list(length=1)
    attendance_payments = attendance_result[0]["total"] if attendance_result else 0
    
    # 3. Sum from employers' pending_payment field
    pipeline_pending = [
        {
            "$group": {
                "_id": None,
                "total": {"$sum": "$pending_payment"}
            }
        }
    ]
    pending_result = await db.employers.aggregate(pipeline_pending).to_list(length=1)
    pending_payments = pending_result[0]["total"] if pending_result else 0
    
    # Use the maximum value from all sources
    total_payments = max(employer_payments, attendance_payments, pending_payments + employer_payments)
    
    return {
        "total_contractors": total_contractors,
        "total_workers": total_workers,
        "total_employers": total_employers,
        "total_payments": total_payments
    }

@api_router.get("/gateway/public-key")
async def get_public_key():
    """Get Razorpay public key for frontend (no auth required) - Production: from environment variables"""
    key_id = os.getenv('RAZORPAY_KEY_ID')
    
    if not key_id:
        raise HTTPException(status_code=404, detail="Payment gateway not configured")
    
    return {
        "key_id": key_id,
        "test_mode": key_id.startswith('rzp_test_')
    }

async def get_razorpay_client():
    """Get configured Razorpay client - Production: from environment variables"""
    key_id = os.getenv('RAZORPAY_KEY_ID')
    key_secret = os.getenv('RAZORPAY_KEY_SECRET')
    
    if not key_id or not key_secret:
        raise HTTPException(status_code=400, detail="Payment gateway not configured. Please set RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET environment variables.")
    
    return razorpay.Client(auth=(key_id, key_secret))

@api_router.post("/payment/create-order")
async def create_payment_order(
    order_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Create a Razorpay order for subscription payment"""
    
    amount = order_data.get('amount')  # Amount in rupees
    plan_type = order_data.get('plan_type', 'Contractor Plus')  # Default plan name
    duration_days = order_data.get('duration_days', 30)
    promo_code = order_data.get('promo_code')  # Get promo code if applied
    
    if not amount or amount <= 0:
        raise HTTPException(status_code=400, detail="Invalid amount")
    
    try:
        client = await get_razorpay_client()
        
        # Create Razorpay order (amount in paise)
        razorpay_order = client.order.create({
            "amount": int(amount * 100),  # Convert to paise
            "currency": "INR",
            "payment_capture": 1,
            "notes": {
                "contractor_id": current_user.id,
                "plan_type": plan_type,
                "duration_days": str(duration_days)
            }
        })
        
        # Save order in database
        payment_order = PaymentOrder(
            contractor_id=current_user.id,
            razorpay_order_id=razorpay_order['id'],
            amount=amount,
            currency="INR",
            status="created",
            plan_type=plan_type,
            duration_days=duration_days
        )
        
        order_dict = payment_order.model_dump()
        order_dict['created_at'] = order_dict['created_at'].isoformat()
        order_dict['updated_at'] = order_dict['updated_at'].isoformat()
        
        # Add promo code if provided
        if promo_code:
            order_dict['promo_code'] = promo_code
        
        await db.payment_orders.insert_one(order_dict)
        
        return {
            "order_id": razorpay_order['id'],
            "amount": amount,
            "currency": "INR"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create order: {str(e)}")

@api_router.post("/payment/verify")
async def verify_payment(
    payment_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Verify payment and activate subscription"""
    
    razorpay_order_id = payment_data.get('razorpay_order_id')
    razorpay_payment_id = payment_data.get('razorpay_payment_id')
    razorpay_signature = payment_data.get('razorpay_signature')
    
    if not all([razorpay_order_id, razorpay_payment_id, razorpay_signature]):
        raise HTTPException(status_code=400, detail="Missing payment verification data")
    
    try:
        client = await get_razorpay_client()
        
        # Verify signature
        params_dict = {
            'razorpay_order_id': razorpay_order_id,
            'razorpay_payment_id': razorpay_payment_id,
            'razorpay_signature': razorpay_signature
        }
        
        client.utility.verify_payment_signature(params_dict)
        
        # Update order in database
        order = await db.payment_orders.find_one({
            "razorpay_order_id": razorpay_order_id,
            "contractor_id": current_user.id
        })
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Update order status
        await db.payment_orders.update_one(
            {"razorpay_order_id": razorpay_order_id},
            {"$set": {
                "status": "paid",
                "payment_id": razorpay_payment_id,
                "signature": razorpay_signature,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Check if this is a plan change payment
        if order.get('plan_change_transaction_id'):
            # This is a plan change payment - complete the plan change
            transaction_id = order['plan_change_transaction_id']
            transaction = await db.plan_change_transactions.find_one({"id": transaction_id})
            
            if not transaction:
                raise HTTPException(status_code=404, detail="Plan change transaction not found")
            
            # Update user subscription with plan change details
            update_fields = {
                "subscription_plan": transaction['new_plan'],
                "subscription_status": "active",
                "plan_end_date": transaction['new_end_date'],
                "subscription_end_date": transaction['new_end_date'],
                "payment_method": "razorpay",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.users.update_one(
                {"id": current_user.id},
                {"$set": update_fields}
            )
            
            # Mark transaction as completed
            await db.plan_change_transactions.update_one(
                {"id": transaction_id},
                {"$set": {
                    "status": "completed",
                    "payment_id": razorpay_payment_id,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Log security event
            await log_security_event(
                event_type=SecurityEventType.PLAN_CHANGE,
                severity=SecuritySeverity.LOW,
                ip_address="",
                user_id=current_user.id,
                email="",
                details={
                    "action": "plan_change_payment_completed",
                    "transaction_type": transaction['transaction_type'],
                    "old_plan": transaction['old_plan'],
                    "new_plan": transaction['new_plan'],
                    "amount_paid": transaction['amount_to_pay'],
                    "transaction_id": transaction_id,
                    "payment_id": razorpay_payment_id
                },
                endpoint="/payment/verify"
            )
            
            return {
                "success": True,
                "message": f"Payment verified! Plan changed to {transaction['new_plan']}",
                "subscription_end_date": transaction['new_end_date'],
                "plan_change": True,
                "new_plan": transaction['new_plan'],
                "prorated_days": transaction['prorated_days']
            }
        
        # Regular subscription payment (not plan change)
        user_dict = await db.users.find_one({"id": current_user.id})
        current_end = user_dict.get('subscription_end_date')
        if current_end and isinstance(current_end, str):
            current_end_dt = datetime.fromisoformat(current_end)
        else:
            current_end_dt = datetime.now(timezone.utc)
        
        # If current subscription is still active, extend from end date
        if current_end_dt > datetime.now(timezone.utc):
            new_end = current_end_dt + timedelta(days=order['duration_days'])
        else:
            new_end = datetime.now(timezone.utc) + timedelta(days=order['duration_days'])
        
        # Update user subscription
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {
                "subscription_plan": order['plan_type'],
                "subscription_status": "active",
                "subscription_start_date": datetime.now(timezone.utc).isoformat(),
                "subscription_end_date": new_end.isoformat(),
                "payment_method": "razorpay"
            }}
        )
        
        # If promo code was used, update usage count
        if order.get('promo_code'):
            promo_code = order['promo_code']
            await db.promo_codes.update_one(
                {"code": promo_code},
                {
                    "$inc": {"times_used": 1},
                    "$push": {"used_by": current_user.id},
                    "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
                }
            )
        
        return {
            "success": True,
            "message": "Payment verified and subscription activated",
            "subscription_end_date": new_end.isoformat()
        }
        
    except razorpay.errors.SignatureVerificationError:
        # Update order as failed
        await db.payment_orders.update_one(
            {"razorpay_order_id": razorpay_order_id},
            {"$set": {
                "status": "failed",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        raise HTTPException(status_code=400, detail="Payment verification failed")
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Verification error: {str(e)}")

@api_router.post("/payment/create-subscription")
async def create_razorpay_subscription(
    subscription_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Create a Razorpay subscription for auto-renewal"""
    
    plan_id = subscription_data.get('plan_id')
    
    if not plan_id:
        raise HTTPException(status_code=400, detail="Plan ID is required")
    
    try:
        # Get plan details from database
        plan = await db.subscription_plans.find_one({"id": plan_id, "is_active": True})
        
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")
        
        # Check if plan has razorpay_plan_id configured
        razorpay_plan_id = plan.get('razorpay_plan_id')
        
        if not razorpay_plan_id:
            raise HTTPException(
                status_code=400, 
                detail="This plan does not support auto-renewal. Please contact support or use one-time payment."
            )
        
        client = await get_razorpay_client()
        
        # Create Razorpay subscription
        razorpay_subscription = client.subscription.create({
            "plan_id": razorpay_plan_id,
            "customer_notify": 1,  # Send email/SMS to customer
            "total_count": 0,  # 0 = infinite renewals until cancelled
            "quantity": 1,
            "notes": {
                "contractor_id": current_user.id,
                "contractor_email": current_user.email,
                "plan_name": plan.get('name', 'Subscription Plan')
            }
        })
        
        subscription_id = razorpay_subscription['id']
        
        # Save subscription record in database
        subscription_record = {
            "id": str(uuid.uuid4()),
            "contractor_id": current_user.id,
            "razorpay_subscription_id": subscription_id,
            "razorpay_plan_id": razorpay_plan_id,
            "plan_name": plan.get('name'),
            "plan_id": plan_id,
            "amount": plan.get('price', 0),
            "currency": "INR",
            "status": "created",  # created, authenticated, active, paused, cancelled, completed
            "duration_days": plan.get('duration_days', 30),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.razorpay_subscriptions.insert_one(subscription_record)
        
        return {
            "subscription_id": subscription_id,
            "plan_name": plan.get('name'),
            "amount": plan.get('price', 0),
            "currency": "INR",
            "short_url": razorpay_subscription.get('short_url')  # Payment link if available
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Failed to create subscription: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create subscription: {str(e)}")

@api_router.post("/payment/verify-subscription")
async def verify_subscription_payment(
    payment_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Verify Razorpay subscription payment and activate subscription"""
    
    razorpay_payment_id = payment_data.get('razorpay_payment_id')
    razorpay_subscription_id = payment_data.get('razorpay_subscription_id')
    razorpay_signature = payment_data.get('razorpay_signature')
    
    if not all([razorpay_payment_id, razorpay_subscription_id, razorpay_signature]):
        raise HTTPException(status_code=400, detail="Missing payment verification data")
    
    try:
        client = await get_razorpay_client()
        
        # Verify signature
        params_dict = {
            'razorpay_payment_id': razorpay_payment_id,
            'razorpay_subscription_id': razorpay_subscription_id,
            'razorpay_signature': razorpay_signature
        }
        
        client.utility.verify_payment_signature(params_dict)
        
        # Get subscription details from database
        subscription = await db.razorpay_subscriptions.find_one({
            "razorpay_subscription_id": razorpay_subscription_id,
            "contractor_id": current_user.id
        })
        
        if not subscription:
            raise HTTPException(status_code=404, detail="Subscription not found")
        
        # Fetch subscription details from Razorpay to get status
        razorpay_sub_details = client.subscription.fetch(razorpay_subscription_id)
        
        # Update subscription status in database
        await db.razorpay_subscriptions.update_one(
            {"razorpay_subscription_id": razorpay_subscription_id},
            {"$set": {
                "status": razorpay_sub_details.get('status', 'active'),
                "payment_id": razorpay_payment_id,
                "signature": razorpay_signature,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        
        # Calculate subscription end date
        duration_days = subscription.get('duration_days', 30)
        now = datetime.now(timezone.utc)
        
        # Get current user subscription status
        user_dict = await db.users.find_one({"id": current_user.id})
        current_end = user_dict.get('subscription_end_date')
        
        if current_end and isinstance(current_end, str):
            current_end_dt = datetime.fromisoformat(current_end)
        else:
            current_end_dt = now
        
        # If current subscription is still active, extend from end date
        if current_end_dt > now:
            new_end = current_end_dt + timedelta(days=duration_days)
        else:
            new_end = now + timedelta(days=duration_days)
        
        # Update user subscription with auto-renewal enabled
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {
                "subscription_plan": subscription.get('plan_name'),
                "subscription_status": "active",
                "subscription_start_date": now.isoformat(),
                "subscription_end_date": new_end.isoformat(),
                "payment_method": "razorpay_subscription",
                "razorpay_subscription_id": razorpay_subscription_id,
                "auto_renew": True,
                "updated_at": now.isoformat()
            }}
        )
        
        # Create payment record for this transaction
        payment_record = {
            "id": str(uuid.uuid4()),
            "contractor_id": current_user.id,
            "razorpay_subscription_id": razorpay_subscription_id,
            "razorpay_payment_id": razorpay_payment_id,
            "amount": subscription.get('amount', 0),
            "currency": "INR",
            "status": "paid",
            "payment_method": "razorpay_subscription",
            "plan_name": subscription.get('plan_name'),
            "duration_days": duration_days,
            "created_at": now.isoformat()
        }
        
        await db.payment_orders.insert_one(payment_record)
        
        # Log security event
        await log_security_event(
            event_type=SecurityEventType.SUBSCRIPTION_ACTIVATED,
            severity=SecuritySeverity.LOW,
            ip_address="",
            user_id=current_user.id,
            email=current_user.email,
            details={
                "action": "subscription_payment_verified",
                "subscription_id": razorpay_subscription_id,
                "plan": subscription.get('plan_name'),
                "auto_renew": True,
                "payment_id": razorpay_payment_id
            },
            endpoint="/payment/verify-subscription"
        )
        
        return {
            "success": True,
            "message": "Subscription activated successfully with auto-renewal",
            "subscription_end_date": new_end.isoformat(),
            "auto_renew": True,
            "next_billing_date": new_end.isoformat()
        }
        
    except razorpay.errors.SignatureVerificationError:
        # Update subscription as failed
        await db.razorpay_subscriptions.update_one(
            {"razorpay_subscription_id": razorpay_subscription_id},
            {"$set": {
                "status": "failed",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
        raise HTTPException(status_code=400, detail="Payment verification failed")
    
    except Exception as e:
        logging.error(f"Subscription verification error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Verification error: {str(e)}")

@api_router.get("/admin/payment-orders")
async def get_all_payment_orders(
    admin: AdminUser = Depends(get_current_admin),
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """Get all payment orders"""
    query = {}
    if status:
        query["status"] = status
    
    total = await db.payment_orders.count_documents(query)
    orders = await db.payment_orders.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with user data
    for order in orders:
        user = await db.users.find_one({"id": order['contractor_id']}, {"_id": 0, "name": 1, "email": 1})
        if user:
            order['contractor_name'] = user.get('name', 'Unknown')
            order['contractor_email'] = user.get('email', 'Unknown')
    
    return {
        "orders": orders,
        "total": total,
        "skip": skip,
        "limit": limit
    }

@api_router.get("/gateway/status")
async def check_gateway_status():
    """Check if payment gateway is configured - Production: checks environment variables"""
    key_id = os.getenv('RAZORPAY_KEY_ID')
    key_secret = os.getenv('RAZORPAY_KEY_SECRET')
    return {"configured": bool(key_id and key_secret)}

@api_router.post("/payment/webhook")
async def razorpay_webhook(request: Request):
    """Handle Razorpay webhook notifications"""
    
    # Get webhook signature from headers
    webhook_signature = request.headers.get('X-Razorpay-Signature')
    
    if not webhook_signature:
        raise HTTPException(status_code=400, detail="Missing webhook signature")
    
    # Get webhook body
    body = await request.body()
    
    try:
        # Get webhook secret from settings
        settings = await db.gateway_settings.find_one({"is_active": True})
        
        if not settings or not settings.get('webhook_secret'):
            logging.warning("Webhook received but no webhook secret configured")
            return {"status": "webhook_secret_not_configured"}
        
        webhook_secret = decrypt_data(settings['webhook_secret'])
        
        # Verify webhook signature
        client = await get_razorpay_client()
        client.utility.verify_webhook_signature(body.decode(), webhook_signature, webhook_secret)
        
        # Parse webhook data
        import json
        webhook_data = json.loads(body.decode())
        
        event = webhook_data.get('event')
        
        # Handle different webhook events
        if event == 'payment.captured':
            payload = webhook_data.get('payload', {}).get('payment', {}).get('entity', {})
            order_id = payload.get('order_id')
            payment_id = payload.get('id')
            
            # Update order in database
            order = await db.payment_orders.find_one({"razorpay_order_id": order_id})
            
            if order and order['status'] != 'paid':
                # Update order status
                await db.payment_orders.update_one(
                    {"razorpay_order_id": order_id},
                    {"$set": {
                        "status": "paid",
                        "payment_id": payment_id,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                
                # Activate subscription
                user = await db.users.find_one({"id": order['contractor_id']})
                if user:
                    current_end = user.get('subscription_end_date')
                    if current_end and isinstance(current_end, str):
                        current_end_dt = datetime.fromisoformat(current_end)
                    else:
                        current_end_dt = datetime.now(timezone.utc)
                    
                    if current_end_dt > datetime.now(timezone.utc):
                        new_end = current_end_dt + timedelta(days=order['duration_days'])
                    else:
                        new_end = datetime.now(timezone.utc) + timedelta(days=order['duration_days'])
                    
                    await db.users.update_one(
                        {"id": order['contractor_id']},
                        {"$set": {
                            "subscription_plan": order['plan_type'],
                            "subscription_status": "active",
                            "subscription_start_date": datetime.now(timezone.utc).isoformat(),
                            "subscription_end_date": new_end.isoformat()
                        }}
                    )
                
                logging.info(f"Webhook: Payment captured for order {order_id}")
        
        elif event == 'payment.failed':
            payload = webhook_data.get('payload', {}).get('payment', {}).get('entity', {})
            order_id = payload.get('order_id')
            
            if order_id:
                await db.payment_orders.update_one(
                    {"razorpay_order_id": order_id},
                    {"$set": {
                        "status": "failed",
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
                logging.info(f"Webhook: Payment failed for order {order_id}")
        
        # Handle subscription.charged event (auto-renewal payment)
        elif event == 'subscription.charged':
            payload = webhook_data.get('payload', {}).get('subscription', {}).get('entity', {})
            subscription_id = payload.get('id')
            payment_id = payload.get('payment_id')
            
            # Get subscription from database
            subscription = await db.razorpay_subscriptions.find_one({"razorpay_subscription_id": subscription_id})
            
            if subscription:
                contractor_id = subscription.get('contractor_id')
                duration_days = subscription.get('duration_days', 30)
                
                # Extend subscription
                user = await db.users.find_one({"id": contractor_id})
                if user:
                    current_end = user.get('subscription_end_date')
                    if current_end and isinstance(current_end, str):
                        current_end_dt = datetime.fromisoformat(current_end)
                    else:
                        current_end_dt = datetime.now(timezone.utc)
                    
                    # Extend from current end date
                    new_end = current_end_dt + timedelta(days=duration_days)
                    
                    await db.users.update_one(
                        {"id": contractor_id},
                        {"$set": {
                            "subscription_end_date": new_end.isoformat(),
                            "subscription_status": "active",
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }}
                    )
                    
                    # Record the renewal payment
                    payment_record = {
                        "id": str(uuid.uuid4()),
                        "contractor_id": contractor_id,
                        "razorpay_subscription_id": subscription_id,
                        "razorpay_payment_id": payment_id,
                        "amount": subscription.get('amount', 0),
                        "currency": "INR",
                        "status": "paid",
                        "payment_method": "razorpay_subscription",
                        "plan_name": subscription.get('plan_name'),
                        "duration_days": duration_days,
                        "is_renewal": True,
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    
                    await db.payment_orders.insert_one(payment_record)
                    
                    logging.info(f"Webhook: Subscription renewed for {subscription_id}, new end: {new_end.isoformat()}")
        
        # Handle subscription.cancelled event
        elif event == 'subscription.cancelled':
            payload = webhook_data.get('payload', {}).get('subscription', {}).get('entity', {})
            subscription_id = payload.get('id')
            
            # Update subscription in database
            await db.razorpay_subscriptions.update_one(
                {"razorpay_subscription_id": subscription_id},
                {"$set": {
                    "status": "cancelled",
                    "cancelled_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Update user auto_renew status
            await db.users.update_one(
                {"razorpay_subscription_id": subscription_id},
                {"$set": {
                    "auto_renew": False,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            logging.info(f"Webhook: Subscription cancelled {subscription_id}")
        
        # Handle subscription.paused event
        elif event == 'subscription.paused':
            payload = webhook_data.get('payload', {}).get('subscription', {}).get('entity', {})
            subscription_id = payload.get('id')
            
            await db.razorpay_subscriptions.update_one(
                {"razorpay_subscription_id": subscription_id},
                {"$set": {
                    "status": "paused",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            logging.info(f"Webhook: Subscription paused {subscription_id}")
        
        # Handle subscription.resumed event
        elif event == 'subscription.resumed':
            payload = webhook_data.get('payload', {}).get('subscription', {}).get('entity', {})
            subscription_id = payload.get('id')
            
            await db.razorpay_subscriptions.update_one(
                {"razorpay_subscription_id": subscription_id},
                {"$set": {
                    "status": "active",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            logging.info(f"Webhook: Subscription resumed {subscription_id}")
        
        # Handle subscription.completed event
        elif event == 'subscription.completed':
            payload = webhook_data.get('payload', {}).get('subscription', {}).get('entity', {})
            subscription_id = payload.get('id')
            
            await db.razorpay_subscriptions.update_one(
                {"razorpay_subscription_id": subscription_id},
                {"$set": {
                    "status": "completed",
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            logging.info(f"Webhook: Subscription completed {subscription_id}")
        
        return {"status": "success"}
        
    except Exception as e:
        logging.error(f"Webhook error: {str(e)}")
        raise HTTPException(status_code=400, detail="Webhook verification failed")

# ============ SUBSCRIPTION MANAGEMENT ENDPOINTS ============

@api_router.get("/subscription/transactions")
async def get_my_transactions(current_user: User = Depends(get_current_user)):
    """Get user's transaction history including payments and activation keys"""
    
    transactions = []
    
    # Get all payment orders (includes both Razorpay payments and activation key redemptions)
    payment_orders = await db.payment_orders.find({
        "contractor_id": current_user.id
    }).sort("created_at", -1).to_list(100)
    
    for order in payment_orders:
        payment_method = order.get("payment_method", "razorpay")
        
        if payment_method == "activation_key":
            # Activation key redemption
            is_paid = order.get("is_paid", False)
            amount = order.get("amount", 0) if is_paid else 0
            
            transactions.append({
                "id": order.get("id"),
                "description": f"Activated with Key - {order.get('plan_name', 'Plan')}" + (" (Paid)" if is_paid else ""),
                "amount": amount,
                "payment_method": "activation_key",
                "payment_method_label": "Activation Key (Paid)" if is_paid else "Activation Key",
                "status": "success",
                "created_at": order.get("created_at"),
                "is_paid": is_paid
            })
        else:
            # Razorpay payment
            amount = order.get("amount", 0)
            # Handle both paise (if > 10000 assume it's in paise) and rupees
            if amount > 10000:
                amount = amount / 100  # Convert paise to rupees
            transactions.append({
                "id": order.get("id"),
                "description": f"Subscription Payment - {order.get('plan_name', 'Monthly Plan')}",
                "amount": amount,
                "payment_method": "razorpay",
                "payment_method_label": "Razorpay",
                "status": order.get("status", "pending"),
                "created_at": order.get("created_at")
            })
    
    # Include trial activation, if any
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if user_doc and user_doc.get("trial_activated_at"):
        created = user_doc.get("trial_activated_at")
        if isinstance(created, datetime):
            created = created.isoformat()
        elif isinstance(created, str):
            # Already a string, use as is
            pass
        else:
            created = datetime.now(timezone.utc).isoformat()
        transactions.append({
            "id": f"trial_{user_doc.get('id')}",
            "description": "Trial Activated - Contractor Plus Trial",
            "amount": 0,
            "payment_method": "trial",
            "payment_method_label": "Free Trial",
            "status": "success",
            "created_at": created
        })
    
    # Get extension key usage
    extension_keys = await db.extension_key_usage.find({
        "user_id": current_user.id
    }).sort("applied_at", -1).to_list(100)
    
    for ext in extension_keys:
        transactions.append({
            "id": f"ext_{ext.get('id')}",
            "description": f"Validity Extended - {ext.get('days_extended', 30)} days",
            "amount": 0,
            "payment_method": "extension_key",
            "payment_method_label": "Extension Key",
            "status": "success",
            "created_at": ext.get("applied_at")
        })
    
    # Sort by date (most recent first)
    transactions.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    
    return transactions

@api_router.get("/subscription/trial-status")
async def get_trial_status(current_user: User = Depends(get_current_user)):
    """
    Check trial status and whether payment prompt should be shown
    Returns trial information and days remaining
    """
    # Check if user is on trial
    if current_user.payment_method != "trial":
        return {
            "is_trial": False,
            "show_payment_prompt": False
        }
    
    # Get user document for additional fields
    user_doc = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_doc:
        return {"is_trial": False, "show_payment_prompt": False}
    
    # Calculate days remaining
    if not current_user.plan_end_date:
        return {"is_trial": True, "show_payment_prompt": False, "days_remaining": 0}
    
    try:
        end_date = datetime.fromisoformat(current_user.plan_end_date)
        now = datetime.now(timezone.utc)
        days_remaining = (end_date - now).days
        
        # Show prompt if 2 days or less remaining and not dismissed
        show_prompt = (
            days_remaining <= 2 and 
            days_remaining >= 0 and 
            not user_doc.get("trial_payment_prompt_dismissed", False) and
            not user_doc.get("trial_payment_setup", False)
        )
        
        return {
            "is_trial": True,
            "days_remaining": max(0, days_remaining),
            "show_payment_prompt": show_prompt,
            "trial_end_date": current_user.plan_end_date,
            "payment_setup": user_doc.get("trial_payment_setup", False)
        }
    except Exception as e:
        print(f"Error calculating trial status: {e}")
        return {"is_trial": True, "show_payment_prompt": False, "days_remaining": 0}

@api_router.post("/subscription/dismiss-trial-prompt")
async def dismiss_trial_prompt(current_user: User = Depends(get_current_user)):
    """
    User dismissed the trial payment prompt
    """
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"trial_payment_prompt_dismissed": True}}
    )
    return {"message": "Prompt dismissed"}

@api_router.post("/subscription/setup-trial-payment")
async def setup_trial_payment(
    plan_data: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Setup Razorpay subscription to start after trial ends
    Creates a subscription that will auto-charge when trial expires
    """
    if current_user.payment_method != "trial":
        raise HTTPException(status_code=400, detail="Not on trial")
    
    # Get plan details
    plan_id = plan_data.get("plan_id")
    if not plan_id:
        raise HTTPException(status_code=400, detail="Plan ID required")
    
    plan = await db.subscription_plans.find_one({"id": plan_id})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Check if plan has Razorpay plan ID
    razorpay_plan_id = plan.get("razorpay_plan_id")
    if not razorpay_plan_id:
        raise HTTPException(
            status_code=400, 
            detail="This plan doesn't support auto-renewal. Please use one-time payment instead."
        )
    
    # Calculate start date (when trial ends)
    try:
        trial_end = datetime.fromisoformat(current_user.plan_end_date)
    except:
        raise HTTPException(status_code=400, detail="Invalid trial end date")
    
    # Create Razorpay subscription that starts after trial
    try:
        subscription = razorpay_client.subscription.create({
            "plan_id": razorpay_plan_id,
            "customer_notify": 1,
            "quantity": 1,
            "start_at": int(trial_end.timestamp()),  # Start when trial ends
            "total_count": 12,  # 12 months
            "notes": {
                "user_id": current_user.id,
                "email": current_user.email,
                "name": current_user.name,
                "conversion_type": "trial_to_paid"
            }
        })
        
        # Update user with subscription details
        await db.users.update_one(
            {"id": current_user.id},
            {"$set": {
                "razorpay_subscription_id": subscription["id"],
                "auto_renew": True,
                "trial_payment_setup": True,
                "trial_conversion_plan": plan.get("name"),
                "trial_payment_prompt_dismissed": True  # Don't show prompt anymore
            }}
        )
        
        return {
            "message": "Payment method added successfully! Your subscription will start automatically when trial ends.",
            "subscription_id": subscription["id"],
            "starts_at": trial_end.isoformat(),
            "plan_name": plan.get("name"),
            "amount": plan.get("price")
        }
        
    except Exception as e:
        print(f"Error creating Razorpay subscription: {e}")
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to setup payment method: {str(e)}"
        )

@api_router.get("/subscription/limits")
async def get_subscription_limits(
    current_user: User = Depends(get_current_user)
):
    """
    Get current subscription plan limits and usage.
    Returns: current usage counts and plan limits
    """
    if current_user.role == "admin":
        return {
            "plan_name": "Admin",
            "current_workers": 0,
            "current_employers": 0,
            "max_workers": None,  # Unlimited
            "max_employers": None,  # Unlimited
            "worker_limit_reached": False,
            "employer_limit_reached": False
        }
    
    # Get plan limits
    limits = await get_plan_limits(current_user.subscription_plan)
    max_workers = limits.get("max_workers")
    max_employers = limits.get("max_employers")
    
    # Get current usage
    worker_count = await db.workers.count_documents({"contractor_id": current_user.id})
    employer_count = await db.employers.count_documents({"contractor_id": current_user.id})
    
    return {
        "plan_name": current_user.subscription_plan,
        "current_workers": worker_count,
        "current_employers": employer_count,
        "max_workers": max_workers,  # None = unlimited
        "max_employers": max_employers,  # None = unlimited
        "worker_limit_reached": max_workers is not None and worker_count >= max_workers,
        "employer_limit_reached": max_employers is not None and employer_count >= max_employers
    }

@api_router.post("/subscription/apply-extension-key")
async def apply_extension_key(
    key_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Apply a validity extension key to user's subscription"""
    
    extension_key = key_data.get("key", "").strip()
    
    if not extension_key:
        raise HTTPException(status_code=400, detail="Extension key is required")
    
    # Find the key in database
    key_doc = await db.extension_keys.find_one({
        "key": extension_key,
        "is_active": True
    })
    
    if not key_doc:
        raise HTTPException(status_code=404, detail="Invalid or expired extension key")
    
    # Check if key has been used
    if key_doc.get("used", False):
        raise HTTPException(status_code=400, detail="This key has already been used")
    
    # Check expiry date
    if key_doc.get("expiry_date"):
        expiry_date = datetime.fromisoformat(key_doc["expiry_date"])
        if datetime.now(timezone.utc) > expiry_date:
            raise HTTPException(status_code=400, detail="This key has expired")
    
    # SECURITY: Check if extension key plan matches user's current plan
    key_plan = key_doc.get("plan")  # Can be None for all-plans keys
    user_plan = (await db.users.find_one({"id": current_user.id})).get("subscription_plan")
    
    if key_plan and user_plan and key_plan != user_plan:
        raise HTTPException(
            status_code=400,
            detail=f"This extension key is for '{key_plan}' plan, but your current plan is '{user_plan}'. Extension keys can only be used for the same plan or for all-plans keys."
        )
    
    # Get days to extend
    days_to_extend = key_doc.get("duration_days", 30)
    
    # Get current user's subscription end date
    user_dict = await db.users.find_one({"id": current_user.id})
    current_end = user_dict.get("plan_end_date")
    
    # Calculate new end date
    if current_end:
        try:
            end_dt = datetime.fromisoformat(current_end)
        except:
            end_dt = datetime.now(timezone.utc)
    else:
        end_dt = datetime.now(timezone.utc)
    
    # Extend from current end date or now (whichever is later)
    if end_dt < datetime.now(timezone.utc):
        new_end = datetime.now(timezone.utc) + timedelta(days=days_to_extend)
    else:
        new_end = end_dt + timedelta(days=days_to_extend)
    
    # Update user's subscription
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {
            "plan_end_date": new_end.isoformat(),
            "subscription_status": "active",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Mark key as used
    await db.extension_keys.update_one(
        {"key": extension_key},
        {"$set": {
            "used": True,
            "used_by": current_user.id,
            "used_by_email": current_user.email,
            "used_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Record extension key usage
    usage_record = {
        "id": str(uuid.uuid4()),
        "user_id": current_user.id,
        "user_email": current_user.email,
        "extension_key": extension_key,
        "days_extended": days_to_extend,
        "previous_end_date": current_end,
        "new_end_date": new_end.isoformat(),
        "applied_at": datetime.now(timezone.utc).isoformat()
    }
    await db.extension_key_usage.insert_one(usage_record)
    
    return {
        "message": f"Successfully extended subscription by {days_to_extend} days",
        "new_end_date": new_end.isoformat(),
        "days_extended": days_to_extend
    }

@api_router.post("/subscription/cancel")
async def cancel_subscription(current_user: User = Depends(get_current_user)):
    """Cancel user's subscription (will remain active until end date, auto-renewal disabled for Razorpay)"""
    
    # Check if subscription is active or cancelled
    user_dict = await db.users.find_one({"id": current_user.id})
    if not user_dict:
        raise HTTPException(status_code=404, detail="User not found")
    
    current_status = user_dict.get("subscription_status")
    
    if current_status == "cancelled":
        return {
            "message": "Subscription is already cancelled.",
            "access_until": user_dict.get("plan_end_date"),
            "auto_renew": False
        }
    
    if current_status != "active":
        raise HTTPException(
            status_code=400,
            detail="No active subscription to cancel"
        )
    
    # If user has a Razorpay subscription, cancel it on Razorpay's side
    razorpay_subscription_id = user_dict.get("razorpay_subscription_id")
    
    if razorpay_subscription_id:
        try:
            client = await get_razorpay_client()
            
            # Cancel subscription on Razorpay (user retains access until end of billing period)
            client.subscription.cancel(razorpay_subscription_id, {
                "cancel_at_cycle_end": 1  # Cancel at end of current billing cycle
            })
            
            # Update subscription status in database
            await db.razorpay_subscriptions.update_one(
                {"razorpay_subscription_id": razorpay_subscription_id},
                {"$set": {
                    "status": "cancelled",
                    "cancelled_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            logging.info(f"Razorpay subscription {razorpay_subscription_id} cancelled for user {current_user.id}")
            
        except Exception as e:
            logging.error(f"Failed to cancel Razorpay subscription: {str(e)}")
            # Continue with local cancellation even if Razorpay API fails
    
    # Update subscription status to cancelled (will still work until end date)
    # For Razorpay subscriptions, auto-renewal will be disabled
    update_fields = {
        "subscription_status": "cancelled",
        "cancellation_date": datetime.now(timezone.utc).isoformat(),
        "auto_renew": False,  # Disable auto-renewal
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": update_fields}
    )
    
    # Log security event
    await log_security_event(
        event_type=SecurityEventType.SUBSCRIPTION_CANCELLED,
        severity=SecuritySeverity.LOW,
        ip_address="",
        user_id=current_user.id,
        email=current_user.email,
        details={
            "action": "subscription_cancelled",
            "razorpay_subscription_id": razorpay_subscription_id,
            "access_until": user_dict.get("plan_end_date")
        },
        endpoint="/subscription/cancel"
    )
    
    return {
        "message": "Subscription cancelled successfully. You will retain access until your current period ends.",
        "access_until": user_dict.get("plan_end_date"),
        "auto_renew": False
    }

@api_router.post("/subscription/restart")
async def restart_subscription(current_user: User = Depends(get_current_user)):
    """Restart a cancelled subscription (re-enable auto-renewal for Razorpay)"""
    
    user_dict = await db.users.find_one({"id": current_user.id})
    if not user_dict:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Only allow restart if subscription is cancelled
    if user_dict.get("subscription_status") != "cancelled":
        raise HTTPException(
            status_code=400, 
            detail="Subscription is not cancelled. Cannot restart."
        )
    
    # Check if still within billing period
    plan_end_date = user_dict.get("plan_end_date")
    if plan_end_date:
        if isinstance(plan_end_date, str):
            end_date = datetime.fromisoformat(plan_end_date)
        else:
            end_date = plan_end_date
        if getattr(end_date, 'tzinfo', None) is None:
            end_date = end_date.replace(tzinfo=timezone.utc)
        
        if end_date < datetime.now(timezone.utc):
            raise HTTPException(
                status_code=400,
                detail="Subscription has expired. Please purchase a new subscription."
            )
    
    # Re-enable subscription and auto-renewal (for Razorpay subscriptions)
    payment_method = user_dict.get("payment_method", "")
    update_fields = {
        "subscription_status": "active",
        "cancellation_date": None,  # Clear cancellation date
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Only enable auto-renewal for Razorpay subscriptions
    if payment_method == "razorpay":
        update_fields["auto_renew"] = True
    
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": update_fields}
    )
    
    message = "Subscription restarted successfully."
    if payment_method == "razorpay":
        message += " Auto-renewal has been enabled."
    
    return {
        "message": message,
        "subscription_status": "active",
        "access_until": plan_end_date
    }

@api_router.post("/subscription/apply-promo")
async def apply_promo_code(
    promo_data: dict,
    current_user: User = Depends(get_current_user)
):
    """Apply a promo/discount code and return discount details"""
    
    promo_code = promo_data.get("code", "").strip().upper()
    
    if not promo_code:
        raise HTTPException(status_code=400, detail="Promo code is required")
    
    # Find the promo code in database
    promo_doc = await db.promo_codes.find_one({
        "code": promo_code,
        "is_active": True
    })
    
    if not promo_doc:
        raise HTTPException(status_code=404, detail="Invalid or expired promo code")
    
    # Check expiry date
    if promo_doc.get("expiry_date"):
        expiry_date = datetime.fromisoformat(promo_doc["expiry_date"])
        if datetime.now(timezone.utc) > expiry_date:
            raise HTTPException(status_code=400, detail="This promo code has expired")
    
    # Check usage limit
    max_uses = promo_doc.get("max_uses", 0)
    current_uses = promo_doc.get("times_used", 0)
    
    if max_uses > 0 and current_uses >= max_uses:
        raise HTTPException(status_code=400, detail="This promo code has reached its usage limit")
    
    # Check if user has already used this code
    used_by = promo_doc.get("used_by", [])
    if current_user.id in used_by:
        raise HTTPException(status_code=400, detail="You have already used this promo code")
    
    discount_type = promo_doc.get("discount_type", "percentage")  # percentage or fixed
    discount_value = promo_doc.get("discount_value", 0)
    
    return {
        "valid": True,
        "code": promo_code,
        "discount_type": discount_type,
        "discount_value": discount_value,
        "description": promo_doc.get("description", "Discount applied")
    }


# ============ NETFLIX-STYLE PLAN CHANGE SYSTEM ============

class PlanChangeRequest(BaseModel):
    plan_id: str

class PlanChangeTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    old_plan: str
    new_plan: str
    old_plan_price: float
    new_plan_price: float
    days_remaining: int
    unused_credit: float
    amount_to_pay: float
    prorated_days: int
    old_end_date: str
    new_end_date: str
    transaction_type: str  # "upgrade" or "downgrade"
    status: str  # "completed", "failed"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.post("/subscription/change-plan")
async def change_subscription_plan(
    request: PlanChangeRequest,
    current_user: User = Depends(get_current_user)
):
    """
    Netflix-style plan change with prorated billing.
    
    UPGRADE: User pays difference immediately, gets extended time based on credit.
    DOWNGRADE: User gets credit applied to new plan, extends subscription.
    
    Security Features:
    - Rate limiting: Max 1 plan change per 24 hours
    - Validates plan exists and is active
    - Prevents same-plan changes
    - Calculates exact prorated amounts
    - Logs all transactions for audit trail
    - Prevents exploitation through timing attacks
    """
    
    # ===== SECURITY: Rate Limiting - Prevent rapid plan switching =====
    # Check if user changed plan in last 24 hours
    last_change = await db.plan_change_transactions.find_one(
        {
            "user_id": current_user.id,
            "status": "completed",
            "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()}
        },
        {"_id": 0}
    )
    
    if last_change:
        raise HTTPException(
            status_code=429,
            detail="You can only change your plan once every 24 hours. Please try again later."
        )
    
    # ===== Fetch fresh user data from database =====
    user_dict = await db.users.find_one({"id": current_user.id}, {"_id": 0})
    if not user_dict:
        raise HTTPException(status_code=404, detail="User not found")
    
    current_plan_name = user_dict.get("subscription_plan", "none")
    current_status = user_dict.get("subscription_status", "inactive")
    
    # ===== Validate user has an active subscription =====
    if current_status not in ["active", "cancelled"]:
        raise HTTPException(
            status_code=400,
            detail="You must have an active subscription to change plans. Please purchase a plan first."
        )
    
    # ===== Fetch and validate new plan =====
    new_plan = await db.subscription_plans.find_one({"id": request.plan_id}, {"_id": 0})
    if not new_plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    if not new_plan.get("is_active", True):
        raise HTTPException(status_code=400, detail="This plan is no longer available")
    
    new_plan_name = new_plan.get("name")
    new_plan_price = float(new_plan.get("price", 0))
    new_plan_duration = int(new_plan.get("duration_days", 30))
    
    # ===== Prevent same-plan changes =====
    if current_plan_name == new_plan_name:
        raise HTTPException(
            status_code=400,
            detail="You are already on this plan. No changes needed."
        )
    
    # ===== Check if user is on trial =====
    is_trial = user_dict.get("payment_method") == "trial"
    
    # ===== Fetch current plan details =====
    current_plan = await db.subscription_plans.find_one({"name": current_plan_name}, {"_id": 0})
    if not current_plan:
        # Fallback for legacy plans not in database (or trial plans)
        if is_trial:
            # Trial plan details
            current_plan_price = 0.0
            current_plan_duration = user_dict.get("trial_duration_days", 14)
        else:
            # Legacy paid plan
            current_plan_price = 0.0
            current_plan_duration = 30
    else:
        current_plan_price = float(current_plan.get("price", 0))
        current_plan_duration = int(current_plan.get("duration_days", 30))
    
    # ===== Calculate days remaining on current plan =====
    plan_end_date_str = user_dict.get("plan_end_date") or user_dict.get("subscription_end_date")
    if not plan_end_date_str:
        raise HTTPException(
            status_code=400,
            detail="No active subscription end date found. Please contact support."
        )
    
    try:
        plan_end_date = datetime.fromisoformat(plan_end_date_str)
        if getattr(plan_end_date, 'tzinfo', None) is None:
            plan_end_date = plan_end_date.replace(tzinfo=timezone.utc)
    except:
        raise HTTPException(status_code=400, detail="Invalid subscription end date format")
    
    now = datetime.now(timezone.utc)
    
    # Check if plan already expired
    if plan_end_date <= now:
        raise HTTPException(
            status_code=400,
            detail="Your current plan has expired. Please purchase a new subscription instead of changing plans."
        )
    
    # Calculate exact days remaining (including fractional days for precision)
    time_remaining = plan_end_date - now
    days_remaining_exact = time_remaining.total_seconds() / (24 * 60 * 60)
    days_remaining = max(0, int(days_remaining_exact))
    
    # ===== SPECIAL HANDLING FOR TRIAL USERS =====
    if is_trial:
        # Trial users converting to paid plan
        # No credit given (trial was free)
        # Pay full price for new plan
        # Get full duration of new plan
        
        unused_credit = 0.0
        amount_to_pay = new_plan_price
        prorated_days = new_plan_duration
        new_end_date = now + timedelta(days=prorated_days)
        transaction_type = "trial_conversion"
        
        message = f"Trial converted to {new_plan_name}! Welcome to your paid subscription. You now have {prorated_days} days of full access."
        
    else:
        # ===== PRORATED BILLING CALCULATION (Netflix-style) =====
        
        # Calculate daily rate for both plans
        current_daily_rate = current_plan_price / current_plan_duration if current_plan_duration > 0 else 0
        new_daily_rate = new_plan_price / new_plan_duration if new_plan_duration > 0 else 0
        
        # Calculate unused credit from current plan
        unused_credit = current_daily_rate * days_remaining_exact
        
        # Determine if upgrade or downgrade
        is_upgrade = new_plan_price > current_plan_price
        transaction_type = "upgrade" if is_upgrade else "downgrade"
        
        if is_upgrade:
            # ===== UPGRADE: User pays difference, gets immediate access =====
            # Calculate how much of new plan the credit covers
            credit_days = unused_credit / new_daily_rate if new_daily_rate > 0 else 0
            
            # User needs to pay for full new plan period
            amount_to_pay = new_plan_price
            
            # New end date = now + new plan duration + credit days
            prorated_days = new_plan_duration + int(credit_days)
            new_end_date = now + timedelta(days=prorated_days)
            
            message = f"Plan upgraded to {new_plan_name}. Your unused credit of ₹{unused_credit:.2f} has been applied, extending your subscription by {int(credit_days)} extra days."
            
        else:
            # ===== DOWNGRADE: Credit applied to new plan, extends subscription =====
            # Calculate how many days the credit buys on new plan
            credit_days = unused_credit / new_daily_rate if new_daily_rate > 0 else 0
            
            # No immediate payment needed - credit covers it
            amount_to_pay = 0.0
            
            # New end date = now + credit days
            prorated_days = int(credit_days)
            new_end_date = now + timedelta(days=prorated_days)
            
            message = f"Plan changed to {new_plan_name}. Your unused credit of ₹{unused_credit:.2f} has been applied, giving you {prorated_days} days of service."
    
    # ===== SECURITY: Prevent exploitation - Minimum subscription period =====
    # Ensure new subscription is at least 1 day
    if prorated_days < 1:
        raise HTTPException(
            status_code=400,
            detail="Insufficient credit to change to this plan. Please purchase the new plan instead."
        )
    
    # ===== PAYMENT REQUIRED CHECK =====
    # If amount_to_pay > 0, user needs to complete payment first
    if amount_to_pay > 0:
        # Check if Razorpay is configured
        gateway_settings = await db.gateway_settings.find_one({"is_active": True})
        if not gateway_settings:
            raise HTTPException(
                status_code=400,
                detail="Payment gateway not configured. Please contact support or use an activation key."
            )
        
        # Create pending transaction record
        transaction = PlanChangeTransaction(
            user_id=current_user.id,
            old_plan=current_plan_name,
            new_plan=new_plan_name,
            old_plan_price=current_plan_price,
            new_plan_price=new_plan_price,
            days_remaining=days_remaining,
            unused_credit=round(unused_credit, 2),
            amount_to_pay=round(amount_to_pay, 2),
            prorated_days=prorated_days,
            old_end_date=plan_end_date.isoformat(),
            new_end_date=new_end_date.isoformat(),
            transaction_type=transaction_type,
            status="pending_payment"  # Will be updated after payment
        )
        
        transaction_dict = transaction.model_dump()
        transaction_dict['created_at'] = transaction_dict['created_at'].isoformat()
        
        # Save pending transaction
        await db.plan_change_transactions.insert_one(transaction_dict)
        
        # Create Razorpay order
        try:
            client = await get_razorpay_client()
            
            razorpay_order = client.order.create({
                "amount": int(amount_to_pay * 100),  # Convert to paise
                "currency": "INR",
                "payment_capture": 1,
                "notes": {
                    "contractor_id": current_user.id,
                    "transaction_id": transaction.id,
                    "transaction_type": transaction_type,
                    "old_plan": current_plan_name,
                    "new_plan": new_plan_name
                }
            })
            
            # Save payment order
            payment_order = PaymentOrder(
                contractor_id=current_user.id,
                razorpay_order_id=razorpay_order['id'],
                amount=amount_to_pay,
                currency="INR",
                status="created",
                plan_type=new_plan_name,
                duration_days=prorated_days
            )
            
            order_dict = payment_order.model_dump()
            order_dict['created_at'] = order_dict['created_at'].isoformat()
            order_dict['updated_at'] = order_dict['updated_at'].isoformat()
            order_dict['plan_change_transaction_id'] = transaction.id
            
            await db.payment_orders.insert_one(order_dict)
            
            # Return payment required response
            return {
                "success": False,
                "payment_required": True,
                "message": f"Payment of ₹{amount_to_pay:.2f} required to complete plan change",
                "razorpay_order_id": razorpay_order['id'],
                "amount": amount_to_pay,
                "currency": "INR",
                "transaction": {
                    "id": transaction.id,
                    "type": transaction_type,
                    "old_plan": current_plan_name,
                    "new_plan": new_plan_name,
                    "unused_credit": round(unused_credit, 2),
                    "amount_to_pay": round(amount_to_pay, 2),
                    "prorated_days": prorated_days,
                    "new_end_date": new_end_date.isoformat(),
                    "old_end_date": plan_end_date.isoformat()
                }
            }
            
        except Exception as e:
            # Mark transaction as failed
            await db.plan_change_transactions.update_one(
                {"id": transaction.id},
                {"$set": {"status": "failed"}}
            )
            raise HTTPException(status_code=500, detail=f"Failed to create payment order: {str(e)}")
    
    # ===== NO PAYMENT REQUIRED (Downgrade or zero amount) =====
    # Create transaction record for audit trail
    transaction = PlanChangeTransaction(
        user_id=current_user.id,
        old_plan=current_plan_name,
        new_plan=new_plan_name,
        old_plan_price=current_plan_price,
        new_plan_price=new_plan_price,
        days_remaining=days_remaining,
        unused_credit=round(unused_credit, 2),
        amount_to_pay=0.0,
        prorated_days=prorated_days,
        old_end_date=plan_end_date.isoformat(),
        new_end_date=new_end_date.isoformat(),
        transaction_type=transaction_type,
        status="completed"
    )
    
    transaction_dict = transaction.model_dump()
    transaction_dict['created_at'] = transaction_dict['created_at'].isoformat()
    
    # ===== Update user subscription =====
    update_fields = {
        "subscription_plan": new_plan_name,
        "subscription_status": "active",
        "plan_end_date": new_end_date.isoformat(),
        "subscription_end_date": new_end_date.isoformat(),
        "updated_at": now.isoformat()
    }
    
    # If was cancelled, re-enable auto-renewal for Razorpay
    if current_status == "cancelled" and user_dict.get("payment_method") == "razorpay":
        update_fields["auto_renew"] = True
        update_fields["cancellation_date"] = None
    
    # ===== Execute database updates atomically =====
    try:
        # Insert transaction record
        await db.plan_change_transactions.insert_one(transaction_dict)
        
        # Update user subscription
        result = await db.users.update_one(
            {"id": current_user.id},
            {"$set": update_fields}
        )
        
        if result.modified_count == 0:
            # Rollback transaction record
            await db.plan_change_transactions.update_one(
                {"id": transaction.id},
                {"$set": {"status": "failed"}}
            )
            raise HTTPException(status_code=500, detail="Failed to update subscription")
        
        # Log security event for audit trail
        await log_security_event(
            event_type=SecurityEventType.PLAN_CHANGE,
            severity=SecuritySeverity.LOW,
            ip_address="",
            user_id=current_user.id,
            email=user_dict.get("email"),
            details={
                "action": "plan_change",
                "transaction_type": transaction_type,
                "old_plan": current_plan_name,
                "new_plan": new_plan_name,
                "unused_credit": round(unused_credit, 2),
                "prorated_days": prorated_days,
                "transaction_id": transaction.id
            },
            endpoint="/subscription/change-plan"
        )
        
    except Exception as e:
        # Mark transaction as failed
        await db.plan_change_transactions.update_one(
            {"id": transaction.id},
            {"$set": {"status": "failed"}}
        )
        raise HTTPException(status_code=500, detail=f"Plan change failed: {str(e)}")
    
    # ===== Return detailed response =====
    return {
        "success": True,
        "payment_required": False,
        "message": message,
        "transaction": {
            "id": transaction.id,
            "type": transaction_type,
            "old_plan": current_plan_name,
            "new_plan": new_plan_name,
            "unused_credit": round(unused_credit, 2),
            "amount_to_pay": 0.0,
            "prorated_days": prorated_days,
            "new_end_date": new_end_date.isoformat(),
            "old_end_date": plan_end_date.isoformat()
        },
        "subscription": {
            "plan": new_plan_name,
            "status": "active",
            "end_date": new_end_date.isoformat(),
            "days_remaining": prorated_days
        }
    }


# ============ BOOKINGS ROUTES ============

class BookingCreate(BaseModel):
    employer_id: str
    date: str  # DD-MM-YYYY format
    workers_count: int
    selected_workers: List[str] = []

class BookingUpdate(BaseModel):
    employer_id: Optional[str] = None
    date: Optional[str] = None
    workers_count: Optional[int] = None
    selected_workers: Optional[List[str]] = None

class Booking(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    contractor_id: str
    employer_id: str
    employer_name: Optional[str] = None
    employer_phone: Optional[str] = None
    date: str  # DD-MM-YYYY format
    workers_count: int
    selected_workers: List[str] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

@api_router.post("/bookings", response_model=Booking)
async def create_booking(
    booking_data: BookingCreate,
    current_user: User = Depends(get_active_subscription_user)
):
    """Create a new booking"""
    # Verify employer exists and belongs to contractor
    employer = await db.employers.find_one({
        "id": booking_data.employer_id,
        "contractor_id": current_user.id
    })
    if not employer:
        raise HTTPException(status_code=404, detail="Employer not found")
    
    # Create booking
    booking = Booking(
        contractor_id=current_user.id,
        employer_id=booking_data.employer_id,
        employer_name=employer.get("name"),
        employer_phone=employer.get("phone_number"),
        date=booking_data.date,
        workers_count=booking_data.workers_count,
        selected_workers=booking_data.selected_workers
    )
    
    booking_dict = booking.model_dump()
    booking_dict['created_at'] = booking_dict['created_at'].isoformat()
    booking_dict['updated_at'] = booking_dict['updated_at'].isoformat()
    
    await db.bookings.insert_one(booking_dict)
    
    return booking

@api_router.get("/bookings")
async def get_bookings(
    date: Optional[str] = None,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get all bookings for the contractor"""
    query = {"contractor_id": current_user.id}
    if date:
        query["date"] = date
    
    bookings = await db.bookings.find(query, {"_id": 0}).to_list(1000)
    
    # Enrich with employer details
    for booking in bookings:
        if not booking.get("employer_name"):
            employer = await db.employers.find_one({"id": booking["employer_id"]}, {"_id": 0})
            if employer:
                booking["employer_name"] = employer.get("name")
                booking["employer_phone"] = employer.get("phone_number")
    
    return bookings

@api_router.get("/bookings/availability")
async def get_booking_availability(
    date: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """Get worker availability for a specific date"""
    # Get all active workers
    workers = await db.workers.find({
        "contractor_id": current_user.id,
        "status": "Active"
    }, {"_id": 0}).to_list(1000)
    
    total_workers = len(workers)
    
    # Get bookings for this date
    bookings = await db.bookings.find({
        "contractor_id": current_user.id,
        "date": date
    }, {"_id": 0}).to_list(1000)
    
    # Enrich bookings with employer details
    for booking in bookings:
        if not booking.get("employer_name"):
            employer = await db.employers.find_one({"id": booking["employer_id"]}, {"_id": 0})
            if employer:
                booking["employer_name"] = employer.get("name")
                booking["employer_phone"] = employer.get("phone_number")
    
    # Calculate booked workers
    booked_worker_ids = set()
    total_booked = 0
    
    for booking in bookings:
        if booking.get("selected_workers"):
            booked_worker_ids.update(booking["selected_workers"])
        else:
            # If no specific workers selected, count as generic booking
            total_booked += booking.get("workers_count", 0)
    
    # Available count = total - (specific booked + generic booked)
    available_count = total_workers - len(booked_worker_ids) - total_booked
    available_count = max(0, available_count)
    
    return {
        "date": date,
        "total_workers": total_workers,
        "available_count": available_count,
        "booked_count": len(booked_worker_ids) + total_booked,
        "booked_worker_ids": list(booked_worker_ids),
        "bookings": bookings
    }

@api_router.put("/bookings/{booking_id}", response_model=Booking)
async def update_booking(
    booking_id: str,
    booking_data: BookingUpdate,
    current_user: User = Depends(get_active_subscription_user)
):
    """Update a booking"""
    booking = await db.bookings.find_one({
        "id": booking_id,
        "contractor_id": current_user.id
    })
    if not booking:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if booking_data.employer_id:
        employer = await db.employers.find_one({
            "id": booking_data.employer_id,
            "contractor_id": current_user.id
        })
        if not employer:
            raise HTTPException(status_code=404, detail="Employer not found")
        update_fields["employer_id"] = booking_data.employer_id
        update_fields["employer_name"] = employer.get("name")
        update_fields["employer_phone"] = employer.get("phone_number")
    
    if booking_data.date:
        update_fields["date"] = booking_data.date
    if booking_data.workers_count is not None:
        update_fields["workers_count"] = booking_data.workers_count
    if booking_data.selected_workers is not None:
        update_fields["selected_workers"] = booking_data.selected_workers
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": update_fields}
    )
    
    updated_booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    return Booking(**updated_booking)

@api_router.delete("/bookings/{booking_id}")
async def delete_booking(
    booking_id: str,
    current_user: User = Depends(get_active_subscription_user)
):
    """Delete a booking"""
    result = await db.bookings.delete_one({
        "id": booking_id,
        "contractor_id": current_user.id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Booking not found")
    
    return {"message": "Booking deleted successfully"}


# ============ PLATFORM REVENUE TRACKING ============

@api_router.get("/admin/platform-revenue")
async def get_platform_revenue(
    admin: AdminUser = Depends(get_current_admin),
    payment_method: Optional[str] = None,
    skip: int = 0,
    limit: int = 50
):
    """
    Get platform revenue from all sources:
    - Razorpay one-time payments
    - Razorpay subscriptions
    - Paid activation keys (when is_paid=True)
    """
    
    # Build query
    query = {"status": "success"}
    
    # Filter by payment method if specified
    if payment_method:
        query["payment_method"] = payment_method
    
    # Get all successful payment orders
    payment_orders = await db.payment_orders.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with user details
    revenue_items = []
    for order in payment_orders:
        contractor_id = order.get("contractor_id")
        user = await db.users.find_one({"id": contractor_id}, {"_id": 0, "name": 1, "email": 1})
        
        # Determine payment method label
        pm = order.get("payment_method", "razorpay")
        if pm == "activation_key":
            pm_label = "Activation Key (Paid)"
        elif pm == "razorpay_subscription":
            pm_label = "Razorpay (Recurring)"
        else:
            pm_label = "Razorpay"
        
        revenue_items.append({
            "id": order.get("id"),
            "contractor_id": contractor_id,
            "contractor_name": user.get("name") if user else "Unknown",
            "contractor_email": user.get("email") if user else "Unknown",
            "amount": order.get("amount", 0),
            "payment_method": pm,
            "payment_method_label": pm_label,
            "plan_name": order.get("plan_name"),
            "razorpay_payment_id": order.get("razorpay_payment_id"),
            "razorpay_order_id": order.get("razorpay_order_id"),
            "razorpay_subscription_id": order.get("razorpay_subscription_id"),
            "activation_key": order.get("activation_key"),
            "created_at": order.get("created_at"),
            "is_paid": order.get("is_paid", False)
        })
    
    # Calculate total revenue by method
    all_successful = await db.payment_orders.find(
        {"status": "success"},
        {"_id": 0, "amount": 1, "payment_method": 1}
    ).to_list(10000)
    
    total_amount = 0
    by_method = {
        "razorpay": 0,
        "razorpay_subscription": 0,
        "activation_key": 0
    }
    
    for order in all_successful:
        amount = order.get("amount", 0)
        method = order.get("payment_method", "razorpay")
        
        total_amount += amount
        if method in by_method:
            by_method[method] += amount
    
    # Get total count for pagination
    total_count = await db.payment_orders.count_documents(query)
    
    return {
        "total_amount": round(total_amount, 2),
        "by_method": {k: round(v, 2) for k, v in by_method.items()},
        "total_count": total_count,
        "items": revenue_items
    }

@api_router.delete("/admin/platform-revenue/{record_id}")
async def delete_platform_revenue_record(
    record_id: str,
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete a single platform revenue record"""
    result = await db.payment_orders.delete_one({"id": record_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Revenue record not found")
    
    return {"message": "Revenue record deleted successfully"}

@api_router.delete("/admin/platform-revenue")
async def delete_all_platform_revenue_records(
    admin: AdminUser = Depends(get_current_admin)
):
    """Delete all platform revenue records"""
    result = await db.payment_orders.delete_many({})
    
    return {
        "message": "All revenue records deleted successfully",
        "deleted_count": result.deleted_count
    }


# ============ SITE MAINTENANCE ROUTES ============

@api_router.get("/site-maintenance")
async def get_site_maintenance(current_user: User = Depends(get_current_user)):
    """Get site maintenance status"""
    settings = await db.settings.find_one({"key": "site_maintenance"}, {"_id": 0})
    return {
        "maintenance": settings.get("value", False) if settings else False
    }

@api_router.post("/admin/site-maintenance")
async def set_site_maintenance(
    maintenance_data: dict,
    admin: AdminUser = Depends(get_current_admin)
):
    """Set site maintenance mode (admin only)"""
    maintenance = maintenance_data.get("maintenance", False)
    
    await db.settings.update_one(
        {"key": "site_maintenance"},
        {
            "$set": {
                "key": "site_maintenance",
                "value": maintenance,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": admin.id
            }
        },
        upsert=True
    )
    
    return {
        "message": "Site maintenance updated successfully",
        "maintenance": maintenance
    }


# ============ AUTOMATED DATA DELETION TASKS ============

async def archive_and_delete_user_data(contractor_id: str, deletion_type: str, deleted_by: str, reason: Optional[str] = None, requested_at: Optional[datetime] = None):
    """
    Archive user summary and delete all user data from database.
    This is called by scheduled tasks and admin approval.
    """
    try:
        # Fetch user data
        user = await db.users.find_one({"id": contractor_id})
        if not user:
            logging.warning(f"User {contractor_id} not found for deletion")
            return False
        
        # Count data before deletion
        workers_count = await db.workers.count_documents({"contractor_id": contractor_id})
        employers_count = await db.employers.count_documents({"contractor_id": contractor_id})
        attendance_count = await db.attendance.count_documents({"contractor_id": contractor_id})
        payments_count = await db.payments.count_documents({"contractor_id": contractor_id})
        
        # Create archive record
        archive = DeletedUserArchive(
            contractor_id=contractor_id,
            contractor_name=user.get("name", "Unknown"),
            contractor_email=user.get("email", "Unknown"),
            phone=user.get("phone"),
            subscription_plan=user.get("subscription_plan"),
            subscription_status=user.get("subscription_status"),
            deletion_reason=reason,
            deletion_type=deletion_type,
            requested_at=requested_at,
            deleted_by=deleted_by,
            total_workers=workers_count,
            total_employers=employers_count,
            total_attendance_records=attendance_count,
            total_payments=payments_count,
            account_created_at=user.get("created_at") if isinstance(user.get("created_at"), datetime) else None,
            last_login_at=user.get("last_login") if isinstance(user.get("last_login"), datetime) else None
        )
        
        # Save archive
        archive_dict = archive.model_dump()
        archive_dict['deleted_at'] = archive_dict['deleted_at'].isoformat()
        if archive_dict.get('requested_at'):
            archive_dict['requested_at'] = archive_dict['requested_at'].isoformat()
        if archive_dict.get('account_created_at'):
            archive_dict['account_created_at'] = archive_dict['account_created_at'].isoformat()
        if archive_dict.get('last_login_at'):
            archive_dict['last_login_at'] = archive_dict['last_login_at'].isoformat()
        
        await db.deleted_users_archive.insert_one(archive_dict)
        
        # Delete all user data
        await db.workers.delete_many({"contractor_id": contractor_id})
        await db.employers.delete_many({"contractor_id": contractor_id})
        await db.attendance.delete_many({"contractor_id": contractor_id})
        await db.bookings.delete_many({"contractor_id": contractor_id})
        await db.rooms.delete_many({"contractor_id": contractor_id})
        await db.payments.delete_many({"contractor_id": contractor_id})
        await db.advances.delete_many({"contractor_id": contractor_id})
        await db.extra_charges.delete_many({"contractor_id": contractor_id})
        await db.notifications.delete_many({"user_id": contractor_id})
        await db.messages.delete_many({"user_id": contractor_id})
        await db.payment_orders.delete_many({"contractor_id": contractor_id})
        await db.razorpay_subscriptions.delete_many({"contractor_id": contractor_id})
        await db.users.delete_one({"id": contractor_id})
        
        logging.info(f"User {contractor_id} data archived and deleted. Type: {deletion_type}")
        return True
        
    except Exception as e:
        logging.error(f"Error archiving and deleting user {contractor_id}: {e}")
        import traceback
        traceback.print_exc()
        return False

async def process_scheduled_deletions():
    """
    Background task: Process approved deletion requests that are due (30 days after approval).
    Runs daily.
    """
    try:
        now = datetime.now(timezone.utc)
        
        # Find approved deletion requests where scheduled_deletion_date has passed
        due_deletions = await db.deletion_requests.find({
            "status": "approved",
            "scheduled_deletion_date": {"$lte": now.isoformat()}
        }).to_list(100)
        
        for deletion_req in due_deletions:
            contractor_id = deletion_req["contractor_id"]
            
            # Archive and delete user data
            success = await archive_and_delete_user_data(
                contractor_id=contractor_id,
                deletion_type="user_requested",
                deleted_by=deletion_req.get("processed_by", "system"),
                reason=deletion_req.get("reason"),
                requested_at=deletion_req.get("requested_at")
            )
            
            if success:
                # Update deletion request status
                await db.deletion_requests.update_one(
                    {"id": deletion_req["id"]},
                    {"$set": {
                        "status": "auto_deleted",
                        "deleted_at": now.isoformat()
                    }}
                )
                logging.info(f"Scheduled deletion completed for user {contractor_id}")
        
        if due_deletions:
            logging.info(f"Processed {len(due_deletions)} scheduled deletions")
            
    except Exception as e:
        logging.error(f"Error processing scheduled deletions: {e}")
        import traceback
        traceback.print_exc()

async def cleanup_inactive_free_users():
    """
    Background task: Delete free plan users who haven't logged in for 6 months.
    Runs weekly.
    """
    try:
        now = datetime.now(timezone.utc)
        six_months_ago = now - timedelta(days=180)
        
        # Find free plan users inactive for 6 months
        inactive_users = await db.users.find({
            "$or": [
                {"subscription_plan": {"$in": ["free", "Free", None]}},
                {"subscription_status": "inactive"}
            ],
            "$or": [
                {"last_login": {"$lte": six_months_ago.isoformat()}},
                {"last_login": {"$exists": False}}
            ],
            "created_at": {"$lte": six_months_ago.isoformat()}  # Account must be at least 6 months old
        }).to_list(1000)
        
        deleted_count = 0
        for user in inactive_users:
            # Skip if user has active paid subscription
            if user.get("subscription_status") == "active" and user.get("subscription_plan") not in ["free", "Free", None]:
                continue
            
            # Archive and delete
            success = await archive_and_delete_user_data(
                contractor_id=user["id"],
                deletion_type="auto_inactive",
                deleted_by="system",
                reason="Inactive for 6 months (free plan)"
            )
            
            if success:
                deleted_count += 1
        
        if deleted_count > 0:
            logging.info(f"Cleaned up {deleted_count} inactive free users")
            
    except Exception as e:
        logging.error(f"Error cleaning up inactive users: {e}")
        import traceback
        traceback.print_exc()

# ============ ADMIN ENDPOINTS FOR DELETED USERS ============

@api_router.get("/admin/deleted-users")
async def get_deleted_users(
    admin: AdminUser = Depends(get_current_admin),
    deletion_type: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get archived deleted users for admin reference"""
    try:
        query = {}
        if deletion_type:
            query["deletion_type"] = deletion_type
        
        deleted_users = await db.deleted_users_archive.find(
            query,
            {"_id": 0}
        ).sort("deleted_at", -1).skip(skip).limit(limit).to_list(limit)
        
        total = await db.deleted_users_archive.count_documents(query)
        
        return {
            "deleted_users": deleted_users,
            "total": total,
            "skip": skip,
            "limit": limit
        }
    except Exception as e:
        logging.error(f"Error fetching deleted users: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch deleted users: {str(e)}")

@api_router.get("/admin/deleted-users/stats")
async def get_deleted_users_stats(admin: AdminUser = Depends(get_current_admin)):
    """Get statistics about deleted users"""
    try:
        total_deleted = await db.deleted_users_archive.count_documents({})
        user_requested = await db.deleted_users_archive.count_documents({"deletion_type": "user_requested"})
        admin_approved = await db.deleted_users_archive.count_documents({"deletion_type": "admin_approved"})
        auto_inactive = await db.deleted_users_archive.count_documents({"deletion_type": "auto_inactive"})
        
        # Get recent deletions (last 30 days)
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        recent_deletions = await db.deleted_users_archive.count_documents({
            "deleted_at": {"$gte": thirty_days_ago.isoformat()}
        })
        
        return {
            "total_deleted": total_deleted,
            "by_type": {
                "user_requested": user_requested,
                "admin_approved": admin_approved,
                "auto_inactive": auto_inactive
            },
            "recent_deletions_30_days": recent_deletions
        }
    except Exception as e:
        logging.error(f"Error fetching deleted users stats: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch stats: {str(e)}")

# Schedule background tasks
@app.on_event("startup")
async def schedule_background_tasks():
    """Schedule automated cleanup tasks"""
    async def run_daily_tasks():
        while True:
            await asyncio.sleep(86400)  # Run every 24 hours
            await process_scheduled_deletions()
    
    async def run_weekly_tasks():
        while True:
            await asyncio.sleep(604800)  # Run every 7 days
            await cleanup_inactive_free_users()
    
    # Start background tasks
    asyncio.create_task(run_daily_tasks())
    asyncio.create_task(run_weekly_tasks())
    logging.info("Background cleanup tasks scheduled")


# ============================================================
# WORK HISTORY REPORTS (auditable per-day work records)
# ============================================================

def _parse_dd_mm_yyyy(s):
    """Parse a DD-MM-YYYY date string, return datetime or None."""
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d-%m-%Y").replace(tzinfo=timezone.utc)
    except (ValueError, TypeError):
        try:
            return datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=timezone.utc)
        except (ValueError, TypeError):
            return None


def _date_in_range(date_str, start, end):
    """Inclusive range check on DD-MM-YYYY strings."""
    d = _parse_dd_mm_yyyy(date_str)
    if d is None:
        return False
    if start and d < start:
        return False
    if end and d > end:
        return False
    return True


async def _build_work_history(
    contractor_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    worker_id: Optional[str] = None,
    employer_id: Optional[str] = None,
):
    """
    Build a unified work history dataset for a contractor.

    Returns a dict with:
      - filters: echoes the filters
      - summary: aggregated stats
      - records: per-(worker, date) detailed rows
      - day_of_week_breakdown: counts by weekday
      - top_workers: leaderboard by days worked / wages earned
      - top_employers: leaderboard by amount paid / workers received
    """
    start_dt = _parse_dd_mm_yyyy(start_date)
    end_dt = _parse_dd_mm_yyyy(end_date)

    # ---- Pull attendance from both collections, then dedupe ----
    base_q = {"contractor_id": contractor_id}
    if worker_id:
        base_q["worker_id"] = worker_id
    if employer_id:
        base_q["employer_id"] = employer_id

    new_recs = await db.worker_attendance.find(base_q, {"_id": 0}).to_list(50000)
    old_q = dict(base_q)
    # mode=employer aggregates contain selected_workers but no worker_id key, so when filtering
    # by worker_id we can simply match through the worker docs (mode=worker / no mode).
    legacy_recs = await db.attendance.find(old_q, {"_id": 0}).to_list(50000)

    merged = []
    for r in new_recs:
        merged.append({
            "worker_id": r.get("worker_id"),
            "employer_id": r.get("employer_id", "") or "",
            "date": r.get("date"),
            "status": r.get("status", "Present"),
            "wage_earned": float(r.get("wage_earned") or r.get("wage_amount") or 0),
        })

    for r in legacy_recs:
        # Skip employer-mode summary rows; the per-worker rows below already cover them.
        mode = (r.get("mode") or "").lower()
        if mode == "employer":
            continue
        if not r.get("worker_id"):
            continue
        merged.append({
            "worker_id": r.get("worker_id"),
            "employer_id": r.get("employer_id", "") or "",
            "date": r.get("date"),
            "status": r.get("status", "Present"),
            "wage_earned": float(r.get("wage_amount") or r.get("wage_earned") or 0),
        })

    # Dedup by (worker_id, date), prefer non-empty employer_id and explicit Absent/Late status
    by_key = {}
    for r in merged:
        if not r.get("worker_id") or not r.get("date"):
            continue
        if not _date_in_range(r["date"], start_dt, end_dt):
            continue
        key = (r["worker_id"], r["date"])
        existing = by_key.get(key)
        if not existing:
            by_key[key] = r
            continue
        # Merge: prefer record with employer_id set, then prefer explicit Absent/Late
        ex_emp = existing.get("employer_id") or ""
        nw_emp = r.get("employer_id") or ""
        if not ex_emp and nw_emp:
            existing["employer_id"] = nw_emp
        ex_status = (existing.get("status") or "Present").strip().lower()
        nw_status = (r.get("status") or "Present").strip().lower()
        if ex_status == "present" and nw_status in ("absent", "late"):
            existing["status"] = r.get("status")
        # Take max of wage_earned (avoid losing wage info in merge)
        existing["wage_earned"] = max(
            float(existing.get("wage_earned") or 0),
            float(r.get("wage_earned") or 0),
        )

    deduped = list(by_key.values())

    # ---- Pull commissions for the same window/filters ----
    comm_q = {"contractor_id": contractor_id}
    if worker_id:
        comm_q["worker_id"] = worker_id
    if employer_id:
        comm_q["employer_id"] = employer_id
    commissions = await db.commissions.find(comm_q, {"_id": 0}).to_list(50000)
    comm_index = {}
    for c in commissions:
        if not _date_in_range(c.get("date"), start_dt, end_dt):
            continue
        key = (c.get("worker_id"), c.get("date"))
        comm_index[key] = c

    # ---- Pull worker / employer profile lookups ----
    worker_ids = list({r["worker_id"] for r in deduped if r.get("worker_id")})
    employer_ids = list({r["employer_id"] for r in deduped if r.get("employer_id")})
    workers_map = {}
    if worker_ids:
        async for w in db.workers.find({"id": {"$in": worker_ids}}, {"_id": 0}):
            workers_map[w["id"]] = w
    employers_map = {}
    if employer_ids:
        async for e in db.employers.find({"id": {"$in": employer_ids}}, {"_id": 0}):
            employers_map[e["id"]] = e

    # ---- Build detail records & aggregations ----
    records = []
    total_wages = 0.0
    total_collected = 0.0
    total_commission = 0.0
    present_count = 0
    absent_count = 0
    late_count = 0
    dow_breakdown = {"Monday": 0, "Tuesday": 0, "Wednesday": 0,
                     "Thursday": 0, "Friday": 0, "Saturday": 0, "Sunday": 0}
    worker_agg = {}
    employer_agg = {}

    for r in deduped:
        comm = comm_index.get((r["worker_id"], r["date"]), {})
        worker = workers_map.get(r["worker_id"], {})
        employer = employers_map.get(r["employer_id"], {}) if r.get("employer_id") else {}

        wage_earned = float(r.get("wage_earned") or comm.get("wage_to_worker") or 0)
        amount_from_employer = float(comm.get("payment_from_employer") or 0)
        commission_amount = float(comm.get("commission_amount") or 0)
        status = r.get("status") or "Present"

        d = _parse_dd_mm_yyyy(r["date"])
        dow = d.strftime("%A") if d else ""

        rec = {
            "date": r["date"],
            "iso_date": d.strftime("%Y-%m-%d") if d else r["date"],
            "day_of_week": dow,
            "worker_id": r["worker_id"],
            "worker_name": worker.get("name", "Unknown Worker"),
            "worker_phone": worker.get("phone_number", ""),
            "employer_id": r.get("employer_id") or "",
            "employer_name": employer.get("name", "Unassigned") if employer else (
                "" if not r.get("employer_id") else "Unassigned"),
            "status": status,
            "wage_earned": round(wage_earned, 2),
            "amount_from_employer": round(amount_from_employer, 2),
            "commission_amount": round(commission_amount, 2),
        }
        records.append(rec)

        s_lower = status.strip().lower()
        if s_lower == "present":
            present_count += 1
            if dow in dow_breakdown:
                dow_breakdown[dow] += 1
            total_wages += wage_earned
            total_collected += amount_from_employer
            total_commission += commission_amount

            # Worker aggregation
            wagg = worker_agg.setdefault(r["worker_id"], {
                "worker_id": r["worker_id"],
                "worker_name": worker.get("name", "Unknown Worker"),
                "days_worked": 0, "total_wages": 0.0, "total_commission": 0.0,
            })
            wagg["days_worked"] += 1
            wagg["total_wages"] += wage_earned
            wagg["total_commission"] += commission_amount

            # Employer aggregation
            if r.get("employer_id"):
                eagg = employer_agg.setdefault(r["employer_id"], {
                    "employer_id": r["employer_id"],
                    "employer_name": employer.get("name", "Unassigned"),
                    "worker_days": 0, "total_amount_paid": 0.0,
                    "total_commission": 0.0, "unique_workers": set(),
                })
                eagg["worker_days"] += 1
                eagg["total_amount_paid"] += amount_from_employer
                eagg["total_commission"] += commission_amount
                eagg["unique_workers"].add(r["worker_id"])
        elif s_lower == "absent":
            absent_count += 1
        elif s_lower == "late":
            late_count += 1

    # Sort detail records — most recent date first, then by worker name
    records.sort(key=lambda x: (x.get("iso_date", ""), x.get("worker_name", "")), reverse=True)

    # Convert worker / employer aggs to leaderboards
    top_workers = sorted(
        [{**v, "total_wages": round(v["total_wages"], 2),
          "total_commission": round(v["total_commission"], 2)}
         for v in worker_agg.values()],
        key=lambda x: (-x["days_worked"], -x["total_wages"]),
    )
    top_employers = []
    for v in employer_agg.values():
        top_employers.append({
            "employer_id": v["employer_id"],
            "employer_name": v["employer_name"],
            "worker_days": v["worker_days"],
            "unique_workers": len(v["unique_workers"]),
            "total_amount_paid": round(v["total_amount_paid"], 2),
            "total_commission": round(v["total_commission"], 2),
        })
    top_employers.sort(key=lambda x: (-x["worker_days"], -x["total_amount_paid"]))

    unique_dates = len({r["date"] for r in records})

    summary = {
        "total_records": len(records),
        "present_count": present_count,
        "absent_count": absent_count,
        "late_count": late_count,
        "unique_dates": unique_dates,
        "unique_workers": len(worker_agg),
        "unique_employers": len(employer_agg),
        "total_wages_earned": round(total_wages, 2),
        "total_amount_collected": round(total_collected, 2),
        "total_commission": round(total_commission, 2),
        "attendance_rate": round((present_count / len(records) * 100), 2) if records else 0.0,
    }

    return {
        "filters": {
            "start_date": start_date,
            "end_date": end_date,
            "worker_id": worker_id,
            "employer_id": employer_id,
        },
        "summary": summary,
        "records": records,
        "day_of_week_breakdown": dow_breakdown,
        "top_workers": top_workers[:50],
        "top_employers": top_employers[:50],
    }


@api_router.get("/reports/work-history")
async def get_work_history(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    worker_id: Optional[str] = None,
    employer_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    """Audit-grade work history. Filters: start_date, end_date (DD-MM-YYYY), worker_id, employer_id."""
    return await _build_work_history(current_user.id, start_date, end_date, worker_id, employer_id)


def _safe_filename_part(s: str) -> str:
    """Make a string safe to use inside a filename."""
    if not s:
        return ""
    keep = [c if (c.isalnum() or c in "-_") else "_" for c in s.strip()]
    return "".join(keep)[:40].strip("_") or ""


@api_router.get("/reports/work-history/export")
async def export_work_history(
    format: str = "csv",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    worker_id: Optional[str] = None,
    employer_id: Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    """Export work history as CSV, Excel, or PDF."""
    fmt = (format or "csv").lower()
    if fmt not in ("csv", "excel", "xlsx", "pdf"):
        raise HTTPException(status_code=400, detail="Invalid format. Use csv, excel, or pdf.")

    data = await _build_work_history(current_user.id, start_date, end_date, worker_id, employer_id)
    records = data["records"]
    summary = data["summary"]

    # Resolve worker / employer names for richer branding context
    worker_doc = None
    employer_doc = None
    if worker_id:
        worker_doc = await db.workers.find_one(
            {"id": worker_id, "contractor_id": current_user.id}, {"_id": 0}
        )
    if employer_id:
        employer_doc = await db.employers.find_one(
            {"id": employer_id, "contractor_id": current_user.id}, {"_id": 0}
        )

    # IST timestamp for "Generated on"
    ist_now = datetime.now(timezone.utc) + timedelta(hours=5, minutes=30)
    generated_str = ist_now.strftime("%d %b %Y, %I:%M %p IST")

    period_str = f"{start_date or 'All time'} to {end_date or 'today'}"
    contractor_name = current_user.name or "Contractor"
    contractor_email = current_user.email or ""
    contractor_phone = current_user.phone or ""
    plan_name = (current_user.subscription_plan or "Free").title()

    scope_label = "All workers & employers"
    if worker_doc and employer_doc:
        scope_label = f"Worker: {worker_doc.get('name', '')}  •  Employer: {employer_doc.get('name', '')}"
    elif worker_doc:
        scope_label = f"Worker: {worker_doc.get('name', '')} ({worker_doc.get('phone_number', '')})"
    elif employer_doc:
        scope_label = f"Employer: {employer_doc.get('name', '')} ({employer_doc.get('phone_number', '')})"

    # Build a meaningful filename
    filter_parts = []
    if worker_doc:
        filter_parts.append(_safe_filename_part(worker_doc.get("name", "worker")))
    elif worker_id:
        filter_parts.append("worker")
    if employer_doc:
        filter_parts.append(_safe_filename_part(employer_doc.get("name", "employer")))
    elif employer_id:
        filter_parts.append("employer")
    if start_date or end_date:
        filter_parts.append(f"{start_date or 'all'}_to_{end_date or 'all'}")
    filename_base = "GuestWorker_WorkHistory" + ("_" + "_".join(filter_parts) if filter_parts else "")

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        # Branded header block
        writer.writerow(["GuestWorker  —  Work History Report"])
        writer.writerow(["https://guestworker.in"])
        writer.writerow([])
        writer.writerow(["Contractor", contractor_name])
        if contractor_email:
            writer.writerow(["Email", contractor_email])
        if contractor_phone:
            writer.writerow(["Phone", contractor_phone])
        writer.writerow(["Plan", plan_name])
        writer.writerow(["Period", period_str])
        writer.writerow(["Scope", scope_label])
        writer.writerow(["Generated", generated_str])
        writer.writerow([])
        writer.writerow(["SUMMARY"])
        for k, v in summary.items():
            writer.writerow([k.replace("_", " ").title(), v])
        writer.writerow([])
        writer.writerow(["DETAILED RECORDS"])
        writer.writerow([
            "Date", "Day", "Worker", "Worker Phone", "Employer", "Status",
            "Wage Earned (INR)", "Amount From Employer (INR)", "Commission (INR)",
        ])
        for r in records:
            writer.writerow([
                r["date"], r["day_of_week"], r["worker_name"], r["worker_phone"],
                r["employer_name"], r["status"],
                r["wage_earned"], r["amount_from_employer"], r["commission_amount"],
            ])
        writer.writerow([])
        writer.writerow([f"Generated by GuestWorker • {generated_str} • Confidential"])
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8-sig")),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    if fmt in ("excel", "xlsx"):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.properties.creator = "GuestWorker"
        wb.properties.title = f"Work History — {contractor_name}"
        wb.properties.subject = period_str
        wb.properties.description = "Work history report exported from GuestWorker (https://guestworker.in)"
        ws = wb.active
        ws.title = "Work History"

        BRAND = "3B2ED0"           # Indigo
        BRAND_DARK = "2A1FB8"
        ACCENT = "4F46E5"
        SOFT = "EEF2FF"
        TEXT_GREY = "475569"

        title_font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
        sub_font = Font(name="Calibri", bold=False, size=10, color="E0E7FF")
        section_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        label_font = Font(name="Calibri", bold=True, size=10, color=TEXT_GREY)
        value_font = Font(name="Calibri", bold=False, size=10, color="0F172A")
        footer_font = Font(name="Calibri", italic=True, size=9, color="64748B")

        brand_fill = PatternFill(start_color=BRAND, end_color=BRAND_DARK, fill_type="solid")
        section_fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
        header_fill = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
        soft_fill = PatternFill(start_color=SOFT, end_color=SOFT, fill_type="solid")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")

        thin = Side(style="thin", color="E2E8F0")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        NUM_COLS = 9
        last_col = get_column_letter(NUM_COLS)

        # ---- Brand banner ----
        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = "GuestWorker"
        ws["A1"].font = title_font
        ws["A1"].fill = brand_fill
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 32

        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = "Work History Report  •  guestworker.in"
        ws["A2"].font = sub_font
        ws["A2"].fill = brand_fill
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 18

        # ---- Contractor info block ----
        info_rows = [
            ("Contractor", contractor_name),
            ("Email", contractor_email),
            ("Phone", contractor_phone or "—"),
            ("Plan", plan_name),
            ("Period", period_str),
            ("Scope", scope_label),
            ("Generated", generated_str),
        ]
        row = 4
        for label, value in info_rows:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=1).alignment = left
            ws.cell(row=row, column=1).fill = soft_fill
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NUM_COLS)
            vc = ws.cell(row=row, column=2, value=value)
            vc.font = value_font
            vc.alignment = left
            row += 1

        # ---- SUMMARY section banner ----
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        sc = ws.cell(row=row, column=1, value="SUMMARY")
        sc.font = section_font
        sc.fill = section_fill
        sc.alignment = center
        ws.row_dimensions[row].height = 22
        row += 1
        for k, v in summary.items():
            lc = ws.cell(row=row, column=1, value=k.replace("_", " ").title())
            lc.font = label_font
            lc.alignment = left
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NUM_COLS)
            vc = ws.cell(row=row, column=2, value=v)
            vc.font = value_font
            vc.alignment = left
            row += 1

        # ---- DETAILED RECORDS banner ----
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        dc = ws.cell(row=row, column=1, value="DETAILED RECORDS")
        dc.font = section_font
        dc.fill = section_fill
        dc.alignment = center
        ws.row_dimensions[row].height = 22
        row += 1

        headers = ["Date", "Day", "Worker", "Worker Phone", "Employer", "Status",
                   "Wage (INR)", "From Employer (INR)", "Commission (INR)"]
        header_row = row
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = cell_border
        ws.row_dimensions[row].height = 24
        row += 1

        first_data_row = row
        for r in records:
            cells = [
                (1, r["date"], left),
                (2, r["day_of_week"], left),
                (3, r["worker_name"], left),
                (4, r["worker_phone"], left),
                (5, r["employer_name"], left),
                (6, r["status"], center),
                (7, float(r["wage_earned"] or 0), right),
                (8, float(r["amount_from_employer"] or 0), right),
                (9, float(r["commission_amount"] or 0), right),
            ]
            for col, val, align in cells:
                c = ws.cell(row=row, column=col, value=val)
                c.font = value_font
                c.alignment = align
                c.border = cell_border
                if col in (7, 8, 9):
                    c.number_format = '"₹"#,##0'
            row += 1

        # Totals row
        if records:
            total_wages = sum(float(r["wage_earned"] or 0) for r in records)
            total_from_emp = sum(float(r["amount_from_employer"] or 0) for r in records)
            total_comm = sum(float(r["commission_amount"] or 0) for r in records)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            tlabel = ws.cell(row=row, column=1, value="TOTAL")
            tlabel.font = Font(bold=True, color="FFFFFF")
            tlabel.fill = brand_fill
            tlabel.alignment = right
            for col, val in [(7, total_wages), (8, total_from_emp), (9, total_comm)]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = brand_fill
                c.alignment = right
                c.number_format = '"₹"#,##0'
                c.border = cell_border
            ws.row_dimensions[row].height = 22
            row += 1

        # Footer
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        fc = ws.cell(
            row=row, column=1,
            value=f"Generated by GuestWorker on {generated_str}  •  guestworker.in  •  Confidential",
        )
        fc.font = footer_font
        fc.alignment = center

        # Freeze pane on header row of detailed records
        ws.freeze_panes = ws.cell(row=first_data_row, column=1)

        # Column widths (curated)
        widths = {1: 14, 2: 12, 3: 26, 4: 16, 5: 26, 6: 12, 7: 16, 8: 22, 9: 18}
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Print setup
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = f"{header_row}:{header_row}"

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
        )

    # PDF
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether,
    )

    BRAND = colors.HexColor("#3B2ED0")
    BRAND_DARK = colors.HexColor("#2A1FB8")
    ACCENT = colors.HexColor("#4F46E5")
    SOFT = colors.HexColor("#EEF2FF")
    TEXT = colors.HexColor("#0F172A")
    MUTED = colors.HexColor("#64748B")
    BORDER = colors.HexColor("#E2E8F0")
    ROW_ALT = colors.HexColor("#F8FAFC")

    out = io.BytesIO()
    page_w, page_h = landscape(A4)
    doc = SimpleDocTemplate(
        out, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=28 * mm, bottomMargin=18 * mm,
        title=f"GuestWorker — Work History ({contractor_name})",
        author="GuestWorker",
        subject=period_str,
    )
    styles = getSampleStyleSheet()
    h2 = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12,
        textColor=BRAND_DARK, spaceAfter=6, spaceBefore=4,
    )
    body = ParagraphStyle(
        "BodyClean", parent=styles["BodyText"], fontName="Helvetica",
        fontSize=9.5, textColor=TEXT, leading=13,
    )
    label_style = ParagraphStyle(
        "Label", parent=body, fontName="Helvetica-Bold", textColor=MUTED, fontSize=8.5,
    )

    def _draw_chrome(canvas, _doc):
        canvas.saveState()
        # Top brand bar
        canvas.setFillColor(BRAND)
        canvas.rect(0, page_h - 22 * mm, page_w, 22 * mm, fill=1, stroke=0)
        # Logo mark
        canvas.setFillColor(colors.white)
        canvas.roundRect(12 * mm, page_h - 17 * mm, 11 * mm, 11 * mm, 2 * mm, fill=1, stroke=0)
        canvas.setFillColor(BRAND_DARK)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawCentredString(17.5 * mm, page_h - 12.6 * mm, "GW")
        # Wordmark
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 18)
        canvas.drawString(28 * mm, page_h - 12 * mm, "GuestWorker")
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.HexColor("#E0E7FF"))
        canvas.drawString(28 * mm, page_h - 17 * mm, "Work History Report  •  guestworker.in")
        # Right side: contractor + period
        canvas.setFont("Helvetica-Bold", 10)
        canvas.setFillColor(colors.white)
        canvas.drawRightString(page_w - 12 * mm, page_h - 12 * mm, contractor_name[:60])
        canvas.setFont("Helvetica", 8.5)
        canvas.setFillColor(colors.HexColor("#E0E7FF"))
        canvas.drawRightString(page_w - 12 * mm, page_h - 17 * mm, f"{period_str}")

        # Footer line + page number + branding
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.4)
        canvas.line(12 * mm, 14 * mm, page_w - 12 * mm, 14 * mm)
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(MUTED)
        canvas.drawString(12 * mm, 9 * mm, f"Generated by GuestWorker  •  {generated_str}")
        canvas.drawCentredString(page_w / 2, 9 * mm, "guestworker.in  •  Confidential")
        canvas.drawRightString(page_w - 12 * mm, 9 * mm, f"Page {_doc.page}")
        canvas.restoreState()

    elements = []

    # ---- Meta block (contractor + scope) ----
    meta_data = [
        [Paragraph("Contractor", label_style), Paragraph(contractor_name, body),
         Paragraph("Plan", label_style), Paragraph(plan_name, body)],
        [Paragraph("Email", label_style), Paragraph(contractor_email or "—", body),
         Paragraph("Phone", label_style), Paragraph(contractor_phone or "—", body)],
        [Paragraph("Period", label_style), Paragraph(period_str, body),
         Paragraph("Scope", label_style), Paragraph(scope_label, body)],
    ]
    meta_table = Table(meta_data, colWidths=[24 * mm, 95 * mm, 24 * mm, 130 * mm])
    meta_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), SOFT),
        ("BOX", (0, 0), (-1, -1), 0.5, BORDER),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 10))

    # ---- KPI cards row ----
    def _kpi_card(label, value, accent_hex="#3B2ED0", accent_color=None):
        if accent_color is None:
            accent_color = colors.HexColor(accent_hex)
        inner = Table(
            [[Paragraph(f'<font color="#64748B" size=8>{label}</font>', body)],
             [Paragraph(f'<font color="{accent_hex}" size=14><b>{value}</b></font>', body)]],
            colWidths=[58 * mm],
        )
        inner.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("BOX", (0, 0), (-1, -1), 0.6, BORDER),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LINEABOVE", (0, 0), (-1, 0), 2, accent_color),
        ]))
        return inner

    kpi_row = [
        _kpi_card("Work Days", str(summary.get("total_records", 0)), "#3B2ED0", BRAND),
        _kpi_card("Workers", str(summary.get("unique_workers", 0)), "#10B981"),
        _kpi_card("Employers", str(summary.get("unique_employers", 0)), "#8B5CF6"),
        _kpi_card("Wages Paid", f"₹{summary.get('total_wages_earned', 0):,.0f}", "#F59E0B"),
        _kpi_card("Commission", f"₹{summary.get('total_commission', 0):,.0f}", "#EC4899"),
    ]
    kpi_table = Table([kpi_row], colWidths=[(page_w - 24 * mm) / 5] * 5)
    kpi_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 12))

    # ---- Summary breakdown ----
    elements.append(Paragraph("Summary", h2))
    summary_rows = [["Metric", "Value"]] + [
        [k.replace("_", " ").title(), str(v)] for k, v in summary.items()
    ]
    sum_table = Table(summary_rows, colWidths=[90 * mm, 60 * mm])
    sum_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.25, BORDER),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [ROW_ALT, colors.white]),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(KeepTogether(sum_table))
    elements.append(Spacer(1, 14))

    # ---- Detailed records ----
    elements.append(Paragraph("Detailed Records", h2))
    if not records:
        elements.append(Paragraph("<i>No records for the selected filters.</i>", body))
    else:
        det_rows = [["Date", "Day", "Worker", "Employer", "Status", "Wage", "From Emp.", "Comm."]]
        for r in records:
            det_rows.append([
                r["date"], (r["day_of_week"] or "")[:3],
                (r["worker_name"] or "")[:24],
                (r["employer_name"] or "—")[:24],
                r["status"],
                f"₹{(r['wage_earned'] or 0):,.0f}",
                f"₹{(r['amount_from_employer'] or 0):,.0f}",
                f"₹{(r['commission_amount'] or 0):,.0f}",
            ])
        # Totals row
        total_wages = sum(float(r["wage_earned"] or 0) for r in records)
        total_from_emp = sum(float(r["amount_from_employer"] or 0) for r in records)
        total_comm = sum(float(r["commission_amount"] or 0) for r in records)
        det_rows.append([
            "TOTAL", "", "", "", "",
            f"₹{total_wages:,.0f}",
            f"₹{total_from_emp:,.0f}",
            f"₹{total_comm:,.0f}",
        ])

        det_table = Table(
            det_rows,
            colWidths=[24 * mm, 14 * mm, 56 * mm, 56 * mm, 20 * mm, 24 * mm, 30 * mm, 24 * mm],
            repeatRows=1,
        )
        det_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), BRAND),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ALIGN", (5, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (4, 1), (4, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.25, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, ROW_ALT]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            # Totals row
            ("BACKGROUND", (0, -1), (-1, -1), BRAND_DARK),
            ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (0, -1), (4, -1), "RIGHT"),
        ]))
        elements.append(det_table)

    doc.build(elements, onFirstPage=_draw_chrome, onLaterPages=_draw_chrome)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
    )


# Include router - MUST BE AFTER ALL ROUTE DEFINITIONS
app.include_router(api_router)
